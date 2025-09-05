import pandas as pd
import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext
import tkinter.simpledialog
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import numbers, Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import logging
import traceback
import os


# ===== Windows 11 Modern Theme =====
class Windows11Theme:
    """Windows 11 inspired color scheme and styling"""

    # Core Windows 11 Colors
    COLORS = {
        # Primary Theme
        'bg_primary': '#F3F3F3',  # Light gray background
        'bg_secondary': '#FFFFFF',  # Pure white
        'bg_tertiary': '#FAFAFA',  # Very light gray
        'bg_accent': '#005A9F',  # Windows blue
        'bg_card': '#FFFFFF',  # Card background

        # Accent Colors
        'accent_blue': '#0078D4',  # Windows primary blue
        'accent_light_blue': '#40E0FF',  # Light blue
        'accent_green': '#107C10',  # Windows green
        'accent_orange': '#FF8C00',  # Windows orange
        'accent_red': '#D13438',  # Windows red
        'accent_purple': '#8764B8',  # Windows purple

        # Text Colors
        'text_primary': '#323130',  # Dark gray
        'text_secondary': '#605E5C',  # Medium gray
        'text_tertiary': '#8A8886',  # Light gray
        'text_white': '#FFFFFF',  # White text

        # Border and Shadow
        'border_light': '#E1DFDD',  # Light border
        'border_medium': '#C8C6C4',  # Medium border
        'shadow': '#0000001A',  # Light shadow

        # Status Colors
        'success': '#107C10',  # Green
        'warning': '#FF8C00',  # Orange
        'error': '#D13438',  # Red
        'info': '#0078D4',  # Blue
    }

    # Typography
    FONTS = {
        'heading_large': ('Segoe UI', 20, 'bold'),
        'heading_medium': ('Segoe UI', 16, 'bold'),
        'heading_small': ('Segoe UI', 14, 'bold'),
        'body_large': ('Segoe UI', 12),
        'body_medium': ('Segoe UI', 11),
        'body_small': ('Segoe UI', 10),
        'caption': ('Segoe UI', 9),
        'monospace': ('Consolas', 10),
    }

    # Spacing
    SPACING = {
        'xs': 4,
        'sm': 8,
        'md': 16,
        'lg': 24,
        'xl': 32,
        'xxl': 48,
    }

    # Border Radius (simulated with relief)
    RADIUS = {
        'small': 4,
        'medium': 8,
        'large': 12,
    }


class ModernButton(tk.Button):
    """Windows 11 styled button"""

    def __init__(self, parent, text="", command=None, style="primary", **kwargs):
        self.theme = Windows11Theme()

        # Define button styles
        styles = {
            'primary': {
                'bg': self.theme.COLORS['accent_blue'],
                'fg': self.theme.COLORS['text_white'],
                'hover_bg': '#106EBE',
                'active_bg': '#005A9F'
            },
            'secondary': {
                'bg': self.theme.COLORS['bg_secondary'],
                'fg': self.theme.COLORS['text_primary'],
                'hover_bg': '#F8F8F8',
                'active_bg': '#F0F0F0'
            },
            'success': {
                'bg': self.theme.COLORS['accent_green'],
                'fg': self.theme.COLORS['text_white'],
                'hover_bg': '#0E6E0E',
                'active_bg': '#0D5F0D'
            },
            'warning': {
                'bg': self.theme.COLORS['accent_orange'],
                'fg': self.theme.COLORS['text_white'],
                'hover_bg': '#E67E00',
                'active_bg': '#CC7000'
            },
            'danger': {
                'bg': self.theme.COLORS['accent_red'],
                'fg': self.theme.COLORS['text_white'],
                'hover_bg': '#BC2E32',
                'active_bg': '#A7292C'
            }
        }

        config = styles.get(style, styles['primary'])

        # Default button configuration
        default_config = {
            'font': self.theme.FONTS['body_medium'],
            'bg': config['bg'],
            'fg': config['fg'],
            'relief': 'flat',
            'bd': 0,
            'padx': 20,
            'pady': 10,
            'cursor': 'hand2',
            'activebackground': config['active_bg'],
            'activeforeground': config['fg']
        }

        # Merge with user kwargs
        default_config.update(kwargs)

        super().__init__(parent, text=text, command=command, **default_config)

        # Store colors for hover effects
        self.default_bg = config['bg']
        self.hover_bg = config['hover_bg']

        # Bind hover effects
        self.bind("<Enter>", self._on_enter)
        self.bind("<Leave>", self._on_leave)

    def _on_enter(self, event):
        self.configure(bg=self.hover_bg)

    def _on_leave(self, event):
        self.configure(bg=self.default_bg)


class ModernCard(tk.Frame):
    """Windows 11 styled card container"""

    def __init__(self, parent, title="", **kwargs):
        self.theme = Windows11Theme()

        default_config = {
            'bg': self.theme.COLORS['bg_card'],
            'relief': 'solid',
            'bd': 1,
            'highlightthickness': 0,
            'highlightcolor': self.theme.COLORS['border_light'],
            'highlightbackground': self.theme.COLORS['border_light']
        }

        default_config.update(kwargs)
        super().__init__(parent, **default_config)

        if title:
            self._create_title(title)

    def _create_title(self, title):
        title_frame = tk.Frame(self, bg=self.theme.COLORS['bg_card'])
        title_frame.pack(fill='x', padx=self.theme.SPACING['md'],
                         pady=(self.theme.SPACING['md'], self.theme.SPACING['sm']))

        title_label = tk.Label(
            title_frame,
            text=title,
            font=self.theme.FONTS['heading_small'],
            fg=self.theme.COLORS['text_primary'],
            bg=self.theme.COLORS['bg_card']
        )
        title_label.pack(anchor='w')


class ModernEntry(tk.Frame):
    """Windows 11 styled entry with label"""

    def __init__(self, parent, label="", placeholder="", **kwargs):
        self.theme = Windows11Theme()
        super().__init__(parent, bg=self.theme.COLORS['bg_card'])

        if label:
            label_widget = tk.Label(
                self,
                text=label,
                font=self.theme.FONTS['body_medium'],
                fg=self.theme.COLORS['text_primary'],
                bg=self.theme.COLORS['bg_card']
            )
            label_widget.pack(anchor='w', pady=(0, 4))

        self.entry = tk.Entry(
            self,
            font=self.theme.FONTS['body_medium'],
            bg=self.theme.COLORS['bg_secondary'],
            fg=self.theme.COLORS['text_primary'],
            relief='solid',
            bd=1,
            highlightthickness=2,
            highlightcolor=self.theme.COLORS['accent_blue'],
            highlightbackground=self.theme.COLORS['border_light'],
            **kwargs
        )
        self.entry.pack(fill='x', ipady=8)

        if placeholder:
            self._add_placeholder(placeholder)

    def _add_placeholder(self, placeholder):
        self.entry.insert(0, placeholder)
        self.entry.configure(fg=self.theme.COLORS['text_tertiary'])

        def on_focus_in(event):
            if self.entry.get() == placeholder:
                self.entry.delete(0, tk.END)
                self.entry.configure(fg=self.theme.COLORS['text_primary'])

        def on_focus_out(event):
            if not self.entry.get():
                self.entry.insert(0, placeholder)
                self.entry.configure(fg=self.theme.COLORS['text_tertiary'])

        self.entry.bind('<FocusIn>', on_focus_in)
        self.entry.bind('<FocusOut>', on_focus_out)

    def get(self):
        return self.entry.get()

    def set(self, value):
        self.entry.delete(0, tk.END)
        self.entry.insert(0, value)


# ===== Enhanced Configuration with Built-in Data =====
class ConfigManager:
    def __init__(self):
        self.config_file = "mapcost_config.ini"
        self.base_files = {
            'hashira': None,
            'hamada': None
        }
        self.initialize_default_data()
        self.load_config()

    def initialize_default_data(self):
        """สร้างข้อมูล Hamada Cost และ Hashira Cost จากข้อมูลจริง"""

        # ข้อมูลจริงจาก Hamada Cost.xlsx
        hamada_menu_names = [
            "Ebi Tempura Set", "Salmon Tempura Set", "Ebi & Salmon Tempura Set",
            "Ebi Tempura Don Set", "Salmon Tempura Don Set", "Ebi & Salmon Tempura Don Set",
            "Ebi Tempura Udon", "Salmon Tempura Udon", "Ebi & Salmon Tempura Udon",
            "Yakiudon Buta", "Yakiudon Kaisen", "Ebi Fry Curry Rice",
            "Tonkatsu Curry Rice", "Salmonkatsu Curry Rice", "Ebi Tempura（1pc）",
            "Salmon Tempura（1pc）", "Takoyaki", "Sushi Yorokobi",
            "Sushi Tokujo", "Maguro Zanmai", "Salmon Zanmai",
            "Hamachi Zanmai", "Duo Sushi", "Salmon Mountain",
            "Salmon Sushi & Ikura Gunkan", "California Roll", "Salmon Roll",
            "Unagi Roll", "Hokkaido Hotate Roll", "Salmon Chirashi Sushi",
            "Salmon Tobiko Chirashi Sushi", "Barachirashi Sushi", "Salmon Don",
            "Salmon Ikura Don", "Maguro Don", "Negitoro Ikura Don",
            "Kaisen Don", "Unagi Don", "Salmon Sashimi",
            "Maguro Sashimi", "Hamachi Sashimi", "Hotate Sashimi",
            "Salmon Sushi (2pcs)", "Maguro Sushi (2pcs)", "Hamachi Sushi (2pcs)",
            "Tamago Sushi (2pcs)", "Ebi Sushi (2pcs)", "Ika Sushi (2pcs)",
            "Hokki Sushi (2pcs)", "Hotate Sushi (2pcs)", "Unagi Sushi (2pcs)",
            "Anago Sushi (2pcs)", "Negitoro Gunkan (2pcs)", "Tobiko Gunkan (2pcs)",
            "Ikura Gunkan (2pcs)", "GG California Roll", "GG Salmon Sushi",
            "GG Tamago & Ebi Sushi", "GG Ebi & Tobiko Don", "GG Kanikama & Tobiko Don",
            "GG Triple Mix Don", "GG Ebiten Don", "GG Salmon & Ebiten Don",
            "Coca-Cola Original", "Coke-Cola Zero Sugar", "Sprite",
            "Soda", "Mineral Water", "HBD Sparkling Water Lemon (No Sugar, No Calories)",
            "HBD Sparkling Water Honey Yuzu (No Sugar, No Calories)",
            "HBD Sparkling Water Peach (No Sugar, No Calories)",
            "HBD Sparkling Water Kyoho (No Sugar, No Calories)", "Asahi Beer",
            "Heineken Beer ", "Singha Beer", "Drinking Water",
            "Matcha Soft Serve", "Yogurt Soft Serve", "Two Tone Soft Serve",
            "Matcha Japanese", "Matcha Marshmallow", "Mix Berry Yogurt",
            "Brownie Yogurt", "Two Tone Brownie", "Saikyo Matcha",
            "Rice", "Miso Soup"
        ]
        hamada_costs = [
            67.53, 51.84, 70.42, 69.33, 53.64, 69.29,
            56.03, 54.34, 73.57, 37.21, 48.67, 62.31,
            69.17, 60.62, 16.16, 10.92, 33.71, 63.56,
            203.12, 89.93, 98.14, 168.1, 108.73, 200.7,
            193.91, 50.76, 87.59, 104.1, 133.97, 48.53,
            66.95, 77.77, 79.45, 108.68, 73.91, 111.38,
            243.48, 246.42, 42.71, 39.59, 60.89, 52.93,
            29.95, 27.87, 42.07, 9.63, 19.94, 19.87,
            19.96, 51.71, 56.12, 112.92, 21.27, 18.32,
            68.24, 51.43, 93.14, 56.38, 68.9, 48.9,
            58.8, 63.34, 74.35, 12.54, 12.54, 12.54,
            7.59, 7, 17, 17, 17, 17, 41.92,
            38.43, 31.71, 3.74, 15.09, 17.72, 16.41,
            22.6, 30.78, 34.94, 41.79, 37.45, 31.54,
            5.85, 3.58
        ]

        # ข้อมูลจริงจาก Hashira Cost.xlsx (ครบทั้ง 68 รายการ)
        hashira_menu_names = [
             "Pork Cut Steak Set", "Tonteki Set", "Beef Cut Steak Set (Medium Rare / Well done)",
            "Buta Teriyaki Set", "Tonkatsu Set", "Tonkatsu Tamagotoji Set",
            "Beef Hamburg Set", "Salmon Set (Teriyaki / Shioyaki)", "Kaisen Teppan Yaki Set",
            "Topping Fried Egg", "Gokuatsu Tonkatsu Set", "Gokuatsu Katsu Don Set",
            "Gokuatsu Tonkatsu Tamagotoji Set", "Tonkotsu Chashumen", "Tonkotsu Niku-Niku Chashumen",
            "Tonkotsu Spicy Ramen", "Tonkotsu Hashira Ramen", "Miso Chashumen",
            "Miso Hokkaido Ramen", "Shoyu Chashumen", "Shoyu Tokyo Ramen",
            "Tom Yum Tonkotsu Kaisen Ramen", "Tom Yum Kung Ramen", "Tori Paitan Karaage Ramen",
            "Tori Paitan Kaisen Ramen", "Tonkatsu", "Hotate",
            "Ebi Tempura", "Chashu", "Spicy Nikumiso",
            "Karashinegi", "Negi", "Naruto",
            "Ajitama", "Menma", "Nori",
            "Corn", "Butter", "Boiled Cabbage",
            "Boiled Bean Sprout", "Karashi Takana", "Kikurage",
            "Kanikama", "Kaedama", "GG Sauce Katsudon",
            "GG Buta Teriyaki Don", "GG Tonkatsu Bento", "GG Yakisoba",
            "GG Yakisoba with Sunny Side Egg", "GG Pork Okonomiyaki", "Coca-Cola Original",
            "Coca-Cola Zero Sugar", "Sprite", "Soda",
            "Mineral Water", "HBD Sparkling Water Lemon (No Sugar, No Calories)",
            "HBD Sparkling Water Honey Yuzu (No Sugar, No Calories)",
            "HBD Sparkling WatCalorieser Peach (No Sugar, No )",
            "HBD Sparkling Water Kyoho (No Sugar, No Calories)", "Asahi Beer",
            "Heineken Beer ", "Singha Beer", "Drinking Water",
            "Melon Ball", "Vanilla Monaka", "Azuki Bar",
            "Rice", "Miso Soup"
        ]

        hashira_costs = [
            46.53, 131.99, 107.87, 41.77, 58.39, 60.91,
            76.4, 68.02, 121.23, 3.5, 131.77, 133,
            131.46, 64.23, 77.98, 54.05, 74.33, 56.52,
            70.08, 63.1, 64.44, 55.55, 116.08, 42.88,
            50.08, 37.4, 33.33, 26, 27.3, 7.97,
            10.54, 6.43, 12.92, 3.84, 9.6, 5.4,
            6, 1.31, 2.73, 1.98, 6.45, 3.41,
            7.92, 7.5, 67.94, 45.93, 79.95, 45.61,
            49.11, 52.4, 12.54, 12.54, 12.54, 7.59,
            7, 17, 17, 17, 17, 41.92,
            38.43, 31.71, 3.74, 25, 27, 27,
            5.85, 3.04
        ]

        # สร้าง DataFrame
        hamada_real_data = {
            "MENU NAME": hamada_menu_names,
            "Material Cost": hamada_costs
        }

        hashira_real_data = {
            "MENU NAME": hashira_menu_names,
            "Material Cost": hashira_costs
        }

        # สร้าง DataFrame และ set index
        self.base_files['hamada'] = pd.DataFrame(hamada_real_data).set_index('MENU NAME')
        self.base_files['hashira'] = pd.DataFrame(hashira_real_data).set_index('MENU NAME')

        logger.info("โหลดข้อมูลจริง Base Cost เรียบร้อย")
        logger.info(f"Hamada Cost: {len(self.base_files['hamada'])} เมนู")
        logger.info(f"Hashira Cost: {len(self.base_files['hashira'])} เมนู")

    def load_config(self):
        """โหลดการตั้งค่าจากไฟล์"""
        # โหลดไฟล์ Base Cost.xlsx หากมี (จะ override ข้อมูลตัวอย่าง)
        self.load_base_cost_file()

    def load_base_cost_file(self):
        """โหลดไฟล์ Base Cost.xlsx"""
        base_file = "Base Cost.xlsx"
        if os.path.exists(base_file):
            try:
                # อ่าน Hashira Cost sheet
                hashira_df = pd.read_excel(base_file, sheet_name="Hashira Cost")
                if 'MENU NAME' in hashira_df.columns:
                    self.base_files['hashira'] = hashira_df.set_index("MENU NAME")

                # อ่าน Hamada Cost sheet
                hamada_df = pd.read_excel(base_file, sheet_name="Hamada Cost")
                if 'MENU NAME' in hamada_df.columns:
                    self.base_files['hamada'] = hamada_df.set_index("MENU NAME")

                logger.info(f"โหลดไฟล์ {base_file} สำเร็จ (แทนที่ข้อมูลตัวอย่าง)")
                logger.info(f"Hashira Cost: {len(self.base_files['hashira'])} เมนู")
                logger.info(f"Hamada Cost: {len(self.base_files['hamada'])} เมนู")

            except Exception as e:
                logger.error(f"Error loading base file: {e}")
                logger.info("ใช้ข้อมูลตัวอย่างต่อไป")

    def save_base_cost_file(self):
        """บันทึกไฟล์ Base Cost.xlsx"""
        try:
            wb = Workbook()
            # ลบ sheet default
            wb.remove(wb.active)

            # สร้าง Hashira Cost sheet
            if self.base_files['hashira'] is not None:
                ws_hashira = wb.create_sheet("Hashira Cost")
                df_hashira = self.base_files['hashira'].reset_index()

                for r in dataframe_to_rows(df_hashira, index=False, header=True):
                    ws_hashira.append(r)

                # จัดรูปแบบ header
                for cell in ws_hashira[1]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")

            # สร้าง Hamada Cost sheet
            if self.base_files['hamada'] is not None:
                ws_hamada = wb.create_sheet("Hamada Cost")
                df_hamada = self.base_files['hamada'].reset_index()

                for r in dataframe_to_rows(df_hamada, index=False, header=True):
                    ws_hamada.append(r)

                # จัดรูปแบบ header
                for cell in ws_hamada[1]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")

            # ปรับความกว้างคอลัมน์
            for ws in wb.worksheets:
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 60)
                    ws.column_dimensions[column_letter].width = adjusted_width

            wb.save("Base Cost.xlsx")
            logger.info("บันทึก Base Cost.xlsx สำเร็จ")
            return True

        except Exception as e:
            logger.error(f"Error saving base file: {e}")
            return False

    def update_single_base(self, base_type, new_df):
        """อัพเดท Base Cost แยกกันตามประเภท"""
        try:
            # ลบช่องว่างจากชื่อคอลัมน์
            new_df.columns = new_df.columns.str.strip()

            # ตรวจสอบคอลัมน์ที่จำเป็น
            if 'MENU NAME' not in new_df.columns:
                raise ValueError(f"ไฟล์ไม่มีคอลัมน์ 'MENU NAME'")

            if 'Material Cost' not in new_df.columns:
                raise ValueError(f"ไฟล์ไม่มีคอลัมน์ 'Material Cost'")

            # ตั้งค่า index
            new_df = new_df.set_index("MENU NAME")

            # อัพเดทข้อมูล
            self.base_files[base_type] = new_df

            # บันทึกไฟล์ Base Cost.xlsx
            self.save_base_cost_file()

            logger.info(f"อัพเดท {base_type.upper()} Cost สำเร็จ: {len(new_df)} เมนู")
            return True

        except Exception as e:
            logger.error(f"Error updating {base_type} base: {e}")
            raise e


# ===== Logging Setup =====
logging.basicConfig(
    level=logging.DEBUG,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('hamada_calculator_debug.log', encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# Initialize configuration
config_manager = ConfigManager()


class Windows11CostCalculatorApp:
    def __init__(self, root):
        self.root = root
        self.theme = Windows11Theme()

        # Configure root window
        self._setup_window()

        # Application variables
        self.import_file = None
        self.template_file = None
        self.selected_base_type = tk.StringVar(value='hamada')
        self.df_result = None
        self.df_import = None
        self.df_template = None
        self.debug_mode = tk.BooleanVar(value=False)

        # Debug data
        self.debug_data = {
            'not_found_menus': [],
            'zero_qty_items': [],
            'invalid_qty_items': [],
            'nan_cost_items': [],
            'matched_menus': [],
            'not_sold_menus': [],
            'processing_summary': {}
        }

        self._build_modern_ui()
        logger.info("Windows 11 Cost Calculator initialized")

    def _setup_window(self):
        """Setup main window with Windows 11 style"""
        self.root.title("🍽️ Japanese Cost Calculator - Mcost_V1.0")
        self.root.geometry("1600x1000")
        self.root.configure(bg=self.theme.COLORS['bg_primary'])
        self.root.minsize(1400, 900)

        # Center window
        self.root.update_idletasks()
        x = (self.root.winfo_screenwidth() // 2) - (1600 // 2)
        y = (self.root.winfo_screenheight() // 2) - (1000 // 2)
        self.root.geometry(f"1600x1000+{x}+{y}")

    def _build_modern_ui(self):
        """Build modern Windows 11 style UI"""
        # Main container with padding
        main_container = tk.Frame(self.root, bg=self.theme.COLORS['bg_primary'])
        main_container.pack(fill=tk.BOTH, expand=True, padx=self.theme.SPACING['lg'],
                            pady=self.theme.SPACING['lg'])

        # Header section
        self._create_header(main_container)

        # Content area with cards
        content_frame = tk.Frame(main_container, bg=self.theme.COLORS['bg_primary'])
        content_frame.pack(fill=tk.BOTH, expand=True, pady=(self.theme.SPACING['lg'], 0))

        # Left panel (controls)
        left_panel = tk.Frame(content_frame, bg=self.theme.COLORS['bg_primary'], width=400)
        left_panel.pack(side=tk.LEFT, fill=tk.Y, padx=(0, self.theme.SPACING['md']))
        left_panel.pack_propagate(False)

        # Right panel (results)
        right_panel = tk.Frame(content_frame, bg=self.theme.COLORS['bg_primary'])
        right_panel.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)

        # Build panels
        self._create_control_panel(left_panel)
        self._create_results_panel(right_panel)

    def _create_header(self, parent):
        """Create modern header with gradient-like appearance"""
        header_frame = tk.Frame(parent, bg=self.theme.COLORS['accent_blue'], height=120)
        header_frame.pack(fill=tk.X)
        header_frame.pack_propagate(False)

        # Header content
        header_content = tk.Frame(header_frame, bg=self.theme.COLORS['accent_blue'])
        header_content.pack(expand=True, fill=tk.BOTH, padx=self.theme.SPACING['xl'],
                            pady=self.theme.SPACING['lg'])

        # Title
        title_label = tk.Label(
            header_content,
            text="🍽️ Japanese Cost Calculator",
            font=self.theme.FONTS['heading_large'],
            fg=self.theme.COLORS['text_white'],
            bg=self.theme.COLORS['accent_blue']
        )
        title_label.pack(anchor='w')

        # Subtitle
        subtitle_label = tk.Label(
            header_content,
            text="JP.Mcost V1.0• Hamada & Hashira Cost Management By Japanese Restaurant Group [IMPACT]",
            font=self.theme.FONTS['body_large'],
            fg='#E3F2FD',
            bg=self.theme.COLORS['accent_blue']
        )
        subtitle_label.pack(anchor='w', pady=(4, 0))

    def _create_control_panel(self, parent):
        """Create left control panel with modern cards"""
        # Base Type Selection Card
        base_card = ModernCard(parent, title="🏪 Restaurant Selection")
        base_card.pack(fill=tk.X, pady=(0, self.theme.SPACING['md']))

        base_content = tk.Frame(base_card, bg=self.theme.COLORS['bg_card'])
        base_content.pack(fill=tk.X, padx=self.theme.SPACING['md'],
                          pady=(0, self.theme.SPACING['md']))

        # Radio buttons with modern styling
        self._create_radio_section(base_content)

        # File Management Card
        file_card = ModernCard(parent, title="📁 File Management")
        file_card.pack(fill=tk.X, pady=(0, self.theme.SPACING['md']))

        file_content = tk.Frame(file_card, bg=self.theme.COLORS['bg_card'])
        file_content.pack(fill=tk.X, padx=self.theme.SPACING['md'],
                          pady=(0, self.theme.SPACING['md']))

        self._create_file_section(file_content)

        # Actions Card
        action_card = ModernCard(parent, title="⚡ Actions")
        action_card.pack(fill=tk.X, pady=(0, self.theme.SPACING['md']))

        action_content = tk.Frame(action_card, bg=self.theme.COLORS['bg_card'])
        action_content.pack(fill=tk.X, padx=self.theme.SPACING['md'],
                            pady=(0, self.theme.SPACING['md']))

        self._create_action_section(action_content)

        # Status Card
        status_card = ModernCard(parent, title="📊 System Status")
        status_card.pack(fill=tk.BOTH, expand=True)

        status_content = tk.Frame(status_card, bg=self.theme.COLORS['bg_card'])
        status_content.pack(fill=tk.BOTH, expand=True, padx=self.theme.SPACING['md'],
                            pady=(0, self.theme.SPACING['md']))

        self._create_status_section(status_content)

    def _create_radio_section(self, parent):
        """Create modern radio button section"""
        radio_frame = tk.Frame(parent, bg=self.theme.COLORS['bg_card'])
        radio_frame.pack(fill=tk.X, pady=self.theme.SPACING['sm'])

        # Custom radio buttons
        hashira_frame = tk.Frame(radio_frame, bg=self.theme.COLORS['bg_secondary'],
                                 relief='solid', bd=1)
        hashira_frame.pack(fill=tk.X, pady=(0, self.theme.SPACING['sm']))

        self.hashira_radio = tk.Radiobutton(
            hashira_frame,
            text="🏯 Hashira Sky Kitchen",
            variable=self.selected_base_type,
            value='hashira',
            command=self.on_base_type_change,
            font=self.theme.FONTS['body_medium'],
            bg=self.theme.COLORS['bg_secondary'],
            fg=self.theme.COLORS['text_primary'],
            selectcolor=self.theme.COLORS['accent_blue'],
            activebackground=self.theme.COLORS['bg_secondary'],
            relief='flat',
            bd=0
        )
        self.hashira_radio.pack(anchor='w', padx=self.theme.SPACING['md'],
                                pady=self.theme.SPACING['sm'])

        hamada_frame = tk.Frame(radio_frame, bg=self.theme.COLORS['bg_secondary'],
                                relief='solid', bd=1)
        hamada_frame.pack(fill=tk.X)

        self.hamada_radio = tk.Radiobutton(
            hamada_frame,
            text="🍜 Hamada Sky Kitchen",
            variable=self.selected_base_type,
            value='hamada',
            command=self.on_base_type_change,
            font=self.theme.FONTS['body_medium'],
            bg=self.theme.COLORS['bg_secondary'],
            fg=self.theme.COLORS['text_primary'],
            selectcolor=self.theme.COLORS['accent_blue'],
            activebackground=self.theme.COLORS['bg_secondary'],
            relief='flat',
            bd=0
        )
        self.hamada_radio.pack(anchor='w', padx=self.theme.SPACING['md'],
                               pady=self.theme.SPACING['sm'])

    def _create_file_section(self, parent):
        """Create file management section"""
        # Import file button
        ModernButton(
            parent,
            text="📄 Select Import File",
            command=self.load_import_file,
            style="primary"
        ).pack(fill=tk.X, pady=(0, self.theme.SPACING['sm']))

        # Template file button
        ModernButton(
            parent,
            text="📋 Select Template File",
            command=self.load_template_file,
            style="secondary"
        ).pack(fill=tk.X, pady=(0, self.theme.SPACING['sm']))

        # Check base file button
        ModernButton(
            parent,
            text="🔍 Check Base File",
            command=self.check_base_file,
            style="secondary"
        ).pack(fill=tk.X, pady=(0, self.theme.SPACING['md']))

        # File status labels
        self.import_status_label = tk.Label(
            parent,
            text="📂 Import: No file selected",
            font=self.theme.FONTS['caption'],
            fg=self.theme.COLORS['text_secondary'],
            bg=self.theme.COLORS['bg_card'],
            anchor='w'
        )
        self.import_status_label.pack(fill=tk.X, pady=(0, 4))

        self.template_status_label = tk.Label(
            parent,
            text="📋 Template: No file selected",
            font=self.theme.FONTS['caption'],
            fg=self.theme.COLORS['text_secondary'],
            bg=self.theme.COLORS['bg_card'],
            anchor='w'
        )
        self.template_status_label.pack(fill=tk.X)

    def _create_action_section(self, parent):
        """Create action buttons section"""
        # Debug mode checkbox
        debug_frame = tk.Frame(parent, bg=self.theme.COLORS['bg_card'])
        debug_frame.pack(fill=tk.X, pady=(0, self.theme.SPACING['md']))

        self.debug_checkbox = tk.Checkbutton(
            debug_frame,
            text="🔍 Debug Mode",
            variable=self.debug_mode,
            font=self.theme.FONTS['body_medium'],
            bg=self.theme.COLORS['bg_card'],
            fg=self.theme.COLORS['text_primary'],
            selectcolor=self.theme.COLORS['accent_blue'],
            activebackground=self.theme.COLORS['bg_card']
        )
        self.debug_checkbox.pack(anchor='w')

        # Calculate button
        ModernButton(
            parent,
            text="⚡ Calculate Costs",
            command=self.calculate,
            style="success"
        ).pack(fill=tk.X, pady=(0, self.theme.SPACING['sm']))

        # Export button
        ModernButton(
            parent,
            text="💾 Export Excel Report",
            command=self.export_complete_excel,
            style="primary"
        ).pack(fill=tk.X, pady=(0, self.theme.SPACING['sm']))

        # Clear debug button
        ModernButton(
            parent,
            text="🗑️ Clear Debug Log",
            command=self.clear_debug,
            style="warning"
        ).pack(fill=tk.X)

    def _create_status_section(self, parent):
        """Create status information section"""
        # Base files status
        self.hashira_status_label = tk.Label(
            parent,
            text="🏯 Hashira: Checking...",
            font=self.theme.FONTS['body_small'],
            fg=self.theme.COLORS['text_secondary'],
            bg=self.theme.COLORS['bg_card'],
            anchor='w'
        )
        self.hashira_status_label.pack(fill=tk.X, pady=(0, 4))

        self.hamada_status_label = tk.Label(
            parent,
            text="🍜 Hamada: Checking...",
            font=self.theme.FONTS['body_small'],
            fg=self.theme.COLORS['text_secondary'],
            bg=self.theme.COLORS['bg_card'],
            anchor='w'
        )
        self.hamada_status_label.pack(fill=tk.X, pady=(0, self.theme.SPACING['md']))

        # Statistics label
        self.stats_label = tk.Label(
            parent,
            text="📈 Ready to calculate",
            font=self.theme.FONTS['body_medium'],
            fg=self.theme.COLORS['text_primary'],
            bg=self.theme.COLORS['bg_card'],
            anchor='w',
            wraplength=350
        )
        self.stats_label.pack(fill=tk.X)

    def _create_results_panel(self, parent):
        """Create results panel with tabbed interface"""
        # Results card
        results_card = ModernCard(parent, title="📊 Results & Analysis")
        results_card.pack(fill=tk.BOTH, expand=True)

        results_content = tk.Frame(results_card, bg=self.theme.COLORS['bg_card'])
        results_content.pack(fill=tk.BOTH, expand=True, padx=self.theme.SPACING['md'],
                             pady=(0, self.theme.SPACING['md']))

        # Create modern notebook
        self.notebook = ttk.Notebook(results_content)
        self.notebook.pack(fill=tk.BOTH, expand=True)

        # Configure modern notebook style
        self._configure_notebook_style()

        # Results tab
        self._create_results_tab()

        # Debug tab
        self._create_debug_tab()

    def _configure_notebook_style(self):
        """Configure modern style for notebook"""
        style = ttk.Style()

        # Configure notebook style
        style.configure('Modern.TNotebook',
                        background=self.theme.COLORS['bg_card'],
                        borderwidth=0)

        style.configure('Modern.TNotebook.Tab',
                        padding=[20, 10],
                        font=self.theme.FONTS['body_medium'],
                        focuscolor='none')

        style.map('Modern.TNotebook.Tab',
                  background=[('selected', self.theme.COLORS['accent_blue']),
                              ('!selected', self.theme.COLORS['bg_secondary'])],
                  foreground=[('selected', self.theme.COLORS['text_white']),
                              ('!selected', self.theme.COLORS['text_primary'])])

        self.notebook.configure(style='Modern.TNotebook')

    def _create_results_tab(self):
        """Create results tab with modern treeview"""
        results_frame = ttk.Frame(self.notebook)
        self.notebook.add(results_frame, text="📋 Calculation Results")

        # Table container
        table_container = tk.Frame(results_frame, bg=self.theme.COLORS['bg_card'])
        table_container.pack(fill=tk.BOTH, expand=True, padx=self.theme.SPACING['md'],
                             pady=self.theme.SPACING['md'])

        # Create modern treeview
        self._create_modern_treeview(table_container)

    def _create_modern_treeview(self, parent):
        """Create modern styled treeview"""
        # Treeview frame
        tree_frame = tk.Frame(parent, bg=self.theme.COLORS['bg_card'])
        tree_frame.pack(fill=tk.BOTH, expand=True)

        # Create treeview
        self.tree = ttk.Treeview(
            tree_frame,
            columns=("menu", "qty", "cost", "total"),
            show="headings",
            height=20
        )

        # Configure columns
        self.tree.heading("menu", text="🍽️ Menu Name")
        self.tree.heading("qty", text="📊 Quantity")
        self.tree.heading("cost", text="💵 Unit Cost")
        self.tree.heading("total", text="💰 Total Cost")

        self.tree.column("menu", width=300, anchor="w")
        self.tree.column("qty", width=100, anchor="center")
        self.tree.column("cost", width=120, anchor="e")
        self.tree.column("total", width=120, anchor="e")

        # Scrollbar
        tree_scroll = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=tree_scroll.set)

        # Pack treeview and scrollbar
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        tree_scroll.pack(side=tk.RIGHT, fill=tk.Y)

        # Configure modern treeview style
        self._configure_treeview_style()

    def _configure_treeview_style(self):
        """Configure modern treeview style"""
        style = ttk.Style()

        style.configure('Modern.Treeview',
                        background=self.theme.COLORS['bg_secondary'],
                        foreground=self.theme.COLORS['text_primary'],
                        fieldbackground=self.theme.COLORS['bg_secondary'],
                        font=self.theme.FONTS['body_small'],
                        rowheight=25)

        style.configure('Modern.Treeview.Heading',
                        background=self.theme.COLORS['accent_blue'],
                        foreground=self.theme.COLORS['text_white'],
                        font=self.theme.FONTS['body_medium'],
                        relief='flat')

        style.map('Modern.Treeview',
                  background=[('selected', self.theme.COLORS['accent_light_blue'])])

        self.tree.configure(style='Modern.Treeview')

    def _create_debug_tab(self):
        """Create debug tab with modern text area"""
        debug_frame = ttk.Frame(self.notebook)
        self.notebook.add(debug_frame, text="🔍 Debug Console")

        # Debug container
        debug_container = tk.Frame(debug_frame, bg=self.theme.COLORS['bg_card'])
        debug_container.pack(fill=tk.BOTH, expand=True, padx=self.theme.SPACING['md'],
                             pady=self.theme.SPACING['md'])

        # Debug text area
        self.debug_text = scrolledtext.ScrolledText(
            debug_container,
            wrap=tk.WORD,
            height=20,
            font=self.theme.FONTS['monospace'],
            bg='#1E1E1E',  # Dark theme for console
            fg='#FFFFFF',
            insertbackground='#FFFFFF',
            selectbackground=self.theme.COLORS['accent_blue'],
            relief='flat',
            borderwidth=0
        )
        self.debug_text.pack(fill=tk.BOTH, expand=True)

    # ===== Event Handlers =====
    def on_base_type_change(self):
        """Handle base type change"""
        selected = self.selected_base_type.get()
        self.debug_log(f"Changed base type to: {selected.upper()}")
        self.update_all_base_status()

    def debug_log(self, message, level="INFO"):
        """Add message to debug log"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        log_message = f"[{timestamp}] {level}: {message}\n"

        self.debug_text.insert(tk.END, log_message)
        self.debug_text.see(tk.END)

        if level == "ERROR":
            logger.error(message)
        elif level == "WARNING":
            logger.warning(message)
        else:
            logger.info(message)

    def clear_debug(self):
        """Clear debug log"""
        self.debug_text.delete(1.0, tk.END)
        self.debug_log("Debug log cleared")
        self.debug_data = {
            'not_found_menus': [],
            'zero_qty_items': [],
            'invalid_qty_items': [],
            'nan_cost_items': [],
            'matched_menus': [],
            'not_sold_menus': [],
            'processing_summary': {}
        }

    def update_all_base_status(self):
        """Update status of all base files"""
        # Hashira Status
        hashira_df = config_manager.base_files.get('hashira')
        if hashira_df is not None and not hashira_df.empty:
            hashira_count = len(hashira_df)
            hashira_text = f"🏯 Hashira: Ready ({hashira_count} menus)"
            hashira_color = self.theme.COLORS['success']
        else:
            hashira_text = "🏯 Hashira: No data found"
            hashira_color = self.theme.COLORS['error']

        # Hamada Status
        hamada_df = config_manager.base_files.get('hamada')
        if hamada_df is not None and not hamada_df.empty:
            hamada_count = len(hamada_df)
            hamada_text = f"🍜 Hamada: Ready ({hamada_count} menus)"
            hamada_color = self.theme.COLORS['success']
        else:
            hamada_text = "🍜 Hamada: No data found"
            hamada_color = self.theme.COLORS['error']

        self.hashira_status_label.config(text=hashira_text, fg=hashira_color)
        self.hamada_status_label.config(text=hamada_text, fg=hamada_color)

    def load_import_file(self):
        """Load import file"""
        from tkinter import filedialog

        try:
            file_path = filedialog.askopenfilename(
                title="Select Import File",
                filetypes=[("Excel Files", "*.xlsx"), ("Excel Files", "*.xls")]
            )

            if not file_path:
                return

            # Load and store file data
            self.df_import = pd.read_excel(file_path)
            self.import_file = file_path

            filename = os.path.basename(file_path)
            self.import_status_label.config(
                text=f"📂 Import: {filename} ({len(self.df_import)} rows)",
                fg=self.theme.COLORS['success']
            )

            self.debug_log(f"Import file loaded: {file_path}")
            self._preview_file(self.df_import, "Import")

        except Exception as e:
            self.debug_log(f"Error loading import file: {e}", "ERROR")
            messagebox.showerror("Error", f"Cannot load file: {str(e)}")

    def load_template_file(self):
        """Load template file"""
        from tkinter import filedialog

        try:
            file_path = filedialog.askopenfilename(
                title="Select Template File",
                filetypes=[("Excel Files", "*.xlsx"), ("Excel Files", "*.xls")]
            )

            if not file_path:
                return

            # Load and store template file data
            self.df_template = pd.read_excel(file_path)
            self.template_file = file_path

            filename = os.path.basename(file_path)
            self.template_status_label.config(
                text=f"📋 Template: {filename} ({len(self.df_template)} rows)",
                fg=self.theme.COLORS['success']
            )

            self.debug_log(f"Template file loaded: {file_path}")
            self._preview_file(self.df_template, "Template")

        except Exception as e:
            self.debug_log(f"Error loading template file: {e}", "ERROR")
            messagebox.showerror("Error", f"Cannot load file: {str(e)}")

    def _preview_file(self, df_data, file_type):
        """Preview file data in debug log"""
        try:
            self.debug_log(f"{file_type} file: {len(df_data)} rows, {len(df_data.columns)} columns")
            self.debug_log(f"Columns: {df_data.columns.tolist()}")

            # Show sample data
            self.debug_log(f"Sample data from {file_type} (first 5 rows):")
            for i, (idx, row) in enumerate(df_data.head(5).iterrows()):
                menu = row.get("MENU NAME", "N/A")
                qty = row.get("Qty", "N/A")
                self.debug_log(f"  Row {i + 1}: {menu} = {qty}")

        except Exception as e:
            self.debug_log(f"Error previewing {file_type} file: {str(e)}", "ERROR")

    def check_base_file(self):
        """Check and display base file information"""
        selected = self.selected_base_type.get()
        df_base = config_manager.base_files.get(selected)

        if df_base is None:
            messagebox.showerror("Error", f"Cannot load {selected.upper()} Cost data")
            return

        self.debug_log(f"=== Checking {selected.upper()} Cost ===")
        self.debug_log(f"Total menus: {len(df_base)}")
        self.debug_log(f"Columns: {df_base.columns.tolist()}")

        # Price statistics
        try:
            if "Material Cost" in df_base.columns:
                cost_stats = pd.to_numeric(df_base["Material Cost"], errors='coerce').describe()
                self.debug_log("Material Cost Statistics:")
                self.debug_log(f"  Minimum: {cost_stats['min']:.2f} THB")
                self.debug_log(f"  Maximum: {cost_stats['max']:.2f} THB")
                self.debug_log(f"  Average: {cost_stats['mean']:.2f} THB")
                self.debug_log(f"  Total items: {int(cost_stats['count'])}")
        except Exception as e:
            self.debug_log(f"Error calculating statistics: {str(e)}", "ERROR")

    def calculate(self):
        """Calculate costs - Enhanced Version"""
        df_base = config_manager.base_files.get(self.selected_base_type.get())
        selected_base = self.selected_base_type.get()

        # Validation checks
        if df_base is None or df_base.empty:
            messagebox.showerror("Error", f"{selected_base.upper()} Cost file cannot be loaded")
            return

        if self.df_import is None:
            messagebox.showwarning("Warning", "Please select an Import file first")
            return

        self.debug_log(f"=== Starting calculation with {selected_base.upper()} Cost ===")

        try:
            # Clean data
            df_base.columns = df_base.columns.str.strip()
            self.df_import.columns = self.df_import.columns.str.strip()

            # Check required columns
            required_columns = ["MENU NAME", "Qty"]
            missing_columns = [col for col in required_columns if col not in self.df_import.columns]

            if missing_columns:
                error_msg = f"Import file missing columns: {', '.join(missing_columns)}"
                self.debug_log(error_msg, "ERROR")
                messagebox.showerror("Error", error_msg)
                return

            self.debug_log(f"Import file loaded successfully: {len(self.df_import)} rows")

        except Exception as e:
            error_msg = f"Cannot read file: {str(e)}"
            self.debug_log(error_msg, "ERROR")
            messagebox.showerror("Error", error_msg)
            return

        # Start calculation
        results = []
        matched_count = 0
        self.debug_data['not_found_menus'] = []
        self.debug_data['matched_menus'] = []

        for idx, row in self.df_import.iterrows():
            menu = row.get("MENU NAME")
            qty = row.get("Qty", 0)

            if pd.isna(menu) or menu == "":
                continue

            try:
                qty = float(qty) if not pd.isna(qty) else 0
            except:
                qty = 0
                self.debug_data['invalid_qty_items'].append(menu)

            if menu in df_base.index:
                try:
                    material_cost = df_base.at[menu, "Material Cost"]
                    if pd.isna(material_cost):
                        material_cost = 0
                    else:
                        material_cost = float(material_cost)

                    total_cost = qty * material_cost
                    results.append([menu, qty, material_cost, total_cost])
                    matched_count += 1
                    self.debug_data['matched_menus'].append(menu)

                    if self.debug_mode.get():
                        self.debug_log(f"✓ Found menu: {menu} = {material_cost:.2f} x {qty} = {total_cost:.2f}")

                except Exception as e:
                    self.debug_log(f"Error calculating {menu}: {str(e)}", "ERROR")
                    self.debug_data['nan_cost_items'].append(menu)
            else:
                self.debug_data['not_found_menus'].append(menu)
                if self.debug_mode.get():
                    self.debug_log(f"❌ Menu not found: {menu}", "WARNING")

        # Create results DataFrame
        self.df_result = pd.DataFrame(results, columns=["MENU NAME", "Qty", "Material Cost", "Total Cost"])

        if self.df_result.empty:
            error_msg = f"No matching menu names found in {selected_base.upper()} Cost"
            self.debug_log(error_msg, "WARNING")
            messagebox.showwarning("No Data Found", error_msg)
            return

        # Calculate totals
        total_qty = self.df_result["Qty"].sum()
        total_material_cost = self.df_result["Material Cost"].sum()
        grand_total = self.df_result["Total Cost"].sum()

        # Add TOTAL row
        total_row = pd.DataFrame([["TOTAL", total_qty, total_material_cost, grand_total]],
                                 columns=["MENU NAME", "Qty", "Material Cost", "Total Cost"])
        self.df_result = pd.concat([self.df_result, total_row], ignore_index=True)

        # Save summary in debug_data
        self.debug_data['processing_summary'] = {
            'total_import_items': len(self.df_import),
            'matched_items': matched_count,
            'not_found_items': len(self.debug_data['not_found_menus']),
            'total_qty': total_qty,
            'total_material_cost': total_material_cost,
            'grand_total': grand_total,
            'base_type_used': selected_base.upper()
        }

        # Update display
        self._update_result_table()
        self._update_statistics(matched_count, grand_total, selected_base)

        success_msg = f"Calculation completed - Found {matched_count} items (using {selected_base.upper()} Cost)"
        self.debug_log(success_msg)
        messagebox.showinfo("Success", success_msg)

    def _update_result_table(self):
        """Update results table"""
        # Clear existing items
        for row in self.tree.get_children():
            self.tree.delete(row)

        # Add new items
        for _, r in self.df_result.iterrows():
            qty_display = int(r["Qty"]) if r["Qty"] != "" else ""
            cost_display = f"{r['Material Cost']:.2f}" if r["Material Cost"] != "" else ""
            total_display = f"{r['Total Cost']:.2f}" if r["Total Cost"] != "" else ""

            tags = ("total_row",) if r["MENU NAME"] == "TOTAL" else ()

            self.tree.insert("", tk.END,
                             values=(r["MENU NAME"], qty_display, cost_display, total_display),
                             tags=tags)

        # Configure total row styling
        self.tree.tag_configure("total_row",
                                background=self.theme.COLORS['accent_blue'],
                                foreground=self.theme.COLORS['text_white'])

    def _update_statistics(self, matched, grand_total, base_type):
        """Update statistics display"""
        stats_text = (f"✅ Found: {matched} items | 💰 Total: {grand_total:,.2f} THB | "
                      f"🏪 Base: {base_type.upper()}")
        self.stats_label.config(text=stats_text, fg=self.theme.COLORS['success'])

    def export_complete_excel(self):
        """Export complete Excel file with 4 sheets"""
        self.debug_log("🔄 Starting Excel export...")

        # Check basic data
        if self.df_result is None or self.df_result.empty:
            messagebox.showwarning("Warning", "Please calculate before exporting")
            return

        selected_base = self.selected_base_type.get()
        filename = f"Cost_Analysis_Complete_{selected_base.upper()}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        file_path = os.path.join(os.getcwd(), filename)

        try:
            self.debug_log(f"📂 Creating Excel file at: {file_path}")

            # Create Excel Workbook
            wb = Workbook()
            wb.remove(wb.active)  # Remove default sheet

            sheets_created = 0

            # Sheet 1: Calculation results
            if self.df_result is not None:
                ws_result = wb.create_sheet("Calculation results")

                # Add header and summary info
                ws_result.append(["📊 Cost Calculation Results"])
                ws_result.append([])
                ws_result.append([f"🏪 Base Cost Used: {selected_base.upper()}"])
                ws_result.append([f"📅 Calculation Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
                ws_result.append([f"📋 Total Items: {len(self.df_result) - 1}"])  # Exclude TOTAL row
                ws_result.append([])

                # Add results table
                for r in dataframe_to_rows(self.df_result, index=False, header=True):
                    ws_result.append(r)

                # Format header
                ws_result['A1'].font = Font(bold=True, size=14, color="FFFFFF")
                ws_result['A1'].fill = PatternFill(start_color="0078D4", end_color="0078D4", fill_type="solid")

                # Format table headers
                header_row = 7
                for cell in ws_result[header_row]:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="107C10", end_color="107C10", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")

                # Highlight TOTAL row
                for row in range(header_row + 1, ws_result.max_row + 1):
                    if ws_result[f'A{row}'].value == "TOTAL":
                        for col in range(1, 5):
                            cell = ws_result.cell(row=row, column=col)
                            cell.font = Font(bold=True, color="FFFFFF")
                            cell.fill = PatternFill(start_color="D13438", end_color="D13438", fill_type="solid")

                sheets_created += 1
                self.debug_log(f"✅ Sheet 1: Calculation results ({len(self.df_result)} items)")

            # Sheet 2: Base Cost
            df_base = config_manager.base_files.get(selected_base)
            if df_base is not None:
                ws_base = wb.create_sheet(f"Base Cost ({selected_base.upper()})")
                df_base_export = df_base.reset_index()

                # Add header
                ws_base.append([f"🏪 {selected_base.upper()} Base Cost Data"])
                ws_base.append([])
                ws_base.append([f"📅 Export Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
                ws_base.append([f"📋 Total Menus: {len(df_base_export)}"])
                ws_base.append([])

                # Add data
                for r in dataframe_to_rows(df_base_export, index=False, header=True):
                    ws_base.append(r)

                # Format header
                ws_base['A1'].font = Font(bold=True, size=14, color="FFFFFF")
                if selected_base == 'hashira':
                    ws_base['A1'].fill = PatternFill(start_color="8764B8", end_color="8764B8", fill_type="solid")
                else:
                    ws_base['A1'].fill = PatternFill(start_color="FF8C00", end_color="FF8C00", fill_type="solid")

                # Format table headers
                header_row = 6
                for cell in ws_base[header_row]:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="605E5C", end_color="605E5C", fill_type="solid")

                sheets_created += 1
                self.debug_log(f"✅ Sheet 2: Base Cost {selected_base.upper()} ({len(df_base_export)} menus)")

            # Sheet 3: Import (By User)
            if self.df_import is not None:
                ws_import = wb.create_sheet("Import (By User)")

                # Add header
                ws_import.append(["📁 Import (By User) Data"])
                ws_import.append([])
                ws_import.append([f"📂 File: {os.path.basename(self.import_file) if self.import_file else 'N/A'}"])
                ws_import.append([f"📅 Import Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
                ws_import.append([f"📋 Total Rows: {len(self.df_import)}"])
                ws_import.append([])

                # Add data
                for r in dataframe_to_rows(self.df_import, index=False, header=True):
                    ws_import.append(r)

                # Format header
                ws_import['A1'].font = Font(bold=True, size=14, color="FFFFFF")
                ws_import['A1'].fill = PatternFill(start_color="107C10", end_color="107C10", fill_type="solid")

                # Format table headers
                header_row = 7
                for cell in ws_import[header_row]:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="40E0FF", end_color="40E0FF", fill_type="solid")

                sheets_created += 1
                self.debug_log(f"✅ Sheet 3: Import (By User) ({len(self.df_import)} rows)")

            # Sheet 4: Raw file(POS)
            if self.df_template is not None:
                ws_template = wb.create_sheet("Raw file(POS)")

                # Add header
                ws_template.append(["📄 Raw file(POS) Data"])
                ws_template.append([])
                ws_template.append([f"📂 File: {os.path.basename(self.template_file) if self.template_file else 'N/A'}"])
                ws_template.append([f"📅 Import Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
                ws_template.append([f"📋 Total Rows: {len(self.df_template)}"])
                ws_template.append([])

                # Add data
                for r in dataframe_to_rows(self.df_template, index=False, header=True):
                    ws_template.append(r)

                # Format header
                ws_template['A1'].font = Font(bold=True, size=14, color="FFFFFF")
                ws_template['A1'].fill = PatternFill(start_color="FF8C00", end_color="FF8C00", fill_type="solid")

                # Format table headers
                header_row = 7
                for cell in ws_template[header_row]:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="D13438", end_color="D13438", fill_type="solid")

                sheets_created += 1
                self.debug_log(f"✅ Sheet 4: Raw file(POS) ({len(self.df_template)} rows)")
            else:
                # Create empty sheet if no template
                ws_template = wb.create_sheet("Raw file(POS)")
                ws_template.append(["📄 Raw file(POS) Data"])
                ws_template.append([])
                ws_template.append(["ℹ️ No template file imported"])

                ws_template['A1'].font = Font(bold=True, size=14, color="FFFFFF")
                ws_template['A1'].fill = PatternFill(start_color="8A8886", end_color="8A8886", fill_type="solid")

                sheets_created += 1
                self.debug_log("✅ Sheet 4: Raw file(POS) (no data)")

            # Adjust column widths for all sheets
            for ws in wb.worksheets:
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width

            # Save file
            wb.save(file_path)

            # Export summary
            summary_msg = f"""✅ Excel Export Successful!

📁 Filename: {filename}
📍 Location: {file_path}
📊 Total Sheets: {sheets_created}

📋 Sheet Details:
1. Calculation results ({len(self.df_result)} items) ✨ With totals
2. Base Cost ({selected_base.upper()}) ({len(df_base)} menus)
3. Import (By User) ({len(self.df_import) if self.df_import is not None else 0} rows)
4. Raw file(POS) ({len(self.df_template) if self.df_template is not None else 0} rows)

💰 Summary:
📊 Total Qty: {self.debug_data['processing_summary']['total_qty']:,.0f}
💵 Total Material Cost: {self.debug_data['processing_summary']['total_material_cost']:,.2f} THB
💰 Grand Total: {self.debug_data['processing_summary']['grand_total']:,.2f} THB
🏪 Base: {selected_base.upper()}

✅ File ready for use"""

            self.debug_log(f"✅ Export successful: {filename}")
            self.debug_log(f"📊 Sheets: {sheets_created}")
            self.debug_log(f"📊 Total Qty: {self.debug_data['processing_summary']['total_qty']:,.0f}")
            self.debug_log(
                f"💵 Total Material Cost: {self.debug_data['processing_summary']['total_material_cost']:,.2f} THB")
            self.debug_log(f"💰 Grand Total: {self.debug_data['processing_summary']['grand_total']:,.2f} THB")

            # Show modern success dialog
            self._show_success_dialog("Export Successful!", summary_msg)

        except Exception as e:
            error_msg = f"Error exporting file: {str(e)}"
            self.debug_log(error_msg, "ERROR")
            self.debug_log(f"Traceback: {traceback.format_exc()}", "ERROR")
            messagebox.showerror("Export Error", error_msg)

    def _show_success_dialog(self, title, message):
        """Show modern success dialog"""
        dialog = tk.Toplevel(self.root)
        dialog.title(title)
        dialog.geometry("600x500")
        dialog.configure(bg=self.theme.COLORS['bg_primary'])
        dialog.transient(self.root)
        dialog.grab_set()

        # Center dialog
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (600 // 2)
        y = (dialog.winfo_screenheight() // 2) - (500 // 2)
        dialog.geometry(f"600x500+{x}+{y}")

        # Header
        header_frame = tk.Frame(dialog, bg=self.theme.COLORS['success'], height=80)
        header_frame.pack(fill='x')
        header_frame.pack_propagate(False)

        header_content = tk.Frame(header_frame, bg=self.theme.COLORS['success'])
        header_content.pack(expand=True, fill='both', padx=30, pady=20)

        tk.Label(
            header_content,
            text="✅ " + title,
            font=self.theme.FONTS['heading_medium'],
            fg=self.theme.COLORS['text_white'],
            bg=self.theme.COLORS['success']
        ).pack(anchor='w')

        # Content
        content_frame = tk.Frame(dialog, bg=self.theme.COLORS['bg_primary'])
        content_frame.pack(fill='both', expand=True, padx=30, pady=20)

        # Message text
        text_widget = tk.Text(
            content_frame,
            wrap=tk.WORD,
            font=self.theme.FONTS['body_medium'],
            bg=self.theme.COLORS['bg_secondary'],
            fg=self.theme.COLORS['text_primary'],
            relief='flat',
            borderwidth=0,
            padx=20,
            pady=20
        )
        text_widget.pack(fill='both', expand=True)
        text_widget.insert('1.0', message)
        text_widget.config(state='disabled')

        # Button frame
        button_frame = tk.Frame(dialog, bg=self.theme.COLORS['bg_primary'])
        button_frame.pack(fill='x', padx=30, pady=(0, 30))

        ModernButton(
            button_frame,
            text="Close",
            command=dialog.destroy,
            style="primary"
        ).pack(side='right')


def main():
    """Main function"""
    try:
        root = tk.Tk()
        app = Windows11CostCalculatorApp(root)

        # Initialize status
        app.debug_log("=== Windows 11 Cost Calculator Started ===")
        app.debug_log("✅ Real data loaded - Hamada (87 menus) and Hashira (68 menus)")
        app.debug_log("💾 Export: Single Excel file with 4 sheets")
        app.debug_log("📊 Sheet 1 includes totals for Qty, Material Cost, Total Cost")
        app.debug_log("📂 Files will be saved in the same folder as the program")
        app.debug_log("🎨 Windows 11 Modern UI Theme Applied")

        # Update initial status
        app.update_all_base_status()

        root.mainloop()

    except Exception as e:
        logger.error(f"Critical error: {str(e)}")
        messagebox.showerror("Error", f"Application error: {str(e)}")


if __name__ == "__main__":
    main()