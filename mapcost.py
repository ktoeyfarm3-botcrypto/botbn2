import pandas as pd
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import numbers, Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import logging
import traceback
import os


# ==== Modern GUI Styling ====
def configure_modern_style():
    """กำหนดธีมสีสวยงามและทันสมัย"""
    style = ttk.Style()

    # เลือก theme ที่สวยงาม
    try:
        style.theme_use('clam')  # ใช้ clam theme ที่ดูทันสมัย
    except:
        style.theme_use('default')

    # กำหนดสีสำหรับ theme
    colors = {
        'bg_primary': '#2c3e50',  # Dark blue-gray
        'bg_secondary': '#34495e',  # Lighter blue-gray
        'accent': '#3498db',  # Bright blue
        'success': '#27ae60',  # Green
        'warning': '#f39c12',  # Orange
        'danger': '#e74c3c',  # Red
        'light': '#ecf0f1',  # Light gray
        'white': '#ffffff',  # White
        'text_dark': '#2c3e50',  # Dark text
        'text_light': '#ffffff'  # Light text
    }

    # Configure styles
    style.configure('Title.TLabel', font=('Segoe UI', 16, 'bold'), foreground=colors['text_dark'])
    style.configure('Heading.TLabel', font=('Segoe UI', 12, 'bold'), foreground=colors['accent'])
    style.configure('Status.TLabel', font=('Segoe UI', 10), padding=(5, 2))

    # Button styles
    style.configure('Accent.TButton',
                    font=('Segoe UI', 9, 'bold'),
                    padding=(10, 5))

    style.configure('Success.TButton',
                    font=('Segoe UI', 9, 'bold'),
                    padding=(10, 5))

    style.configure('Warning.TButton',
                    font=('Segoe UI', 9, 'bold'),
                    padding=(10, 5))

    # Frame styles
    style.configure('Card.TFrame', relief='solid', borderwidth=1, padding=10)
    style.configure('Sidebar.TFrame', relief='solid', borderwidth=1, padding=8)

    # Notebook styles
    style.configure('Modern.TNotebook', tabposition='n')
    style.configure('Modern.TNotebook.Tab',
                    font=('Segoe UI', 10, 'bold'),
                    padding=(20, 8))

    return colors


# ==== ตั้งค่า Logging ====
logging.basicConfig(
    level=logging.DEBUG,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('hamada_calculator_debug.log', encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# ==== ระบบภาษา ====
LANGUAGES = {
    'th': {
        'app_title': 'Hamada Cost Calculator with Advanced Debug',
        'select_import_file': 'เลือกไฟล์ Import',
        'calculate': 'คำนวณ',
        'save_excel': 'บันทึก Excel',
        'save_excel_debug': 'บันทึก Excel + Debug',
        'check_base_file': 'ตรวจสอบไฟล์ Base',
        'export_base_file': 'Export ไฟล์ Base',
        'import_new_base': 'Import Base ใหม่',
        'clear_debug': 'เคลียร์ Debug',
        'debug_mode': 'เปิด Debug Mode',
        'no_file_selected': 'ยังไม่ได้เลือกไฟล์',
        'file_selected': 'ไฟล์',
        'base_file_ready': 'ไฟล์ Base: พร้อมใช้งาน',
        'base_file_not_found': 'ไฟล์ Base: ไม่พบ',
        'results_tab': 'ผลลัพธ์',
        'debug_tab': 'Debug Log',
        'statistics': 'สถิติ',
        'no_data': 'ยังไม่มีข้อมูล',
        'menu_name': 'MENU NAME',
        'qty': 'Qty',
        'material_cost': 'Material Cost',
        'total_cost': 'Total Cost',
        'grand_total': 'Grand Total',
        'matched_menus': 'พบเมนู',
        'not_found_menus': 'ไม่พบ',
        'total_amount': 'รวม',
        'baht': 'บาท',
        'error': 'ผิดพลาด',
        'warning': 'คำเตือน',
        'success': 'สำเร็จ',
        'please_select_file': 'กรุณาเลือกไฟล์ Import ก่อน',
        'calculation_complete': 'คำนวณเสร็จสิ้น - พบ',
        'items': 'รายการ',
        # Debug sheets
        'sheet_results': 'ผลลัพธ์',
        'sheet_not_found': 'เมนูที่ไม่พบ',
        'sheet_zero_qty': 'Qty เป็นศูนย์',
        'sheet_invalid_qty': 'Qty ไม่ถูกต้อง',
        'sheet_nan_cost': 'Material Cost NaN',
        'sheet_summary': 'สรุปการประมวลผล',
        'processing_summary': 'สรุปการประมวลผล',
        'total_items': 'รายการทั้งหมด',
        'matched_items': 'พบเมนูที่ตรงกัน',
        'not_found_items': 'ไม่พบเมนู',
        'zero_qty_items': 'Qty = 0',
        'invalid_qty_items': 'Qty ไม่ถูกต้อง',
        'nan_cost_items': 'Material Cost NaN',
        'processing_time': 'เวลาประมวลผล',
        'base_file': 'ไฟล์ Base',
        'import_file': 'ไฟล์ Import',
        'matched_menus_list': 'รายการเมนูที่พบ',
        'not_found_menus_list': 'รายการเมนูที่ไม่พบ',
        'not_sold_menus': 'รายการที่ไม่ได้ขาย',
        'not_sold_menus_list': 'รายการที่ไม่ได้ขาย (ไม่มีใน Import)',
        'row_number': 'แถวที่',
        'menu_name_col': 'ชื่อเมนู',
        'quantity': 'จำนวน',
        'original_qty': 'Qty เดิม',
        'invalid_qty': 'Qty ที่ไม่ถูกต้อง'
    },
    'en': {
        'app_title': 'Hamada Cost Calculator with Advanced Debug',
        'select_import_file': 'Select Import File',
        'calculate': 'Calculate',
        'save_excel': 'Save Excel',
        'save_excel_debug': 'Save Excel + Debug',
        'check_base_file': 'Check Base File',
        'export_base_file': 'Export Base File',
        'import_new_base': 'Import New Base',
        'clear_debug': 'Clear Debug',
        'debug_mode': 'Enable Debug Mode',
        'no_file_selected': 'No file selected',
        'file_selected': 'File',
        'base_file_ready': 'Base File: Ready',
        'base_file_not_found': 'Base File: Not Found',
        'results_tab': 'Results',
        'debug_tab': 'Debug Log',
        'statistics': 'Statistics',
        'no_data': 'No data available',
        'menu_name': 'MENU NAME',
        'qty': 'Qty',
        'material_cost': 'Material Cost',
        'total_cost': 'Total Cost',
        'grand_total': 'Grand Total',
        'matched_menus': 'Found',
        'not_found_menus': 'Not Found',
        'total_amount': 'Total',
        'baht': 'Baht',
        'error': 'Error',
        'warning': 'Warning',
        'success': 'Success',
        'please_select_file': 'Please select import file first',
        'calculation_complete': 'Calculation complete - found',
        'items': 'items',
        # Debug sheets
        'sheet_results': 'Results',
        'sheet_not_found': 'Menus Not Found',
        'sheet_zero_qty': 'Zero Quantity',
        'sheet_invalid_qty': 'Invalid Quantity',
        'sheet_nan_cost': 'Material Cost NaN',
        'sheet_summary': 'Processing Summary',
        'processing_summary': 'Processing Summary',
        'total_items': 'Total Items',
        'matched_items': 'Matched Items',
        'not_found_items': 'Not Found Items',
        'zero_qty_items': 'Zero Qty Items',
        'invalid_qty_items': 'Invalid Qty Items',
        'nan_cost_items': 'NaN Cost Items',
        'processing_time': 'Processing Time',
        'base_file': 'Base File',
        'import_file': 'Import File',
        'matched_menus_list': 'Matched Menus List',
        'not_found_menus_list': 'Not Found Menus List',
        'not_sold_menus': 'Not Sold Items',
        'not_sold_menus_list': 'Not Sold Items (Missing from Import)',
        'row_number': 'Row Number',
        'menu_name_col': 'Menu Name',
        'quantity': 'Quantity',
        'original_qty': 'Original Qty',
        'invalid_qty': 'Invalid Qty'
    },
    'jp': {
        'app_title': 'Hamada コスト計算機（高度デバッグ付き）',
        'select_import_file': 'インポートファイル選択',
        'calculate': '計算',
        'save_excel': 'Excel保存',
        'save_excel_debug': 'Excel + デバッグ保存',
        'check_base_file': 'ベースファイル確認',
        'export_base_file': 'ベースファイル出力',
        'import_new_base': '新ベース取込',
        'clear_debug': 'デバッグクリア',
        'debug_mode': 'デバッグモード有効',
        'no_file_selected': 'ファイル未選択',
        'file_selected': 'ファイル',
        'base_file_ready': 'ベースファイル: 準備完了',
        'base_file_not_found': 'ベースファイル: 見つかりません',
        'results_tab': '結果',
        'debug_tab': 'デバッグログ',
        'statistics': '統計',
        'no_data': 'データなし',
        'menu_name': 'メニュー名',
        'qty': '数量',
        'material_cost': '材料費',
        'total_cost': '合計金額',
        'grand_total': '総合計',
        'matched_menus': '一致',
        'not_found_menus': '未発見',
        'total_amount': '合計',
        'baht': 'バーツ',
        'error': 'エラー',
        'warning': '警告',
        'success': '成功',
        'please_select_file': '最初にインポートファイルを選択してください',
        'calculation_complete': '計算完了 - 発見',
        'items': '項目',
        # Debug sheets
        'sheet_results': '結果',
        'sheet_not_found': '未発見メニュー',
        'sheet_zero_qty': 'ゼロ数量',
        'sheet_invalid_qty': '無効数量',
        'sheet_nan_cost': '材料費NaN',
        'sheet_summary': '処理サマリー',
        'processing_summary': '処理サマリー',
        'total_items': '全項目',
        'matched_items': '一致項目',
        'not_found_items': '未発見項目',
        'zero_qty_items': 'ゼロ数量項目',
        'invalid_qty_items': '無効数量項目',
        'nan_cost_items': 'NaN費用項目',
        'processing_time': '処理時間',
        'base_file': 'ベースファイル',
        'import_file': 'インポートファイル',
        'matched_menus_list': '一致メニューリスト',
        'not_found_menus_list': '未発見メニューリスト',
        'not_sold_menus': '未販売項目',
        'not_sold_menus_list': '未販売項目（インポートに含まれない）',
        'row_number': '行番号',
        'menu_name_col': 'メニュー名',
        'quantity': '数量',
        'original_qty': '元の数量',
        'invalid_qty': '無効な数量'
    }
}

# ==== โหลดไฟล์ตั้งต้น ====
BASE_FILE = "Hamada Main cost.xlsx"


def load_base_file():
    """โหลดไฟล์ base พร้อม error handling"""
    try:
        if not os.path.exists(BASE_FILE):
            logger.error(f"ไม่พบไฟล์ base: {BASE_FILE}")
            return None

        df = pd.read_excel(BASE_FILE)
        logger.info(f"โหลดไฟล์ base สำเร็จ: {BASE_FILE}")
        logger.debug(f"Columns in base file: {df.columns.tolist()}")
        logger.debug(f"Shape: {df.shape}")

        # ตรวจสอบคอลัมน์ที่จำเป็น
        required_columns = ["MENU NAME", "Material Cost"]
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            logger.error(f"ไฟล์ base ขาดคอลัมน์: {missing_columns}")
            return None

        df = df.set_index("MENU NAME")
        logger.debug(f"Index set to MENU NAME. Total menus: {len(df)}")
        logger.debug(f"Sample menu names: {df.index[:5].tolist()}")

        return df
    except Exception as e:
        logger.error(f"Error loading base file: {str(e)}")
        logger.error(traceback.format_exc())
        return None


df_base = load_base_file()


class CostCalculatorApp:
    def __init__(self, root):
        self.root = root
        self.current_language = 'th'  # Default language

        # Configure modern styling
        self.colors = configure_modern_style()

        self.root.title(self.t('app_title'))
        self.root.geometry("1200x850")
        self.root.configure(bg=self.colors['light'])

        # Set minimum size
        self.root.minsize(1000, 700)

        self.import_file = None
        self.df_result = None
        self.debug_mode = tk.BooleanVar(value=False)
        self.debug_data = {
            'not_found_menus': [],
            'zero_qty_items': [],
            'invalid_qty_items': [],
            'nan_cost_items': [],
            'matched_menus': [],
            'not_sold_menus': [],
            'processing_summary': {}
        }

        # UI
        self._build_ui()

        logger.info("Application initialized")

    def t(self, key):
        """แปลข้อความตามภาษาที่เลือก"""
        return LANGUAGES.get(self.current_language, {}).get(key, key)

    def change_language(self, lang_code):
        """เปลี่ยนภาษา"""
        self.current_language = lang_code
        self.lang_var.set(lang_code)
        self.debug_log(f"เปลี่ยนภาษาเป็น: {lang_code}")
        self._refresh_ui_text()

    def _refresh_ui_text(self):
        """รีเฟรช text ใน UI ตามภาษาที่เลือก"""
        self.root.title(self.t('app_title'))

        # อัพเดท text ของปุ่ม
        for widget in self.root.winfo_children():
            self._update_widget_text(widget)

        # อัพเดท column headers with icons
        self.tree.heading("menu", text=f"🍽️ {self.t('menu_name')}")
        self.tree.heading("qty", text=f"📊 {self.t('qty')}")
        self.tree.heading("cost", text=f"💵 {self.t('material_cost')}")
        self.tree.heading("total", text=f"💰 {self.t('total_cost')}")

        # อัพเดท status labels
        if hasattr(self, 'status_label'):
            current_text = self.status_label.cget("text")
            if "ยังไม่ได้เลือกไฟล์" in current_text or "No file selected" in current_text or "ファイル未選択" in current_text:
                self.status_label.config(text=self.t('no_file_selected'))

    def _update_widget_text(self, widget):
        """อัพเดท text ของ widget แบบ recursive"""
        try:
            widget_type = widget.winfo_class()

            if widget_type == 'TButton':
                current_text = widget.cget("text")
                # Map button texts
                button_map = {
                    'เลือกไฟล์ Import': 'select_import_file',
                    'Select Import File': 'select_import_file',
                    'インポートファイル選択': 'select_import_file',
                    'คำนวณ': 'calculate',
                    'Calculate': 'calculate',
                    '計算': 'calculate',
                    'บันทึก Excel': 'save_excel',
                    'Save Excel': 'save_excel',
                    'Excel保存': 'save_excel',
                    'บันทึก Excel + Debug': 'save_excel_debug',
                    'Save Excel + Debug': 'save_excel_debug',
                    'Excel + デバッグ保存': 'save_excel_debug',
                    'ตรวจสอบไฟล์ Base': 'check_base_file',
                    'Check Base File': 'check_base_file',
                    'ベースファイル確認': 'check_base_file',
                    'Export ไฟล์ Base': 'export_base_file',
                    'Export Base File': 'export_base_file',
                    'ベースファイル出力': 'export_base_file',
                    'Import Base ใหม่': 'import_new_base',
                    'Import New Base': 'import_new_base',
                    '新ベース取込': 'import_new_base',
                    'เคลียร์ Debug': 'clear_debug',
                    'Clear Debug': 'clear_debug',
                    'デバッグクリア': 'clear_debug'
                }

                for text, key in button_map.items():
                    if text in current_text:
                        widget.config(text=self.t(key))
                        break

            elif widget_type == 'TCheckbutton':
                current_text = widget.cget("text")
                if any(x in current_text for x in ['เปิด Debug Mode', 'Enable Debug Mode', 'デバッグモード有効']):
                    widget.config(text=self.t('debug_mode'))

            elif widget_type == 'TLabelframe':
                current_text = widget.cget("text")
                if any(x in current_text for x in ['สถิติ', 'Statistics', '統計']):
                    widget.config(text=self.t('statistics'))

            # อัพเดท children widgets
            for child in widget.winfo_children():
                self._update_widget_text(child)

        except:
            pass

    def _build_ui(self):
        # Main container with gradient-like effect
        main_frame = ttk.Frame(self.root, style='Card.TFrame')
        main_frame.pack(fill=tk.BOTH, expand=True, padx=15, pady=15)

        # Header section
        header_frame = ttk.Frame(main_frame)
        header_frame.pack(fill=tk.X, pady=(0, 20))

        # App title
        title_label = ttk.Label(header_frame, text="🍱 Hamada Cost Calculator", style='Title.TLabel')
        title_label.pack(pady=(0, 10))

        # Subtitle
        subtitle = ttk.Label(header_frame, text="Advanced Cost Management System",
                             font=('Segoe UI', 11), foreground=self.colors['bg_secondary'])
        subtitle.pack()

        # Language and Debug section
        control_frame = ttk.LabelFrame(main_frame, text="🌐 Settings", padding=15, style='Card.TFrame')
        control_frame.pack(fill=tk.X, pady=(0, 15))

        # Language selection with modern styling
        lang_frame = ttk.Frame(control_frame)
        lang_frame.pack(fill=tk.X, pady=(0, 10))

        ttk.Label(lang_frame, text="🌍 Language:", font=('Segoe UI', 10, 'bold')).pack(side=tk.LEFT, padx=(0, 10))

        self.lang_var = tk.StringVar(value='th')
        lang_buttons = [
            ('🇹🇭 ไทย', 'th', 'Accent.TButton'),
            ('🇺🇸 English', 'en', 'Accent.TButton'),
            ('🇯🇵 日本語', 'jp', 'Accent.TButton')
        ]

        for text, code, style in lang_buttons:
            btn = ttk.Button(lang_frame, text=text, style=style,
                             command=lambda c=code: self.change_language(c))
            btn.pack(side=tk.LEFT, padx=3)

        # Debug mode with modern checkbox
        debug_frame = ttk.Frame(control_frame)
        debug_frame.pack(fill=tk.X)

        self.debug_checkbox = ttk.Checkbutton(debug_frame, text="🔍 " + self.t('debug_mode'),
                                              variable=self.debug_mode,
                                              style='Modern.TCheckbutton')
        self.debug_checkbox.pack(side=tk.LEFT)

        # Action buttons section
        action_frame = ttk.LabelFrame(main_frame, text="🚀 Actions", padding=15, style='Card.TFrame')
        action_frame.pack(fill=tk.X, pady=(0, 15))

        # Primary actions (row 1)
        btn_frame1 = ttk.Frame(action_frame)
        btn_frame1.pack(fill=tk.X, pady=(0, 8))

        primary_buttons = [
            (self.t('select_import_file'), self.load_file, '📁', 'Accent.TButton'),
            (self.t('calculate'), self.calculate, '⚡', 'Success.TButton'),
            (self.t('save_excel'), self.export_excel, '💾', 'Success.TButton'),
            (self.t('save_excel_debug'), self.export_excel_with_debug, '🔍💾', 'Warning.TButton')
        ]

        for text, command, icon, style in primary_buttons:
            btn = ttk.Button(btn_frame1, text=f"{icon} {text}", command=command, style=style)
            btn.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)

        # Secondary actions (row 2)
        btn_frame2 = ttk.Frame(action_frame)
        btn_frame2.pack(fill=tk.X)

        secondary_buttons = [
            (self.t('check_base_file'), self.check_base_file, '🔍', 'Accent.TButton'),
            (self.t('export_base_file'), self.export_base_file, '📤', 'Accent.TButton'),
            (self.t('import_new_base'), self.import_new_base, '📥', 'Warning.TButton'),
            (self.t('clear_debug'), self.clear_debug, '🗑️', 'Warning.TButton')
        ]

        for text, command, icon, style in secondary_buttons:
            btn = ttk.Button(btn_frame2, text=f"{icon} {text}", command=command, style=style)
            btn.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)

        # Status section
        status_frame = ttk.LabelFrame(main_frame, text="📊 Status", padding=10, style='Card.TFrame')
        status_frame.pack(fill=tk.X, pady=(0, 15))

        # File status with icons
        file_status_frame = ttk.Frame(status_frame)
        file_status_frame.pack(fill=tk.X, pady=2)

        ttk.Label(file_status_frame, text="📂", font=('Segoe UI', 12)).pack(side=tk.LEFT, padx=(0, 5))
        self.status_label = ttk.Label(file_status_frame, text=self.t('no_file_selected'),
                                      style='Status.TLabel', foreground=self.colors['danger'])
        self.status_label.pack(side=tk.LEFT)

        # Base file status
        base_status_frame = ttk.Frame(status_frame)
        base_status_frame.pack(fill=tk.X, pady=2)

        ttk.Label(base_status_frame, text="🗂️", font=('Segoe UI', 12)).pack(side=tk.LEFT, padx=(0, 5))
        base_status = self.t(
            'base_file_ready') if df_base is not None else f"{self.t('base_file_not_found')} {BASE_FILE}"
        color = self.colors['success'] if df_base is not None else self.colors['danger']
        self.base_status_label = ttk.Label(base_status_frame, text=base_status,
                                           style='Status.TLabel', foreground=color)
        self.base_status_label.pack(side=tk.LEFT)

        # Main content area
        content_frame = ttk.Frame(main_frame, style='Card.TFrame')
        content_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 15))

        # สร้าง Notebook สำหรับแท็บ
        self.notebook = ttk.Notebook(content_frame, style='Modern.TNotebook')
        self.notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # แท็บผลลัพธ์
        result_frame = ttk.Frame(self.notebook)
        self.notebook.add(result_frame, text=f"📋 {self.t('results_tab')}")

        # Modern table styling
        table_frame = ttk.Frame(result_frame, padding=10)
        table_frame.pack(fill=tk.BOTH, expand=True)

        # Table header
        table_header = ttk.Label(table_frame, text="💰 Calculation Results",
                                 style='Heading.TLabel')
        table_header.pack(pady=(0, 10))

        # ตารางผลลัพธ์ with modern styling
        self.tree = ttk.Treeview(table_frame, columns=("menu", "qty", "cost", "total"),
                                 show="headings", height=16, style='Modern.Treeview')

        # Configure headers with icons
        self.tree.heading("menu", text=f"🍽️ {self.t('menu_name')}")
        self.tree.heading("qty", text=f"📊 {self.t('qty')}")
        self.tree.heading("cost", text=f"💵 {self.t('material_cost')}")
        self.tree.heading("total", text=f"💰 {self.t('total_cost')}")

        self.tree.column("menu", width=400, anchor="w")
        self.tree.column("qty", width=100, anchor="center")
        self.tree.column("cost", width=150, anchor="e")
        self.tree.column("total", width=150, anchor="e")

        # Modern scrollbar
        tree_scroll = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=tree_scroll.set)

        tree_scroll.pack(side=tk.RIGHT, fill=tk.Y, padx=(5, 0))
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # แท็บ Debug
        debug_frame = ttk.Frame(self.notebook)
        self.notebook.add(debug_frame, text=f"🔍 {self.t('debug_tab')}")

        # Debug header
        debug_header_frame = ttk.Frame(debug_frame, padding=10)
        debug_header_frame.pack(fill=tk.X)

        debug_title = ttk.Label(debug_header_frame, text="🔧 Debug Console", style='Heading.TLabel')
        debug_title.pack()

        # Debug text area with modern styling
        debug_text_frame = ttk.Frame(debug_frame, padding=(10, 0, 10, 10))
        debug_text_frame.pack(fill=tk.BOTH, expand=True)

        self.debug_text = scrolledtext.ScrolledText(debug_text_frame, wrap=tk.WORD, height=18,
                                                    font=('Consolas', 10),
                                                    bg='#1e1e1e', fg='#ffffff',
                                                    insertbackground='#ffffff',
                                                    selectbackground='#3498db',
                                                    relief='flat', borderwidth=0)
        self.debug_text.pack(fill=tk.BOTH, expand=True)

        # Statistics section with modern card design
        self.stats_frame = ttk.LabelFrame(main_frame, text="📈 Statistics",
                                          padding=15, style='Card.TFrame')
        self.stats_frame.pack(fill=tk.X)

        self.stats_label = ttk.Label(self.stats_frame, text=self.t('no_data'),
                                     style='Status.TLabel',
                                     font=('Segoe UI', 11, 'bold'))
        # Configure modern treeview styling
        self._configure_treeview_style()

    def _configure_treeview_style(self):
        """กำหนดสไตล์ modern สำหรับ treeview"""
        style = ttk.Style()

        # Configure treeview
        style.configure('Modern.Treeview',
                        background='#ffffff',
                        foreground='#2c3e50',
                        fieldbackground='#ffffff',
                        font=('Segoe UI', 10))

        style.configure('Modern.Treeview.Heading',
                        background='#3498db',
                        foreground='#ffffff',
                        font=('Segoe UI', 10, 'bold'),
                        relief='flat')

        # Alternate row colors
        style.map('Modern.Treeview',
                  background=[('selected', '#3498db')],
                  foreground=[('selected', '#ffffff')])

        # Apply style to tree
        self.tree.configure(style='Modern.Treeview')

    def _create_modern_button(self, parent, text, command, icon="", btn_type="primary"):
        """สร้างปุ่มสไตล์ modern"""
        styles = {
            'primary': 'Accent.TButton',
            'success': 'Success.TButton',
            'warning': 'Warning.TButton'
        }

        full_text = f"{icon} {text}" if icon else text
        return ttk.Button(parent, text=full_text, command=command,
                          style=styles.get(btn_type, 'Accent.TButton'))

        # Configure modern treeview styling
        self._configure_treeview_style()

    def debug_log(self, message, level="INFO"):
        """เพิ่มข้อความใน debug log"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        log_message = f"[{timestamp}] {level}: {message}\n"

        self.debug_text.insert(tk.END, log_message)
        self.debug_text.see(tk.END)

        # Log ลงไฟล์ด้วย
        if level == "ERROR":
            logger.error(message)
        elif level == "WARNING":
            logger.warning(message)
        else:
            logger.info(message)

    def clear_debug(self):
        """เคลียร์ debug log"""
        self.debug_text.delete(1.0, tk.END)
        self.debug_log("Debug log cleared")
        # เคลียร์ debug data
        self.debug_data = {
            'not_found_menus': [],
            'zero_qty_items': [],
            'invalid_qty_items': [],
            'nan_cost_items': [],
            'matched_menus': [],
            'not_sold_menus': [],
            'processing_summary': {}
        }

    def export_base_file(self):
        """Export ไฟล์ base ปัจจุบัน"""
        global df_base
        if df_base is None:
            messagebox.showerror(self.t('error'), f"ไม่มีไฟล์ base ให้ export")
            return

        file_path = filedialog.asksaveasfilename(
            title=f"{self.t('export_base_file')}",
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")]
        )
        if not file_path:
            return

        try:
            # Reset index เพื่อให้ MENU NAME กลับมาเป็นคอลัมน์
            df_export = df_base.reset_index()
            df_export.to_excel(file_path, index=False)

            self.debug_log(f"Export ไฟล์ base สำเร็จ: {os.path.basename(file_path)}")
            messagebox.showinfo(self.t('success'), f"Export ไฟล์ base แล้ว:\n{file_path}")

        except Exception as e:
            error_msg = f"ไม่สามารถ export ไฟล์ base ได้: {str(e)}"
            self.debug_log(error_msg, "ERROR")
            messagebox.showerror(self.t('error'), error_msg)

    def import_new_base(self):
        """Import ไฟล์ base ใหม่"""
        file_path = filedialog.askopenfilename(
            title=f"{self.t('import_new_base')}",
            filetypes=[("Excel Files", "*.xlsx *.xls")]
        )
        if not file_path:
            return

        try:
            # อ่านไฟล์ใหม่
            df_new = pd.read_excel(file_path)
            self.debug_log(f"อ่านไฟล์ base ใหม่: {os.path.basename(file_path)}")
            self.debug_log(f"Columns: {df_new.columns.tolist()}")
            self.debug_log(f"Shape: {df_new.shape}")

            # ตรวจสอบคอลัมน์ที่จำเป็น
            required_columns = ["MENU NAME", "Material Cost"]
            missing_columns = [col for col in required_columns if col not in df_new.columns]
            if missing_columns:
                error_msg = f"ไฟล์ขาดคอลัมน์ที่จำเป็น: {missing_columns}"
                self.debug_log(error_msg, "ERROR")
                messagebox.showerror(self.t('error'), error_msg)
                return

            # ตรวจสอบข้อมูลซ้ำใน MENU NAME
            duplicates = df_new[df_new["MENU NAME"].duplicated()]
            if not duplicates.empty:
                self.debug_log(f"พบข้อมูลซ้ำ {len(duplicates)} รายการ", "WARNING")
                for _, dup_row in duplicates.iterrows():
                    self.debug_log(f"  ซ้ำ: {dup_row['MENU NAME']}", "WARNING")

            # Set index และอัพเดท global variable
            global df_base
            df_base = df_new.set_index("MENU NAME")

            # อัพเดท status
            self.base_status_label.config(text=f"🗂️ {self.t('base_file')}: {os.path.basename(file_path)}",
                                          foreground=self.colors['success'])

            success_msg = f"อัพเดทไฟล์ base สำเร็จ - มี {len(df_base)} เมนู"
            self.debug_log(success_msg)
            messagebox.showinfo(self.t('success'), success_msg)

        except Exception as e:
            error_msg = f"ไม่สามารถ import ไฟล์ base ใหม่ได้: {str(e)}"
            self.debug_log(error_msg, "ERROR")
            self.debug_log(traceback.format_exc(), "ERROR")
            messagebox.showerror(self.t('error'), error_msg)

    def check_base_file(self):
        """ตรวจสอบและแสดงข้อมูลไฟล์ base"""
        if df_base is None:
            messagebox.showerror(self.t('error'), f"ไม่สามารถโหลดไฟล์ {BASE_FILE}")
            return

        self.debug_log("=== ตรวจสอบไฟล์ Base ===")
        self.debug_log(f"ไฟล์: {BASE_FILE}")
        self.debug_log(f"จำนวนเมนู: {len(df_base)}")
        self.debug_log(f"คอลัมน์: {df_base.columns.tolist()}")

        # สถิติราคา
        try:
            cost_stats = df_base["Material Cost"].describe()
            self.debug_log("สถิติราคา Material Cost:")
            self.debug_log(f"  ต่ำสุด: {cost_stats['min']:.2f}")
            self.debug_log(f"  สูงสุด: {cost_stats['max']:.2f}")
            self.debug_log(f"  เฉลี่ย: {cost_stats['mean']:.2f}")
            self.debug_log(f"  จำนวนรายการ: {cost_stats['count']}")
        except Exception as e:
            self.debug_log(f"Error คำนวณสถิติ: {str(e)}", "ERROR")

        # ตัวอย่างเมนู
        self.debug_log("ตัวอย่างเมนู 10 รายการแรก:")
        for i, (menu_name, row) in enumerate(df_base.head(10).iterrows()):
            material_cost = row.get("Material Cost", 0)
            self.debug_log(f"  {i + 1}. {menu_name}: {material_cost:.2f}")

    def load_file(self):
        file_path = filedialog.askopenfilename(
            title=self.t('select_import_file'),
            filetypes=[("Excel Files", "*.xlsx *.xls")]
        )
        if not file_path:
            return

        self.import_file = file_path
        self.status_label.config(text=f"📂 {self.t('file_selected')}: {os.path.basename(file_path)}",
                                 foreground=self.colors['success'])

        self.debug_log(f"เลือกไฟล์: {file_path}")

        # ตรวจสอบไฟล์ที่เลือก
        try:
            df_preview = pd.read_excel(file_path)
            self.debug_log(f"ไฟล์ import มี {len(df_preview)} แถว, {len(df_preview.columns)} คอลัมน์")
            self.debug_log(f"คอลัมน์: {df_preview.columns.tolist()}")

            # ตรวจสอบคอลัมน์ที่จำเป็น
            required_cols = ["MENU NAME", "Qty"]
            missing_cols = [col for col in required_cols if col not in df_preview.columns]
            if missing_cols:
                self.debug_log(f"คำเตือน: ไฟล์ขาดคอลัมน์ {missing_cols}", "WARNING")

            # แสดงตัวอย่างข้อมูล
            self.debug_log("ตัวอย่างข้อมูล 5 แถวแรก:")
            for i, row in df_preview.head(5).iterrows():
                menu = row.get("MENU NAME", "N/A")
                qty = row.get("Qty", "N/A")
                self.debug_log(f"  แถว {i + 1}: {menu} = {qty}")

        except Exception as e:
            self.debug_log(f"Error ตรวจสอบไฟล์: {str(e)}", "ERROR")

        messagebox.showinfo("นำเข้าไฟล์", f"เลือกไฟล์แล้ว:\n{os.path.basename(file_path)}")

    def calculate(self):
        if df_base is None:
            messagebox.showerror(self.t('error'), f"ไฟล์ base ({BASE_FILE}) ไม่สามารถโหลดได้")
            return

        if not self.import_file:
            messagebox.showwarning(self.t('warning'), self.t('please_select_file'))
            return

        self.debug_log("=== เริ่มการคำนวณ ===")

        # เคลียร์ debug data เก่า
        self.debug_data = {
            'not_found_menus': [],
            'zero_qty_items': [],
            'invalid_qty_items': [],
            'nan_cost_items': [],
            'matched_menus': [],
            'not_sold_menus': [],
            'processing_summary': {}
        }

        try:
            df_import = pd.read_excel(self.import_file)
            self.debug_log(f"อ่านไฟล์ import สำเร็จ: {len(df_import)} แถว")
        except Exception as e:
            error_msg = f"ไม่สามารถอ่านไฟล์ได้: {str(e)}"
            self.debug_log(error_msg, "ERROR")
            messagebox.showerror(self.t('error'), error_msg)
            return

        results = []
        matched_count = 0

        self.debug_log(f"เริ่มประมวลผล {len(df_import)} รายการ")

        for idx, row in df_import.iterrows():
            menu = row.get("MENU NAME")
            qty = row.get("Qty", 0)

            if self.debug_mode.get():
                self.debug_log(f"ประมวลผลแถว {idx + 1}: {menu} = {qty}")

            # ตรวจสอบ menu name
            if pd.isna(menu) or menu == "":
                self.debug_log(f"แถว {idx + 1}: ชื่อเมนูว่าง", "WARNING")
                continue

            # ตรวจสอบ quantity
            original_qty = qty
            try:
                qty = float(qty) if not pd.isna(qty) else 0
                if qty == 0:
                    self.debug_data['zero_qty_items'].append({
                        'row': idx + 1,
                        'menu': menu,
                        'original_qty': original_qty
                    })
                    if self.debug_mode.get():
                        self.debug_log(f"แถว {idx + 1}: Qty = 0 สำหรับ {menu}", "WARNING")
            except (ValueError, TypeError):
                self.debug_data['invalid_qty_items'].append({
                    'row': idx + 1,
                    'menu': menu,
                    'invalid_qty': original_qty
                })
                self.debug_log(f"แถว {idx + 1}: Qty ไม่ใช่ตัวเลข ({original_qty}) สำหรับ {menu}", "WARNING")
                qty = 0

            # ค้นหาใน base file
            if menu in df_base.index:
                try:
                    material_cost = df_base.at[menu, "Material Cost"]

                    # ตรวจสอบ material cost
                    if pd.isna(material_cost):
                        self.debug_data['nan_cost_items'].append({
                            'menu': menu,
                            'row': idx + 1
                        })
                        self.debug_log(f"แถว {idx + 1}: Material Cost เป็น NaN สำหรับ {menu}", "WARNING")
                        material_cost = 0
                    else:
                        material_cost = float(material_cost)

                    total_cost = qty * material_cost
                    results.append([menu, qty, material_cost, total_cost])

                    # เก็บรายการที่พบ
                    self.debug_data['matched_menus'].append({
                        'row': idx + 1,
                        'menu': menu,
                        'qty': qty,
                        'material_cost': material_cost,
                        'total_cost': total_cost
                    })

                    matched_count += 1

                    if self.debug_mode.get():
                        self.debug_log(f"  ✓ พบเมนู: {material_cost:.2f} x {qty} = {total_cost:.2f}")

                except Exception as e:
                    self.debug_log(f"Error คำนวณ {menu}: {str(e)}", "ERROR")
            else:
                self.debug_data['not_found_menus'].append({
                    'row': idx + 1,
                    'menu': menu,
                    'qty': qty
                })
                if self.debug_mode.get():
                    self.debug_log(f"  ✗ ไม่พบเมนู: {menu}", "WARNING")

        # หาเมนูที่ไม่ได้ขาย (อยู่ใน base แต่ไม่มีใน import)
        import_menus = set(df_import["MENU NAME"].dropna().tolist())
        base_menus = set(df_base.index.tolist())
        not_sold_menus = base_menus - import_menus

        self.debug_data['not_sold_menus'] = []
        for menu in not_sold_menus:
            self.debug_data['not_sold_menus'].append({
                'menu': menu,
                'material_cost': df_base.at[menu, "Material Cost"] if menu in df_base.index else 0
            })

        # บันทึกสรุปการประมวลผล
        self.debug_data['processing_summary'] = {
            'total_rows': len(df_import),
            'matched_count': matched_count,
            'not_found_count': len(self.debug_data['not_found_menus']),
            'zero_qty_count': len(self.debug_data['zero_qty_items']),
            'invalid_qty_count': len(self.debug_data['invalid_qty_items']),
            'nan_cost_count': len(self.debug_data['nan_cost_items']),
            'not_sold_count': len(self.debug_data['not_sold_menus']),
            'timestamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }

        # สรุปผลการประมวลผล
        self.debug_log(f"=== สรุปผลการประมวลผล ===")
        summary = self.debug_data['processing_summary']
        self.debug_log(f"รายการทั้งหมด: {summary['total_rows']}")
        self.debug_log(f"พบเมนูที่ตรงกัน: {summary['matched_count']}")
        self.debug_log(f"ไม่พบเมนู: {summary['not_found_count']}")
        self.debug_log(f"Qty = 0: {summary['zero_qty_count']}")
        self.debug_log(f"Qty ไม่ถูกต้อง: {summary['invalid_qty_count']}")
        self.debug_log(f"Material Cost NaN: {summary['nan_cost_count']}")
        self.debug_log(f"รายการที่ไม่ได้ขาย: {summary['not_sold_count']}")

        # สร้าง DataFrame ผลลัพธ์
        self.df_result = pd.DataFrame(results, columns=["MENU NAME", "Qty", "Material Cost", "Total Cost"])

        if self.df_result.empty:
            error_msg = "ไม่มีชื่อเมนูที่ตรงกับไฟล์ตั้งต้น"
            self.debug_log(error_msg, "WARNING")
            messagebox.showwarning("ไม่พบข้อมูล", error_msg)
            return

        # คำนวณ Grand Total
        grand_total = self.df_result["Total Cost"].sum()
        self.debug_log(f"Grand Total: {grand_total:.2f}")

        # เพิ่มแถว Grand Total
        grand_total_row = pd.DataFrame([[self.t('grand_total'), "", "", grand_total]],
                                       columns=["MENU NAME", "Qty", "Material Cost", "Total Cost"])
        self.df_result = pd.concat([self.df_result, grand_total_row], ignore_index=True)

        # แสดงในตาราง
        self._update_result_table()

        # อัพเดทสถิติ
        self._update_statistics(matched_count, summary['not_found_count'], grand_total)

        success_msg = f"{self.t('calculation_complete')} {matched_count} {self.t('items')}"
        self.debug_log(success_msg)
        messagebox.showinfo(self.t('success'), success_msg)

    def export_excel_with_debug(self):
        """บันทึก Excel พร้อม Debug sheets"""
        if self.df_result is None or self.df_result.empty:
            messagebox.showwarning(self.t('warning'), "กรุณาคำนวณก่อนบันทึก")
            return

        file_path = filedialog.asksaveasfilename(
            title=self.t('save_excel_debug'),
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")]
        )
        if not file_path:
            return

        self.debug_log("=== เริ่มการบันทึกไฟล์พร้อม Debug ===")

        try:
            # สร้าง workbook ใหม่
            wb = Workbook()

            # Sheet 1: ผลลัพธ์หลัก
            ws_main = wb.active
            ws_main.title = self.t('sheet_results')

            df_export = self.df_result.copy()
            df_export["Material Cost"] = pd.to_numeric(df_export["Material Cost"], errors="coerce").round(2)
            df_export["Total Cost"] = pd.to_numeric(df_export["Total Cost"], errors="coerce").round(2)

            # เขียนข้อมูลลง sheet
            for r in dataframe_to_rows(df_export, index=False, header=True):
                ws_main.append(r)

            # จัดรูปแบบ header
            for cell in ws_main[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

            # จัดรูปแบบตัวเลข
            for row in ws_main.iter_rows(min_row=2):
                if len(row) >= 3 and isinstance(row[2].value, (int, float)):  # Material Cost
                    row[2].number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
                if len(row) >= 4 and isinstance(row[3].value, (int, float)):  # Total Cost
                    row[3].number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1

            # ไฮไลท์แถว Grand Total
            for row in ws_main.iter_rows(min_row=2):
                if row[0].value == self.t('grand_total'):
                    for cell in row:
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")

            # Sheet 2: เมนูที่ไม่พบ
            if self.debug_data['not_found_menus']:
                ws_not_found = wb.create_sheet(self.t('sheet_not_found'))
                ws_not_found.append([self.t('row_number'), self.t('menu_name_col'), self.t('quantity')])

                # Header formatting
                for cell in ws_not_found[1]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

                for item in self.debug_data['not_found_menus']:
                    ws_not_found.append([item['row'], item['menu'], item['qty']])

                self.debug_log(f"เขียน sheet เมนูที่ไม่พบ: {len(self.debug_data['not_found_menus'])} รายการ")

            # Sheet 3: รายการ Qty = 0
            if self.debug_data['zero_qty_items']:
                ws_zero_qty = wb.create_sheet(self.t('sheet_zero_qty'))
                ws_zero_qty.append([self.t('row_number'), self.t('menu_name_col'), self.t('original_qty')])

                # Header formatting
                for cell in ws_zero_qty[1]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

                for item in self.debug_data['zero_qty_items']:
                    ws_zero_qty.append([item['row'], item['menu'], item['original_qty']])

                self.debug_log(f"เขียน sheet Qty เป็นศูนย์: {len(self.debug_data['zero_qty_items'])} รายการ")

            # Sheet 4: รายการ Qty ไม่ถูกต้อง
            if self.debug_data['invalid_qty_items']:
                ws_invalid_qty = wb.create_sheet(self.t('sheet_invalid_qty'))
                ws_invalid_qty.append([self.t('row_number'), self.t('menu_name_col'), self.t('invalid_qty')])

                # Header formatting
                for cell in ws_invalid_qty[1]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

                for item in self.debug_data['invalid_qty_items']:
                    ws_invalid_qty.append([item['row'], item['menu'], str(item['invalid_qty'])])

                self.debug_log(f"เขียน sheet Qty ไม่ถูกต้อง: {len(self.debug_data['invalid_qty_items'])} รายการ")

            # Sheet 5: รายการ Material Cost = NaN
            if self.debug_data['nan_cost_items']:
                ws_nan_cost = wb.create_sheet(self.t('sheet_nan_cost'))
                ws_nan_cost.append([self.t('row_number'), self.t('menu_name_col')])

                # Header formatting
                for cell in ws_nan_cost[1]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

                for item in self.debug_data['nan_cost_items']:
                    ws_nan_cost.append([item['row'], item['menu']])

                self.debug_log(f"เขียน sheet Material Cost NaN: {len(self.debug_data['nan_cost_items'])} รายการ")

            # Sheet 6: สรุปการประมวลผล (แบบละเอียด)
            ws_summary = wb.create_sheet(self.t('sheet_summary'))

            # ส่วนสรุป
            summary_data = [
                [self.t('processing_summary'), ""],
                ["", ""],
                [self.t('total_items'), self.debug_data['processing_summary']['total_rows']],
                [self.t('matched_items'), self.debug_data['processing_summary']['matched_count']],
                [self.t('not_found_items'), self.debug_data['processing_summary']['not_found_count']],
                [self.t('zero_qty_items'), self.debug_data['processing_summary']['zero_qty_count']],
                [self.t('invalid_qty_items'), self.debug_data['processing_summary']['invalid_qty_count']],
                [self.t('nan_cost_items'), self.debug_data['processing_summary']['nan_cost_count']],
                [self.t('not_sold_menus'), self.debug_data['processing_summary']['not_sold_count']],
                ["", ""],
                [self.t('processing_time'), self.debug_data['processing_summary']['timestamp']],
                [self.t('base_file'), BASE_FILE],
                [self.t('import_file'), os.path.basename(self.import_file) if self.import_file else "N/A"],
                ["", ""],
                ["", ""]
            ]

            # เพิ่มรายการเมนูที่พบ
            if self.debug_data['matched_menus']:
                summary_data.append([self.t('matched_menus_list'), ""])
                summary_data.append([self.t('row_number'), f"{self.t('menu_name_col')} ({self.t('qty')})"])
                for item in self.debug_data['matched_menus']:
                    summary_data.append([item['row'], f"{item['menu']} ({item['qty']})"])
                summary_data.append(["", ""])

            # เพิ่มรายการเมนูที่ไม่พบ
            if self.debug_data['not_found_menus']:
                summary_data.append([self.t('not_found_menus_list'), ""])
                summary_data.append([self.t('row_number'), f"{self.t('menu_name_col')} ({self.t('qty')})"])
                for item in self.debug_data['not_found_menus']:
                    summary_data.append([item['row'], f"{item['menu']} ({item['qty']})"])
                summary_data.append(["", ""])

            # เพิ่มรายการที่ไม่ได้ขาย
            if self.debug_data['not_sold_menus']:
                summary_data.append([self.t('not_sold_menus_list'), ""])
                summary_data.append([self.t('menu_name_col'), self.t('material_cost')])
                for item in self.debug_data['not_sold_menus']:
                    summary_data.append([item['menu'], f"{item['material_cost']:.2f}"])

            for row_data in summary_data:
                ws_summary.append(row_data)

            # Format summary sheet
            # Header สรุป
            ws_summary[1][0].font = Font(bold=True, size=14)
            ws_summary[1][0].fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")

            # Headers รายการเมนู
            for row_idx, row in enumerate(ws_summary.iter_rows(), 1):
                if row[0].value in [self.t('matched_menus_list'), self.t('not_found_menus_list')]:
                    row[0].font = Font(bold=True, size=12)
                    row[0].fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
                elif row[0].value == self.t('not_sold_menus_list'):
                    row[0].font = Font(bold=True, size=12)
                    row[0].fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")

                # ทำสีแดงสำหรับรายการที่ไม่ได้ขาย
                if row_idx > 1:  # ข้าม header row
                    prev_row = ws_summary[row_idx - 1]
                    if prev_row[0].value == self.t('not_sold_menus_list') or (
                            row_idx > 2 and any(ws_summary[i][0].value == self.t('not_sold_menus_list')
                                                for i in range(max(1, row_idx - 20), row_idx)) and
                            row[0].value != "" and row[0].value not in [self.t('menu_name_col')]
                    ):
                        # ทำสีแดงสำหรับชื่อเมนูที่ไม่ได้ขาย
                        if row[0].value != self.t('menu_name_col') and row[0].value != "":
                            row[0].font = Font(color="FF0000", bold=True)  # สีแดง

            # ปรับความกว้างคอลัมน์
            for ws in wb.worksheets:
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 60)
                    ws.column_dimensions[column_letter].width = adjusted_width

            wb.save(file_path)

            success_msg = f"บันทึกไฟล์พร้อม Debug สำเร็จ: {os.path.basename(file_path)}"
            self.debug_log(success_msg)

            # แสดงสรุป sheets ที่สร้าง
            sheet_info = f"สร้าง {len(wb.worksheets)} sheets:\n"
            for ws in wb.worksheets:
                sheet_info += f"- {ws.title}\n"

            messagebox.showinfo(self.t('success'), f"บันทึกผลลัพธ์พร้อม Debug แล้ว:\n{file_path}\n\n{sheet_info}")

        except Exception as e:
            error_msg = f"ไม่สามารถบันทึกไฟล์ได้: {str(e)}"
            self.debug_log(error_msg, "ERROR")
            self.debug_log(traceback.format_exc(), "ERROR")
            messagebox.showerror(self.t('error'), error_msg)

    def _update_result_table(self):
        """อัพเดทตารางผลลัพธ์"""
        # เคลียร์ตารางเก่า
        for row in self.tree.get_children():
            self.tree.delete(row)

        # เพิ่มข้อมูลใหม่
        for _, r in self.df_result.iterrows():
            qty_display = int(r["Qty"]) if r["Qty"] != "" else ""
            cost_display = f"{r['Material Cost']:.2f}" if r["Material Cost"] != "" else ""
            total_display = f"{r['Total Cost']:.2f}" if r["Total Cost"] != "" else ""

            # ไฮไลท์แถว Grand Total
            tags = ("grand_total",) if r["MENU NAME"] == self.t('grand_total') else ()

            self.tree.insert("", tk.END, values=(r["MENU NAME"], qty_display, cost_display, total_display), tags=tags)

        # จัดรูปแบบแถว Grand Total with modern styling
        self.tree.tag_configure("grand_total",
                                background="#3498db",
                                foreground="#ffffff",
                                font=("Segoe UI", 10, "bold"))

    def _update_statistics(self, matched, not_found, grand_total):
        """อัพเดทสถิติ"""
        stats_text = f"✅ {self.t('matched_menus')}: {matched} | ❌ {self.t('not_found_menus')}: {not_found} | 💰 {self.t('total_amount')}: {grand_total:,.2f} {self.t('baht')}"
        self.stats_label.config(text=stats_text)

    def export_excel(self):
        """บันทึก Excel แบบปกติ (ไม่มี debug sheets)"""
        if self.df_result is None or self.df_result.empty:
            messagebox.showwarning(self.t('warning'), "กรุณาคำนวณก่อนบันทึก")
            return

        file_path = filedialog.asksaveasfilename(
            title=self.t('save_excel'),
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")]
        )
        if not file_path:
            return

        self.debug_log("=== เริ่มการบันทึกไฟล์ (แบบปกติ) ===")

        try:
            # ปัดเศษทศนิยม 2 ตำแหน่ง
            df_export = self.df_result.copy()

            # แปล headers ตามภาษาที่เลือก
            df_export.columns = [self.t('menu_name'), self.t('qty'), self.t('material_cost'), self.t('total_cost')]

            if self.t('material_cost') in df_export.columns:
                df_export[self.t('material_cost')] = pd.to_numeric(df_export[self.t('material_cost')],
                                                                   errors="coerce").round(2)
            if self.t('total_cost') in df_export.columns:
                df_export[self.t('total_cost')] = pd.to_numeric(df_export[self.t('total_cost')], errors="coerce").round(
                    2)

            # เขียน DataFrame ออกเป็น Excel
            df_export.to_excel(file_path, index=False)

            # เปิดไฟล์อีกครั้งเพื่อจัดรูปแบบตัวเลข
            wb = load_workbook(file_path)
            ws = wb.active

            # หาคอลัมน์ที่เกี่ยวข้อง
            col_cost = None
            col_total = None
            for col_idx, cell in enumerate(ws[1], 1):
                if cell.value == self.t('material_cost'):
                    col_cost = col_idx
                elif cell.value == self.t('total_cost'):
                    col_total = col_idx

            # จัดรูปแบบตัวเลขเป็น #,##0.00
            if col_cost:
                for row in ws.iter_rows(min_row=2, min_col=col_cost, max_col=col_cost):
                    for cell in row:
                        if isinstance(cell.value, (int, float)):
                            cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1

            if col_total:
                for row in ws.iter_rows(min_row=2, min_col=col_total, max_col=col_total):
                    for cell in row:
                        if isinstance(cell.value, (int, float)):
                            cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1

            wb.save(file_path)

            success_msg = f"บันทึกไฟล์สำเร็จ: {os.path.basename(file_path)}"
            self.debug_log(success_msg)
            messagebox.showinfo(self.t('success'), f"บันทึกผลลัพธ์แล้วที่:\n{file_path}")

        except Exception as e:
            error_msg = f"ไม่สามารถบันทึกไฟล์ได้: {str(e)}"
            self.debug_log(error_msg, "ERROR")
            self.debug_log(traceback.format_exc(), "ERROR")
            messagebox.showerror(self.t('error'), error_msg)


def main():
    """ฟังก์ชันหลักพร้อม error handling"""
    try:
        root = tk.Tk()

        # Set modern window icon and styling
        try:
            root.iconphoto(True, tk.PhotoImage(data=''))  # Modern icon placeholder
        except:
            pass

        app = CostCalculatorApp(root)
        logger.info("Starting application")
        root.mainloop()
    except Exception as e:
        logger.error(f"Critical error in main: {str(e)}")
        logger.error(traceback.format_exc())
        messagebox.showerror("ผิดพลาดร้ายแรง", f"เกิดข้อผิดพลาดร้ายแรง:\n{str(e)}")


if __name__ == "__main__":
    main()