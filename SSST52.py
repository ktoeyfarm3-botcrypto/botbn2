"""
Japanese Group - Sales Intelligence Dashboard  v5
=======================================================
CHANGES from v4:
  TAB 1  :  Live Dashboard (existing v4 behaviour, unchanged)
  TAB 2  :  Config
              1.1  Budget editor  – 7-year range (−3 / +3 from today)
                   per-branch per-month; inline grid; Excel template download/import
              1.2  Event manager  – add/edit/delete; Excel import/export
              1.3  Promotion manager – add/edit/delete; Excel import/export
              1.4  Real-time refresh interval (minutes)
  TAB 3  :  3-Year Comparison
              choose 3 date ranges; pick columns; see active Promos + Events
  TAB 4  :  Budget vs Actual  (date / month / year picker)
              table + bar chart; export Excel
  ALL    :  Export Excel works from every tab;
              includes Promo/Event sheets where relevant
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pyodbc
import pandas as pd
from datetime import datetime, timedelta
import calendar
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import sys, os, subprocess, copy, json

# ============================================================
# Column Mapping
# ============================================================
COLUMN_LABELS = {
    'GrossSale':         'Gross Sales',
    'EntAmt':            'ENT',
    'TotalInitDiscount': 'Discount',
    'NetSaleBFVat':      'NET',
    'SVCBFTax':          'Svc Chg',
    'VAT':               'VAT',
    'TotalREVENUE':      'Total Rev',
    'EITotal':           'EI Total',
    'TATotal':           'TA Total',
    'CancelTotal':       'Cancel',
}
COLUMNS = list(COLUMN_LABELS.keys())

# ============================================================
# Branch Data
# ============================================================
BRANCH_IDS = [19, 23, 22, 17, 18, 21, 20, 16, 12]
BRANCH_NAMES = {
    19: "TSUBOHACHI-BEEHIVE",
    23: "TSUBOHACHI-THE PROMENADE",
    22: "TSUBOHACHI-ZPELL",
    17: "TSUBOHACHI-THE PORTAL",
    18: "TSUBOHACHI-COSMO BAZAAR",
    21: "TSUBOHACHI-ROBINSON RP",
    20: "TSUBOHACHI-CENTRAL WESTGATE",
    16: "TAISHO-TEI-IMPACT",
    12: "NIPPONYOKOCHO-IMPACT",
}
BRANCH_SHORT = {
    "TSUBOHACHI-BEEHIVE":          "BEEHIVE",
    "TSUBOHACHI-THE PROMENADE":    "PROMENADE",
    "TSUBOHACHI-ZPELL":            "ZPELL",
    "TSUBOHACHI-THE PORTAL":       "THE PORTAL",
    "TSUBOHACHI-COSMO BAZAAR":     "COSMO BAZAAR",
    "TSUBOHACHI-ROBINSON RP":      "ROBINSON RP",
    "TSUBOHACHI-CENTRAL WESTGATE": "CENTRAL WG",
    "TAISHO-TEI-IMPACT":           "TAISHO-TEI",
    "NIPPONYOKOCHO-IMPACT":        "NIPPON YC",
}

MONTH_NAMES = ["Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec","Jan","Feb","Mar"]
MONTH_IDX   = {4:0,5:1,6:2,7:3,8:4,9:5,10:6,11:7,12:8,1:9,2:10,3:11}

# ============================================================
# Budget  (multi-year)
# ============================================================
BASE_BUDGET = {
    19: [780000,760000,824200,823970,856970,1208970,741470,741470,768970,741470,759070,1321500],
    23: [2000000,2200000,1800000,1900000,1900000,1900000,2200000,2100000,2200000,2000000,1900000,2200000],
    22: [1500000,1500000,1500000,1500000,1700000,1500000,1500000,1500000,1550000,1600000,1500000,1600000],
    17: [870000,2200000,1020000,590000,930000,1150000,770000,1170000,1490000,410000,520000,2335000],
    18: [450000,720000,495000,550000,495000,462000,462000,550000,682000,440000,440000,650000],
    21: [1500000,1500000,1600000,1500000,1600000,1600000,1600000,1600000,1600000,1600000,1400000,1500000],
    20: [800000,750000,750000,750000,850000,750000,800000,800000,800000,800000,730000,800000],
    16: [1450000,2500000,1595000,1100000,1760000,1320000,990000,1650000,2090000,700000,770000,3392300],
    12: [3392000,4600000,1980000,1210000,1507000,1210000,1320000,3080000,3960000,715000,715000,4181100],
}

# BUDGET_STORE: {year: {branch_id: [12 monthly values]}}
# Populated at startup; editable via Config tab
BUDGET_STORE = {}

# ============================================================
# Events & Promotions  (in-memory, editable via Config tab)
# ============================================================
# Each event: {id, name, start_date (date), end_date (date), branches [list of branch names or "ALL"], note}
EVENTS = []
# Each promo: {id, name, start_date (date), end_date (date), branches [list], detail}
PROMOTIONS = []
_next_event_id = 1
_next_promo_id = 1

# ============================================================
# Refresh interval (minutes)
# ============================================================
REFRESH_INTERVAL_MIN = 5   # editable in Config

# ============================================================
# Database Configuration
# ============================================================
ODBC_DRIVER = "ODBC Driver 17 for SQL Server"
HQ_SERVER   = "172.28.116.11,4443"
HQ_DATABASE = "HQ_JAPANESE_RESTAURANT_GROUP"
HQ_USER     = "japanese"
HQ_PASS     = "Password@jpsql1"

REALTIME_DBS = [
    {"server": "172.28.141.118,5906", "database": "TSUBOHACHI_BEEHIVE",          "user": "japanesefront", "password": "Password@jpsql1", "branch": "TSUBOHACHI-BEEHIVE"},
    {"server": "172.28.141.123,5906", "database": "TSUBOHACHI_PROMENADE",        "user": "japanesefront", "password": "Password@jpsql1", "branch": "TSUBOHACHI-THE PROMENADE"},
    {"server": "172.28.141.124,5909", "database": "TSUBOHACHI_ZPELL",            "user": "japanesefront", "password": "Password@jpsql1", "branch": "TSUBOHACHI-ZPELL"},
    {"server": "172.28.141.116,5910", "database": "FB_Tsubohachi",               "user": "japanesefront", "password": "Password@jpsql1", "branch": "TSUBOHACHI-THE PORTAL"},
    {"server": "172.28.141.126,5906", "database": "TsubohachiCOSMO",             "user": "japanesefront", "password": "Password@jpsql1", "branch": "TSUBOHACHI-COSMO BAZAAR"},
    {"server": "172.28.141.122,5906", "database": "TSUBOHACHI_ROBINSON_RP",      "user": "japanesefront", "password": "Password@jpsql1", "branch": "TSUBOHACHI-ROBINSON RP"},
    {"server": "172.28.141.120,5906", "database": "TSUBOHACHI_CENTRAL_WESTGATE", "user": "japanesefront", "password": "Password@jpsql1", "branch": "TSUBOHACHI-CENTRAL WESTGATE"},
    {"server": "172.28.116.201,1433", "database": "FB_Taishotei_Ramen",          "user": "japanesefront", "password": "Password@jpsql1", "branch": "TAISHO-TEI-IMPACT"},
    {"server": "172.28.141.125,5906", "database": "FB_Nippon_Yokocho",           "user": "japanesefront", "password": "Password@jpsql1", "branch": "NIPPONYOKOCHO-IMPACT"},
]

# ============================================================
# Colors
# ============================================================
C = {
    "bg":         "#0A0E1A",
    "panel":      "#0F1523",
    "card":       "#151D2E",
    "row_a":      "#111827",
    "row_b":      "#0F1520",
    "total_bg":   "#0D1F35",
    "sep":        "#1E2D45",
    "text":       "#E8EDF5",
    "muted":      "#7A8BA8",
    "gold":       "#FFD166",
    "rt":         "#00D4FF",
    "dy":         "#00FFB2",
    "mtd":        "#4F8EF7",
    "bud":        "#B57BFF",
    "hit_hi":     "#00382B",
    "hit_mid":    "#352800",
    "hit_lo":     "#350010",
    "blink_a":    "#FF3860",
    "blink_b":    "#FF8C00",
    "ok":         "#00FFB2",
    "err":        "#FF3860",
    "titlebar":   "#080C18",
    "btn_min":    "#2D3748",
    "btn_max":    "#2D3748",
    "btn_close":  "#8B1A1A",
    "kpi_rt":     "#00D4FF",
    "kpi_dy":     "#00FFB2",
    "kpi_mtd":    "#4F8EF7",
    "kpi_bud":    "#B57BFF",
    "kpi_rev":    "#FFD166",
    "bar_hi":     "#00FFB2",
    "bar_mid":    "#FFD166",
    "bar_lo":     "#FF6B8A",
    "bar_track":  "#1E2D45",
    "live_on":    "#00FFB2",
    "live_off":   "#FF6B8A",
    "hist":       "#FFD166",
    "hist_bg":    "#1A1200",
    "btn_hist":   "#2A1F00",
    "btn_live":   "#002A1F",
    # Config / new tabs
    "cfg":        "#FF9F43",
    "cmp":        "#A29BFE",
    "bva":        "#FD79A8",
    "tab_sel":    "#151D2E",
    "tab_bg":     "#0A0E1A",
    "inp_bg":     "#0F1A2E",
    "inp_fg":     "#E8EDF5",
    "inp_bd":     "#1E2D45",
    "event_bg":   "#0D1A0D",
    "event_fg":   "#00FFB2",
    "promo_bg":   "#1A0D1A",
    "promo_fg":   "#B57BFF",
    "del_btn":    "#4A0010",
    "del_fg":     "#FF6B8A",
    "add_btn":    "#00382B",
    "add_fg":     "#00FFB2",
    "save_btn":   "#1A1200",
    "save_fg":    "#FFD166",
    "imp_btn":    "#1A0A00",
    "imp_fg":     "#FF9F43",
}

# ============================================================
# Auto-Scale
# ============================================================
_tmp = tk.Tk(); _tmp.withdraw()
SCREEN_W = _tmp.winfo_screenwidth()
SCREEN_H = _tmp.winfo_screenheight()
_tmp.destroy()

REF_W, REF_H = 1920, 1080
SCALE = max(0.6, min(min(SCREEN_W / REF_W, SCREEN_H / REF_H), 3.0))

def S(px): return max(1, int(px * SCALE))

FS_TITLE   = max(12, int(22 * SCALE))
FS_SECTION = max(10, int(13 * SCALE))
FS_COLHEAD = max(9,  int(10 * SCALE))
FS_DATA    = max(9,  int(11 * SCALE))
FS_TOTAL   = max(9,  int(11 * SCALE))
FS_CLOCK   = max(10, int(13 * SCALE))
FS_STATUS  = max(9,  int(10 * SCALE))
FS_BTN     = max(9,  int(10 * SCALE))
FS_KPI_VAL = max(14, int(20 * SCALE))
FS_KPI_LBL = max(8,  int(9  * SCALE))
ROW_H      = S(26)
COL_BRANCH = S(130)
COL_NUM    = S(100)

FONT_TITLE   = ("Segoe UI", FS_TITLE,   "bold")
FONT_SECTION = ("Segoe UI", FS_SECTION, "bold")
FONT_DATA    = ("Consolas", FS_DATA)
FONT_TOTAL   = ("Consolas", FS_TOTAL,   "bold")
FONT_CLOCK   = ("Segoe UI", FS_CLOCK,   "bold")
FONT_STATUS  = ("Segoe UI", FS_STATUS)
FONT_BTN     = ("Segoe UI", FS_BTN,     "bold")
FONT_KPI_VAL = ("Consolas", FS_KPI_VAL, "bold")
FONT_KPI_LBL = ("Segoe UI", FS_KPI_LBL)

# ============================================================
# Budget helpers
# ============================================================
def _get_budget(year, branch_id, month_idx):
    """Return budget value for fiscal-month index (0=Apr … 11=Mar)."""
    return BUDGET_STORE.get(year, {}).get(branch_id, BASE_BUDGET.get(branch_id, [0]*12))[month_idx]

def _init_budget_store():
    """Populate BUDGET_STORE for current year ±3."""
    cur_year = datetime.now().year
    for yr in range(cur_year - 3, cur_year + 4):
        BUDGET_STORE[yr] = {}
        for bid in BRANCH_IDS:
            BUDGET_STORE[yr][bid] = list(BASE_BUDGET.get(bid, [0]*12))

_init_budget_store()

# ============================================================
# DB helpers
# ============================================================
def _conn_str(server, database, user, password):
    return (
        f"DRIVER={{{ODBC_DRIVER}}};"
        f"SERVER={server};"
        f"DATABASE={database};"
        f"UID={user};"
        f"PWD={password};"
        f"Encrypt=no;"
        f"TrustServerCertificate=yes;"
        f"Connection Timeout=5;"
    )

def fetch_realtime(from_dt, to_dt):
    all_data = []
    for cfg in REALTIME_DBS:
        try:
            conn   = pyodbc.connect(_conn_str(cfg["server"], cfg["database"], cfg["user"], cfg["password"]), timeout=5)
            cursor = conn.cursor()
            cursor.execute("""
                SET NOCOUNT ON;
                EXEC srptGetSystemSaleReport
                    @FromDate=?, @ToDate=?,
                    @TerminalId=N'', @PaymentCategory=N'', @CashCountId=N''
            """, from_dt, to_dt)
            if cursor.description:
                cols = [c[0] for c in cursor.description]
                rows = cursor.fetchall()
                if rows:
                    df = pd.DataFrame.from_records(rows, columns=cols)
                    df["BranchName"] = cfg["branch"]
                    for c in COLUMNS:
                        if c in df.columns:
                            df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0.0).round(2)
                    all_data.append(df)
            cursor.close(); conn.close()
        except Exception:
            pass
    if all_data:
        return pd.concat(all_data, ignore_index=True), "ok"
    return pd.DataFrame(), "err"

def fetch_hq(from_dt, to_dt):
    try:
        conn   = pyodbc.connect(_conn_str(HQ_SERVER, HQ_DATABASE, HQ_USER, HQ_PASS), timeout=5)
        cursor = conn.cursor()
    except Exception:
        return pd.DataFrame(), "err"
    all_data = []
    for bid in BRANCH_IDS:
        try:
            cursor.execute("""
                SET NOCOUNT ON;
                EXEC srptGetSystemSaleReport
                @FromDate=?, @ToDate=?, @TerminalId=N'',
                @PaymentCategory=N'', @BranchId=?, @CashCountId=N''
            """, from_dt, to_dt, bid)
            if cursor.description:
                cols = [c[0] for c in cursor.description]
                rows = cursor.fetchall()
                if rows:
                    df = pd.DataFrame.from_records(rows, columns=cols)
                    df["BranchName"] = BRANCH_NAMES.get(bid, f"Branch {bid}")
                    for c in COLUMNS:
                        if c in df.columns:
                            df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0.0).round(2)
                    all_data.append(df)
        except Exception:
            pass
    cursor.close(); conn.close()
    if all_data:
        return pd.concat(all_data, ignore_index=True), "ok"
    return pd.DataFrame(), "err"

# ============================================================
# Root Window
# ============================================================
root = tk.Tk()
root.title("Japanese Group - Sales Intelligence v5")
root.configure(bg=C["bg"])
root.overrideredirect(True)

WIN_W = int(SCREEN_W * 0.97)
WIN_H = int(SCREEN_H * 0.97)
WIN_X = (SCREEN_W - WIN_W) // 2
WIN_Y = (SCREEN_H - WIN_H) // 2
root.geometry(f"{WIN_W}x{WIN_H}+{WIN_X}+{WIN_Y}")

_is_maximized = False
_restore_geo  = f"{WIN_W}x{WIN_H}+{WIN_X}+{WIN_Y}"

def maximize_window():
    global _is_maximized, _restore_geo
    if not _is_maximized:
        _restore_geo  = root.geometry()
        _is_maximized = True
        root.geometry(f"{SCREEN_W}x{SCREEN_H}+0+0")
    else:
        _is_maximized = False
        root.geometry(_restore_geo)

def minimize_window():
    root.overrideredirect(False)
    root.iconify()

def on_deiconify(event):
    root.overrideredirect(True)
root.bind("<Map>", on_deiconify)

_drag_x = _drag_y = 0
def drag_start(event):
    global _drag_x, _drag_y
    _drag_x = event.x_root - root.winfo_x()
    _drag_y = event.y_root - root.winfo_y()

def drag_move(event):
    if not _is_maximized:
        root.geometry(f"+{event.x_root - _drag_x}+{event.y_root - _drag_y}")

# ============================================================
# Custom Title Bar
# ============================================================
titlebar = tk.Frame(root, bg=C["titlebar"], height=S(40))
titlebar.pack(fill="x", side="top")
titlebar.pack_propagate(False)
titlebar.bind("<ButtonPress-1>",   drag_start)
titlebar.bind("<B1-Motion>",       drag_move)
titlebar.bind("<Double-Button-1>", lambda e: maximize_window())

tb_left = tk.Frame(titlebar, bg=C["titlebar"])
tb_left.pack(side="left", padx=S(10), pady=S(6))
for w in [tb_left]:
    w.bind("<ButtonPress-1>", drag_start)
    w.bind("<B1-Motion>",     drag_move)

tk.Frame(tb_left, bg=C["rt"], width=S(3)).pack(side="left", fill="y", padx=(0, S(8)))
lbl_wintitle = tk.Label(tb_left,
    text="JAPANESE GROUP  ·  Sales Intelligence Dashboard  v5",
    bg=C["titlebar"], fg=C["text"],
    font=("Segoe UI", FS_SECTION, "bold"))
lbl_wintitle.pack(side="left")
lbl_wintitle.bind("<ButtonPress-1>", drag_start)
lbl_wintitle.bind("<B1-Motion>",     drag_move)

tb_right = tk.Frame(titlebar, bg=C["titlebar"])
tb_right.pack(side="right")

def _tb_btn(parent, text, bg, command, hover_bg):
    b = tk.Label(parent, text=text, bg=bg, fg=C["text"],
                 font=("Segoe UI", FS_BTN), width=S(4),
                 cursor="hand2", pady=S(5))
    b.pack(side="left")
    b.bind("<Button-1>", lambda e: command())
    b.bind("<Enter>", lambda e: b.config(bg=hover_bg))
    b.bind("<Leave>", lambda e: b.config(bg=bg))
    return b

_tb_btn(tb_right, "-", C["btn_min"],   minimize_window, "#3D4F65")
_tb_btn(tb_right, "+", C["btn_max"],   maximize_window, "#3D4F65")
_tb_btn(tb_right, "x", C["btn_close"], root.destroy,    "#C0392B")

tb_info = tk.Frame(titlebar, bg=C["titlebar"])
tb_info.pack(side="right", padx=(0, S(6)))

lbl_status = tk.Label(tb_info, text="● LIVE", bg=C["titlebar"], fg=C["ok"],
                      font=("Segoe UI", FS_STATUS, "bold"))
lbl_status.pack(side="left", padx=(0, S(10)))

lbl_clock = tk.Label(tb_info, text="", bg=C["titlebar"], fg=C["gold"], font=FONT_CLOCK)
lbl_clock.pack(side="left", padx=(0, S(10)))

lbl_reset_info = tk.Label(tb_info, text="", bg=C["titlebar"], fg=C["muted"],
                           font=("Segoe UI", FS_STATUS))
lbl_reset_info.pack(side="left")

def _tick():
    now = datetime.now()
    lbl_clock.config(text=now.strftime("%d %b %Y   %H:%M:%S"))
    next_reset = now.replace(hour=4, minute=0, second=0, microsecond=0)
    if now >= next_reset:
        next_reset += timedelta(days=1)
    diff = next_reset - now
    h, rem = divmod(int(diff.total_seconds()), 3600)
    m, s   = divmod(rem, 60)
    lbl_reset_info.config(text=f"[R] reset in  {h:02d}:{m:02d}:{s:02d}")
    root.after(1000, _tick)
_tick()

tk.Frame(root, bg=C["rt"], height=2).pack(fill="x")

# ============================================================
# Tab Bar  (custom, no ttk.Notebook)
# ============================================================
TABS = [
    ("tab_live", "  📡  LIVE DASHBOARD",  C["rt"]),
    ("tab_cfg",  "  ⚙  CONFIG",           C["cfg"]),
    ("tab_cmp",  "  📊  3-YEAR COMPARE",  C["cmp"]),
    ("tab_bva",  "  📈  BUDGET vs ACTUAL", C["bva"]),
]
_active_tab = tk.StringVar(value="tab_live")
_tab_frames = {}
_tab_btns   = {}

tab_bar = tk.Frame(root, bg=C["titlebar"], height=S(34))
tab_bar.pack(fill="x")
tab_bar.pack_propagate(False)

tab_content = tk.Frame(root, bg=C["bg"])
tab_content.pack(fill="both", expand=True)

def _show_tab(tab_id):
    _active_tab.set(tab_id)
    for tid, frame in _tab_frames.items():
        if tid == tab_id:
            frame.pack(fill="both", expand=True)
        else:
            frame.pack_forget()
    for tid, (btn, accent) in _tab_btns.items():
        if tid == tab_id:
            btn.config(bg=C["card"], fg=accent)
        else:
            btn.config(bg=C["titlebar"], fg=C["muted"])

for tab_id, label, accent in TABS:
    btn = tk.Label(tab_bar, text=label, bg=C["titlebar"], fg=C["muted"],
                   font=("Segoe UI", FS_BTN, "bold"),
                   padx=S(14), pady=S(6), cursor="hand2")
    btn.pack(side="left")
    _tab_btns[tab_id] = (btn, accent)
    btn.bind("<Button-1>", lambda e, tid=tab_id: _show_tab(tid))
    btn.bind("<Enter>",    lambda e, b=btn, a=accent: b.config(fg=a)
                           if _active_tab.get() != tab_id else None)

# ============================================================
# Scrollable helper
# ============================================================
def _make_scrollable(parent):
    outer = tk.Frame(parent, bg=C["bg"])
    outer.pack(fill="both", expand=True)
    canvas  = tk.Canvas(outer, bg=C["bg"], highlightthickness=0, bd=0)
    vscroll = ttk.Scrollbar(outer, orient="vertical", command=canvas.yview)
    canvas.configure(yscrollcommand=vscroll.set)
    vscroll.pack(side="right", fill="y")
    canvas.pack(side="left", fill="both", expand=True)
    inner = tk.Frame(canvas, bg=C["bg"])
    win   = canvas.create_window((0, 0), window=inner, anchor="nw")
    inner.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
    canvas.bind("<Configure>", lambda e: canvas.itemconfig(win, width=e.width))
    canvas.bind_all("<MouseWheel>",
        lambda e: canvas.yview_scroll(int(-1*(e.delta/120)), "units")
        if str(canvas) in str(root.focus_get()) or True else None)
    return inner

# ============================================================
# TTK Style
# ============================================================
style = ttk.Style()
style.theme_use("clam")
style.configure("TV.Treeview",
    background=C["row_a"], foreground=C["text"],
    fieldbackground=C["row_a"], rowheight=ROW_H, font=FONT_DATA,
    borderwidth=0, relief="flat")
style.configure("TV.Treeview.Heading",
    background=C["card"], foreground=C["muted"],
    font=("Segoe UI", FS_COLHEAD, "bold"),
    borderwidth=0, relief="flat", padding=(S(4), S(4)))
style.map("TV.Treeview",
    background=[("selected", C["mtd"])],
    foreground=[("selected", "#ffffff")])
style.configure("Cfg.Treeview",
    background=C["inp_bg"], foreground=C["text"],
    fieldbackground=C["inp_bg"], rowheight=S(24), font=("Consolas", FS_DATA),
    borderwidth=0, relief="flat")
style.configure("Cfg.Treeview.Heading",
    background=C["card"], foreground=C["gold"],
    font=("Segoe UI", FS_COLHEAD, "bold"),
    borderwidth=0, relief="flat", padding=(S(3), S(3)))

# ============================================================
# Utility
# ============================================================
def _fmt(v):
    try: return f"{float(v):,.0f}"
    except: return str(v)

def _short(name):
    return BRANCH_SHORT.get(name, name)

def _make_label_entry(parent, label, row, col, width=14, default="",
                       label_fg=None, entry_bg=None, entry_fg=None):
    fg = label_fg or C["muted"]
    ebg = entry_bg or C["inp_bg"]
    efg = entry_fg or C["inp_fg"]
    tk.Label(parent, text=label, bg=C["panel"], fg=fg,
             font=FONT_KPI_LBL).grid(row=row, column=col, sticky="w",
                                     padx=(S(6),S(2)), pady=S(2))
    var = tk.StringVar(value=str(default))
    ent = tk.Entry(parent, textvariable=var, width=width,
                   bg=ebg, fg=efg, insertbackground=efg,
                   relief="flat", font=("Consolas", FS_DATA),
                   highlightthickness=1, highlightcolor=C["sep"],
                   highlightbackground=C["sep"])
    ent.grid(row=row, column=col+1, sticky="ew", padx=(0, S(8)), pady=S(2))
    return var

def _make_spinbox(parent, from_, to_, default, width=6, **kwargs):
    var = tk.StringVar(value=str(default))
    sp  = tk.Spinbox(parent, from_=from_, to=to_, textvariable=var, width=width,
                     bg=C["inp_bg"], fg=C["hist"],
                     font=("Consolas", FS_KPI_LBL, "bold"),
                     relief="flat", bd=0,
                     buttonbackground=C["sep"],
                     highlightthickness=0,
                     justify="center", **kwargs)
    sp.pack(side="left", padx=S(2))
    def get_val():
        try: return int(float(var.get()))
        except: return default
    sp.get_val = get_val
    return sp

def _date_picker_frame(parent, default_date, label=""):
    f = tk.Frame(parent, bg=C["panel"])
    if label:
        tk.Label(f, text=label, bg=C["panel"], fg=C["muted"],
                 font=FONT_KPI_LBL).pack(side="left", padx=(0, S(4)))
    var_d = tk.StringVar(value=f"{default_date.day:02d}")
    var_m = tk.StringVar(value=f"{default_date.month:02d}")
    var_y = tk.StringVar(value=f"{default_date.year:04d}")
    sp_style = dict(bg=C["btn_hist"], fg=C["hist"],
                    font=("Consolas", FS_KPI_LBL, "bold"),
                    relief="flat", bd=0,
                    buttonbackground=C["sep"],
                    highlightthickness=0, justify="center")
    tk.Spinbox(f, from_=1, to=31,   width=2, textvariable=var_d, format="%02.0f", **sp_style).pack(side="left")
    tk.Label(f, text="/", bg=C["panel"], fg=C["hist"], font=("Consolas", FS_KPI_LBL)).pack(side="left")
    tk.Spinbox(f, from_=1, to=12,   width=2, textvariable=var_m, format="%02.0f", **sp_style).pack(side="left")
    tk.Label(f, text="/", bg=C["panel"], fg=C["hist"], font=("Consolas", FS_KPI_LBL)).pack(side="left")
    tk.Spinbox(f, from_=2020, to=2099, width=4, textvariable=var_y, format="%4.0f", **sp_style).pack(side="left")
    def get_date():
        try: return datetime(int(var_y.get()), int(var_m.get()), int(var_d.get())).date()
        except: return datetime.now().date()
    f.get_date = get_date
    return f

def _accent_btn(parent, text, bg, fg, cmd, padx=S(12), pady=S(4)):
    b = tk.Label(parent, text=text, bg=bg, fg=fg,
                 font=("Segoe UI", FS_BTN, "bold"),
                 padx=padx, pady=pady, cursor="hand2", relief="flat")
    b.bind("<Button-1>", lambda e: cmd())
    hover = "#" + "".join(f"{min(255,int(bg[i:i+2],16)+30):02X}" for i in [1,3,5])
    b.bind("<Enter>", lambda e: b.config(bg=hover))
    b.bind("<Leave>", lambda e: b.config(bg=bg))
    return b

def _section_header(parent, text, accent):
    f = tk.Frame(parent, bg=C["card"])
    f.pack(fill="x")
    tk.Frame(f, bg=accent, height=S(2)).pack(fill="x")
    tk.Label(f, text=text, bg=C["card"], fg=accent,
             font=FONT_SECTION, padx=S(10), pady=S(5)).pack(side="left")
    return f

# ============================================================
# ============================================================
#   TAB 1: LIVE DASHBOARD  (v4 logic, fully preserved)
# ============================================================
# ============================================================
frame_live = tk.Frame(tab_content, bg=C["bg"])
_tab_frames["tab_live"] = frame_live

# --- live state ---
_live_running   = True
_hist_mode      = False
_refresh_job    = None

# --- scrollable main ---
live_scroll_outer = tk.Frame(frame_live, bg=C["bg"])
live_scroll_outer.pack(fill="both", expand=True)
live_canvas  = tk.Canvas(live_scroll_outer, bg=C["bg"], highlightthickness=0, bd=0)
live_vscroll = ttk.Scrollbar(live_scroll_outer, orient="vertical", command=live_canvas.yview)
live_canvas.configure(yscrollcommand=live_vscroll.set)
live_vscroll.pack(side="right", fill="y")
live_canvas.pack(side="left", fill="both", expand=True)
main = tk.Frame(live_canvas, bg=C["bg"])
_main_win = live_canvas.create_window((0, 0), window=main, anchor="nw")
main.bind("<Configure>", lambda e: live_canvas.configure(scrollregion=live_canvas.bbox("all")))
live_canvas.bind("<Configure>", lambda e: live_canvas.itemconfig(_main_win, width=e.width))
live_canvas.bind_all("<MouseWheel>",
    lambda e: live_canvas.yview_scroll(int(-1*(e.delta/120)), "units"))

# KPI Strip
kpi_frame = tk.Frame(main, bg=C["bg"])
kpi_frame.pack(fill="x", padx=S(10), pady=(S(8), S(4)))

kpi_cards = {}
KPI_DEFS = [
    ("rt_net",  "Real-time NET",   C["kpi_rt"],  "Today so far"),
    ("dy_net",  "Yesterday NET",   C["kpi_dy"],  "All branches"),
    ("mtd_net", "MTD NET",         C["kpi_mtd"], "Month to date"),
    ("bud_pct", "Budget Achieved", C["kpi_bud"], "vs monthly target"),
    ("tot_rev", "Total Revenue",   C["kpi_rev"], "MTD incl. svc+VAT"),
]
for i, (key, label, accent, sub) in enumerate(KPI_DEFS):
    kpi_frame.columnconfigure(i, weight=1)
    card = tk.Frame(kpi_frame, bg=C["panel"])
    card.grid(row=0, column=i, sticky="nsew", padx=S(4))
    tk.Frame(card, bg=accent, height=S(2)).pack(fill="x")
    inner = tk.Frame(card, bg=C["panel"])
    inner.pack(fill="both", expand=True, padx=S(10), pady=S(7))
    tk.Label(inner, text=label, bg=C["panel"], fg=C["muted"], font=FONT_KPI_LBL).pack(anchor="w")
    lv = tk.Label(inner, text="-", bg=C["panel"], fg=accent, font=FONT_KPI_VAL)
    lv.pack(anchor="w")
    tk.Label(inner, text=sub, bg=C["panel"], fg=C["muted"], font=FONT_KPI_LBL).pack(anchor="w")
    kpi_cards[key] = lv

tk.Frame(main, bg=C["sep"], height=1).pack(fill="x", padx=S(10), pady=(S(2), 0))

# Control bar
ctrl_bar = tk.Frame(main, bg=C["panel"])
ctrl_bar.pack(fill="x", padx=S(10), pady=(S(4), S(2)))

ctrl_left = tk.Frame(ctrl_bar, bg=C["panel"])
ctrl_left.pack(side="left", padx=S(8), pady=S(4))

lbl_live_status = tk.Label(ctrl_left, text="● LIVE", bg=C["panel"], fg=C["live_on"],
                            font=("Segoe UI", FS_STATUS, "bold"))
lbl_live_status.pack(side="left", padx=(0, S(8)))

def _make_ctrl_btn(parent, text, bg, fg, cmd):
    b = tk.Label(parent, text=text, bg=bg, fg=fg,
                 font=("Segoe UI", FS_BTN, "bold"),
                 padx=S(12), pady=S(4), cursor="hand2", relief="flat")
    b.pack(side="left", padx=S(3))
    b.bind("<Button-1>", lambda e: cmd())
    return b

btn_stop_run = _make_ctrl_btn(ctrl_left, "  ⏸  PAUSE LIVE", C["btn_min"], C["live_on"], lambda: toggle_live())

def toggle_live():
    global _live_running, _refresh_job
    if _hist_mode: return
    if _live_running:
        _live_running = False
        if _refresh_job:
            root.after_cancel(_refresh_job); _refresh_job = None
        btn_stop_run.config(text="  ▶  RESUME LIVE", fg=C["live_off"], bg="#1A0000")
        lbl_live_status.config(text="⏸ PAUSED", fg=C["live_off"])
        lbl_status.config(text="* PAUSED", fg=C["live_off"])
    else:
        _live_running = True
        btn_stop_run.config(text="  ⏸  PAUSE LIVE", fg=C["live_on"], bg=C["btn_min"])
        lbl_live_status.config(text="● LIVE", fg=C["live_on"])
        lbl_status.config(text="* LIVE", fg=C["ok"])
        refresh_all()

tk.Frame(ctrl_bar, bg=C["sep"], width=2).pack(side="left", fill="y", padx=S(10), pady=S(4))

ctrl_right = tk.Frame(ctrl_bar, bg=C["panel"])
ctrl_right.pack(side="left", padx=S(4), pady=S(4))

tk.Label(ctrl_right, text="HISTORICAL VIEW", bg=C["panel"], fg=C["hist"],
         font=("Segoe UI", FS_KPI_LBL, "bold")).pack(side="left", padx=(0, S(8)))
tk.Label(ctrl_right, text="From:", bg=C["panel"], fg=C["muted"], font=FONT_KPI_LBL).pack(side="left")

def _make_date_spinbox(parent, default_date):
    f = tk.Frame(parent, bg=C["btn_hist"], padx=S(3), pady=S(2))
    f.pack(side="left", padx=S(3))
    sp_style = dict(bg=C["btn_hist"], fg=C["hist"],
                    font=("Consolas", FS_KPI_LBL, "bold"),
                    relief="flat", bd=0, buttonbackground=C["sep"],
                    highlightthickness=0, justify="center")
    var_d = tk.StringVar(value=f"{default_date.day:02d}")
    var_m = tk.StringVar(value=f"{default_date.month:02d}")
    var_y = tk.StringVar(value=f"{default_date.year:04d}")
    tk.Spinbox(f, from_=1, to=31, width=2, textvariable=var_d, format="%02.0f", **sp_style).pack(side="left")
    tk.Label(f, text="/", bg=C["btn_hist"], fg=C["hist"], font=("Consolas", FS_KPI_LBL)).pack(side="left")
    tk.Spinbox(f, from_=1, to=12, width=2, textvariable=var_m, format="%02.0f", **sp_style).pack(side="left")
    tk.Label(f, text="/", bg=C["btn_hist"], fg=C["hist"], font=("Consolas", FS_KPI_LBL)).pack(side="left")
    tk.Spinbox(f, from_=2020, to=2099, width=4, textvariable=var_y, format="%4.0f", **sp_style).pack(side="left")
    def get_date():
        try: return datetime(int(var_y.get()), int(var_m.get()), int(var_d.get())).date()
        except: return datetime.now().date()
    f.get_date = get_date
    return f

today_d = datetime.now().date()
yest_d  = today_d - timedelta(days=1)
hist_from_picker = _make_date_spinbox(ctrl_right, yest_d)
tk.Label(ctrl_right, text="  To:", bg=C["panel"], fg=C["muted"], font=FONT_KPI_LBL).pack(side="left")
hist_to_picker = _make_date_spinbox(ctrl_right, yest_d)

_hist_time_var = tk.StringVar(value="operational")
tk.Label(ctrl_right, text="  Time:", bg=C["panel"], fg=C["muted"], font=FONT_KPI_LBL).pack(side="left", padx=(S(6),0))
for txt, val in [("05:00-23:59","operational"),("00:00-23:59","fullday")]:
    tk.Radiobutton(ctrl_right, text=txt, variable=_hist_time_var, value=val,
                   bg=C["panel"], fg=C["muted"], selectcolor=C["bg"],
                   activebackground=C["panel"], activeforeground=C["hist"],
                   font=FONT_KPI_LBL).pack(side="left", padx=S(2))

btn_load_hist = _make_ctrl_btn(ctrl_right, "  🔍  LOAD", "#2A1800", C["hist"], lambda: load_historical())
btn_load_hist.pack(side="left", padx=(S(6), S(2)))
btn_back_live = _make_ctrl_btn(ctrl_right, "  ↩  BACK TO LIVE", C["btn_live"], C["dy"], lambda: back_to_live())
btn_back_live.pack(side="left", padx=S(2))
btn_back_live.config(state="disabled")

hist_banner = tk.Frame(main, bg=C["hist_bg"])
hist_banner_lbl = tk.Label(hist_banner, text="", bg=C["hist_bg"], fg=C["hist"],
                            font=("Segoe UI", FS_STATUS, "bold"),
                            padx=S(14), pady=S(5))
hist_banner_lbl.pack(side="left")

def _show_hist_banner(from_d, to_d):
    hist_banner_lbl.config(
        text=f"  ⚠  HISTORICAL MODE  |  {from_d.strftime('%d %b %Y')}  →  {to_d.strftime('%d %b %Y')}"
             f"  |  HQ database  |  Auto-refresh PAUSED")
    hist_banner.pack(fill="x", padx=S(10), pady=(S(2), 0))

def _hide_hist_banner():
    hist_banner.pack_forget()

# Section helpers
_section_states = {}
_section_trees  = {}
_section_data   = {}

def _configure_tree(tree):
    tree.tag_configure("evenrow",  background=C["row_a"], foreground=C["text"])
    tree.tag_configure("oddrow",   background=C["row_b"], foreground=C["text"])
    tree.tag_configure("totalrow", background=C["total_bg"], foreground=C["gold"], font=FONT_TOTAL)
    tree.tag_configure("blink_a",  background=C["blink_a"], foreground="#ffffff", font=FONT_TOTAL)
    tree.tag_configure("blink_b",  background=C["blink_b"], foreground="#ffffff", font=FONT_TOTAL)

def _redraw_collapsed(section_id):
    tree = _section_trees[section_id]
    tree.delete(*tree.get_children())
    data = _section_data.get(section_id, [])
    if data:
        tr = data[-1]
        tree.insert("", "end", values=tr[0], tags=(tr[1],))

def _redraw_expanded(section_id):
    tree = _section_trees[section_id]
    tree.delete(*tree.get_children())
    for values, tag in _section_data.get(section_id, []):
        tree.insert("", "end", values=values, tags=(tag,))

def section_store_and_draw(section_id, rows_data):
    _section_data[section_id] = rows_data
    if _section_states.get(section_id, True):
        _redraw_expanded(section_id)
    else:
        _redraw_collapsed(section_id)

def _build_tree_section(parent, section_id, title, accent, columns,
                         col_widths=None, rows=9, extra_header_widget_fn=None):
    _section_states[section_id] = True
    card = tk.Frame(parent, bg=C["card"])
    card.pack(fill="both", expand=True)
    tk.Frame(card, bg=accent, height=S(2)).pack(fill="x")
    hdr = tk.Frame(card, bg=C["card"])
    hdr.pack(fill="x", padx=S(8), pady=(S(5), S(3)))
    tk.Label(hdr, text=title, bg=C["card"], fg=accent, font=FONT_SECTION).pack(side="left")
    if extra_header_widget_fn:
        extra_header_widget_fn(hdr)
    btn_toggle = tk.Label(hdr, text="  -  ", bg=C["sep"], fg=C["text"],
                          font=("Segoe UI", FS_BTN, "bold"),
                          cursor="hand2", padx=S(5), pady=S(1))
    btn_toggle.pack(side="right")
    tf = tk.Frame(card, bg=C["card"])
    tf.pack(fill="both", expand=True, padx=S(6), pady=(0, S(6)))
    tree = ttk.Treeview(tf, columns=columns, show="headings",
                         height=rows, style="TV.Treeview", selectmode="none")
    vsb = ttk.Scrollbar(tf, orient="vertical", command=tree.yview)
    tree.configure(yscrollcommand=vsb.set)
    tree.pack(side="left", fill="both", expand=True)
    vsb.pack(side="right", fill="y")
    _section_trees[section_id] = tree
    _configure_tree(tree)
    for col in columns:
        label  = COLUMN_LABELS.get(col, col)
        anchor = "w" if col == "BranchName" else "e"
        w = (col_widths or {}).get(col, COL_BRANCH if col == "BranchName" else COL_NUM)
        tree.heading(col, text=label, anchor="center")
        tree.column(col, width=w, anchor=anchor, stretch=True, minwidth=S(50))
    def toggle(event=None):
        if _section_states[section_id]:
            _section_states[section_id] = False
            btn_toggle.config(text="  +  ")
            tree.config(height=1)
            _redraw_collapsed(section_id)
        else:
            _section_states[section_id] = True
            btn_toggle.config(text="  -  ")
            tree.config(height=rows)
            _redraw_expanded(section_id)
    btn_toggle.bind("<Button-1>", toggle)
    return tree, card

# RT bar chart
_rt_bar_canvas     = None
_rt_bar_h          = 0
_rt_bar_data_store = []
RT_BAR_ROW_H = S(20); RT_BAR_PAD_Y = S(3)
RT_BAR_NAME_W = S(88); RT_BAR_AMT_W = S(90); RT_BAR_PAD_X = S(8)

def _build_rt_bar_canvas(parent):
    global _rt_bar_canvas, _rt_bar_h
    n = len(BRANCH_IDS)
    _rt_bar_h = (RT_BAR_ROW_H + RT_BAR_PAD_Y) * n + RT_BAR_PAD_Y * 2
    _rt_bar_canvas = tk.Canvas(parent, bg=C["card"], height=_rt_bar_h,
                                highlightthickness=0, bd=0)
    _rt_bar_canvas.pack(fill="x", padx=S(6), pady=(0, S(4)))
    _rt_bar_canvas.bind("<Configure>", lambda e: _draw_rt_bars(_rt_bar_data_store, _rt_bar_canvas.winfo_width()))

def _draw_rt_bars(branch_data, canvas_w):
    global _rt_bar_data_store
    _rt_bar_data_store = branch_data
    if not _rt_bar_canvas or canvas_w < 10: return
    _rt_bar_canvas.delete("all")
    avail_w = canvas_w - RT_BAR_NAME_W - RT_BAR_AMT_W - RT_BAR_PAD_X * 3
    for i, (name, pct, actual_str, target_str) in enumerate(branch_data):
        y_top = RT_BAR_PAD_Y + i * (RT_BAR_ROW_H + RT_BAR_PAD_Y)
        y_mid = y_top + RT_BAR_ROW_H // 2
        _rt_bar_canvas.create_text(RT_BAR_PAD_X, y_mid, text=name, anchor="w",
                                    fill=C["muted"], font=("Segoe UI", FS_KPI_LBL))
        tx = RT_BAR_PAD_X + RT_BAR_NAME_W + RT_BAR_PAD_X
        ty = y_top + S(3); th = RT_BAR_ROW_H - S(6)
        _rt_bar_canvas.create_rectangle(tx, ty, tx+avail_w, ty+th, fill=C["bar_track"], outline="")
        _rt_bar_canvas.create_line(tx+avail_w, ty, tx+avail_w, ty+th, fill=C["muted"], width=1, dash=(2,2))
        pct_clamped = min(pct, 115.0)
        fill_w = int(avail_w * pct_clamped / 115.0)
        bar_color = C["bar_hi"] if pct>=100 else C["bar_mid"] if pct>=70 else C["bar_lo"]
        if fill_w > 0:
            _rt_bar_canvas.create_rectangle(tx, ty, tx+fill_w, ty+th, fill=bar_color, outline="")
        pct_text = f"{pct:.0f}%"
        if fill_w > S(28):
            _rt_bar_canvas.create_text(tx+fill_w-S(4), y_mid, text=pct_text, anchor="e",
                                        fill="#000000", font=("Segoe UI", FS_KPI_LBL, "bold"))
        else:
            _rt_bar_canvas.create_text(tx+fill_w+S(4), y_mid, text=pct_text, anchor="w",
                                        fill=bar_color, font=("Segoe UI", FS_KPI_LBL))
        ax = tx + avail_w + RT_BAR_PAD_X
        _rt_bar_canvas.create_text(ax+RT_BAR_AMT_W, y_mid,
                                    text=f"{actual_str} / {target_str}", anchor="e",
                                    fill=bar_color, font=("Consolas", FS_KPI_LBL))

def _compute_rt_bar_data(df_rt, ref_date=None):
    ref  = ref_date or datetime.now().date()
    midx = MONTH_IDX[ref.month]
    days = calendar.monthrange(ref.year, ref.month)[1]
    result = []
    for bid in BRANCH_IDS:
        bname  = BRANCH_NAMES[bid]
        sname  = BRANCH_SHORT.get(bname, bname)
        monthly_bgt = _get_budget(ref.year, bid, midx)
        daily_target = monthly_bgt / days if days else 1
        if not df_rt.empty and bname in df_rt["BranchName"].values:
            row    = df_rt[df_rt["BranchName"] == bname].iloc[0]
            actual = float(row.get("NetSaleBFVat", 0) or 0) + float(row.get("SVCBFTax", 0) or 0)
        else:
            actual = 0.0
        pct = (actual / daily_target * 100) if daily_target else 0
        result.append((sname, pct, f"{actual:,.0f}", f"{daily_target:,.0f}"))
    return result

# RT countdown
_rt_countdown_secs = 300
_rt_countdown_lbl  = None
_rt_countdown_job  = None

def _reset_rt_countdown():
    global _rt_countdown_secs
    _rt_countdown_secs = REFRESH_INTERVAL_MIN * 60
    _tick_rt_countdown()

def _tick_rt_countdown():
    global _rt_countdown_secs, _rt_countdown_job
    if _rt_countdown_lbl is None: return
    if _hist_mode or not _live_running:
        _rt_countdown_lbl.config(text="  ⏸ paused", fg=C["muted"])
        _rt_countdown_job = root.after(2000, _tick_rt_countdown)
        return
    mins, secs = divmod(_rt_countdown_secs, 60)
    _rt_countdown_lbl.config(
        text=f"  next refresh in {mins:01d}:{secs:02d}",
        fg=C["rt"] if _rt_countdown_secs > 30 else C["blink_a"])
    if _rt_countdown_secs > 0:
        _rt_countdown_secs -= 1
    _rt_countdown_job = root.after(1000, _tick_rt_countdown)

def _inject_rt_countdown(hdr_frame):
    global _rt_countdown_lbl
    _rt_countdown_lbl = tk.Label(hdr_frame, text="", bg=C["card"], fg=C["rt"],
                                  font=("Segoe UI", FS_STATUS))
    _rt_countdown_lbl.pack(side="left", padx=(S(12), 0))

# Budget bar
_bud_bar_canvas = None
_bud_bar_h      = 0
BAR_ROW_H = S(22); BAR_PAD_Y = S(4); BAR_NAME_W = S(88); BAR_AMT_W = S(72); BAR_PAD_X = S(8)

def _draw_bars(branch_data, canvas_w):
    if not _bud_bar_canvas or canvas_w < 10: return
    _bud_bar_canvas.delete("all")
    avail_w = canvas_w - BAR_NAME_W - BAR_AMT_W - BAR_PAD_X * 3
    for i, (name, pct, amt_str) in enumerate(branch_data):
        y_top = BAR_PAD_Y + i * (BAR_ROW_H + BAR_PAD_Y)
        y_mid = y_top + BAR_ROW_H // 2
        _bud_bar_canvas.create_text(BAR_PAD_X, y_mid, text=name, anchor="w",
                                     fill=C["muted"], font=("Segoe UI", FS_KPI_LBL))
        tx = BAR_PAD_X + BAR_NAME_W + BAR_PAD_X
        ty = y_top + S(3); th = BAR_ROW_H - S(6)
        _bud_bar_canvas.create_rectangle(tx, ty, tx+avail_w, ty+th, fill=C["bar_track"], outline="")
        pct_clamped = min(pct, 115.0)
        fill_w = int(avail_w * pct_clamped / 115.0)
        bar_color = C["bar_hi"] if pct>=100 else C["bar_mid"] if pct>=90 else C["bar_lo"]
        if fill_w > 0:
            _bud_bar_canvas.create_rectangle(tx, ty, tx+fill_w, ty+th, fill=bar_color, outline="")
        pct_text = f"{pct:.0f}%"
        if fill_w > S(30):
            _bud_bar_canvas.create_text(tx+fill_w-S(4), y_mid, text=pct_text, anchor="e",
                                         fill="#000000", font=("Segoe UI", FS_KPI_LBL, "bold"))
        else:
            _bud_bar_canvas.create_text(tx+fill_w+S(4), y_mid, text=pct_text, anchor="w",
                                         fill=bar_color, font=("Segoe UI", FS_KPI_LBL))
        ax = tx + avail_w + BAR_PAD_X
        _bud_bar_canvas.create_text(ax+BAR_AMT_W, y_mid, text=amt_str, anchor="e",
                                     fill=bar_color, font=("Consolas", FS_KPI_LBL))

def _build_budget_section(parent):
    global _bud_bar_canvas, _bud_bar_h
    _section_states["bud"] = True
    card = tk.Frame(parent, bg=C["card"])
    card.pack(fill="both", expand=True)
    tk.Frame(card, bg=C["bud"], height=S(2)).pack(fill="x")
    hdr = tk.Frame(card, bg=C["card"])
    hdr.pack(fill="x", padx=S(8), pady=(S(5), S(3)))
    tk.Label(hdr, text="  BUDGET vs ACTUAL", bg=C["card"], fg=C["bud"], font=FONT_SECTION).pack(side="left")
    btn_toggle = tk.Label(hdr, text="  -  ", bg=C["sep"], fg=C["text"],
                          font=("Segoe UI", FS_BTN, "bold"), cursor="hand2", padx=S(5), pady=S(1))
    btn_toggle.pack(side="right")
    tk.Label(hdr, text=datetime.now().strftime("%b %Y"), bg=C["card"], fg=C["muted"],
             font=FONT_KPI_LBL).pack(side="right", padx=S(6))
    n = len(BRANCH_IDS)
    _bud_bar_h = (BAR_ROW_H + BAR_PAD_Y) * n + BAR_PAD_Y
    _bud_bar_canvas = tk.Canvas(card, bg=C["card"], height=_bud_bar_h, highlightthickness=0, bd=0)
    _bud_bar_canvas.pack(fill="x", padx=S(6), pady=(S(2), 0))
    _bud_bar_canvas.bind("<Configure>",
        lambda e: _draw_bars(_section_data.get("bud_bars", []), _bud_bar_canvas.winfo_width()))
    leg = tk.Frame(card, bg=C["card"])
    leg.pack(fill="x", padx=S(8), pady=(S(2), S(3)))
    for txt, fg in [("# >=100%", C["bar_hi"]),("# 90-99%", C["bar_mid"]),("# <90%", C["bar_lo"])]:
        tk.Label(leg, text=txt, bg=C["card"], fg=fg,
                 font=("Segoe UI", FS_KPI_LBL, "bold"), padx=S(6)).pack(side="left")
    BUD_COLS   = ["BranchName","Budget","Actual","% Hit","Variance","% Var"]
    BUD_WIDTHS = {"BranchName":COL_BRANCH,"Budget":COL_NUM,"Actual":COL_NUM,
                  "% Hit":S(60),"Variance":COL_NUM,"% Var":S(55)}
    tf = tk.Frame(card, bg=C["card"])
    tf.pack(fill="both", expand=True, padx=S(6), pady=(0, S(6)))
    tree = ttk.Treeview(tf, columns=BUD_COLS, show="headings",
                         height=10, style="TV.Treeview", selectmode="none")
    vsb = ttk.Scrollbar(tf, orient="vertical", command=tree.yview)
    tree.configure(yscrollcommand=vsb.set)
    tree.pack(side="left", fill="both", expand=True)
    vsb.pack(side="right", fill="y")
    _section_trees["bud"] = tree
    _configure_tree(tree)
    tree.tag_configure("hit_hi",  background=C["hit_hi"],  foreground="#00FFB2", font=FONT_DATA)
    tree.tag_configure("hit_mid", background=C["hit_mid"], foreground=C["gold"], font=FONT_DATA)
    tree.tag_configure("hit_lo",  background=C["hit_lo"],  foreground="#FF6B8A", font=FONT_DATA)
    tree.tag_configure("totalrow",background=C["total_bg"],foreground=C["gold"], font=FONT_TOTAL)
    BLAB = {"BranchName":"Branch","% Hit":"% Hit","% Var":"% Var"}
    for col in BUD_COLS:
        label  = BLAB.get(col, COLUMN_LABELS.get(col, col))
        anchor = "w" if col == "BranchName" else "e"
        tree.heading(col, text=label, anchor="center")
        tree.column(col, width=BUD_WIDTHS.get(col, COL_NUM), anchor=anchor, stretch=True, minwidth=S(40))
    body_frames = [_bud_bar_canvas, leg, tf]
    def toggle(event=None):
        if _section_states["bud"]:
            _section_states["bud"] = False
            btn_toggle.config(text="  +  ")
            for f in body_frames: f.pack_forget()
            tree.config(height=1)
            _redraw_collapsed("bud")
            tf.pack(fill="x", padx=S(6), pady=(0, S(6)))
        else:
            _section_states["bud"] = True
            btn_toggle.config(text="  -  ")
            tf.pack_forget()
            _bud_bar_canvas.pack(fill="x", padx=S(6), pady=(S(2), 0))
            leg.pack(fill="x", padx=S(8), pady=(S(2), S(3)))
            tree.config(height=10)
            tf.pack(fill="both", expand=True, padx=S(6), pady=(0, S(6)))
            _redraw_expanded("bud")
    btn_toggle.bind("<Button-1>", toggle)
    return tree

# 2-col layout
def _make_two_col(parent):
    row = tk.Frame(parent, bg=C["bg"])
    row.pack(fill="both", expand=True, padx=S(10), pady=S(4))
    left  = tk.Frame(row, bg=C["bg"])
    right = tk.Frame(row, bg=C["bg"])
    left.pack(side="left",  fill="both", expand=True, padx=(0, S(4)))
    right.pack(side="left", fill="both", expand=True, padx=(S(4), 0))
    return left, right

ALL_COLS = ["BranchName"] + COLUMNS

row1_l, row1_r = _make_two_col(main)
tree_rt, rt_card = _build_tree_section(
    row1_l, "rt", "  REAL-TIME SALES", C["rt"], ALL_COLS, rows=10,
    extra_header_widget_fn=_inject_rt_countdown)
_build_rt_bar_canvas(rt_card)
tree_bud = _build_budget_section(row1_r)

row2_l, row2_r = _make_two_col(main)
tree_dy,  _ = _build_tree_section(row2_l, "dy",  "  DAILY SALES - YESTERDAY", C["dy"],  ALL_COLS, rows=10)
tree_mtd, _ = _build_tree_section(row2_r, "mtd", "  MONTH-TO-DATE SALES",     C["mtd"], ALL_COLS, rows=10)

# Status bar
tk.Frame(main, bg=C["sep"], height=1).pack(fill="x")
status_bar = tk.Frame(main, bg=C["panel"])
status_bar.pack(fill="x")
lbl_updated = tk.Label(status_bar, text="Last updated: -",
                        bg=C["panel"], fg=C["muted"], font=FONT_STATUS, padx=S(14), pady=S(6))
lbl_updated.pack(side="left")

btn_export_live = tk.Button(status_bar, text="  EXPORT TO EXCEL",
    bg=C["dy"], fg="#000000", font=FONT_BTN, relief="flat", bd=0,
    padx=S(16), pady=S(6), cursor="hand2")
btn_export_live.pack(side="right", padx=S(12), pady=S(5))

# Blink engine
_blink_jobs = {}
def _blink_item(tree, item_id, orig_tag, step=0):
    try:
        if not tree.exists(item_id):
            _blink_jobs.pop(item_id, None); return
    except: return
    if step < 6:
        tree.item(item_id, tags=("blink_a" if step%2==0 else "blink_b",))
        job = root.after(350, lambda: _blink_item(tree, item_id, orig_tag, step+1))
        _blink_jobs[item_id] = (tree, job)
    else:
        tree.item(item_id, tags=(orig_tag,))
        _blink_jobs.pop(item_id, None)

def _trigger_blink(tree, row_idx, orig_tag):
    children = tree.get_children()
    if row_idx >= len(children): return
    item = children[row_idx]
    if item in _blink_jobs:
        _, job = _blink_jobs[item]; root.after_cancel(job)
    _blink_item(tree, item, orig_tag)

def _detect_changes_and_blink(tree, new_rows_data, old_rows_data):
    if not old_rows_data: return
    old_map = {(v[0] if v else ""): (v, t) for v, t in old_rows_data}
    def _do():
        for i, (new_vals, new_tag) in enumerate(new_rows_data):
            key = new_vals[0] if new_vals else ""
            if key in old_map:
                old_vals, _ = old_map[key]
                if any(str(nv)!=str(ov) for nv, ov in zip(new_vals[1:], old_vals[1:])):
                    _trigger_blink(tree, i, new_tag)
    root.after(80, _do)

_prev_data = {sid: [] for sid in ["rt","dy","mtd","bud"]}
_last_rt_df = pd.DataFrame()

def _fill_tree_with_blink(df, section_id, tree):
    if df.empty:
        new_rows = [(["No Data"]+["-"]*len(COLUMNS), "evenrow")]
    else:
        new_rows = []
        for i, row in enumerate(df.itertuples(index=False)):
            vals = [_short(row.BranchName)] + [_fmt(getattr(row, c, 0.0)) for c in COLUMNS]
            new_rows.append((vals, "evenrow" if i%2==0 else "oddrow"))
        totals = [_fmt(df[c].sum()) if c in df.columns else "0" for c in COLUMNS]
        new_rows.append((["--- TOTAL ---"]+totals, "totalrow"))
    old_rows = _prev_data.get(section_id, [])
    _detect_changes_and_blink(tree, new_rows, old_rows)
    _prev_data[section_id] = new_rows
    section_store_and_draw(section_id, new_rows)

def _fill_rt(df, status):
    global _last_rt_df
    lbl_status.config(
        text="* LIVE"       if (status=="ok" and _live_running and not _hist_mode) else
             "* HISTORICAL" if _hist_mode else
             "* PAUSED"     if not _live_running else "* ERROR",
        fg=C["ok"] if (status=="ok" and _live_running and not _hist_mode) else
           C["hist"] if _hist_mode else C["live_off"])
    if df.empty:
        section_store_and_draw("rt", [(["No Data"]+["-"]*len(COLUMNS), "evenrow")])
        kpi_cards["rt_net"].config(text="-")
        _last_rt_df = df.copy(); return
    new_rows = []
    for i, row in enumerate(df.itertuples(index=False)):
        vals = [_short(row.BranchName)] + [_fmt(getattr(row, c, 0.0)) for c in COLUMNS]
        new_rows.append((vals, "evenrow" if i%2==0 else "oddrow"))
    totals = [_fmt(df[c].sum()) if c in df.columns else "0" for c in COLUMNS]
    new_rows.append((["--- TOTAL ---"]+totals, "totalrow"))
    old_rows = _prev_data.get("rt", [])
    _detect_changes_and_blink(tree_rt, new_rows, old_rows)
    _prev_data["rt"] = new_rows
    section_store_and_draw("rt", new_rows)
    _last_rt_df = df.copy()
    net_total = df["NetSaleBFVat"].sum() if "NetSaleBFVat" in df.columns else 0
    kpi_cards["rt_net"].config(text=f"{net_total:,.0f}")
    ref_d = hist_to_picker.get_date() if _hist_mode else None
    bar_data = _compute_rt_bar_data(df, ref_date=ref_d)
    root.after(60, lambda: _draw_rt_bars(bar_data,
                                          _rt_bar_canvas.winfo_width() if _rt_bar_canvas else 1))

def _fill_budget(df, ref_date=None):
    ref_d = ref_date or (hist_to_picker.get_date() if _hist_mode else datetime.now().date())
    midx  = MONTH_IDX[ref_d.month]
    yr    = ref_d.year
    if df.empty:
        section_store_and_draw("bud", [(["No Data","-","-","-","-","-"], "evenrow")])
        kpi_cards["bud_pct"].config(text="-"); return
    df2 = df.copy()
    df2["Sales"] = df2.get("NetSaleBFVat", 0) + df2.get("SVCBFTax", 0)
    new_rows = []; bar_data = []; tot_s = tot_b = 0
    for row in df2.itertuples(index=False):
        bid   = next((b for b,n in BRANCH_NAMES.items() if n==row.BranchName), None)
        sales = getattr(row, "Sales", 0.0)
        bgt   = _get_budget(yr, bid, midx)
        pct   = (sales/bgt*100) if bgt else 0
        var   = bgt - sales
        tag   = "hit_hi" if pct>=100 else "hit_mid" if pct>=90 else "hit_lo"
        new_rows.append(([_short(row.BranchName),
                          _fmt(bgt),_fmt(sales),f"{pct:.1f}%",_fmt(var),
                          f"{(var/bgt*100):.1f}%" if bgt else "0.0%"], tag))
        bar_data.append((_short(row.BranchName), pct, _fmt(sales)))
        tot_s += sales; tot_b += bgt
    tp = (tot_s/tot_b*100) if tot_b else 0
    tv = tot_b - tot_s
    new_rows.append((["--- TOTAL ---",_fmt(tot_b),_fmt(tot_s),
                       f"{tp:.1f}%",_fmt(tv),
                       f"{(tv/tot_b*100):.1f}%" if tot_b else "0.0%"],"totalrow"))
    old_rows = _prev_data.get("bud", [])
    _detect_changes_and_blink(tree_bud, new_rows, old_rows)
    _prev_data["bud"] = new_rows
    section_store_and_draw("bud", new_rows)
    _section_data["bud_bars"] = bar_data
    root.after(50, lambda: _draw_bars(bar_data, _bud_bar_canvas.winfo_width() if _bud_bar_canvas else 1))
    on_target = sum(1 for _,p,_ in bar_data if p>=100)
    kpi_cards["bud_pct"].config(text=f"{tp:.1f}%  ({on_target}/{len(bar_data)})")

def _update_kpi_dy(df):
    if df.empty: kpi_cards["dy_net"].config(text="-"); return
    net = df["NetSaleBFVat"].sum() if "NetSaleBFVat" in df.columns else 0
    kpi_cards["dy_net"].config(text=f"{net:,.0f}")

def _update_kpi_mtd(df):
    if df.empty:
        kpi_cards["mtd_net"].config(text="-"); kpi_cards["tot_rev"].config(text="-"); return
    net = df["NetSaleBFVat"].sum() if "NetSaleBFVat" in df.columns else 0
    rev = df["TotalREVENUE"].sum() if "TotalREVENUE" in df.columns else 0
    kpi_cards["mtd_net"].config(text=f"{net:,.0f}")
    kpi_cards["tot_rev"].config(text=f"{rev:,.0f}")

def load_historical():
    global _hist_mode, _live_running, _refresh_job
    from_d = hist_from_picker.get_date()
    to_d   = hist_to_picker.get_date()
    if from_d > to_d:
        messagebox.showwarning("Date Error", "From date must be ≤ To date."); return
    if (to_d - from_d).days > 366:
        messagebox.showwarning("Range Too Large", "Select 366 days or less."); return
    if _refresh_job:
        root.after_cancel(_refresh_job); _refresh_job = None
    _live_running = False; _hist_mode = True
    btn_stop_run.config(state="disabled", fg=C["muted"])
    btn_back_live.config(state="normal")
    lbl_live_status.config(text="⏸ HISTORICAL", fg=C["hist"])
    lbl_status.config(text="* HISTORICAL", fg=C["hist"])
    t_from = "05:00:00" if _hist_time_var.get()=="operational" else "00:00:00"
    from_str = f"{from_d} {t_from}"; to_str = f"{to_d} 23:59:59"
    _show_hist_banner(from_d, to_d)
    lbl_updated.config(text=f"Loading historical …")
    root.update_idletasks()
    df_hist, st = fetch_hq(from_str, to_str)
    _fill_rt(df_hist, st)
    _fill_tree_with_blink(df_hist, "dy",  tree_dy)
    _update_kpi_dy(df_hist)
    _fill_tree_with_blink(df_hist, "mtd", tree_mtd)
    _update_kpi_mtd(df_hist)
    _fill_budget(df_hist, ref_date=to_d)
    lbl_updated.config(
        text=f"Historical: {from_d.strftime('%d %b %Y')} → {to_d.strftime('%d %b %Y')}"
             f"  ({st.upper()})  |  Fetched {datetime.now().strftime('%H:%M:%S')}")

def back_to_live():
    global _hist_mode, _live_running
    _hist_mode = False; _live_running = True
    _hide_hist_banner()
    btn_stop_run.config(state="normal", text="  ⏸  PAUSE LIVE", fg=C["live_on"], bg=C["btn_min"])
    btn_back_live.config(state="disabled")
    lbl_live_status.config(text="● LIVE", fg=C["live_on"])
    lbl_status.config(text="* LIVE", fg=C["ok"])
    refresh_all()

def refresh_all():
    global _refresh_job
    if _hist_mode or not _live_running: return
    today     = datetime.now().date()
    yesterday = today - timedelta(days=1)
    first_dom = today.replace(day=1)
    df_rt, st = fetch_realtime(f"{today} 05:00:00", f"{today} 23:59:00")
    _fill_rt(df_rt, st)
    df_mtd, _ = fetch_realtime(f"{first_dom} 05:00:00", f"{today} 23:59:00")
    _fill_budget(df_mtd)
    _fill_tree_with_blink(df_mtd, "mtd", tree_mtd)
    _update_kpi_mtd(df_mtd)
    df_y, _   = fetch_hq(f"{yesterday} 00:00:00", f"{yesterday} 23:59:59")
    _fill_tree_with_blink(df_y, "dy", tree_dy)
    _update_kpi_dy(df_y)
    lbl_updated.config(
        text=f"Last updated: {datetime.now().strftime('%d %b %Y  %H:%M:%S')}"
             f"  |  Auto-refresh every {REFRESH_INTERVAL_MIN} min  |  Daily reset 04:00")
    _reset_rt_countdown()
    _refresh_job = root.after(REFRESH_INTERVAL_MIN * 60_000, refresh_all)

# ============================================================
# Export Excel (shared helper)
# ============================================================
def _excel_write_section(ws, sr, tree, title, columns, col_labels_override=None, tab_color="1E2D45"):
    MF = "#,##0"; PF = "0.00%"
    ws.merge_cells(start_row=sr, start_column=1, end_row=sr, end_column=len(columns))
    tc = ws.cell(row=sr, column=1, value=title)
    tc.font = Font(bold=True, color="FFFFFF", size=12)
    tc.fill = PatternFill("solid", fgColor=tab_color)
    tc.alignment = Alignment(horizontal="center")
    sr += 1
    for ci, col in enumerate(columns, 1):
        lbl = (col_labels_override or {}).get(col, COLUMN_LABELS.get(col, col))
        c = ws.cell(row=sr, column=ci, value=lbl)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1E2D45")
        c.alignment = Alignment(horizontal="center")
    sr += 1
    for i, item in enumerate(tree.get_children()):
        vals = tree.item(item)["values"]
        z = "111827" if i%2==0 else "0F1520"
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=sr, column=ci, value=v)
            cell.fill = PatternFill("solid", fgColor=z)
            s = str(v).strip()
            try:
                if "%" in s:
                    cell.value = float(s.replace("%","").replace(",",""))/100
                    cell.number_format = PF
                else:
                    cell.value = float(s.replace(",",""))
                    cell.number_format = MF
            except: pass
        sr += 1
    # autofit
    for ci in range(1, len(columns)+1):
        ml = 0
        start = sr - (i+3 if i>=0 else 3)
        for r in range(max(1, start), sr):
            cv = ws.cell(row=r, column=ci).value
            if cv is not None: ml = max(ml, len(str(cv)))
        cd = get_column_letter(ci)
        ws.column_dimensions[cd].width = max(ws.column_dimensions[cd].width or 0, ml+3)
    sr += 2
    return sr

def _export_events_promos_sheet(wb):
    # Events sheet
    we = wb.create_sheet("Events")
    we.cell(1,1,"Name").font = Font(bold=True, color="FFFFFF")
    we.cell(1,1).fill = PatternFill("solid", fgColor="003015")
    for ci, h in enumerate(["Name","Start Date","End Date","Branches","Note"],1):
        c = we.cell(1, ci, h)
        c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor="003015")
    for ri, ev in enumerate(EVENTS, 2):
        we.cell(ri,1, ev.get("name",""))
        we.cell(ri,2, str(ev.get("start_date","")))
        we.cell(ri,3, str(ev.get("end_date","")))
        we.cell(ri,4, ", ".join(ev.get("branches",[])))
        we.cell(ri,5, ev.get("note",""))
    # Promos sheet
    wp = wb.create_sheet("Promotions")
    for ci, h in enumerate(["Name","Start Date","End Date","Branches","Detail"],1):
        c = wp.cell(1, ci, h)
        c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor="2D0030")
    for ri, pr in enumerate(PROMOTIONS, 2):
        wp.cell(ri,1, pr.get("name",""))
        wp.cell(ri,2, str(pr.get("start_date","")))
        wp.cell(ri,3, str(pr.get("end_date","")))
        wp.cell(ri,4, ", ".join(pr.get("branches",[])))
        wp.cell(ri,5, pr.get("detail",""))

def export_live_excel():
    fp = filedialog.asksaveasfilename(defaultextension=".xlsx",
        filetypes=[("Excel files","*.xlsx")], title="Save Live Report")
    if not fp: return
    wb = Workbook()
    ws = wb.active; ws.title = "Sales Report"; sr = 1
    mode_tag = ""
    if _hist_mode:
        fd = hist_from_picker.get_date().strftime('%d %b %Y')
        td = hist_to_picker.get_date().strftime('%d %b %Y')
        mode_tag = f" [HISTORICAL {fd} - {td}]"
    BUD_COLS = ["BranchName","Budget","Actual","% Hit","Variance","% Var"]
    sr = _excel_write_section(ws, sr, tree_rt,  f"Real-time Sales{mode_tag}",          ALL_COLS, tab_color="006080")
    sr = _excel_write_section(ws, sr, tree_dy,  f"Daily Sales Yesterday{mode_tag}",     ALL_COLS, tab_color="006045")
    sr = _excel_write_section(ws, sr, tree_mtd, f"Month-To-Date Sales{mode_tag}",       ALL_COLS, tab_color="003580")
    sr = _excel_write_section(ws, sr, tree_bud, f"Budget vs Actual{mode_tag}",          BUD_COLS,
                               col_labels_override={"BranchName":"Branch","% Hit":"% Hit","% Var":"% Var"},
                               tab_color="4B0082")
    _export_events_promos_sheet(wb)
    wb.save(fp)
    messagebox.showinfo("Export Complete", f"Saved:\n{fp}")

btn_export_live.config(command=export_live_excel)
root.bind("<e>", lambda e: export_live_excel())
root.bind("<E>", lambda e: export_live_excel())
root.bind("<F11>", lambda e: maximize_window())

# ============================================================
# ============================================================
#   TAB 2: CONFIG
# ============================================================
# ============================================================
frame_cfg = tk.Frame(tab_content, bg=C["bg"])
_tab_frames["tab_cfg"] = frame_cfg

cfg_scroll = _make_scrollable(frame_cfg)

def _cfg_section_hdr(parent, text):
    f = tk.Frame(parent, bg=C["card"])
    f.pack(fill="x", padx=S(8), pady=(S(8), 0))
    tk.Frame(f, bg=C["cfg"], height=S(2)).pack(fill="x")
    tk.Label(f, text=text, bg=C["card"], fg=C["cfg"],
             font=FONT_SECTION, padx=S(10), pady=S(5)).pack(side="left")
    return f

# ---- 1.4  Refresh interval (put first, simple) ----
_cfg_section_hdr(cfg_scroll, "  ⏱  1.4  REAL-TIME REFRESH INTERVAL")
ref_int_f = tk.Frame(cfg_scroll, bg=C["panel"])
ref_int_f.pack(fill="x", padx=S(16), pady=S(6))
tk.Label(ref_int_f, text="Refresh every", bg=C["panel"], fg=C["muted"],
         font=FONT_KPI_LBL).pack(side="left", padx=S(6))
ref_int_var = tk.StringVar(value=str(REFRESH_INTERVAL_MIN))
ref_int_sp = tk.Spinbox(ref_int_f, from_=1, to=60, width=4, textvariable=ref_int_var,
                         bg=C["inp_bg"], fg=C["hist"], font=("Consolas",FS_DATA,"bold"),
                         relief="flat", bd=0, buttonbackground=C["sep"],
                         highlightthickness=0, justify="center")
ref_int_sp.pack(side="left", padx=S(4))
tk.Label(ref_int_f, text="minutes", bg=C["panel"], fg=C["muted"], font=FONT_KPI_LBL).pack(side="left")

def _apply_refresh_interval():
    global REFRESH_INTERVAL_MIN, _refresh_job, _rt_countdown_secs
    try: new_val = max(1, min(60, int(float(ref_int_var.get()))))
    except: messagebox.showwarning("Invalid", "Enter a number 1-60."); return
    REFRESH_INTERVAL_MIN = new_val
    _rt_countdown_secs   = new_val * 60
    if _refresh_job:
        root.after_cancel(_refresh_job); _refresh_job = None
    if _live_running and not _hist_mode:
        _refresh_job = root.after(REFRESH_INTERVAL_MIN * 60_000, refresh_all)
    messagebox.showinfo("Saved", f"Refresh interval set to {new_val} minute(s).")

b = _accent_btn(ref_int_f, "  APPLY", C["save_btn"], C["save_fg"], _apply_refresh_interval)
b.pack(side="left", padx=S(10))

tk.Frame(cfg_scroll, bg=C["sep"], height=1).pack(fill="x", padx=S(8), pady=S(6))

# ---- 1.1  Budget editor ----
_cfg_section_hdr(cfg_scroll, "  💰  1.1  BUDGET EDITOR  (7-year range)")

# Year picker row
bud_yr_f = tk.Frame(cfg_scroll, bg=C["panel"])
bud_yr_f.pack(fill="x", padx=S(16), pady=(S(4),0))
cur_year = datetime.now().year
tk.Label(bud_yr_f, text="Year:", bg=C["panel"], fg=C["muted"], font=FONT_KPI_LBL).pack(side="left", padx=S(4))
bud_year_var = tk.StringVar(value=str(cur_year))
bud_year_sp  = tk.Spinbox(bud_yr_f, from_=cur_year-3, to=cur_year+3, width=5,
                            textvariable=bud_year_var,
                            bg=C["inp_bg"], fg=C["gold"], font=("Consolas",FS_DATA,"bold"),
                            relief="flat", bd=0, buttonbackground=C["sep"],
                            highlightthickness=0, justify="center")
bud_year_sp.pack(side="left", padx=S(4))

tk.Label(bud_yr_f, text="Branch:", bg=C["panel"], fg=C["muted"], font=FONT_KPI_LBL).pack(side="left", padx=(S(10),S(4)))
bud_branch_var = tk.StringVar()
all_branch_names = [BRANCH_NAMES[bid] for bid in BRANCH_IDS]
bud_branch_cb = ttk.Combobox(bud_yr_f, textvariable=bud_branch_var,
                               values=all_branch_names, state="readonly",
                               width=28, font=FONT_KPI_LBL)
bud_branch_cb.current(0)
bud_branch_cb.pack(side="left", padx=S(4))

# Grid of 12 months
bud_grid_f = tk.Frame(cfg_scroll, bg=C["panel"])
bud_grid_f.pack(fill="x", padx=S(16), pady=S(4))
bud_month_vars = []
for mi, mname in enumerate(MONTH_NAMES):
    tk.Label(bud_grid_f, text=mname, bg=C["panel"], fg=C["muted"],
             font=FONT_KPI_LBL, width=6, anchor="e").grid(row=0, column=mi*2, padx=S(2), pady=S(2))
    var = tk.StringVar(value="0")
    ent = tk.Entry(bud_grid_f, textvariable=var, width=10,
                   bg=C["inp_bg"], fg=C["gold"], insertbackground=C["gold"],
                   relief="flat", font=("Consolas",FS_DATA),
                   highlightthickness=1, highlightcolor=C["sep"],
                   highlightbackground=C["sep"], justify="right")
    ent.grid(row=0, column=mi*2+1, padx=S(2), pady=S(2))
    bud_month_vars.append(var)

def _load_bud_to_grid(*args):
    try:
        yr  = int(bud_year_var.get())
        bid = next(b for b,n in BRANCH_NAMES.items() if n==bud_branch_var.get())
        vals = BUDGET_STORE.get(yr, {}).get(bid, BASE_BUDGET.get(bid,[0]*12))
        for mi, var in enumerate(bud_month_vars):
            var.set(f"{vals[mi]:,.0f}")
    except: pass

bud_year_sp.config(command=_load_bud_to_grid)
bud_branch_cb.bind("<<ComboboxSelected>>", _load_bud_to_grid)
_load_bud_to_grid()

def _save_budget_grid():
    try:
        yr  = int(bud_year_var.get())
        bid = next(b for b,n in BRANCH_NAMES.items() if n==bud_branch_var.get())
        vals = []
        for var in bud_month_vars:
            vals.append(float(var.get().replace(",","")))
        if yr not in BUDGET_STORE: BUDGET_STORE[yr] = {}
        BUDGET_STORE[yr][bid] = vals
        messagebox.showinfo("Saved", f"Budget saved for {BRANCH_NAMES[bid]} ({yr})")
    except Exception as ex:
        messagebox.showerror("Error", str(ex))

def _download_bud_template():
    fp = filedialog.asksaveasfilename(defaultextension=".xlsx",
        filetypes=[("Excel","*.xlsx")], title="Save Budget Template")
    if not fp: return
    wb = Workbook(); ws = wb.active; ws.title = "Budget"
    headers = ["Year","Branch"] + MONTH_NAMES
    for ci, h in enumerate(headers,1):
        c = ws.cell(1, ci, h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="2A1800")
    row = 2
    for yr in range(cur_year-3, cur_year+4):
        for bid in BRANCH_IDS:
            ws.cell(row, 1, yr)
            ws.cell(row, 2, BRANCH_NAMES[bid])
            vals = BUDGET_STORE.get(yr, {}).get(bid, BASE_BUDGET.get(bid,[0]*12))
            for mi, v in enumerate(vals):
                ws.cell(row, 3+mi, v)
            row += 1
    wb.save(fp)
    messagebox.showinfo("Template Downloaded", fp)

def _import_bud_excel():
    fp = filedialog.askopenfilename(filetypes=[("Excel","*.xlsx")], title="Import Budget Excel")
    if not fp: return
    try:
        wb = load_workbook(fp, data_only=True)
        ws = wb.active
        headers = [str(ws.cell(1,c).value).strip() for c in range(1, ws.max_column+1)]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None: continue
            try:
                yr  = int(row[0])
                bname = str(row[1]).strip()
                bid = next((b for b,n in BRANCH_NAMES.items() if n==bname), None)
                if bid is None: continue
                vals = [float(row[i] or 0) for i in range(2, 14)]
                if yr not in BUDGET_STORE: BUDGET_STORE[yr] = {}
                BUDGET_STORE[yr][bid] = vals
            except: continue
        _load_bud_to_grid()
        messagebox.showinfo("Import OK", "Budget imported successfully.")
    except Exception as ex:
        messagebox.showerror("Import Error", str(ex))

bud_btn_f = tk.Frame(cfg_scroll, bg=C["panel"])
bud_btn_f.pack(fill="x", padx=S(16), pady=(S(2), S(6)))
for txt, bg, fg, cmd in [
    ("  SAVE",              C["save_btn"], C["save_fg"], _save_budget_grid),
    ("  DOWNLOAD TEMPLATE", C["imp_btn"],  C["imp_fg"],  _download_bud_template),
    ("  IMPORT EXCEL",      C["imp_btn"],  C["imp_fg"],  _import_bud_excel),
]:
    b = _accent_btn(bud_btn_f, txt, bg, fg, cmd)
    b.pack(side="left", padx=S(4))

tk.Frame(cfg_scroll, bg=C["sep"], height=1).pack(fill="x", padx=S(8), pady=S(6))

# ---- 1.2  Event Manager ----
_cfg_section_hdr(cfg_scroll, "  📅  1.2  EVENT MANAGER")

ev_list_f = tk.Frame(cfg_scroll, bg=C["panel"])
ev_list_f.pack(fill="x", padx=S(16), pady=(S(4),0))

EV_COLS = ("ID","Name","Start","End","Branches","Note")
ev_tree = ttk.Treeview(ev_list_f, columns=EV_COLS, show="headings",
                        height=5, style="Cfg.Treeview")
for col, w in zip(EV_COLS, [S(30),S(160),S(90),S(90),S(200),S(200)]):
    ev_tree.heading(col, text=col)
    ev_tree.column(col, width=w, minwidth=S(20))
ev_tree.pack(fill="x")

def _refresh_ev_tree():
    ev_tree.delete(*ev_tree.get_children())
    for ev in EVENTS:
        ev_tree.insert("","end", values=(
            ev["id"], ev["name"],
            str(ev["start_date"]), str(ev["end_date"]),
            ", ".join(ev["branches"]), ev.get("note","")
        ))

# Event form
ev_form_f = tk.Frame(cfg_scroll, bg=C["panel"])
ev_form_f.pack(fill="x", padx=S(16), pady=S(4))
ev_name_var = tk.StringVar()
ev_note_var = tk.StringVar()
ev_branches_var = tk.StringVar()
ev_id_editing = tk.StringVar(value="")

tk.Label(ev_form_f, text="Name:", bg=C["panel"], fg=C["muted"], font=FONT_KPI_LBL).grid(row=0,column=0,sticky="w",padx=S(4),pady=S(2))
tk.Entry(ev_form_f, textvariable=ev_name_var, width=30, bg=C["inp_bg"], fg=C["event_fg"],
         insertbackground=C["event_fg"], relief="flat", font=("Consolas",FS_DATA)).grid(row=0,column=1,padx=S(4),pady=S(2))

today_d2 = datetime.now().date()
ev_start_f = _date_picker_frame(ev_form_f, today_d2, "Start:")
ev_start_f.grid(row=0, column=2, padx=S(4), pady=S(2))
ev_end_f = _date_picker_frame(ev_form_f, today_d2, "End:")
ev_end_f.grid(row=0, column=3, padx=S(4), pady=S(2))

tk.Label(ev_form_f, text="Branches (ALL or comma list):", bg=C["panel"], fg=C["muted"],
         font=FONT_KPI_LBL).grid(row=1,column=0,sticky="w",padx=S(4),pady=S(2))
tk.Entry(ev_form_f, textvariable=ev_branches_var, width=40, bg=C["inp_bg"], fg=C["event_fg"],
         insertbackground=C["event_fg"], relief="flat", font=("Consolas",FS_DATA)).grid(row=1,column=1,columnspan=2,padx=S(4),pady=S(2))
tk.Label(ev_form_f, text="Note:", bg=C["panel"], fg=C["muted"], font=FONT_KPI_LBL).grid(row=2,column=0,sticky="w",padx=S(4),pady=S(2))
tk.Entry(ev_form_f, textvariable=ev_note_var, width=60, bg=C["inp_bg"], fg=C["event_fg"],
         insertbackground=C["event_fg"], relief="flat", font=("Consolas",FS_DATA)).grid(row=2,column=1,columnspan=3,padx=S(4),pady=S(2))

def _add_event():
    global _next_event_id
    name = ev_name_var.get().strip()
    if not name: messagebox.showwarning("Missing", "Event name required."); return
    br_raw = ev_branches_var.get().strip()
    branches = ["ALL"] if not br_raw or br_raw.upper()=="ALL" else [b.strip() for b in br_raw.split(",")]
    ev = {"id":_next_event_id, "name":name,
          "start_date":ev_start_f.get_date(), "end_date":ev_end_f.get_date(),
          "branches":branches, "note":ev_note_var.get()}
    EVENTS.append(ev); _next_event_id += 1
    _refresh_ev_tree()

def _edit_event():
    sel = ev_tree.selection()
    if not sel: messagebox.showinfo("Select","Select an event to edit."); return
    ev_id = int(ev_tree.item(sel[0])["values"][0])
    ev    = next((e for e in EVENTS if e["id"]==ev_id), None)
    if not ev: return
    name = ev_name_var.get().strip()
    if not name: messagebox.showwarning("Missing","Event name required."); return
    br_raw = ev_branches_var.get().strip()
    ev["name"]       = name
    ev["start_date"] = ev_start_f.get_date()
    ev["end_date"]   = ev_end_f.get_date()
    ev["branches"]   = ["ALL"] if not br_raw or br_raw.upper()=="ALL" else [b.strip() for b in br_raw.split(",")]
    ev["note"]       = ev_note_var.get()
    _refresh_ev_tree()

def _delete_event():
    sel = ev_tree.selection()
    if not sel: messagebox.showinfo("Select","Select an event to delete."); return
    ev_id = int(ev_tree.item(sel[0])["values"][0])
    EVENTS[:] = [e for e in EVENTS if e["id"]!=ev_id]
    _refresh_ev_tree()

def _load_ev_to_form(event=None):
    sel = ev_tree.selection()
    if not sel: return
    vals = ev_tree.item(sel[0])["values"]
    ev_name_var.set(vals[1])
    ev_branches_var.set(vals[4])
    ev_note_var.set(vals[5])

ev_tree.bind("<<TreeviewSelect>>", _load_ev_to_form)

def _download_ev_template():
    fp = filedialog.asksaveasfilename(defaultextension=".xlsx",
        filetypes=[("Excel","*.xlsx")], title="Event Template")
    if not fp: return
    wb = Workbook(); ws = wb.active; ws.title = "Events"
    for ci, h in enumerate(["Name","Start (YYYY-MM-DD)","End (YYYY-MM-DD)","Branches","Note"],1):
        c = ws.cell(1,ci,h); c.font=Font(bold=True,color="FFFFFF")
        c.fill=PatternFill("solid",fgColor="003015")
    ws.cell(2,1,"Sample Event"); ws.cell(2,2,str(today_d2)); ws.cell(2,3,str(today_d2))
    ws.cell(2,4,"ALL"); ws.cell(2,5,"Example note")
    wb.save(fp); messagebox.showinfo("Template", fp)

def _import_ev_excel():
    global _next_event_id
    fp = filedialog.askopenfilename(filetypes=[("Excel","*.xlsx")])
    if not fp: return
    try:
        wb = load_workbook(fp, data_only=True); ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None: continue
            try:
                sd = datetime.strptime(str(row[1]).strip(), "%Y-%m-%d").date()
                ed = datetime.strptime(str(row[2]).strip(), "%Y-%m-%d").date()
                br_raw = str(row[3]).strip()
                branches = ["ALL"] if not br_raw or br_raw.upper()=="ALL" else [b.strip() for b in br_raw.split(",")]
                EVENTS.append({"id":_next_event_id,"name":str(row[0]),"start_date":sd,
                                "end_date":ed,"branches":branches,"note":str(row[4] or "")})
                _next_event_id += 1
            except: continue
        _refresh_ev_tree()
        messagebox.showinfo("Import OK","Events imported.")
    except Exception as ex:
        messagebox.showerror("Import Error", str(ex))

ev_btn_f = tk.Frame(cfg_scroll, bg=C["panel"])
ev_btn_f.pack(fill="x", padx=S(16), pady=(S(2),S(6)))
for txt, bg, fg, cmd in [
    ("  ADD",              C["add_btn"],  C["add_fg"],  _add_event),
    ("  SAVE EDIT",        C["save_btn"], C["save_fg"], _edit_event),
    ("  DELETE",           C["del_btn"],  C["del_fg"],  _delete_event),
    ("  DOWNLOAD TEMPLATE",C["imp_btn"],  C["imp_fg"],  _download_ev_template),
    ("  IMPORT EXCEL",     C["imp_btn"],  C["imp_fg"],  _import_ev_excel),
]:
    b = _accent_btn(ev_btn_f, txt, bg, fg, cmd); b.pack(side="left", padx=S(3))

tk.Frame(cfg_scroll, bg=C["sep"], height=1).pack(fill="x", padx=S(8), pady=S(6))

# ---- 1.3  Promotion Manager ----
_cfg_section_hdr(cfg_scroll, "  🎯  1.3  PROMOTION MANAGER")

pr_list_f = tk.Frame(cfg_scroll, bg=C["panel"])
pr_list_f.pack(fill="x", padx=S(16), pady=(S(4),0))

PR_COLS = ("ID","Name","Start","End","Branches","Detail")
pr_tree = ttk.Treeview(pr_list_f, columns=PR_COLS, show="headings",
                        height=5, style="Cfg.Treeview")
for col, w in zip(PR_COLS, [S(30),S(160),S(90),S(90),S(200),S(200)]):
    pr_tree.heading(col, text=col)
    pr_tree.column(col, width=w, minwidth=S(20))
pr_tree.pack(fill="x")

def _refresh_pr_tree():
    pr_tree.delete(*pr_tree.get_children())
    for pr in PROMOTIONS:
        pr_tree.insert("","end", values=(
            pr["id"], pr["name"],
            str(pr["start_date"]), str(pr["end_date"]),
            ", ".join(pr["branches"]), pr.get("detail","")
        ))

pr_form_f = tk.Frame(cfg_scroll, bg=C["panel"])
pr_form_f.pack(fill="x", padx=S(16), pady=S(4))
pr_name_var   = tk.StringVar()
pr_detail_var = tk.StringVar()
pr_branches_var = tk.StringVar()

tk.Label(pr_form_f, text="Name:", bg=C["panel"], fg=C["muted"], font=FONT_KPI_LBL).grid(row=0,column=0,sticky="w",padx=S(4),pady=S(2))
tk.Entry(pr_form_f, textvariable=pr_name_var, width=30, bg=C["inp_bg"], fg=C["promo_fg"],
         insertbackground=C["promo_fg"], relief="flat", font=("Consolas",FS_DATA)).grid(row=0,column=1,padx=S(4),pady=S(2))
pr_start_f = _date_picker_frame(pr_form_f, today_d2, "Start:")
pr_start_f.grid(row=0, column=2, padx=S(4), pady=S(2))
pr_end_f = _date_picker_frame(pr_form_f, today_d2, "End:")
pr_end_f.grid(row=0, column=3, padx=S(4), pady=S(2))
tk.Label(pr_form_f, text="Branches:", bg=C["panel"], fg=C["muted"], font=FONT_KPI_LBL).grid(row=1,column=0,sticky="w",padx=S(4),pady=S(2))
tk.Entry(pr_form_f, textvariable=pr_branches_var, width=40, bg=C["inp_bg"], fg=C["promo_fg"],
         insertbackground=C["promo_fg"], relief="flat", font=("Consolas",FS_DATA)).grid(row=1,column=1,columnspan=2,padx=S(4),pady=S(2))
tk.Label(pr_form_f, text="Detail:", bg=C["panel"], fg=C["muted"], font=FONT_KPI_LBL).grid(row=2,column=0,sticky="w",padx=S(4),pady=S(2))
tk.Entry(pr_form_f, textvariable=pr_detail_var, width=60, bg=C["inp_bg"], fg=C["promo_fg"],
         insertbackground=C["promo_fg"], relief="flat", font=("Consolas",FS_DATA)).grid(row=2,column=1,columnspan=3,padx=S(4),pady=S(2))

def _add_promo():
    global _next_promo_id
    name = pr_name_var.get().strip()
    if not name: messagebox.showwarning("Missing","Promo name required."); return
    br_raw = pr_branches_var.get().strip()
    branches = ["ALL"] if not br_raw or br_raw.upper()=="ALL" else [b.strip() for b in br_raw.split(",")]
    PROMOTIONS.append({"id":_next_promo_id,"name":name,
                        "start_date":pr_start_f.get_date(),"end_date":pr_end_f.get_date(),
                        "branches":branches,"detail":pr_detail_var.get()})
    _next_promo_id += 1; _refresh_pr_tree()

def _edit_promo():
    sel = pr_tree.selection()
    if not sel: messagebox.showinfo("Select","Select a promotion to edit."); return
    pr_id = int(pr_tree.item(sel[0])["values"][0])
    pr    = next((p for p in PROMOTIONS if p["id"]==pr_id), None)
    if not pr: return
    name = pr_name_var.get().strip()
    if not name: messagebox.showwarning("Missing","Promo name required."); return
    br_raw = pr_branches_var.get().strip()
    pr["name"]       = name
    pr["start_date"] = pr_start_f.get_date()
    pr["end_date"]   = pr_end_f.get_date()
    pr["branches"]   = ["ALL"] if not br_raw or br_raw.upper()=="ALL" else [b.strip() for b in br_raw.split(",")]
    pr["detail"]     = pr_detail_var.get()
    _refresh_pr_tree()

def _delete_promo():
    sel = pr_tree.selection()
    if not sel: messagebox.showinfo("Select","Select a promotion to delete."); return
    pr_id = int(pr_tree.item(sel[0])["values"][0])
    PROMOTIONS[:] = [p for p in PROMOTIONS if p["id"]!=pr_id]
    _refresh_pr_tree()

def _load_pr_to_form(event=None):
    sel = pr_tree.selection()
    if not sel: return
    vals = pr_tree.item(sel[0])["values"]
    pr_name_var.set(vals[1]); pr_branches_var.set(vals[4]); pr_detail_var.set(vals[5])

pr_tree.bind("<<TreeviewSelect>>", _load_pr_to_form)

def _download_pr_template():
    fp = filedialog.asksaveasfilename(defaultextension=".xlsx",
        filetypes=[("Excel","*.xlsx")], title="Promo Template")
    if not fp: return
    wb = Workbook(); ws = wb.active; ws.title = "Promotions"
    for ci, h in enumerate(["Name","Start (YYYY-MM-DD)","End (YYYY-MM-DD)","Branches","Detail"],1):
        c = ws.cell(1,ci,h); c.font=Font(bold=True,color="FFFFFF")
        c.fill=PatternFill("solid",fgColor="2D0030")
    ws.cell(2,1,"Summer Special"); ws.cell(2,2,str(today_d2)); ws.cell(2,3,str(today_d2))
    ws.cell(2,4,"ALL"); ws.cell(2,5,"Buy 2 get 1 free")
    wb.save(fp); messagebox.showinfo("Template", fp)

def _import_pr_excel():
    global _next_promo_id
    fp = filedialog.askopenfilename(filetypes=[("Excel","*.xlsx")])
    if not fp: return
    try:
        wb = load_workbook(fp, data_only=True); ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None: continue
            try:
                sd = datetime.strptime(str(row[1]).strip(), "%Y-%m-%d").date()
                ed = datetime.strptime(str(row[2]).strip(), "%Y-%m-%d").date()
                br_raw = str(row[3]).strip()
                branches = ["ALL"] if not br_raw or br_raw.upper()=="ALL" else [b.strip() for b in br_raw.split(",")]
                PROMOTIONS.append({"id":_next_promo_id,"name":str(row[0]),"start_date":sd,
                                    "end_date":ed,"branches":branches,"detail":str(row[4] or "")})
                _next_promo_id += 1
            except: continue
        _refresh_pr_tree(); messagebox.showinfo("Import OK","Promotions imported.")
    except Exception as ex:
        messagebox.showerror("Import Error", str(ex))

pr_btn_f = tk.Frame(cfg_scroll, bg=C["panel"])
pr_btn_f.pack(fill="x", padx=S(16), pady=(S(2),S(10)))
for txt, bg, fg, cmd in [
    ("  ADD",              C["add_btn"],  C["add_fg"],  _add_promo),
    ("  SAVE EDIT",        C["save_btn"], C["save_fg"], _edit_promo),
    ("  DELETE",           C["del_btn"],  C["del_fg"],  _delete_promo),
    ("  DOWNLOAD TEMPLATE",C["imp_btn"],  C["imp_fg"],  _download_pr_template),
    ("  IMPORT EXCEL",     C["imp_btn"],  C["imp_fg"],  _import_pr_excel),
]:
    b = _accent_btn(pr_btn_f, txt, bg, fg, cmd); b.pack(side="left", padx=S(3))

# ============================================================
# ============================================================
#   TAB 3: 3-YEAR COMPARISON
# ============================================================
# ============================================================
frame_cmp = tk.Frame(tab_content, bg=C["bg"])
_tab_frames["tab_cmp"] = frame_cmp

cmp_scroll = _make_scrollable(frame_cmp)

_cmp_section_hdr = lambda t: _cfg_section_hdr.__wrapped__(cmp_scroll, t) if False else None

tk.Frame(cmp_scroll, bg=C["cmp"], height=S(2)).pack(fill="x")
tk.Label(cmp_scroll, text="  📊  3-YEAR COMPARISON", bg=C["card"], fg=C["cmp"],
         font=FONT_SECTION, padx=S(10), pady=S(5)).pack(anchor="w")

# Date range pickers for 3 periods
cmp_pickers_f = tk.Frame(cmp_scroll, bg=C["panel"])
cmp_pickers_f.pack(fill="x", padx=S(8), pady=S(6))

cmp_pickers = []
_cmp_period_colors = [C["rt"], C["dy"], C["cfg"]]
today_for_cmp = datetime.now().date()

for i in range(3):
    pf = tk.Frame(cmp_pickers_f, bg=C["panel"])
    pf.pack(side="left", padx=S(8))
    default_from = today_for_cmp.replace(year=today_for_cmp.year-i, day=1)
    default_to   = default_from.replace(day=calendar.monthrange(default_from.year, default_from.month)[1])
    color = _cmp_period_colors[i]
    tk.Frame(pf, bg=color, height=S(2)).pack(fill="x")
    tk.Label(pf, text=f"Period {i+1}", bg=C["panel"], fg=color, font=FONT_SECTION).pack(anchor="w", pady=(S(2),0))
    lbl_f = tk.Label(pf, text="From:", bg=C["panel"], fg=C["muted"], font=FONT_KPI_LBL)
    lbl_f.pack(anchor="w")
    pf_from = _date_picker_frame(pf, default_from)
    pf_from.pack(anchor="w", pady=(0, S(2)))
    lbl_t = tk.Label(pf, text="To:", bg=C["panel"], fg=C["muted"], font=FONT_KPI_LBL)
    lbl_t.pack(anchor="w")
    pf_to = _date_picker_frame(pf, default_to)
    pf_to.pack(anchor="w")
    cmp_pickers.append((pf_from, pf_to))

# Column selection
cmp_col_f = tk.Frame(cmp_scroll, bg=C["panel"])
cmp_col_f.pack(fill="x", padx=S(8), pady=(S(4),0))
tk.Label(cmp_col_f, text="Columns to compare:", bg=C["panel"], fg=C["muted"],
         font=FONT_KPI_LBL).pack(side="left", padx=S(6))
cmp_col_vars = {}
for col in COLUMNS:
    var = tk.BooleanVar(value=(col in ["NetSaleBFVat","SVCBFTax","TotalREVENUE"]))
    cb  = tk.Checkbutton(cmp_col_f, text=COLUMN_LABELS[col], variable=var,
                          bg=C["panel"], fg=C["text"], selectcolor=C["bg"],
                          activebackground=C["panel"], activeforeground=C["cmp"],
                          font=FONT_KPI_LBL)
    cb.pack(side="left", padx=S(3))
    cmp_col_vars[col] = var

# Time option
cmp_time_f = tk.Frame(cmp_scroll, bg=C["panel"])
cmp_time_f.pack(fill="x", padx=S(8), pady=(S(2),S(4)))
tk.Label(cmp_time_f, text="Time:", bg=C["panel"], fg=C["muted"], font=FONT_KPI_LBL).pack(side="left", padx=S(6))
cmp_time_var = tk.StringVar(value="operational")
for txt, val in [("05:00-23:59","operational"),("00:00-23:59","fullday")]:
    tk.Radiobutton(cmp_time_f, text=txt, variable=cmp_time_var, value=val,
                   bg=C["panel"], fg=C["muted"], selectcolor=C["bg"],
                   activebackground=C["panel"], font=FONT_KPI_LBL).pack(side="left", padx=S(3))

# Results area
cmp_result_f = tk.Frame(cmp_scroll, bg=C["bg"])
cmp_result_f.pack(fill="both", expand=True, padx=S(8), pady=S(4))

# Promo/Event info label
cmp_promo_lbl = tk.Label(cmp_scroll, text="", bg=C["bg"], fg=C["promo_fg"],
                           font=("Segoe UI", FS_STATUS), wraplength=WIN_W-100, justify="left")
cmp_promo_lbl.pack(anchor="w", padx=S(10), pady=(0, S(4)))
cmp_event_lbl = tk.Label(cmp_scroll, text="", bg=C["bg"], fg=C["event_fg"],
                           font=("Segoe UI", FS_STATUS), wraplength=WIN_W-100, justify="left")
cmp_event_lbl.pack(anchor="w", padx=S(10), pady=(0, S(6)))

_cmp_trees = []

def _get_active_items(items, from_d, to_d):
    result = []
    for it in items:
        sd = it.get("start_date"); ed = it.get("end_date")
        if sd and ed:
            # overlap
            if not (ed < from_d or sd > to_d):
                result.append(it)
    return result

def run_comparison():
    # Clear previous results
    for w in cmp_result_f.winfo_children():
        w.destroy()
    _cmp_trees.clear()

    selected_cols = [c for c in COLUMNS if cmp_col_vars[c].get()]
    if not selected_cols:
        messagebox.showwarning("No Columns","Select at least one column."); return

    t_from = "05:00:00" if cmp_time_var.get()=="operational" else "00:00:00"

    all_dfs = []
    all_labels = []
    for i, (pf_from, pf_to) in enumerate(cmp_pickers):
        fd = pf_from.get_date(); td = pf_to.get_date()
        all_labels.append(f"P{i+1}: {fd.strftime('%d %b %Y')} → {td.strftime('%d %b %Y')}")
        lbl_val = tk.Label(cmp_result_f, text=f"Loading period {i+1} …",
                            bg=C["bg"], fg=_cmp_period_colors[i], font=FONT_KPI_LBL)
        lbl_val.pack(anchor="w", padx=S(4))
        root.update_idletasks()
        df, st = fetch_hq(f"{fd} {t_from}", f"{td} 23:59:59")
        all_dfs.append((df, fd, td, st))
        lbl_val.destroy()

    # Build comparison tree per branch
    display_cols = ["BranchName"] + [c for p in range(3) for c in selected_cols]
    col_ids = []
    col_ids.append("Branch")
    for pi, lbl in enumerate(all_labels):
        for c in selected_cols:
            col_ids.append(f"P{pi+1}_{c}")

    tree_f = tk.Frame(cmp_result_f, bg=C["card"])
    tree_f.pack(fill="both", expand=True)
    tk.Frame(tree_f, bg=C["cmp"], height=S(2)).pack(fill="x")
    lbl_header_f = tk.Frame(tree_f, bg=C["card"])
    lbl_header_f.pack(fill="x", padx=S(8), pady=S(3))
    for i, (_, fd, td, st) in enumerate(all_dfs):
        color = _cmp_period_colors[i]
        tk.Label(lbl_header_f, text=f"  P{i+1}: {fd.strftime('%d %b %Y')} → {td.strftime('%d %b %Y')}  ({st.upper()})",
                 bg=C["card"], fg=color, font=FONT_KPI_LBL).pack(side="left", padx=S(6))

    # Compound headings
    tv_cols = ["Branch"]
    for pi in range(3):
        for c in selected_cols:
            tv_cols.append(f"P{pi+1}_{c}")

    tv = ttk.Treeview(tree_f, columns=tv_cols, show="headings",
                       height=12, style="TV.Treeview")
    vsb = ttk.Scrollbar(tree_f, orient="vertical", command=tv.yview)
    tv.configure(yscrollcommand=vsb.set)
    tv.pack(side="left", fill="both", expand=True, padx=S(6), pady=S(4))
    vsb.pack(side="right", fill="y")
    _configure_tree(tv)
    tv.tag_configure("up",   foreground="#00FFB2")
    tv.tag_configure("down", foreground="#FF6B8A")

    tv.heading("Branch", text="Branch", anchor="w")
    tv.column("Branch", width=COL_BRANCH, anchor="w", minwidth=S(60))
    for pi in range(3):
        c = _cmp_period_colors[pi]
        for col in selected_cols:
            cid = f"P{pi+1}_{col}"
            tv.heading(cid, text=f"P{pi+1} {COLUMN_LABELS[col]}", anchor="e")
            tv.column(cid, width=COL_NUM, anchor="e", minwidth=S(50))

    # Fill data
    for i, row_bid in enumerate(BRANCH_IDS):
        bname = BRANCH_NAMES[row_bid]
        sname = BRANCH_SHORT.get(bname, bname)
        vals  = [sname]
        for df, fd, td, st in all_dfs:
            for col in selected_cols:
                if not df.empty and bname in df["BranchName"].values:
                    r = df[df["BranchName"]==bname].iloc[0]
                    vals.append(_fmt(getattr(r, col, 0.0)))
                else:
                    vals.append("-")
        tag = "evenrow" if i%2==0 else "oddrow"
        tv.insert("","end", values=vals, tags=(tag,))

    # Totals
    tot_vals = ["--- TOTAL ---"]
    for df, _, _, _ in all_dfs:
        for col in selected_cols:
            if not df.empty and col in df.columns:
                tot_vals.append(_fmt(df[col].sum()))
            else:
                tot_vals.append("-")
    tv.insert("","end", values=tot_vals, tags=("totalrow",))
    _cmp_trees.append(tv)

    # Show Promos & Events active during any period
    all_promos_active = set(); all_events_active = set()
    for df, fd, td, st in all_dfs:
        for pr in _get_active_items(PROMOTIONS, fd, td):
            all_promos_active.add(pr["name"])
        for ev in _get_active_items(EVENTS, fd, td):
            all_events_active.add(ev["name"])
    if all_promos_active:
        cmp_promo_lbl.config(text="🎯 Active Promotions in selected periods: " + " | ".join(sorted(all_promos_active)))
    else:
        cmp_promo_lbl.config(text="")
    if all_events_active:
        cmp_event_lbl.config(text="📅 Active Events in selected periods: " + " | ".join(sorted(all_events_active)))
    else:
        cmp_event_lbl.config(text="")

cmp_btn_f = tk.Frame(cmp_scroll, bg=C["bg"])
cmp_btn_f.pack(fill="x", padx=S(8), pady=(0, S(4)))

def _export_cmp_excel():
    if not _cmp_trees:
        messagebox.showinfo("No Data","Run comparison first."); return
    fp = filedialog.asksaveasfilename(defaultextension=".xlsx",
        filetypes=[("Excel","*.xlsx")], title="Save Comparison")
    if not fp: return
    wb = Workbook(); ws = wb.active; ws.title = "Comparison"
    tv = _cmp_trees[0]
    cols = tv["columns"]
    for ci, h in enumerate(cols, 1):
        c = ws.cell(1, ci, tv.heading(h)["text"])
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="2A0050")
    for ri, item in enumerate(tv.get_children(), 2):
        vals = tv.item(item)["values"]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(ri, ci, v)
            s = str(v).strip()
            try:
                if s not in ("-","--- TOTAL ---"):
                    cell.value = float(s.replace(",",""))
                    cell.number_format = "#,##0"
            except: pass
    _export_events_promos_sheet(wb)
    # Promos active note
    wn = wb.create_sheet("Active Periods Note")
    wn.cell(1,1,"Active Promotions in comparison periods")
    wn.cell(1,1).font = Font(bold=True)
    for ri, pr in enumerate(PROMOTIONS, 2):
        wn.cell(ri,1,pr["name"]); wn.cell(ri,2,str(pr["start_date"])); wn.cell(ri,3,str(pr["end_date"]))
    wb.save(fp); messagebox.showinfo("Exported", fp)

b_run = _accent_btn(cmp_btn_f, "  ▶  RUN COMPARISON", "#002040", C["cmp"], run_comparison)
b_run.pack(side="left", padx=S(4))
b_exp = _accent_btn(cmp_btn_f, "  EXPORT EXCEL", C["save_btn"], C["save_fg"], _export_cmp_excel)
b_exp.pack(side="left", padx=S(4))

# ============================================================
# ============================================================
#   TAB 4: BUDGET vs ACTUAL (custom date picker)
# ============================================================
# ============================================================
frame_bva = tk.Frame(tab_content, bg=C["bg"])
_tab_frames["tab_bva"] = frame_bva

bva_scroll = _make_scrollable(frame_bva)

tk.Frame(bva_scroll, bg=C["bva"], height=S(2)).pack(fill="x")
tk.Label(bva_scroll, text="  📈  BUDGET vs ACTUAL  (Custom Date Range)",
         bg=C["card"], fg=C["bva"], font=FONT_SECTION, padx=S(10), pady=S(5)).pack(anchor="w")

bva_ctrl_f = tk.Frame(bva_scroll, bg=C["panel"])
bva_ctrl_f.pack(fill="x", padx=S(10), pady=S(6))
tk.Label(bva_ctrl_f, text="From:", bg=C["panel"], fg=C["muted"], font=FONT_KPI_LBL).pack(side="left", padx=S(4))
bva_from_picker = _date_picker_frame(bva_ctrl_f, today_d2.replace(day=1))
bva_from_picker.pack(side="left", padx=S(4))
tk.Label(bva_ctrl_f, text="  To:", bg=C["panel"], fg=C["muted"], font=FONT_KPI_LBL).pack(side="left")
bva_to_picker = _date_picker_frame(bva_ctrl_f, today_d2)
bva_to_picker.pack(side="left", padx=S(4))
tk.Label(bva_ctrl_f, text="  Time:", bg=C["panel"], fg=C["muted"], font=FONT_KPI_LBL).pack(side="left", padx=(S(6),0))
bva_time_var = tk.StringVar(value="operational")
for txt, val in [("05:00-23:59","operational"),("00:00-23:59","fullday")]:
    tk.Radiobutton(bva_ctrl_f, text=txt, variable=bva_time_var, value=val,
                   bg=C["panel"], fg=C["muted"], selectcolor=C["bg"],
                   activebackground=C["panel"], font=FONT_KPI_LBL).pack(side="left", padx=S(2))

bva_status_lbl = tk.Label(bva_scroll, text="", bg=C["bg"], fg=C["muted"], font=FONT_STATUS)
bva_status_lbl.pack(anchor="w", padx=S(12))

# Summary KPI strip
bva_kpi_f = tk.Frame(bva_scroll, bg=C["bg"])
bva_kpi_f.pack(fill="x", padx=S(10), pady=(S(4),0))
_bva_kpi_lbls = {}
for i, (key, label, accent) in enumerate([
    ("total_bgt","Total Budget",C["bud"]),
    ("total_act","Total Actual",C["rt"]),
    ("overall_pct","Overall %",C["bva"]),
    ("on_target","On Target",C["dy"]),
]):
    bva_kpi_f.columnconfigure(i, weight=1)
    card = tk.Frame(bva_kpi_f, bg=C["panel"])
    card.grid(row=0, column=i, sticky="nsew", padx=S(4), pady=S(2))
    tk.Frame(card, bg=accent, height=S(2)).pack(fill="x")
    inner = tk.Frame(card, bg=C["panel"])
    inner.pack(fill="both", expand=True, padx=S(10), pady=S(6))
    tk.Label(inner, text=label, bg=C["panel"], fg=C["muted"], font=FONT_KPI_LBL).pack(anchor="w")
    lv = tk.Label(inner, text="-", bg=C["panel"], fg=accent, font=FONT_KPI_VAL)
    lv.pack(anchor="w")
    _bva_kpi_lbls[key] = lv

# Bar canvas
_bva_bar_canvas = None
_bva_bar_h      = 0
_bva_bar_data   = []

def _build_bva_bar():
    global _bva_bar_canvas, _bva_bar_h
    n = len(BRANCH_IDS)
    _bva_bar_h = (BAR_ROW_H + BAR_PAD_Y) * n + BAR_PAD_Y
    _bva_bar_canvas = tk.Canvas(bva_scroll, bg=C["card"], height=_bva_bar_h,
                                  highlightthickness=0, bd=0)
    _bva_bar_canvas.pack(fill="x", padx=S(10), pady=(S(4),0))
    _bva_bar_canvas.bind("<Configure>",
        lambda e: _draw_bva_bars(_bva_bar_data, _bva_bar_canvas.winfo_width()))

_build_bva_bar()

def _draw_bva_bars(branch_data, canvas_w):
    global _bva_bar_data
    _bva_bar_data = branch_data
    if not _bva_bar_canvas or canvas_w < 10: return
    _bva_bar_canvas.delete("all")
    avail_w = canvas_w - BAR_NAME_W - BAR_AMT_W - BAR_PAD_X * 3
    for i, (name, pct, amt_str, bgt_str) in enumerate(branch_data):
        y_top = BAR_PAD_Y + i*(BAR_ROW_H+BAR_PAD_Y)
        y_mid = y_top + BAR_ROW_H//2
        _bva_bar_canvas.create_text(BAR_PAD_X, y_mid, text=name, anchor="w",
                                     fill=C["muted"], font=("Segoe UI", FS_KPI_LBL))
        tx = BAR_PAD_X + BAR_NAME_W + BAR_PAD_X
        ty = y_top + S(3); th = BAR_ROW_H - S(6)
        _bva_bar_canvas.create_rectangle(tx, ty, tx+avail_w, ty+th, fill=C["bar_track"], outline="")
        pct_clamped = min(pct, 115.0)
        fill_w = int(avail_w * pct_clamped / 115.0)
        bar_color = C["bar_hi"] if pct>=100 else C["bar_mid"] if pct>=90 else C["bar_lo"]
        if fill_w > 0:
            _bva_bar_canvas.create_rectangle(tx, ty, tx+fill_w, ty+th, fill=bar_color, outline="")
        pct_text = f"{pct:.1f}%"
        if fill_w > S(30):
            _bva_bar_canvas.create_text(tx+fill_w-S(4), y_mid, text=pct_text, anchor="e",
                                         fill="#000000", font=("Segoe UI", FS_KPI_LBL, "bold"))
        else:
            _bva_bar_canvas.create_text(tx+fill_w+S(4), y_mid, text=pct_text, anchor="w",
                                         fill=bar_color, font=("Segoe UI", FS_KPI_LBL))
        ax = tx + avail_w + BAR_PAD_X
        _bva_bar_canvas.create_text(ax+BAR_AMT_W, y_mid,
                                     text=f"{amt_str} / {bgt_str}", anchor="e",
                                     fill=bar_color, font=("Consolas", FS_KPI_LBL))

# Table
BVA_TREE_COLS = ["BranchName","Budget","Actual","% Hit","Variance","% Var"]
bva_tree_f = tk.Frame(bva_scroll, bg=C["card"])
bva_tree_f.pack(fill="both", expand=True, padx=S(10), pady=(S(4),0))
bva_tree = ttk.Treeview(bva_tree_f, columns=BVA_TREE_COLS, show="headings",
                          height=11, style="TV.Treeview")
bva_vsb = ttk.Scrollbar(bva_tree_f, orient="vertical", command=bva_tree.yview)
bva_tree.configure(yscrollcommand=bva_vsb.set)
bva_tree.pack(side="left", fill="both", expand=True)
bva_vsb.pack(side="right", fill="y")
_configure_tree(bva_tree)
bva_tree.tag_configure("hit_hi",  background=C["hit_hi"],  foreground="#00FFB2", font=FONT_DATA)
bva_tree.tag_configure("hit_mid", background=C["hit_mid"], foreground=C["gold"], font=FONT_DATA)
bva_tree.tag_configure("hit_lo",  background=C["hit_lo"],  foreground="#FF6B8A", font=FONT_DATA)
bva_tree.tag_configure("totalrow",background=C["total_bg"],foreground=C["gold"], font=FONT_TOTAL)
for col in BVA_TREE_COLS:
    lbl = {"BranchName":"Branch","% Hit":"% Hit","% Var":"% Var"}.get(col, COLUMN_LABELS.get(col,col))
    anchor = "w" if col=="BranchName" else "e"
    bva_tree.heading(col, text=lbl, anchor="center")
    bva_tree.column(col, width=COL_BRANCH if col=="BranchName" else COL_NUM, anchor=anchor, stretch=True, minwidth=S(40))

# Active promos/events display
bva_promo_lbl = tk.Label(bva_scroll, text="", bg=C["bg"], fg=C["promo_fg"],
                           font=("Segoe UI",FS_STATUS), wraplength=WIN_W-100, justify="left")
bva_promo_lbl.pack(anchor="w", padx=S(12), pady=(S(4),0))
bva_event_lbl = tk.Label(bva_scroll, text="", bg=C["bg"], fg=C["event_fg"],
                           font=("Segoe UI",FS_STATUS), wraplength=WIN_W-100, justify="left")
bva_event_lbl.pack(anchor="w", padx=S(12), pady=(0,S(6)))

def run_bva():
    fd  = bva_from_picker.get_date()
    td  = bva_to_picker.get_date()
    if fd > td:
        messagebox.showwarning("Date Error","From ≤ To required."); return
    t_from = "05:00:00" if bva_time_var.get()=="operational" else "00:00:00"
    bva_status_lbl.config(text=f"Fetching {fd} → {td} …", fg=C["muted"])
    root.update_idletasks()
    df, st = fetch_hq(f"{fd} {t_from}", f"{td} 23:59:59")
    bva_status_lbl.config(
        text=f"Data: {fd.strftime('%d %b %Y')} → {td.strftime('%d %b %Y')}  ({st.upper()})  "
             f"|  Fetched {datetime.now().strftime('%H:%M:%S')}",
        fg=C["ok"] if st=="ok" else C["err"])

    # Use mid-point month/year for budget reference
    mid = fd + (td - fd)//2
    midx = MONTH_IDX.get(mid.month, 0)
    yr   = mid.year
    days_in_range = max(1, (td - fd).days + 1)
    days_in_month = calendar.monthrange(mid.year, mid.month)[1]
    ratio = days_in_range / days_in_month   # scale budget proportionally

    bva_tree.delete(*bva_tree.get_children())
    bar_data = []
    tot_s = tot_b = 0; on_tgt = 0

    for i, bid in enumerate(BRANCH_IDS):
        bname = BRANCH_NAMES[bid]
        sname = BRANCH_SHORT.get(bname, bname)
        monthly_bgt = _get_budget(yr, bid, midx)
        bgt   = monthly_bgt * ratio
        if not df.empty and bname in df["BranchName"].values:
            row   = df[df["BranchName"]==bname].iloc[0]
            sales = float(row.get("NetSaleBFVat",0) or 0) + float(row.get("SVCBFTax",0) or 0)
        else:
            sales = 0.0
        pct = (sales/bgt*100) if bgt else 0
        var = bgt - sales
        tag = "hit_hi" if pct>=100 else "hit_mid" if pct>=90 else "hit_lo"
        bva_tree.insert("","end", values=(
            sname, _fmt(bgt), _fmt(sales), f"{pct:.1f}%",
            _fmt(var), f"{(var/bgt*100):.1f}%" if bgt else "0.0%"
        ), tags=(tag,))
        bar_data.append((sname, pct, _fmt(sales), _fmt(bgt)))
        tot_s += sales; tot_b += bgt
        if pct >= 100: on_tgt += 1

    tp = (tot_s/tot_b*100) if tot_b else 0
    tv = tot_b - tot_s
    bva_tree.insert("","end", values=(
        "--- TOTAL ---",_fmt(tot_b),_fmt(tot_s),
        f"{tp:.1f}%",_fmt(tv),f"{(tv/tot_b*100):.1f}%" if tot_b else "0.0%"
    ), tags=("totalrow",))

    root.after(50, lambda: _draw_bva_bars(bar_data, _bva_bar_canvas.winfo_width() if _bva_bar_canvas else 1))

    _bva_kpi_lbls["total_bgt"].config(text=f"{tot_b:,.0f}")
    _bva_kpi_lbls["total_act"].config(text=f"{tot_s:,.0f}")
    _bva_kpi_lbls["overall_pct"].config(text=f"{tp:.1f}%")
    _bva_kpi_lbls["on_target"].config(text=f"{on_tgt}/{len(BRANCH_IDS)}")

    # Promos / Events
    ap = _get_active_items(PROMOTIONS, fd, td)
    ae = _get_active_items(EVENTS, fd, td)
    bva_promo_lbl.config(
        text="🎯 Active Promotions: " + " | ".join(p["name"] for p in ap) if ap else "")
    bva_event_lbl.config(
        text="📅 Active Events: " + " | ".join(e["name"] for e in ae) if ae else "")

def _export_bva_excel():
    fp = filedialog.asksaveasfilename(defaultextension=".xlsx",
        filetypes=[("Excel","*.xlsx")], title="Save BvA Report")
    if not fp: return
    wb = Workbook(); ws = wb.active; ws.title = "Budget vs Actual"
    fd_str = bva_from_picker.get_date().strftime('%d %b %Y')
    td_str = bva_to_picker.get_date().strftime('%d %b %Y')
    _excel_write_section(ws, 1, bva_tree, f"Budget vs Actual  {fd_str} → {td_str}",
                          BVA_TREE_COLS,
                          col_labels_override={"BranchName":"Branch","% Hit":"% Hit","% Var":"% Var"},
                          tab_color="4B0082")
    _export_events_promos_sheet(wb)
    wb.save(fp); messagebox.showinfo("Exported", fp)

bva_btn_f = tk.Frame(bva_scroll, bg=C["bg"])
bva_btn_f.pack(fill="x", padx=S(10), pady=(S(4),S(8)))
b_run = _accent_btn(bva_btn_f, "  ▶  LOAD DATA", "#001030", C["bva"], run_bva)
b_run.pack(side="left", padx=S(4))
b_exp = _accent_btn(bva_btn_f, "  EXPORT EXCEL", C["save_btn"], C["save_fg"], _export_bva_excel)
b_exp.pack(side="left", padx=S(4))

# ============================================================
# Daily Reset at 04:00
# ============================================================
def _do_reset():
    try: root.destroy()
    except: pass
    subprocess.Popen([sys.executable, os.path.abspath(__file__)])

def schedule_daily_reset():
    now = datetime.now()
    next_reset = now.replace(hour=4, minute=0, second=0, microsecond=0)
    if now >= next_reset: next_reset += timedelta(days=1)
    ms_until = int((next_reset - now).total_seconds() * 1000)
    root.after(ms_until, _do_reset)

# ============================================================
# Initial show + start
# ============================================================
_show_tab("tab_live")
schedule_daily_reset()
_tick_rt_countdown()
refresh_all()
root.mainloop()
