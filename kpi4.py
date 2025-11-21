import tkinter as tk
from tkinter import ttk, messagebox, filedialog, scrolledtext
from datetime import datetime
import json
import os
import base64
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
import io
from PIL import Image
import copy


class KPIManagementSystem:
    def __init__(self, root):
        self.root = root
        self.root.title("🎯 KPI Management System - Multi-Employee Edition")
        self.root.geometry("1800x950")
        self.root.configure(bg='#ecf0f1')

        # Modern style
        style = ttk.Style()
        style.theme_use('clam')
        style.configure('TNotebook.Tab', padding=[20, 10], font=('Segoe UI', 10))

        # Default KPI Templates
        self.kpi_templates = {
            "IT Manager": {
                "KPI1": {"name": "Add new item to POS (5 days before launch)", "weight": 30, "enabled": True},
                "KPI2": {"name": "Special case add new item (urgent)", "weight": 10, "enabled": True},
                "KPI3": {"name": "Equipment condition (max 15 items/year/branch)", "weight": 30, "enabled": True},
                "KPI4": {"name": "Technical support (within 30 minutes)", "weight": 20, "enabled": True},
                "KPI5": {"name": "Profitability (EBITDA)", "weight": 10, "enabled": True}
            },
            "Sales Manager": {
                "KPI1": {"name": "Monthly Sales Target Achievement", "weight": 40, "enabled": True},
                "KPI2": {"name": "New Customer Acquisition", "weight": 25, "enabled": True},
                "KPI3": {"name": "Customer Satisfaction Score", "weight": 20, "enabled": True},
                "KPI4": {"name": "Sales Report Accuracy", "weight": 15, "enabled": True}
            }
        }

        # Data storage
        self.employees = {}
        self.current_employee_id = None

        self.load_all_data()
        self.create_main_interface()

    def create_main_interface(self):
        """สร้าง interface หลัก"""
        # Top bar - Employee selector
        top_frame = tk.Frame(self.root, bg='#2c3e50', height=100)
        top_frame.pack(fill='x', side='top')
        top_frame.pack_propagate(False)

        # Employee selector
        selector_frame = tk.Frame(top_frame, bg='#2c3e50')
        selector_frame.pack(side='left', padx=30, pady=20)

        tk.Label(selector_frame, text="เลือกพนักงาน:",
                 font=('Segoe UI', 12, 'bold'), bg='#2c3e50', fg='white').pack(side='left', padx=10)

        self.employee_var = tk.StringVar()
        self.employee_combo = ttk.Combobox(selector_frame, textvariable=self.employee_var,
                                           font=('Segoe UI', 11), width=35, state='readonly')
        self.employee_combo.pack(side='left', padx=5)
        self.employee_combo.bind('<<ComboboxSelected>>', self.on_employee_selected)

        # Buttons
        btn_frame = tk.Frame(top_frame, bg='#2c3e50')
        btn_frame.pack(side='right', padx=30, pady=20)

        tk.Button(btn_frame, text="➕ เพิ่มพนักงาน", bg='#27ae60', fg='white',
                  font=('Segoe UI', 10, 'bold'), relief='flat', padx=20, pady=8,
                  command=self.add_new_employee).pack(side='left', padx=5)

        tk.Button(btn_frame, text="⚙️ Config KPI", bg='#3498db', fg='white',
                  font=('Segoe UI', 10, 'bold'), relief='flat', padx=20, pady=8,
                  command=self.config_employee_kpi).pack(side='left', padx=5)

        tk.Button(btn_frame, text="📊 Dashboard", bg='#9b59b6', fg='white',
                  font=('Segoe UI', 10, 'bold'), relief='flat', padx=20, pady=8,
                  command=self.show_dashboard).pack(side='left', padx=5)

        # Main content area
        self.content_frame = tk.Frame(self.root, bg='#ecf0f1')
        self.content_frame.pack(fill='both', expand=True)

        # Load employee list
        self.refresh_employee_list()

        # Show welcome screen if no employee
        if not self.employees:
            self.show_welcome_screen()

    def show_welcome_screen(self):
        """แสดงหน้าจอต้อนรับ"""
        for widget in self.content_frame.winfo_children():
            widget.destroy()

        welcome_frame = tk.Frame(self.content_frame, bg='white')
        welcome_frame.place(relx=0.5, rely=0.5, anchor='center')

        tk.Label(welcome_frame, text="👋 ยินดีต้อนรับสู่ระบบ KPI Management",
                 font=('Segoe UI', 20, 'bold'), bg='white', fg='#2c3e50').pack(pady=20, padx=50)

        tk.Label(welcome_frame, text="กรุณาเพิ่มพนักงานเพื่อเริ่มต้นใช้งาน",
                 font=('Segoe UI', 12), bg='white', fg='#7f8c8d').pack(pady=10)

        tk.Button(welcome_frame, text="➕ เพิ่มพนักงานคนแรก", bg='#27ae60', fg='white',
                  font=('Segoe UI', 12, 'bold'), relief='flat', padx=30, pady=12,
                  command=self.add_new_employee).pack(pady=20)

    def add_new_employee(self):
        """เพิ่มพนักงานใหม่"""
        dialog = tk.Toplevel(self.root)
        dialog.title("➕ เพิ่มพนักงานใหม่")
        dialog.geometry("600x550")
        dialog.configure(bg='white')
        dialog.transient(self.root)
        dialog.grab_set()

        # Center dialog
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (600 // 2)
        y = (dialog.winfo_screenheight() // 2) - (550 // 2)
        dialog.geometry(f'600x550+{x}+{y}')

        main_frame = tk.Frame(dialog, bg='white')
        main_frame.pack(fill='both', expand=True, padx=20, pady=20)

        # Header
        header_frame = tk.Frame(main_frame, bg='#9b59b6', height=70)
        header_frame.pack(fill='x', pady=(0, 20))
        header_frame.pack_propagate(False)

        tk.Label(header_frame, text="📊 Dashboard - สรุปผลงานทั้งหมด",
                 font=('Segoe UI', 16, 'bold'), bg='#9b59b6', fg='white').pack(pady=20)

        # Canvas with scrollbar
        canvas = tk.Canvas(main_frame, bg='white', highlightthickness=0)
        scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg='white')

        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        # Employee cards
        for emp_id, emp_data in self.employees.items():
            self.create_employee_card(scrollable_frame, emp_id, emp_data)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # Export button
        btn_frame = tk.Frame(main_frame, bg='white')
        btn_frame.pack(pady=15)

        tk.Button(btn_frame, text="📥 Export ทุกคนเป็น Excel", bg='#27ae60', fg='white',
                  font=('Segoe UI', 11, 'bold'), relief='flat', padx=30, pady=12,
                  command=self.export_all_to_excel).pack()

    def create_employee_card(self, parent, emp_id, emp_data):
        """สร้างการ์ดแสดงข้อมูลพนักงาน"""
        card = tk.Frame(parent, bg='#f8f9fa', relief='solid', borderwidth=1)
        card.pack(fill='x', pady=10, padx=10)

        # Header
        header = tk.Frame(card, bg='#34495e', height=50)
        header.pack(fill='x')
        header.pack_propagate(False)

        tk.Label(header, text=f"{emp_data['info']['name']} ({emp_data['info']['employee_id']})",
                 font=('Segoe UI', 12, 'bold'), bg='#34495e', fg='white').pack(side='left', pady=12, padx=20)

        tk.Label(header, text=emp_data['info']['position'],
                 font=('Segoe UI', 10), bg='#34495e', fg='#ecf0f1').pack(side='left', pady=12)

        # Score
        total_score = self.calculate_total_score_for_employee(emp_data)
        score_label = tk.Label(header, text=f"{total_score:.2f}/5.00",
                               font=('Segoe UI', 14, 'bold'), bg='#34495e')

        if total_score >= 4.5:
            score_label.config(fg='#2ecc71')
        elif total_score >= 3.5:
            score_label.config(fg='#3498db')
        elif total_score >= 2.5:
            score_label.config(fg='#f39c12')
        else:
            score_label.config(fg='#e74c3c')

        score_label.pack(side='right', pady=12, padx=20)

        # KPI Details
        detail_frame = tk.Frame(card, bg='#f8f9fa')
        detail_frame.pack(fill='x', padx=20, pady=15)

        row = 0
        for kpi_id, config in emp_data["kpi_config"].items():
            if not config.get("enabled", True):
                continue

            jobs = emp_data["kpi_jobs"].get(kpi_id, [])
            scored = [j for j in jobs if 'score' in j]
            avg_score = sum(j['score'] for j in scored) / len(scored) if scored else 0

            # KPI row
            tk.Label(detail_frame, text=f"{kpi_id}:", font=('Segoe UI', 9, 'bold'),
                     bg='#f8f9fa', fg='#2c3e50').grid(row=row, column=0, sticky='w', pady=3)

            tk.Label(detail_frame, text=config['name'][:40], font=('Segoe UI', 9),
                     bg='#f8f9fa', fg='#7f8c8d').grid(row=row, column=1, sticky='w', padx=10, pady=3)

            tk.Label(detail_frame, text=f"{len(jobs)} jobs", font=('Segoe UI', 9),
                     bg='#f8f9fa', fg='#7f8c8d').grid(row=row, column=2, sticky='w', padx=10, pady=3)

            # Score bar
            bar_canvas = tk.Canvas(detail_frame, width=100, height=20, bg='#f8f9fa', highlightthickness=0)
            bar_canvas.grid(row=row, column=3, padx=10, pady=3)

            # Background
            bar_canvas.create_rectangle(0, 5, 100, 15, fill='#ecf0f1', outline='')

            # Score bar
            bar_width = (avg_score / 5) * 100
            color = '#2ecc71' if avg_score >= 4 else ('#3498db' if avg_score >= 3 else '#f39c12')
            bar_canvas.create_rectangle(0, 5, bar_width, 15, fill=color, outline='')

            tk.Label(detail_frame, text=f"{avg_score:.2f}/5", font=('Segoe UI', 9, 'bold'),
                     bg='#f8f9fa', fg=color).grid(row=row, column=4, sticky='w', padx=5, pady=3)

            row += 1

    def export_to_excel(self, emp_data):
        """Export ข้อมูลพนักงาน 1 คนเป็น Excel"""
        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=f"KPI_Report_{emp_data['info']['name']}_{datetime.now().strftime('%d%m%Y')}.xlsx"
        )

        if not filename:
            return

        try:
            wb = Workbook()

            # Styles
            header_font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
            subheader_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
            subheader_fill = PatternFill(start_color='3498DB', end_color='3498DB', fill_type='solid')
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Summary Sheet
            ws_summary = wb.active
            ws_summary.title = "สรุปภาพรวม"

            ws_summary['A1'] = "📊 KPI Performance Report"
            ws_summary['A1'].font = Font(name='Calibri', size=16, bold=True, color='2C3E50')
            ws_summary.merge_cells('A1:F1')

            # Employee info
            row = 3
            info_data = [
                ["ชื่อ-นามสกุล:", emp_data['info']['name']],
                ["รหัสพนักงาน:", emp_data['info']['employee_id']],
                ["ตำแหน่ง:", emp_data['info']['position']],
                ["ส่วนงาน:", emp_data['info'].get('section', 'N/A')],
                ["ฝ่าย:", emp_data['info'].get('department', 'N/A')],
                ["วันที่สร้างรายงาน:", datetime.now().strftime('%d/%m/%Y %H:%M')]
            ]

            for label, value in info_data:
                ws_summary[f'A{row}'] = label
                ws_summary[f'A{row}'].font = Font(bold=True)
                ws_summary[f'B{row}'] = value
                row += 1

            row += 2

            # KPI Summary Table
            ws_summary[f'A{row}'] = "KPI"
            ws_summary[f'B{row}'] = "รายการ"
            ws_summary[f'C{row}'] = "น้ำหนัก (%)"
            ws_summary[f'D{row}'] = "จำนวน Jobs"
            ws_summary[f'E{row}'] = "คะแนนเฉลี่ย"
            ws_summary[f'F{row}'] = "คะแนนถ่วงน้ำหนัก"

            for col in ['A', 'B', 'C', 'D', 'E', 'F']:
                cell = ws_summary[f'{col}{row}']
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border

            row += 1
            total_score = 0

            for kpi_id, config in emp_data["kpi_config"].items():
                if not config.get("enabled", True):
                    continue

                jobs = emp_data["kpi_jobs"].get(kpi_id, [])
                scored = [j for j in jobs if 'score' in j]
                avg_score = sum(j['score'] for j in scored) / len(scored) if scored else 0
                weighted = (avg_score * config['weight']) / 100
                total_score += weighted

                ws_summary[f'A{row}'] = kpi_id
                ws_summary[f'B{row}'] = config['name']
                ws_summary[f'C{row}'] = config['weight']
                ws_summary[f'D{row}'] = len(jobs)
                ws_summary[f'E{row}'] = f"{avg_score:.2f}"
                ws_summary[f'F{row}'] = f"{weighted:.4f}"

                for col in ['A', 'B', 'C', 'D', 'E', 'F']:
                    ws_summary[f'{col}{row}'].border = border
                    ws_summary[f'{col}{row}'].alignment = Alignment(horizontal='center')

                row += 1

            # Total
            row += 1
            ws_summary[f'E{row}'] = "คะแนนรวม:"
            ws_summary[f'E{row}'].font = Font(bold=True, size=12)
            ws_summary[f'F{row}'] = f"{total_score:.2f}/5.00"
            ws_summary[f'F{row}'].font = Font(bold=True, size=12, color='27AE60')

            # Rating
            row += 1
            if total_score >= 4.5:
                rating = "Far Exceed Expectations"
                rating_color = '27AE60'
            elif total_score >= 3.5:
                rating = "Exceed Expectations"
                rating_color = '3498DB'
            elif total_score >= 2.5:
                rating = "Meet Expectations"
                rating_color = 'F39C12'
            elif total_score >= 1.5:
                rating = "Partially Meet Expectations"
                rating_color = 'E67E22'
            else:
                rating = "Not Meet Expectations"
                rating_color = 'E74C3C'

            ws_summary[f'E{row}'] = "ระดับการประเมิน:"
            ws_summary[f'E{row}'].font = Font(bold=True, size=12)
            ws_summary[f'F{row}'] = rating
            ws_summary[f'F{row}'].font = Font(bold=True, size=12, color=rating_color)

            # Column widths
            ws_summary.column_dimensions['A'].width = 12
            ws_summary.column_dimensions['B'].width = 45
            ws_summary.column_dimensions['C'].width = 15
            ws_summary.column_dimensions['D'].width = 15
            ws_summary.column_dimensions['E'].width = 18
            ws_summary.column_dimensions['F'].width = 20

            # Detail sheets for each KPI
            for kpi_id, config in emp_data["kpi_config"].items():
                if not config.get("enabled", True):
                    continue

                ws = wb.create_sheet(title=kpi_id)
                ws['A1'] = f"{kpi_id}: {config['name']}"
                ws['A1'].font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
                ws['A1'].fill = subheader_fill
                ws.merge_cells('A1:H1')
                ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

                jobs = emp_data["kpi_jobs"].get(kpi_id, [])

                if not jobs:
                    ws['A3'] = "ยังไม่มีข้อมูล"
                    continue

                # Headers
                row = 3
                if kpi_id == "KPI1":
                    headers = ["ID", "ชื่อ Memo", "วันที่ส่ง", "วันเริ่มโปรโมชั่น", "วันทำเสร็จ", "จำนวนวัน", "คะแนน",
                               "หมายเหตุ", "รูปภาพ"]
                elif kpi_id == "KPI2":
                    headers = ["ID", "ชื่อ Memo", "วันเวลาส่ง", "กำหนดเวลา", "วันเวลาเสร็จ", "ชั่วโมง", "คะแนน",
                               "หมายเหตุ", "รูปภาพ"]
                else:
                    headers = ["ID", "รายละเอียด", "คะแนน", "หมายเหตุ"]

                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=row, column=col_idx, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = border

                row += 1

                # Data
                for job in jobs:
                    if kpi_id == "KPI1":
                        data = [job['id'], job['memo'], job['sent'], job['launch'],
                                job['completed'], job['days'], f"{job['score']}/5", job.get('comment', '')]
                    elif kpi_id == "KPI2":
                        data = [job['id'], job['memo'], job['sent'], job['required'],
                                job['completed'], job['hours'], f"{job['score']}/5", job.get('comment', '')]
                    else:
                        data = [job['id'], job.get('description', 'N/A'),
                                f"{job.get('score', 0)}/5", job.get('comment', '')]

                    for col_idx, value in enumerate(data, 1):
                        cell = ws.cell(row=row, column=col_idx, value=value)
                        cell.border = border
                        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

                    # Add image if exists
                    if kpi_id in ["KPI1", "KPI2"] and job.get('image'):
                        try:
                            img_data = base64.b64decode(job['image'])
                            img = Image.open(io.BytesIO(img_data))
                            img.thumbnail((150, 150), Image.Resampling.LANCZOS)

                            img_byte_arr = io.BytesIO()
                            img.save(img_byte_arr, format='PNG')
                            img_byte_arr.seek(0)

                            xl_img = XLImage(img_byte_arr)
                            xl_img.width = 100
                            xl_img.height = 100

                            img_col = len(headers)
                            ws.add_image(xl_img, f'{get_column_letter(img_col)}{row}')
                            ws.row_dimensions[row].height = 80

                            ws.cell(row=row, column=img_col, value="✓ มีรูปภาพ").font = Font(color='27AE60', bold=True)
                        except:
                            pass

                    row += 1

                # Column widths
                for col_idx in range(1, len(headers) + 1):
                    col_letter = get_column_letter(col_idx)
                    if col_idx == 1:
                        ws.column_dimensions[col_letter].width = 8
                    elif col_idx == len(headers):
                        ws.column_dimensions[col_letter].width = 25
                    else:
                        ws.column_dimensions[col_letter].width = 20

            wb.save(filename)
            messagebox.showinfo("สำเร็จ", f"✓ Export Excel สำเร็จ!\n\nบันทึกที่: {filename}")

            if messagebox.askyesno("เปิดไฟล์", "ต้องการเปิดไฟล์ Excel หรือไม่?"):
                import subprocess
                import platform

                if platform.system() == 'Windows':
                    os.startfile(filename)
                elif platform.system() == 'Darwin':
                    subprocess.Popen(['open', filename])
                else:
                    subprocess.Popen(['xdg-open', filename])

        except Exception as e:
            messagebox.showerror("Error", f"เกิดข้อผิดพลาดในการ Export:\n{str(e)}")

    def export_all_to_excel(self):
        """Export ข้อมูลทุกคนเป็น Excel เดียว"""
        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=f"KPI_All_Employees_{datetime.now().strftime('%d%m%Y')}.xlsx"
        )

        if not filename:
            return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "สรุปภาพรวมทุกคน"

            # Header
            ws['A1'] = "📊 สรุปผลงาน KPI ทุกคน"
            ws['A1'].font = Font(size=16, bold=True, color='2C3E50')
            ws.merge_cells('A1:G1')

            ws['A2'] = f"วันที่: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
            ws.merge_cells('A2:G2')

            # Table header
            row = 4
            headers = ["รหัส", "ชื่อ-นามสกุล", "ตำแหน่ง", "ส่วนงาน", "จำนวน KPI", "คะแนนรวม", "ระดับ"]

            header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
            border = Border(left=Side(style='thin'), right=Side(style='thin'),
                            top=Side(style='thin'), bottom=Side(style='thin'))

            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col_idx, value=header)
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border

            row += 1

            # Data
            for emp_id, emp_data in self.employees.items():
                total_score = self.calculate_total_score_for_employee(emp_data)
                enabled_kpis = sum(1 for c in emp_data["kpi_config"].values() if c.get("enabled", True))

                if total_score >= 4.5:
                    rating = "Far Exceed"
                elif total_score >= 3.5:
                    rating = "Exceed"
                elif total_score >= 2.5:
                    rating = "Meet"
                elif total_score >= 1.5:
                    rating = "Partially Meet"
                else:
                    rating = "Not Meet"

                data = [
                    emp_data['info']['employee_id'],
                    emp_data['info']['name'],
                    emp_data['info']['position'],
                    emp_data['info'].get('section', 'N/A'),
                    enabled_kpis,
                    f"{total_score:.2f}",
                    rating
                ]

                for col_idx, value in enumerate(data, 1):
                    cell = ws.cell(row=row, column=col_idx, value=value)
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center')

                row += 1

            # Column widths
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 25
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 12
            ws.column_dimensions['F'].width = 15
            ws.column_dimensions['G'].width = 20

            wb.save(filename)
            messagebox.showinfo("สำเร็จ", f"✓ Export Excel ทุกคนสำเร็จ!\n\nบันทึกที่: {filename}")

        except Exception as e:
            messagebox.showerror("Error", f"เกิดข้อผิดพลาด:\n{str(e)}")

    def save_all_data(self):
        """บันทึกข้อมูลทั้งหมด"""
        try:
            with open('kpi_multi_employee_data.json', 'w', encoding='utf-8') as f:
                json.dump({
                    "employees": self.employees,
                    "last_updated": datetime.now().strftime("%d/%m/%Y %H:%M:%S")
                }, f, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f"Save error: {e}")

    def load_all_data(self):
        """โหลดข้อมูลทั้งหมด"""
        try:
            if os.path.exists('kpi_multi_employee_data.json'):
                with open('kpi_multi_employee_data.json', 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    self.employees = data.get('employees', {})
                    print(f"✓ โหลดข้อมูล {len(self.employees)} คน - อัพเดทล่าสุด: {data.get('last_updated', 'N/A')}")
        except Exception as e:
            print(f"Load error: {e}")
            self.employees = {}

    def save_all_data(self):
        """บันทึกข้อมูลทั้งหมด"""
        try:
            with open('kpi_multi_employee_data.json', 'w', encoding='utf-8') as f:
                json.dump({
                    "employees": self.employees,
                    "last_updated": datetime.now().strftime("%d/%m/%Y %H:%M:%S")
                }, f, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f"Save error: {e}")

    def load_all_data(self):
        """โหลดข้อมูลทั้งหมด"""
        try:
            if os.path.exists('kpi_multi_employee_data.json'):
                with open('kpi_multi_employee_data.json', 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    self.employees = data.get('employees', {})
                    print(f"✓ โหลดข้อมูล {len(self.employees)} คน - อัพเดทล่าสุด: {data.get('last_updated', 'N/A')}")
        except Exception as e:
            print(f"Load error: {e}")
            self.employees = {}


def main():
    """Main function to run the application"""
    root = tk.Tk()
    app = KPIManagementSystem(root)

    # Center window on screen
    root.update_idletasks()
    width = root.winfo_width()
    height = root.winfo_height()
    x = (root.winfo_screenwidth() // 2) - (width // 2)
    y = (root.winfo_screenheight() // 2) - (height // 2)
    root.geometry(f'{width}x{height}+{x}+{y}')

    root.mainloop()


if __name__ == "__main__":
    main()
    dialog, bg = 'white'
    main_frame.pack(fill='both', expand=True, padx=30, pady=30)

    tk.Label(main_frame, text="ข้อมูลพนักงาน", font=('Segoe UI', 16, 'bold'),
    bg = 'white', fg = '#2c3e50').pack(pady=(0, 20))

    # Form
    form_frame = tk.Frame(main_frame, bg='white')
    form_frame.pack(fill='x', pady=10)

    fields = [
("รหัสพนักงาน:*", "employee_id"),
("ชื่อ-นามสกุล:*", "name"),
("ตำแหน่ง:*", "position"),
("ส่วนงาน:", "section"),
("ฝ่าย:", "department"),
("ระดับ:", "level")
]

vars_dict = {}
for idx, (label, key) in enumerate(fields):
    tk.Label(form_frame, text=label, font=('Segoe UI', 10),
bg = 'white', anchor = 'w').grid(row=idx, column=0, sticky='w', pady=8, padx=5)
vars_dict[key] = tk.StringVar()
ttk.Entry(form_frame, textvariable=vars_dict[key], width=35,
font = ('Segoe UI', 10)).grid(row=idx, column=1, pady=8, padx=5)

# Template selection
tk.Label(form_frame, text="เทมเพลต KPI:", font=('Segoe UI', 10),
bg = 'white', anchor = 'w').grid(row=len(fields), column=0, sticky='w', pady=8, padx=5)
template_var = tk.StringVar(value="IT Manager")
ttk.Combobox(form_frame, textvariable=template_var, width=33,
font = ('Segoe UI', 10), state = 'readonly',
values = list(self.kpi_templates.keys())).grid(row=len(fields), column=1, pady=8, padx=5)

# Buttons
btn_frame = tk.Frame(main_frame, bg='white')
btn_frame.pack(pady=30)


def save_employee():
    emp_id = vars_dict['employee_id'].get().strip()
    name = vars_dict['name'].get().strip()
    position = vars_dict['position'].get().strip()

    if not all([emp_id, name, position]):
        messagebox.showwarning("คำเตือน", "กรุณากรอกข้อมูลที่มีเครื่องหมาย * ให้ครบ")
        return

    if emp_id in self.employees:
        messagebox.showerror("ข้อผิดพลาด", "รหัสพนักงานนี้มีอยู่ในระบบแล้ว")
        return

    # Create new employee
    template_name = template_var.get()
    self.employees[emp_id] = {
        "info": {
            "employee_id": emp_id,
            "name": name,
            "position": position,
            "section": vars_dict['section'].get().strip(),
            "department": vars_dict['department'].get().strip(),
            "level": vars_dict['level'].get().strip()
        },
        "kpi_config": copy.deepcopy(self.kpi_templates[template_name]),
        "kpi_jobs": {}
    }

    # Initialize job storage for enabled KPIs
    for kpi_id, config in self.employees[emp_id]["kpi_config"].items():
        if config.get("enabled", True):
            self.employees[emp_id]["kpi_jobs"][kpi_id] = []

    self.save_all_data()
    self.refresh_employee_list()

    # Select new employee
    self.employee_var.set(f"{emp_id} - {name}")
    self.on_employee_selected(None)

    dialog.destroy()
    messagebox.showinfo("สำเร็จ", f"✓ เพิ่มพนักงาน {name} เรียบร้อยแล้ว")


tk.Button(btn_frame, text="💾 บันทึก", bg='#27ae60', fg='white',
          font=('Segoe UI', 11, 'bold'), relief='flat', padx=30, pady=10,
          command=save_employee).pack(side='left', padx=5)

tk.Button(btn_frame, text="❌ ยกเลิก", bg='#95a5a6', fg='white',
          font=('Segoe UI', 11, 'bold'), relief='flat', padx=30, pady=10,
          command=dialog.destroy).pack(side='left', padx=5)


def config_employee_kpi(self):
    """Config KPI สำหรับพนักงานที่เลือก"""
    if not self.current_employee_id:
        messagebox.showwarning("คำเตือน", "กรุณาเลือกพนักงานก่อน")
        return

    emp_data = self.employees[self.current_employee_id]

    dialog = tk.Toplevel(self.root)
    dialog.title(f"⚙️ Config KPI - {emp_data['info']['name']}")
    dialog.geometry("900x700")
    dialog.configure(bg='white')
    dialog.transient(self.root)
    dialog.grab_set()

    # Center dialog
    dialog.update_idletasks()
    x = (dialog.winfo_screenwidth() // 2) - (900 // 2)
    y = (dialog.winfo_screenheight() // 2) - (700 // 2)
    dialog.geometry(f'900x700+{x}+{y}')

    main_frame = tk.Frame(dialog, bg='white')
    main_frame.pack(fill='both', expand=True, padx=30, pady=30)

    # Header
    header_frame = tk.Frame(main_frame, bg='#3498db', height=60)
    header_frame.pack(fill='x', pady=(0, 20))
    header_frame.pack_propagate(False)

    tk.Label(header_frame, text=f"⚙️ การตั้งค่า KPI - {emp_data['info']['name']}",
             font=('Segoe UI', 14, 'bold'), bg='#3498db', fg='white').pack(pady=15, padx=20)

    # KPI Config Frame
    canvas = tk.Canvas(main_frame, bg='white', highlightthickness=0)
    scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=canvas.yview)
    scrollable_frame = tk.Frame(canvas, bg='white')

    scrollable_frame.bind(
        "<Configure>",
        lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
    )

    canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
    canvas.configure(yscrollcommand=scrollbar.set)

    canvas.pack(side="left", fill="both", expand=True)
    scrollbar.pack(side="right", fill="y")

    # KPI List
    kpi_vars = {}

    # Add new KPI button
    add_frame = tk.Frame(scrollable_frame, bg='#ecf0f1', relief='solid', borderwidth=1)
    add_frame.pack(fill='x', pady=10, padx=10)

    tk.Button(add_frame, text="➕ เพิ่ม KPI ใหม่", bg='#27ae60', fg='white',
              font=('Segoe UI', 10, 'bold'), relief='flat', padx=20, pady=8,
              command=lambda: self.add_new_kpi_to_employee(emp_data, dialog)).pack(pady=15)

    # Existing KPIs
    for kpi_id, config in emp_data["kpi_config"].items():
        kpi_frame = tk.Frame(scrollable_frame, bg='#f8f9fa', relief='solid', borderwidth=1)
        kpi_frame.pack(fill='x', pady=5, padx=10)

        # Header row
        header_row = tk.Frame(kpi_frame, bg='#f8f9fa')
        header_row.pack(fill='x', padx=15, pady=10)

        enabled_var = tk.BooleanVar(value=config.get("enabled", True))
        kpi_vars[kpi_id] = {
            "enabled": enabled_var,
            "name": tk.StringVar(value=config["name"]),
            "weight": tk.IntVar(value=config["weight"])
        }

        ttk.Checkbutton(header_row, variable=enabled_var,
                        text=kpi_id, style='TCheckbutton').pack(side='left')

        tk.Label(header_row, text=config["name"][:50], font=('Segoe UI', 10),
                 bg='#f8f9fa', fg='#2c3e50').pack(side='left', padx=10)

        # Detail row
        detail_row = tk.Frame(kpi_frame, bg='#f8f9fa')
        detail_row.pack(fill='x', padx=15, pady=(0, 10))

        tk.Label(detail_row, text="ชื่อ KPI:", bg='#f8f9fa', font=('Segoe UI', 9)).grid(row=0, column=0, sticky='w',
                                                                                        pady=5)
        ttk.Entry(detail_row, textvariable=kpi_vars[kpi_id]["name"], width=50).grid(row=0, column=1, columnspan=2,
                                                                                    sticky='w', pady=5, padx=5)

        tk.Label(detail_row, text="น้ำหนัก (%):", bg='#f8f9fa', font=('Segoe UI', 9)).grid(row=1, column=0, sticky='w',
                                                                                           pady=5)
        ttk.Spinbox(detail_row, from_=0, to=100, textvariable=kpi_vars[kpi_id]["weight"], width=10).grid(row=1,
                                                                                                         column=1,
                                                                                                         sticky='w',
                                                                                                         pady=5, padx=5)

    # Save button
    btn_frame = tk.Frame(main_frame, bg='white')
    btn_frame.pack(pady=20)

    def save_config():
        total_weight = 0
        for kpi_id, vars in kpi_vars.items():
            if vars["enabled"].get():
                total_weight += vars["weight"].get()

            emp_data["kpi_config"][kpi_id].update({
                "enabled": vars["enabled"].get(),
                "name": vars["name"].get(),
                "weight": vars["weight"].get()
            })

            # Initialize jobs if newly enabled
            if vars["enabled"].get() and kpi_id not in emp_data["kpi_jobs"]:
                emp_data["kpi_jobs"][kpi_id] = []

        if total_weight != 100:
            if not messagebox.askyesno("คำเตือน",
                                       f"น้ำหนักรวม = {total_weight}% (ไม่ใช่ 100%)\nต้องการบันทึกต่อหรือไม่?"):
                return

        self.save_all_data()
        dialog.destroy()
        self.on_employee_selected(None)  # Refresh
        messagebox.showinfo("สำเร็จ", "✓ บันทึกการตั้งค่า KPI เรียบร้อยแล้ว")

    tk.Button(btn_frame, text="💾 บันทึก", bg='#27ae60', fg='white',
              font=('Segoe UI', 11, 'bold'), relief='flat', padx=30, pady=10,
              command=save_config).pack(side='left', padx=5)

    tk.Button(btn_frame, text="❌ ยกเลิก", bg='#95a5a6', fg='white',
              font=('Segoe UI', 11, 'bold'), relief='flat', padx=30, pady=10,
              command=dialog.destroy).pack(side='left', padx=5)


def add_new_kpi_to_employee(self, emp_data, parent_dialog):
    """เพิ่ม KPI ใหม่ให้พนักงาน"""
    dialog = tk.Toplevel(parent_dialog)
    dialog.title("➕ เพิ่ม KPI ใหม่")
    dialog.geometry("500x250")
    dialog.configure(bg='white')
    dialog.transient(parent_dialog)
    dialog.grab_set()

    main_frame = tk.Frame(dialog, bg='white')
    main_frame.pack(fill='both', expand=True, padx=30, pady=30)

    tk.Label(main_frame, text="เพิ่ม KPI ใหม่", font=('Segoe UI', 14, 'bold'),
             bg='white', fg='#2c3e50').pack(pady=(0, 20))

    form_frame = tk.Frame(main_frame, bg='white')
    form_frame.pack(fill='x')

    tk.Label(form_frame, text="รหัส KPI:", bg='white').grid(row=0, column=0, sticky='w', pady=8)
    kpi_id_var = tk.StringVar()
    ttk.Entry(form_frame, textvariable=kpi_id_var, width=30).grid(row=0, column=1, pady=8, padx=10)

    tk.Label(form_frame, text="ชื่อ KPI:", bg='white').grid(row=1, column=0, sticky='w', pady=8)
    kpi_name_var = tk.StringVar()
    ttk.Entry(form_frame, textvariable=kpi_name_var, width=30).grid(row=1, column=1, pady=8, padx=10)

    tk.Label(form_frame, text="น้ำหนัก (%):", bg='white').grid(row=2, column=0, sticky='w', pady=8)
    weight_var = tk.IntVar(value=10)
    ttk.Spinbox(form_frame, from_=0, to=100, textvariable=weight_var, width=28).grid(row=2, column=1, pady=8, padx=10)

    def save_new_kpi():
        kpi_id = kpi_id_var.get().strip()
        kpi_name = kpi_name_var.get().strip()

        if not all([kpi_id, kpi_name]):
            messagebox.showwarning("คำเตือน", "กรุณากรอกข้อมูลให้ครบ")
            return

        if kpi_id in emp_data["kpi_config"]:
            messagebox.showerror("ข้อผิดพลาด", "รหัส KPI นี้มีอยู่แล้ว")
            return

        emp_data["kpi_config"][kpi_id] = {
            "name": kpi_name,
            "weight": weight_var.get(),
            "enabled": True
        }
        emp_data["kpi_jobs"][kpi_id] = []

        dialog.destroy()
        parent_dialog.destroy()
        self.config_employee_kpi()

    btn_frame = tk.Frame(main_frame, bg='white')
    btn_frame.pack(pady=20)

    tk.Button(btn_frame, text="💾 เพิ่ม", bg='#27ae60', fg='white',
              font=('Segoe UI', 10, 'bold'), relief='flat', padx=25, pady=8,
              command=save_new_kpi).pack(side='left', padx=5)

    tk.Button(btn_frame, text="❌ ยกเลิก", bg='#95a5a6', fg='white',
              font=('Segoe UI', 10, 'bold'), relief='flat', padx=25, pady=8,
              command=dialog.destroy).pack(side='left', padx=5)


def refresh_employee_list(self):
    """รีเฟรชรายชื่อพนักงาน"""
    emp_list = [f"{emp_id} - {data['info']['name']}" for emp_id, data in self.employees.items()]
    self.employee_combo['values'] = emp_list

    if emp_list and not self.employee_var.get():
        self.employee_combo.current(0)
        self.on_employee_selected(None)


def on_employee_selected(self, event):
    """เมื่อเลือกพนักงาน"""
    selected = self.employee_var.get()
    if not selected:
        return

    emp_id = selected.split(' - ')[0]
    self.current_employee_id = emp_id
    self.show_employee_workspace()


def show_employee_workspace(self):
    """แสดงพื้นที่ทำงานของพนักงาน"""
    for widget in self.content_frame.winfo_children():
        widget.destroy()

    emp_data = self.employees[self.current_employee_id]

    # Score bar
    self.create_score_bar_for_employee(emp_data)

    # Main notebook
    self.notebook = ttk.Notebook(self.content_frame)
    self.notebook.pack(fill='both', expand=True, padx=10, pady=(0, 10))

    # Create tabs for enabled KPIs
    for kpi_id, config in emp_data["kpi_config"].items():
        if config.get("enabled", True):
            tab = tk.Frame(self.notebook, bg='white')
            self.notebook.add(tab, text=f'{kpi_id} ({config["weight"]}%)')
            self.create_kpi_tab(tab, kpi_id, config, emp_data)

    # Summary tab
    report_tab = tk.Frame(self.notebook, bg='white')
    self.notebook.add(report_tab, text='📊 สรุปและรายงาน')
    self.create_report_tab(report_tab, emp_data)


def create_score_bar_for_employee(self, emp_data):
    """สร้างแถบคะแนนสำหรับพนักงาน"""
    score_frame = tk.Frame(self.content_frame, bg='#34495e', height=90)
    score_frame.pack(fill='x', side='top')
    score_frame.pack_propagate(False)

    # Left - Employee info
    left_frame = tk.Frame(score_frame, bg='#34495e')
    left_frame.pack(side='left', padx=25, pady=15)

    tk.Label(left_frame, text=emp_data['info']['name'],
             font=('Segoe UI', 13, 'bold'), bg='#34495e', fg='white').pack(anchor='w')
    tk.Label(left_frame, text=f"{emp_data['info']['position']} | ID: {emp_data['info']['employee_id']}",
             font=('Segoe UI', 9), bg='#34495e', fg='#ecf0f1').pack(anchor='w')

    # Right - Total score
    right_frame = tk.Frame(score_frame, bg='#34495e')
    right_frame.pack(side='right', padx=25, pady=15)

    tk.Label(right_frame, text="คะแนนรวมทั้งหมด",
             font=('Segoe UI', 10), bg='#34495e', fg='#ecf0f1').pack()

    total_score = self.calculate_total_score_for_employee(emp_data)
    score_label = tk.Label(right_frame, text=f"{total_score:.2f}/5.00",
                           font=('Segoe UI', 26, 'bold'), bg='#34495e')
    score_label.pack()

    # Color based on score
    if total_score >= 4.5:
        score_label.config(fg='#2ecc71')
    elif total_score >= 3.5:
        score_label.config(fg='#3498db')
    elif total_score >= 2.5:
        score_label.config(fg='#f39c12')
    else:
        score_label.config(fg='#e74c3c')

    self.current_score_label = score_label


def calculate_total_score_for_employee(self, emp_data):
    """คำนวณคะแนนรวมของพนักงาน"""
    total = 0
    for kpi_id, config in emp_data["kpi_config"].items():
        if not config.get("enabled", True):
            continue

        jobs = emp_data["kpi_jobs"].get(kpi_id, [])
        if jobs:
            if kpi_id == "KPI3":
                branches = {}
                for j in jobs:
                    branches[j.get('branch', 'Unknown')] = branches.get(j.get('branch', 'Unknown'), 0) + 1
                avg = sum(branches.values()) / len(branches) if branches else 0
                score = 5 if avg < 9 else (4 if avg <= 14 else (3 if avg <= 15 else (2 if avg <= 20 else 1)))
                weighted = (score * config['weight']) / 100
            else:
                scored = [j for j in jobs if 'score' in j]
                if scored:
                    avg_score = sum(j['score'] for j in scored) / len(scored)
                    weighted = (avg_score * config['weight']) / 100
                else:
                    weighted = 0
            total += weighted
    return total


def create_kpi_tab(self, parent, kpi_id, config, emp_data):
    """สร้างแท็บ KPI (ใช้โค้ดเดิม แต่ปรับให้รองรับ emp_data)"""
    container = tk.Frame(parent, bg='white')
    container.pack(fill='both', expand=True, padx=15, pady=15)

    # Header
    header_frame = tk.Frame(container, bg='#3498db', height=60)
    header_frame.pack(fill='x', pady=(0, 15))
    header_frame.pack_propagate(False)

    tk.Label(header_frame, text=f"{kpi_id}: {config['name']}",
             font=('Segoe UI', 12, 'bold'), bg='#3498db', fg='white').pack(pady=15, padx=20, anchor='w')

    # Form Frame
    form_frame = tk.LabelFrame(container, text="📝 บันทึก Job ใหม่",
                               font=('Segoe UI', 11, 'bold'),
                               bg='white', fg='#2c3e50',
                               relief='solid', borderwidth=1)
    form_frame.pack(fill='x', pady=(0, 15), padx=5)

    vars_dict = {}

    # KPI1 and KPI2 forms (with images)
    if kpi_id in ["KPI1", "KPI2"]:
        tk.Label(form_frame, text="ชื่อ Memo:*", bg='white',
                 font=('Segoe UI', 10)).grid(row=0, column=0, sticky=tk.W, pady=8, padx=10)
        vars_dict['memo_name'] = tk.StringVar()
        ttk.Entry(form_frame, textvariable=vars_dict['memo_name'],
                  width=60, font=('Segoe UI', 10)).grid(row=0, column=1, columnspan=3,
                                                        sticky=tk.W, pady=8, padx=5)

        if kpi_id == "KPI1":
            tk.Label(form_frame, text="วันที่ส่ง (DD/MM/YYYY):*", bg='white',
                     font=('Segoe UI', 10)).grid(row=1, column=0, sticky=tk.W, pady=8, padx=10)
            vars_dict['sent_date'] = tk.StringVar(value=datetime.now().strftime("%d/%m/%Y"))
            ttk.Entry(form_frame, textvariable=vars_dict['sent_date'],
                      width=20, font=('Segoe UI', 10)).grid(row=1, column=1, pady=8, padx=5)

            tk.Label(form_frame, text="วันเริ่มโปรโมชั่น (DD/MM/YYYY):*", bg='white',
                     font=('Segoe UI', 10)).grid(row=1, column=2, padx=20, sticky=tk.W)
            vars_dict['launch_date'] = tk.StringVar()
            ttk.Entry(form_frame, textvariable=vars_dict['launch_date'],
                      width=20, font=('Segoe UI', 10)).grid(row=1, column=3, pady=8, padx=5)

            tk.Label(form_frame, text="วันที่ทำเสร็จ (DD/MM/YYYY):*", bg='white',
                     font=('Segoe UI', 10)).grid(row=2, column=0, sticky=tk.W, pady=8, padx=10)
            vars_dict['completed_date'] = tk.StringVar()
            ttk.Entry(form_frame, textvariable=vars_dict['completed_date'],
                      width=20, font=('Segoe UI', 10)).grid(row=2, column=1, pady=8, padx=5)
        else:  # KPI2
            tk.Label(form_frame, text="วันเวลาส่ง (DD/MM/YYYY HH:MM):*", bg='white',
                     font=('Segoe UI', 10)).grid(row=1, column=0, sticky=tk.W, pady=8, padx=10)
            vars_dict['sent_datetime'] = tk.StringVar(value=datetime.now().strftime("%d/%m/%Y %H:%M"))
            ttk.Entry(form_frame, textvariable=vars_dict['sent_datetime'],
                      width=25, font=('Segoe UI', 10)).grid(row=1, column=1, pady=8, padx=5)

            tk.Label(form_frame, text="กำหนดเวลา (DD/MM/YYYY HH:MM):*", bg='white',
                     font=('Segoe UI', 10)).grid(row=1, column=2, padx=20, sticky=tk.W)
            vars_dict['required_datetime'] = tk.StringVar()
            ttk.Entry(form_frame, textvariable=vars_dict['required_datetime'],
                      width=25, font=('Segoe UI', 10)).grid(row=1, column=3, pady=8, padx=5)

            tk.Label(form_frame, text="วันเวลาเสร็จ (DD/MM/YYYY HH:MM):*", bg='white',
                     font=('Segoe UI', 10)).grid(row=2, column=0, sticky=tk.W, pady=8, padx=10)
            vars_dict['completed_datetime'] = tk.StringVar()
            ttk.Entry(form_frame, textvariable=vars_dict['completed_datetime'],
                      width=25, font=('Segoe UI', 10)).grid(row=2, column=1, pady=8, padx=5)

        # Image attachment
        vars_dict['image_data'] = None
        vars_dict['image_path'] = tk.StringVar()
        img_btn = tk.Button(form_frame, text="📎 แนบภาพ", bg='#3498db', fg='white',
                            font=('Segoe UI', 9, 'bold'), relief='flat', padx=15, pady=5,
                            command=lambda: self.select_image(vars_dict))
        img_btn.grid(row=3, column=0, pady=10, padx=10, sticky=tk.W)
        tk.Label(form_frame, textvariable=vars_dict['image_path'], bg='white',
                 font=('Segoe UI', 9), fg='#27ae60').grid(row=3, column=1, columnspan=3, sticky=tk.W, padx=5)

        # Comment
        tk.Label(form_frame, text="หมายเหตุ:", bg='white',
                 font=('Segoe UI', 10)).grid(row=4, column=0, sticky=tk.NW, pady=8, padx=10)
        vars_dict['comment_text'] = tk.Text(form_frame, width=60, height=3,
                                            font=('Segoe UI', 9), relief='solid', borderwidth=1)
        vars_dict['comment_text'].grid(row=4, column=1, columnspan=3, sticky=tk.W, pady=8, padx=5)

    # Generic form for other KPIs
    else:
        tk.Label(form_frame, text="รายละเอียด:*", bg='white',
                 font=('Segoe UI', 10)).grid(row=0, column=0, sticky=tk.W, pady=8, padx=10)
        vars_dict['description'] = tk.StringVar()
        ttk.Entry(form_frame, textvariable=vars_dict['description'],
                  width=60, font=('Segoe UI', 10)).grid(row=0, column=1, columnspan=3,
                                                        sticky=tk.W, pady=8, padx=5)

        tk.Label(form_frame, text="คะแนน (1-5):*", bg='white',
                 font=('Segoe UI', 10)).grid(row=1, column=0, sticky=tk.W, pady=8, padx=10)
        vars_dict['score'] = tk.IntVar(value=3)
        ttk.Spinbox(form_frame, from_=1, to=5, textvariable=vars_dict['score'],
                    width=18, font=('Segoe UI', 10)).grid(row=1, column=1, pady=8, padx=5)

        tk.Label(form_frame, text="หมายเหตุ:", bg='white',
                 font=('Segoe UI', 10)).grid(row=2, column=0, sticky=tk.NW, pady=8, padx=10)
        vars_dict['comment_text'] = tk.Text(form_frame, width=60, height=3,
                                            font=('Segoe UI', 9), relief='solid', borderwidth=1)
        vars_dict['comment_text'].grid(row=2, column=1, columnspan=3, sticky=tk.W, pady=8, padx=5)

    # Buttons
    button_frame = tk.Frame(form_frame, bg='white')
    button_frame.grid(row=10, column=0, columnspan=4, pady=15)

    save_btn = tk.Button(button_frame, text="💾 บันทึก", bg='#27ae60', fg='white',
                         font=('Segoe UI', 10, 'bold'), relief='flat', padx=25, pady=8,
                         command=lambda: self.save_job(kpi_id, vars_dict, emp_data))
    save_btn.pack(side=tk.LEFT, padx=5)

    clear_btn = tk.Button(button_frame, text="🗑️ ล้าง", bg='#95a5a6', fg='white',
                          font=('Segoe UI', 10, 'bold'), relief='flat', padx=25, pady=8,
                          command=lambda: self.clear_form(vars_dict))
    clear_btn.pack(side=tk.LEFT, padx=5)

    # Jobs List
    list_frame = tk.LabelFrame(container, text="📋 รายการ Jobs",
                               font=('Segoe UI', 11, 'bold'),
                               bg='white', fg='#2c3e50',
                               relief='solid', borderwidth=1)
    list_frame.pack(fill='both', expand=True, padx=5)

    self.create_jobs_tree(list_frame, kpi_id, emp_data)


def create_jobs_tree(self, parent, kpi_id, emp_data):
    """สร้างตารางแสดงรายการ Jobs"""
    if kpi_id == "KPI1":
        columns = ("ID", "Memo", "วันที่ส่ง", "วันเริ่มโปรโมชั่น", "Days", "Score", "Image", "Comment")
    elif kpi_id == "KPI2":
        columns = ("ID", "Memo", "วันเวลาส่ง", "Hours", "Score", "Image", "Comment")
    else:
        columns = ("ID", "รายละเอียด", "คะแนน", "หมายเหตุ")

    tree = ttk.Treeview(parent, columns=columns, show='headings', height=10)
    for col in columns:
        tree.heading(col, text=col)
        if col == "Comment" or col == "หมายเหตุ":
            tree.column(col, width=150)
        elif col in ["Memo", "รายละเอียด"]:
            tree.column(col, width=250)
        else:
            tree.column(col, width=100)

    vsb = ttk.Scrollbar(parent, orient="vertical", command=tree.yview)
    tree.configure(yscrollcommand=vsb.set)
    tree.grid(row=0, column=0, sticky='nsew', padx=10, pady=10)
    vsb.grid(row=0, column=1, sticky='ns', pady=10)

    parent.grid_rowconfigure(0, weight=1)
    parent.grid_columnconfigure(0, weight=1)

    btn_frame = tk.Frame(parent, bg='white')
    btn_frame.grid(row=1, column=0, columnspan=2, pady=10)

    del_btn = tk.Button(btn_frame, text="❌ ลบ", bg='#e74c3c', fg='white',
                        font=('Segoe UI', 9, 'bold'), relief='flat', padx=20, pady=5,
                        command=lambda: self.delete_job(kpi_id, tree, emp_data))
    del_btn.pack(side=tk.LEFT, padx=5)

    refresh_btn = tk.Button(btn_frame, text="🔄 รีเฟรช", bg='#3498db', fg='white',
                            font=('Segoe UI', 9, 'bold'), relief='flat', padx=20, pady=5,
                            command=lambda: self.refresh_tree(kpi_id, tree, emp_data))
    refresh_btn.pack(side=tk.LEFT, padx=5)

    self.refresh_tree(kpi_id, tree, emp_data)


def parse_date(self, date_str, fmt="%d/%m/%Y"):
    """แปลงวันที่"""
    try:
        return datetime.strptime(date_str, fmt)
    except:
        return None


def parse_datetime(self, datetime_str, fmt="%d/%m/%Y %H:%M"):
    """แปลงวันเวลา"""
    try:
        return datetime.strptime(datetime_str, fmt)
    except:
        return None


def select_image(self, vars_dict):
    """เลือกรูปภาพ"""
    filename = filedialog.askopenfilename(
        title="เลือกรูปภาพ",
        filetypes=[("Image files", "*.jpg *.jpeg *.png *.gif *.bmp"), ("All files", "*.*")]
    )
    if filename:
        try:
            with open(filename, 'rb') as f:
                vars_dict['image_data'] = base64.b64encode(f.read()).decode()
                vars_dict['image_path'].set(f"✓ {os.path.basename(filename)}")
            messagebox.showinfo("สำเร็จ", "แนบรูปภาพเรียบร้อยแล้ว")
        except Exception as e:
            messagebox.showerror("Error", f"ไม่สามารถอ่านไฟล์ได้: {str(e)}")


def clear_form(self, vars_dict):
    """ล้างฟอร์ม"""
    for key, var in vars_dict.items():
        if isinstance(var, tk.StringVar):
            if 'date' in key.lower() and key in ['sent_date', 'sent_datetime']:
                pass
            else:
                var.set("")
        elif isinstance(var, tk.Text):
            var.delete(1.0, tk.END)
        elif isinstance(var, tk.IntVar):
            var.set(3)
    if 'image_data' in vars_dict:
        vars_dict['image_data'] = None
        if 'image_path' in vars_dict:
            vars_dict['image_path'].set("")


def save_job(self, kpi_id, vars_dict, emp_data):
    """บันทึก Job"""
    try:
        job = {"id": len(emp_data["kpi_jobs"].get(kpi_id, [])) + 1}

        if kpi_id == "KPI1":
            memo = vars_dict['memo_name'].get().strip()
            sent = vars_dict['sent_date'].get().strip()
            launch = vars_dict['launch_date'].get().strip()
            completed = vars_dict['completed_date'].get().strip()
            comment = vars_dict['comment_text'].get(1.0, tk.END).strip()

            if not all([memo, sent, launch, completed]):
                messagebox.showwarning("คำเตือน", "กรุณากรอกข้อมูลที่มีเครื่องหมาย * ให้ครบ")
                return

            sent_dt = self.parse_date(sent)
            launch_dt = self.parse_date(launch)
            completed_dt = self.parse_date(completed)

            if not all([sent_dt, launch_dt, completed_dt]):
                messagebox.showerror("Error", "รูปแบบวันที่ไม่ถูกต้อง กรุณาใช้ DD/MM/YYYY")
                return

            days = (launch_dt - completed_dt).days
            score = 5 if days >= 7 else (4 if days >= 6 else (3 if days >= 5 else (2 if days >= 4 else 1)))

            job.update({
                "memo": memo, "sent": sent, "launch": launch, "completed": completed,
                "days": days, "score": score, "image": vars_dict.get('image_data'),
                "comment": comment
            })

        elif kpi_id == "KPI2":
            memo = vars_dict['memo_name'].get().strip()
            sent = vars_dict['sent_datetime'].get().strip()
            required = vars_dict['required_datetime'].get().strip()
            completed = vars_dict['completed_datetime'].get().strip()
            comment = vars_dict['comment_text'].get(1.0, tk.END).strip()

            if not all([memo, sent, required, completed]):
                messagebox.showwarning("คำเตือน", "กรุณากรอกข้อมูลที่มีเครื่องหมาย * ให้ครบ")
                return

            sent_dt = self.parse_datetime(sent)
            completed_dt = self.parse_datetime(completed)

            if not all([sent_dt, completed_dt]):
                messagebox.showerror("Error", "รูปแบบวันเวลาไม่ถูกต้อง กรุณาใช้ DD/MM/YYYY HH:MM")
                return

            hours = (completed_dt - sent_dt).total_seconds() / 3600
            score = 5 if hours < 6 else (4 if hours <= 11 else (3 if hours <= 12 else (2 if hours <= 24 else 1)))

            job.update({
                "memo": memo, "sent": sent, "required": required, "completed": completed,
                "hours": round(hours, 2), "score": score, "image": vars_dict.get('image_data'),
                "comment": comment
            })

        else:  # Generic KPI
            description = vars_dict['description'].get().strip()
            score = vars_dict['score'].get()
            comment = vars_dict['comment_text'].get(1.0, tk.END).strip()

            if not description:
                messagebox.showwarning("คำเตือน", "กรุณากรอกรายละเอียด")
                return

            job.update({
                "description": description,
                "score": score,
                "comment": comment
            })

        if kpi_id not in emp_data["kpi_jobs"]:
            emp_data["kpi_jobs"][kpi_id] = []

        emp_data["kpi_jobs"][kpi_id].append(job)
        self.save_all_data()

        # Refresh tree
        for widget in self.content_frame.winfo_children():
            if isinstance(widget, ttk.Notebook):
                current_tab = widget.nametowidget(widget.select())
                for child in current_tab.winfo_children():
                    if isinstance(child, tk.Frame):
                        for subchild in child.winfo_children():
                            if isinstance(subchild, tk.LabelFrame) and "รายการ Jobs" in subchild.cget('text'):
                                for tree_widget in subchild.winfo_children():
                                    if isinstance(tree_widget, ttk.Treeview):
                                        self.refresh_tree(kpi_id, tree_widget, emp_data)

        self.clear_form(vars_dict)

        # Update score
        if hasattr(self, 'current_score_label'):
            total_score = self.calculate_total_score_for_employee(emp_data)
            self.current_score_label.config(text=f"{total_score:.2f}/5.00")

        messagebox.showinfo("สำเร็จ", "✓ บันทึกข้อมูลเรียบร้อยแล้ว")

    except Exception as e:
        messagebox.showerror("Error", f"เกิดข้อผิดพลาด: {str(e)}")


def refresh_tree(self, kpi_id, tree, emp_data):
    """รีเฟรชตาราง"""
    for item in tree.get_children():
        tree.delete(item)

    for job in emp_data["kpi_jobs"].get(kpi_id, []):
        if kpi_id == "KPI1":
            comment_preview = job.get('comment', '')[:30] + "..." if len(job.get('comment', '')) > 30 else job.get(
                'comment', '')
            vals = (
                job['id'],
                job['memo'][:25] + "..." if len(job['memo']) > 25 else job['memo'],
                job['sent'],
                job['launch'],
                job['days'],
                f"{job['score']}/5",
                "✓" if job.get('image') else "✗",
                comment_preview
            )
        elif kpi_id == "KPI2":
            comment_preview = job.get('comment', '')[:30] + "..." if len(job.get('comment', '')) > 30 else job.get(
                'comment', '')
            vals = (
                job['id'],
                job['memo'][:25] + "..." if len(job['memo']) > 25 else job['memo'],
                job['sent'],
                job['hours'],
                f"{job['score']}/5",
                "✓" if job.get('image') else "✗",
                comment_preview
            )
        else:
            comment_preview = job.get('comment', '')[:40] + "..." if len(job.get('comment', '')) > 40 else job.get(
                'comment', '')
            vals = (
                job['id'],
                job.get('description', 'N/A')[:40],
                f"{job.get('score', 0)}/5",
                comment_preview
            )

        tree.insert("", tk.END, values=vals)


def delete_job(self, kpi_id, tree, emp_data):
    """ลบ Job"""
    selected = tree.selection()
    if not selected:
        messagebox.showwarning("คำเตือน", "กรุณาเลือกรายการที่ต้องการลบ")
        return

    if messagebox.askyesno("ยืนยัน", "คุณต้องการลบรายการนี้ใช่หรือไม่?"):
        item_id = tree.item(selected[0])['values'][0]
        emp_data["kpi_jobs"][kpi_id] = [j for j in emp_data["kpi_jobs"][kpi_id] if j['id'] != item_id]

        # Re-index
        for i, job in enumerate(emp_data["kpi_jobs"][kpi_id], 1):
            job['id'] = i

        self.save_all_data()
        self.refresh_tree(kpi_id, tree, emp_data)

        # Update score
        if hasattr(self, 'current_score_label'):
            total_score = self.calculate_total_score_for_employee(emp_data)
            self.current_score_label.config(text=f"{total_score:.2f}/5.00")

        messagebox.showinfo("สำเร็จ", "ลบรายการเรียบร้อยแล้ว")


def create_report_tab(self, parent, emp_data):
    """สร้างแท็บรายงาน"""
    container = tk.Frame(parent, bg='white')
    container.pack(fill='both', expand=True, padx=20, pady=20)

    # Header
    header_frame = tk.Frame(container, bg='#3498db', height=60)
    header_frame.pack(fill='x', pady=(0, 15))
    header_frame.pack_propagate(False)

    tk.Label(header_frame, text="📊 สรุปผลการปฏิบัติงาน KPI",
             font=('Segoe UI', 14, 'bold'), bg='#3498db', fg='white').pack(pady=15, padx=20)

    # Summary text
    self.summary_text = scrolledtext.ScrolledText(
        container,
        height=25,
        font=('Courier New', 10),
        relief='solid',
        borderwidth=1,
        bg='#f9f9f9'
    )
    self.summary_text.pack(fill='both', expand=True, pady=(0, 15))

    # Buttons
    btn_frame = tk.Frame(container, bg='white')
    btn_frame.pack(pady=10)

    tk.Button(btn_frame, text="🔄 คำนวณคะแนน", bg='#3498db', fg='white',
              font=('Segoe UI', 10, 'bold'), relief='flat', padx=25, pady=10,
              command=lambda: self.calculate_summary(emp_data)).pack(side=tk.LEFT, padx=5)

    tk.Button(btn_frame, text="📥 Export Excel", bg='#27ae60', fg='white',
              font=('Segoe UI', 10, 'bold'), relief='flat', padx=25, pady=10,
              command=lambda: self.export_to_excel(emp_data)).pack(side=tk.LEFT, padx=5)

    self.calculate_summary(emp_data)


def calculate_summary(self, emp_data):
    """คำนวณสรุปผล"""
    summary = []
    summary.append("=" * 120)
    summary.append(f"📊 สรุปผลการปฏิบัติงาน KPI Performance Report")
    summary.append(f"พนักงาน: {emp_data['info']['name']} | รหัส: {emp_data['info']['employee_id']}")
    summary.append(f"ตำแหน่ง: {emp_data['info']['position']} | ฝ่าย: {emp_data['info'].get('section', 'N/A')}")
    summary.append(f"วันที่สร้างรายงาน: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    summary.append("=" * 120 + "\n")

    total_score = 0
    kpi_scores = {}

    for kpi_id, config in emp_data["kpi_config"].items():
        if not config.get("enabled", True):
            continue

        jobs = emp_data["kpi_jobs"].get(kpi_id, [])
        summary.append(f"{'=' * 50}")
        summary.append(f"{kpi_id}: {config['name']}")
        summary.append(f"น้ำหนัก: {config['weight']}% | จำนวน Jobs: {len(jobs)}")
        summary.append(f"{'=' * 50}")

        if jobs:
            scored = [j for j in jobs if 'score' in j]
            if scored:
                avg_score = sum(j['score'] for j in scored) / len(scored)
                weighted = (avg_score * config['weight']) / 100
                summary.append(f"  ► คะแนนเฉลี่ย: {avg_score:.2f}/5")
                kpi_scores[kpi_id] = avg_score
            else:
                weighted = 0
                kpi_scores[kpi_id] = 0

            summary.append(f"  ► คะแนนถ่วงน้ำหนัก: {weighted:.4f} ({config['weight']}%)")
            total_score += weighted
        else:
            summary.append(f"  ⚠️ ยังไม่มีข้อมูล")
            kpi_scores[kpi_id] = 0

        summary.append("")

    summary.append("\n" + "=" * 120)
    summary.append(f"🎯 คะแนนรวมทั้งหมด: {total_score:.2f}/5.00")

    # Rating
    if total_score >= 4.5:
        rating = "⭐⭐⭐⭐⭐ Far Exceed Expectations"
    elif total_score >= 3.5:
        rating = "⭐⭐⭐⭐ Exceed Expectations"
    elif total_score >= 2.5:
        rating = "⭐⭐⭐ Meet Expectations"
    elif total_score >= 1.5:
        rating = "⭐⭐ Partially Meet Expectations"
    else:
        rating = "⭐ Not Meet Expectations"

    summary.append(f"📈 ระดับการประเมิน: {rating}")
    summary.append("=" * 120)

    self.summary_text.delete(1.0, tk.END)
    self.summary_text.insert(1.0, "\n".join(summary))


def show_dashboard(self):
    """แสดง Dashboard สรุปภาพรวมทุกคน"""
    if not self.employees:
        messagebox.showinfo("แจ้งเตือน", "ยังไม่มีข้อมูลพนักงานในระบบ")
        return

    dialog = tk.Toplevel(self.root)
    dialog.title("📊 Dashboard - สรุปภาพรวมทั้งหมด")
    dialog.geometry("1400x800")
    dialog.configure(bg='white')
    dialog.transient(self.root)

    # Center
    dialog.update_idletasks()
    x = (dialog.winfo_screenwidth() // 2) - (1400 // 2)
    y = (dialog.winfo_screenheight() // 2) - (800 // 2)
    dialog.geometry(f'1400x800+{x}+{y}')

    main_frame = tk.Frame(dialog, bg='white')
    main_frame.pack(fill='both', expand=True, padx=20, pady=20)

    main_frame = tk.Frame(dialog, bg='white')
