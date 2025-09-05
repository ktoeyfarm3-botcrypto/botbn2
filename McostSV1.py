import pandas as pd
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import numbers, Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import logging
import traceback
import os


# ===== Enhanced Configuration with Built-in Data =====
class ConfigManager:
    def __init__(self):
        self.config_file = "mapcost_config.ini"
        self.base_files = {
            'hashira': None,
            'hamada': None
        }
        self.initialize_default_data()
        self.load_config()

    def initialize_default_data(self):
        """สร้างข้อมูล Hamada Cost และ Hashira Cost จากข้อมูลจริง"""

        # ข้อมูลจริงจาก Hamada Cost.xlsx
        hamada_menu_names = [
            "Ebi Tempura Set", "Salmon Tempura Set", "Ebi & Salmon Tempura Set",
            "Ebi Tempura Don Set", "Salmon Tempura Don Set", "Ebi & Salmon Tempura Don Set",
            "Ebi Tempura Udon", "Salmon Tempura Udon", "Ebi & Salmon Tempura Udon",
            "Yakiudon Buta", "Yakiudon Kaisen", "Ebi Fry Curry Rice",
            "Tonkatsu Curry Rice", "Salmonkatsu Curry Rice", "Ebi Tempura（1pc）",
            "Salmon Tempura（1pc）", "Takoyaki", "Sushi Yorokobi",
            "Sushi Tokujo", "Maguro Zanmai", "Salmon Zanmai",
            "Hamachi Zanmai", "Duo Sushi", "Salmon Mountain",
            "Salmon Sushi & Ikura Gunkan", "California Roll", "Salmon Roll",
            "Unagi Roll", "Hokkaido Hotate Roll", "Salmon Chirashi Sushi",
            "Salmon Tobiko Chirashi Sushi", "Barachirashi Sushi", "Salmon Don",
            "Salmon Ikura Don", "Maguro Don", "Negitoro Ikura Don",
            "Kaisen Don", "Unagi Don", "Salmon Sashimi",
            "Maguro Sashimi", "Hamachi Sashimi", "Hotate Sashimi",
            "Salmon Sushi (2pcs)", "Maguro Sushi (2pcs)", "Hamachi Sushi (2pcs)",
            "Tamago Sushi (2pcs)", "Ebi Sushi (2pcs)", "Ika Sushi (2pcs)",
            "Hokki Sushi (2pcs)", "Hotate Sushi (2pcs)", "Unagi Sushi (2pcs)",
            "Anago Sushi (2pcs)", "Negitoro Gunkan (2pcs)", "Tobiko Gunkan (2pcs)",
            "Ikura Gunkan (2pcs)", "GG California Roll", "GG Salmon Sushi",
            "GG Tamago & Ebi Sushi", "GG Ebi & Tobiko Don", "GG Kanikama & Tobiko Don",
            "GG Triple Mix Don", "GG Ebiten Don", "GG Salmon & Ebiten Don",
            "Coca-Cola Original", "Coke-Cola Zero Sugar", "Sprite",
            "Soda", "Mineral Water", "HBD Sparkling Water Lemon (No Sugar, No Calories)",
            "HBD Sparkling Water Honey Yuzu (No Sugar, No Calories)",
            "HBD Sparkling Water Peach (No Sugar, No Calories)",
            "HBD Sparkling Water Kyoho (No Sugar, No Calories)", "Asahi Beer",
            "Heineken Beer ", "Singha Beer", "Drinking Water",
            "Matcha Soft Serve", "Yogurt Soft Serve", "Two Tone Soft Serve",
            "Matcha Japanese", "Matcha Marshmallow", "Mix Berry Yogurt",
            "Brownie Yogurt", "Two Tone Brownie", "Saikyo Matcha",
            "Rice", "Miso Soup"
        ]

        hamada_costs = [
            67.53, 51.84, 70.42, 69.33, 53.64, 69.29,
            56.03, 54.34, 73.57, 37.21, 48.67, 62.31,
            69.17, 60.62, 16.16, 10.92, 33.71, 63.56,
            203.12, 89.93, 98.14, 168.1, 108.73, 200.7,
            193.91, 50.76, 87.59, 104.1, 133.97, 48.53,
            66.95, 77.77, 79.45, 108.68, 73.91, 111.38,
            243.48, 246.42, 42.71, 39.59, 60.89, 52.93,
            29.95, 27.87, 42.07, 9.63, 19.94, 19.87,
            19.96, 51.71, 56.12, 112.92, 21.27, 18.32,
            68.24, 51.43, 93.14, 56.38, 68.9, 48.9,
            58.8, 63.34, 74.35, 12.54, 12.54, 12.54,
            7.59, 7, 17, 17, 17, 17, 41.92,
            38.43, 31.71, 3.74, 15.09, 17.72, 16.41,
            22.6, 30.78, 34.94, 41.79, 37.45, 31.54,
            5.85, 3.58
        ]

        # ข้อมูลจริงจาก Hashira Cost.xlsx
        hashira_menu_names = [
            "Pork Cut Steak Set", "Tonteki Set", "Beef Cut Steak Set (Medium Rare / Well done)",
            "Buta Teriyaki Set", "Tonkatsu Set", "Tonkatsu Tamagotoji Set",
            "Beef Hamburg Set", "Salmon Set (Teriyaki / Shioyaki)", "Kaisen Teppan Yaki Set",
            "Topping Fried Egg", "Gokuatsu Tonkatsu Set", "Gokuatsu Katsu Don Set",
            "Gokuatsu Tonkatsu Tamagotoji Set", "Tonkotsu Chashumen", "Tonkotsu Niku-Niku Chashumen",
            "Tonkotsu Spicy Ramen", "Tonkotsu Hashira Ramen", "Miso Chashumen",
            "Miso Hokkaido Ramen", "Shoyu Chashumen", "Shoyu Tokyo Ramen",
            "Tom Yum Tonkotsu Kaisen Ramen", "Tom Yum Kung Ramen", "Tori Paitan Karaage Ramen",
            "Tori Paitan Kaisen Ramen", "Tonkatsu", "Hotate",
            "Ebi Tempura", "Chashu", "Spicy Nikumiso",
            "Karashinegi", "Negi", "Naruto",
            "Ajitama", "Menma", "Nori",
            "Corn", "Butter", "Boiled Cabbage",
            "Boiled Bean Sprout", "Karashi Takana", "Kikurage",
            "Kanikama", "Kaedama", "GG Sauce Katsudon",
            "GG Buta Teriyaki Don", "GG Tonkatsu Bento", "GG Yakisoba",
            "GG Yakisoba with Sunny Side Egg", "GG Pork Okonomiyaki", "Coca-Cola Original",
            "Coca-Cola Zero Sugar", "Sprite", "Soda",
            "Mineral Water", "HBD Sparkling Water Lemon (No Sugar, No Calories)",
            "HBD Sparkling Water Honey Yuzu (No Sugar, No Calories)",
            "HBD Sparkling Water Peach (No Sugar, No Calories)",
            "HBD Sparkling Water Kyoho (No Sugar, No Calories)", "Asahi Beer",
            "Heineken Beer ", "Singha Beer", "Drinking Water",
            "Melon Ball", "Vanilla Monaka", "Azuki Bar",
            "Rice", "Miso Soup"
        ]

        hashira_costs = [
            46.53, 131.99, 107.87, 41.77, 58.39, 60.91,
            76.4, 68.02, 121.23, 3.5, 131.77, 133,
            131.46, 64.23, 77.98, 54.05, 74.33, 56.52,
            70.08, 63.1, 64.44, 55.55, 116.08, 42.88,
            50.08, 37.4, 33.33, 26, 27.3, 7.97,
            10.54, 6.43, 12.92, 3.84, 9.6, 5.4,
            6, 1.31, 2.73, 1.98, 6.45, 3.41,
            7.92, 7.5, 67.94, 45.93, 79.95, 45.61,
            49.11, 52.4, 12.54, 12.54, 12.54, 7.59,
            7, 17, 17, 17, 17, 41.92,
            38.43, 31.71, 3.74, 25, 27, 27,
            5.85, 3.04
        ]

        # สร้าง DataFrame
        hamada_real_data = {
            "MENU NAME": hamada_menu_names,
            "Material Cost": hamada_costs
        }

        hashira_real_data = {
            "MENU NAME": hashira_menu_names,
            "Material Cost": hashira_costs
        }

        # สร้าง DataFrame และ set index
        self.base_files['hamada'] = pd.DataFrame(hamada_real_data).set_index('MENU NAME')
        self.base_files['hashira'] = pd.DataFrame(hashira_real_data).set_index('MENU NAME')

        logger.info("โหลดข้อมูลจริง Base Cost เรียบร้อย")
        logger.info(f"Hamada Cost: {len(self.base_files['hamada'])} เมนู")
        logger.info(f"Hashira Cost: {len(self.base_files['hashira'])} เมนู")

    def load_config(self):
        """โหลดการตั้งค่าจากไฟล์"""
        # โหลดไฟล์ Base Cost.xlsx หากมี (จะ override ข้อมูลตัวอย่าง)
        self.load_base_cost_file()

    def load_base_cost_file(self):
        """โหลดไฟล์ Base Cost.xlsx"""
        base_file = "Base Cost.xlsx"
        if os.path.exists(base_file):
            try:
                # อ่าน Hashira Cost sheet
                hashira_df = pd.read_excel(base_file, sheet_name="Hashira Cost")
                if 'MENU NAME' in hashira_df.columns:
                    self.base_files['hashira'] = hashira_df.set_index("MENU NAME")

                # อ่าน Hamada Cost sheet
                hamada_df = pd.read_excel(base_file, sheet_name="Hamada Cost")
                if 'MENU NAME' in hamada_df.columns:
                    self.base_files['hamada'] = hamada_df.set_index("MENU NAME")

                logger.info(f"โหลดไฟล์ {base_file} สำเร็จ (แทนที่ข้อมูลตัวอย่าง)")
                logger.info(f"Hashira Cost: {len(self.base_files['hashira'])} เมนู")
                logger.info(f"Hamada Cost: {len(self.base_files['hamada'])} เมนู")

            except Exception as e:
                logger.error(f"Error loading base file: {e}")
                logger.info("ใช้ข้อมูลตัวอย่างต่อไป")

    def save_base_cost_file(self):
        """บันทึกไฟล์ Base Cost.xlsx"""
        try:
            wb = Workbook()
            # ลบ sheet default
            wb.remove(wb.active)

            # สร้าง Hashira Cost sheet
            if self.base_files['hashira'] is not None:
                ws_hashira = wb.create_sheet("Hashira Cost")
                df_hashira = self.base_files['hashira'].reset_index()

                for r in dataframe_to_rows(df_hashira, index=False, header=True):
                    ws_hashira.append(r)

                # จัดรูปแบบ header
                for cell in ws_hashira[1]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")

            # สร้าง Hamada Cost sheet
            if self.base_files['hamada'] is not None:
                ws_hamada = wb.create_sheet("Hamada Cost")
                df_hamada = self.base_files['hamada'].reset_index()

                for r in dataframe_to_rows(df_hamada, index=False, header=True):
                    ws_hamada.append(r)

                # จัดรูปแบบ header
                for cell in ws_hamada[1]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")

            # ปรับความกว้างคอลัมน์
            for ws in wb.worksheets:
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 60)
                    ws.column_dimensions[column_letter].width = adjusted_width

            wb.save("Base Cost.xlsx")
            logger.info("บันทึก Base Cost.xlsx สำเร็จ")
            return True

        except Exception as e:
            logger.error(f"Error saving base file: {e}")
            return False

    def update_single_base(self, base_type, new_df):
        """อัพเดท Base Cost แยกกันตามประเภท"""
        try:
            # ลบช่องว่างจากชื่อคอลัมน์
            new_df.columns = new_df.columns.str.strip()

            # ตรวจสอบคอลัมน์ที่จำเป็น
            if 'MENU NAME' not in new_df.columns:
                raise ValueError(f"ไฟล์ไม่มีคอลัมน์ 'MENU NAME'")

            if 'Material Cost' not in new_df.columns:
                raise ValueError(f"ไฟล์ไม่มีคอลัมน์ 'Material Cost'")

            # ตั้งค่า index
            new_df = new_df.set_index("MENU NAME")

            # อัพเดทข้อมูล
            self.base_files[base_type] = new_df

            # บันทึกไฟล์ Base Cost.xlsx
            self.save_base_cost_file()

            logger.info(f"อัพเดท {base_type.upper()} Cost สำเร็จ: {len(new_df)} เมนู")
            return True

        except Exception as e:
            logger.error(f"Error updating {base_type} base: {e}")
            raise e

    def get_default_base_data(self, base_type):
        """ส่งคืนข้อมูล Base Cost ปัจจุบันเพื่อการ Export"""
        return self.base_files.get(base_type)


# ===== Modern GUI Styling =====
def configure_modern_style():
    """กำหนดธีมสีสวยงามและทันสมัย"""
    style = ttk.Style()

    try:
        style.theme_use('clam')
    except:
        style.theme_use('default')

    colors = {
        'bg_primary': '#2c3e50',
        'bg_secondary': '#34495e',
        'accent': '#3498db',
        'success': '#27ae60',
        'warning': '#f39c12',
        'danger': '#e74c3c',
        'light': '#ecf0f1',
        'white': '#ffffff',
        'text_dark': '#2c3e50',
        'text_light': '#ffffff'
    }

    # Configure styles
    style.configure('Title.TLabel', font=('Segoe UI', 16, 'bold'), foreground=colors['text_dark'])
    style.configure('Heading.TLabel', font=('Segoe UI', 12, 'bold'), foreground=colors['accent'])
    style.configure('Status.TLabel', font=('Segoe UI', 10), padding=(5, 2))

    style.configure('Accent.TButton', font=('Segoe UI', 9, 'bold'), padding=(10, 5))
    style.configure('Success.TButton', font=('Segoe UI', 9, 'bold'), padding=(10, 5))
    style.configure('Warning.TButton', font=('Segoe UI', 9, 'bold'), padding=(10, 5))
    style.configure('Danger.TButton', font=('Segoe UI', 9, 'bold'), padding=(10, 5))

    style.configure('Card.TFrame', relief='solid', borderwidth=1, padding=10)
    style.configure('Modern.TNotebook', tabposition='n')
    style.configure('Modern.TNotebook.Tab', font=('Segoe UI', 10, 'bold'), padding=(20, 8))

    return colors


# ===== Logging Setup =====
logging.basicConfig(
    level=logging.DEBUG,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('hamada_calculator_debug.log', encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# ===== Language System =====
LANGUAGES = {
    'th': {
        'app_title': 'Hamada & Hashira Cost Calculator - Enhanced',
        'select_import_file': 'เลือกไฟล์ Import',
        'calculate': 'คำนวณ',
        'save_excel': 'บันทึก Excel',
        'save_excel_debug': 'บันทึก Excel + Debug',
        'check_base_file': 'ตรวจสอบไฟล์ Base',
        'export_base_file': 'Export ไฟล์ Base',
        'import_new_base': 'Import Base ใหม่',
        'update_hashira': 'อัพเดท Hashira Cost',
        'update_hamada': 'อัพเดท Hamada Cost',
        'import_template_file': 'Import Template เพิ่มเติม',
        'clear_debug': 'เคลียร์ Debug',
        'debug_mode': 'เปิด Debug Mode',
        'select_base_type': 'เลือกประเภท Base Cost',
        'hashira_cost': 'Hashira Cost',
        'hamada_cost': 'Hamada Cost',
        'no_file_selected': 'ยังไม่ได้เลือกไฟล์',
        'file_selected': 'ไฟล์',
        'base_file_ready': 'ไฟล์ Base: พร้อมใช้งาน',
        'base_file_not_found': 'ไฟล์ Base: ไม่พบ',
        'results_tab': 'ผลลัพธ์',
        'debug_tab': 'Debug Log',
        'statistics': 'สถิติ',
        'no_data': 'ยังไม่มีข้อมูล',
        'menu_name': 'MENU NAME',
        'qty': 'Qty',
        'material_cost': 'Material Cost',
        'total_cost': 'Total Cost',
        'grand_total': 'Grand Total',
        'matched_menus': 'พบเมนู',
        'not_found_menus': 'ไม่พบ',
        'total_amount': 'รวม',
        'baht': 'บาท',
        'error': 'ผิดพลาด',
        'warning': 'คำเตือน',
        'success': 'สำเร็จ',
        'please_select_file': 'กรุณาเลือกไฟล์ Import ก่อน',
        'calculation_complete': 'คำนวณเสร็จสิ้น - พบ',
        'items': 'รายการ',
        'sheet_results': 'ผลลัพธ์',
        'sheet_import_data': 'ข้อมูล Import',
        'sheet_template_data': 'ข้อมูล Template',
        'sheet_base_data': 'ข้อมูล Base Cost',
        'sheet_not_found': 'เมนูที่ไม่พบ',
        'sheet_summary': 'สรุปการประมวลผล',
        'template_file': 'ไฟล์ Template'
    }
}

# Initialize configuration
config_manager = ConfigManager()


class EnhancedCostCalculatorApp:
    def __init__(self, root):
        self.root = root
        self.current_language = 'th'
        self.colors = configure_modern_style()

        self.root.title(self.t('app_title'))
        self.root.geometry("1500x950")
        self.root.configure(bg=self.colors['light'])
        self.root.minsize(1300, 850)

        # ตัวแปรสำหรับไฟล์
        self.import_file = None
        self.template_file = None
        self.selected_base_type = tk.StringVar(value='hamada')  # default
        self.df_result = None
        self.debug_mode = tk.BooleanVar(value=False)

        # ข้อมูล debug แยกตามประเภท
        self.debug_data = {
            'not_found_menus': [],
            'zero_qty_items': [],
            'invalid_qty_items': [],
            'nan_cost_items': [],
            'matched_menus': [],
            'not_sold_menus': [],
            'processing_summary': {}
        }

        self._build_ui()
        logger.info("Enhanced Cost Calculator initialized with built-in data")

    def t(self, key):
        """แปลข้อความตามภาษาที่เลือก"""
        return LANGUAGES.get(self.current_language, {}).get(key, key)

    def _build_ui(self):
        """สร้าง UI แบบ Enhanced"""
        # Main container
        main_frame = ttk.Frame(self.root, style='Card.TFrame')
        main_frame.pack(fill=tk.BOTH, expand=True, padx=15, pady=15)

        # Enhanced Header
        header_frame = ttk.Frame(main_frame)
        header_frame.pack(fill=tk.X, pady=(0, 20))

        # Title with modern styling
        title_label = ttk.Label(header_frame, text="🍱 Hamada & Hashira Cost Calculator",
                                style='Title.TLabel')
        title_label.pack(pady=(0, 5))

        subtitle = ttk.Label(header_frame, text="Enhanced Multi-Base Cost Management System (Real Data Loaded)",
                             font=('Segoe UI', 11), foreground=self.colors['bg_secondary'])
        subtitle.pack()

        # ===== Base Type Selection Card =====
        base_selection_frame = ttk.LabelFrame(main_frame, text="🏪 Base Cost Selection",
                                              padding=15, style='Card.TFrame')
        base_selection_frame.pack(fill=tk.X, pady=(0, 15))

        # Base type radio buttons
        base_frame = ttk.Frame(base_selection_frame)
        base_frame.pack(fill=tk.X)

        ttk.Label(base_frame, text="เลือกประเภท Base Cost:",
                  font=('Segoe UI', 11, 'bold')).pack(side=tk.LEFT, padx=(0, 20))

        # Radio buttons สำหรับเลือก base type
        hashira_radio = ttk.Radiobutton(base_frame, text="🏯 Hashira Cost",
                                        variable=self.selected_base_type, value='hashira',
                                        command=self.on_base_type_change)
        hashira_radio.pack(side=tk.LEFT, padx=10)

        hamada_radio = ttk.Radiobutton(base_frame, text="🍜 Hamada Cost",
                                       variable=self.selected_base_type, value='hamada',
                                       command=self.on_base_type_change)
        hamada_radio.pack(side=tk.LEFT, padx=10)

        # Base file status
        self.base_status_label = ttk.Label(base_frame, text="", style='Status.TLabel')
        self.base_status_label.pack(side=tk.RIGHT, padx=20)

        # ===== File Management Section =====
        file_frame = ttk.LabelFrame(main_frame, text="📁 File Management",
                                    padding=15, style='Card.TFrame')
        file_frame.pack(fill=tk.X, pady=(0, 15))

        # File selection buttons - Row 1: Import และ Template
        file_btn_frame1 = ttk.Frame(file_frame)
        file_btn_frame1.pack(fill=tk.X, pady=(0, 8))

        ttk.Button(file_btn_frame1, text="📁 เลือกไฟล์ Import",
                   command=self.load_import_file, style='Accent.TButton').pack(
            side=tk.LEFT, padx=5, fill=tk.X, expand=True)

        ttk.Button(file_btn_frame1, text="📄 Import Template เพิ่มเติม",
                   command=self.load_template_file, style='Warning.TButton').pack(
            side=tk.LEFT, padx=5, fill=tk.X, expand=True)

        ttk.Button(file_btn_frame1, text="🔍 ตรวจสอบไฟล์ Base",
                   command=self.check_base_file, style='Accent.TButton').pack(
            side=tk.LEFT, padx=5, fill=tk.X, expand=True)

        # File selection buttons - Row 2: Base Management
        file_btn_frame2 = ttk.Frame(file_frame)
        file_btn_frame2.pack(fill=tk.X, pady=(0, 8))

        ttk.Button(file_btn_frame2, text="📥 Import Base Cost ใหม่ (ทั้งหมด)",
                   command=self.import_new_base, style='Warning.TButton').pack(
            side=tk.LEFT, padx=5, fill=tk.X, expand=True)

        ttk.Button(file_btn_frame2, text="📤 Export Base Cost",
                   command=self.export_base_file, style='Accent.TButton').pack(
            side=tk.LEFT, padx=5, fill=tk.X, expand=True)

        ttk.Button(file_btn_frame2, text="💾 บันทึก Base เป็นไฟล์",
                   command=self.save_current_base, style='Success.TButton').pack(
            side=tk.LEFT, padx=5, fill=tk.X, expand=True)

        # File selection buttons - Row 3: อัพเดท Base แยกกัน
        file_btn_frame3 = ttk.Frame(file_frame)
        file_btn_frame3.pack(fill=tk.X)

        ttk.Button(file_btn_frame3, text="🏯 อัพเดท Hashira Cost",
                   command=self.update_hashira_cost, style='Danger.TButton').pack(
            side=tk.LEFT, padx=5, fill=tk.X, expand=True)

        ttk.Button(file_btn_frame3, text="🍜 อัพเดท Hamada Cost",
                   command=self.update_hamada_cost, style='Danger.TButton').pack(
            side=tk.LEFT, padx=5, fill=tk.X, expand=True)

        # File status display
        status_frame = ttk.Frame(file_frame)
        status_frame.pack(fill=tk.X, pady=(10, 0))

        self.import_status_label = ttk.Label(status_frame, text="📂 Import: ยังไม่ได้เลือกไฟล์",
                                             style='Status.TLabel', foreground=self.colors['danger'])
        self.import_status_label.pack(anchor="w", pady=2)

        self.template_status_label = ttk.Label(status_frame, text="📄 Template: ยังไม่ได้เลือกไฟล์",
                                               style='Status.TLabel', foreground=self.colors['warning'])
        self.template_status_label.pack(anchor="w", pady=2)

        # Base status display สำหรับแสดงสถานะทั้ง 2 ประเภท
        self.hashira_status_label = ttk.Label(status_frame, text="🏯 Hashira Cost: กำลังตรวจสอบ...",
                                              style='Status.TLabel', foreground=self.colors['warning'])
        self.hashira_status_label.pack(anchor="w", pady=2)

        self.hamada_status_label = ttk.Label(status_frame, text="🍜 Hamada Cost: กำลังตรวจสอบ...",
                                             style='Status.TLabel', foreground=self.colors['warning'])
        self.hamada_status_label.pack(anchor="w", pady=2)

        # ===== Action Buttons =====
        action_frame = ttk.LabelFrame(main_frame, text="🚀 Actions",
                                      padding=15, style='Card.TFrame')
        action_frame.pack(fill=tk.X, pady=(0, 15))

        # Debug mode checkbox
        debug_frame = ttk.Frame(action_frame)
        debug_frame.pack(fill=tk.X, pady=(0, 10))

        self.debug_checkbox = ttk.Checkbutton(debug_frame, text="🔍 เปิด Debug Mode (แสดงรายละเอียด)",
                                              variable=self.debug_mode)
        self.debug_checkbox.pack(side=tk.LEFT)

        ttk.Button(debug_frame, text="🗑️ เคลียร์ Debug",
                   command=self.clear_debug, style='Warning.TButton').pack(side=tk.RIGHT)

        # Action buttons
        btn_frame = ttk.Frame(action_frame)
        btn_frame.pack(fill=tk.X)

        ttk.Button(btn_frame, text="⚡ คำนวณ", command=self.calculate,
                   style='Success.TButton').pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)

        ttk.Button(btn_frame, text="💾 บันทึก Excel", command=self.export_excel,
                   style='Success.TButton').pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)

        ttk.Button(btn_frame, text="🔍💾 บันทึก Excel + Debug", command=self.export_excel_with_debug,
                   style='Warning.TButton').pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)

        # ===== Content Area =====
        content_frame = ttk.Frame(main_frame, style='Card.TFrame')
        content_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 15))

        # Notebook for tabs
        self.notebook = ttk.Notebook(content_frame, style='Modern.TNotebook')
        self.notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # ===== Results Tab =====
        result_frame = ttk.Frame(self.notebook)
        self.notebook.add(result_frame, text=f"📋 {self.t('results_tab')}")

        table_frame = ttk.Frame(result_frame, padding=10)
        table_frame.pack(fill=tk.BOTH, expand=True)

        table_header = ttk.Label(table_frame, text="💰 Calculation Results",
                                 style='Heading.TLabel')
        table_header.pack(pady=(0, 10))

        # Results table
        self.tree = ttk.Treeview(table_frame, columns=("menu", "qty", "cost", "total"),
                                 show="headings", height=16)

        self.tree.heading("menu", text=f"🍽️ {self.t('menu_name')}")
        self.tree.heading("qty", text=f"📊 {self.t('qty')}")
        self.tree.heading("cost", text=f"💵 {self.t('material_cost')}")
        self.tree.heading("total", text=f"💰 {self.t('total_cost')}")

        self.tree.column("menu", width=400, anchor="w")
        self.tree.column("qty", width=100, anchor="center")
        self.tree.column("cost", width=150, anchor="e")
        self.tree.column("total", width=150, anchor="e")

        tree_scroll = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=tree_scroll.set)

        tree_scroll.pack(side=tk.RIGHT, fill=tk.Y, padx=(5, 0))
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # ===== Debug Tab =====
        debug_frame = ttk.Frame(self.notebook)
        self.notebook.add(debug_frame, text=f"🔍 {self.t('debug_tab')}")

        debug_header_frame = ttk.Frame(debug_frame, padding=10)
        debug_header_frame.pack(fill=tk.X)

        debug_title = ttk.Label(debug_header_frame, text="🔧 Enhanced Debug Console",
                                style='Heading.TLabel')
        debug_title.pack()

        debug_text_frame = ttk.Frame(debug_frame, padding=(10, 0, 10, 10))
        debug_text_frame.pack(fill=tk.BOTH, expand=True)

        self.debug_text = scrolledtext.ScrolledText(debug_text_frame, wrap=tk.WORD, height=18,
                                                    font=('Consolas', 10),
                                                    bg='#1e1e1e', fg='#ffffff',
                                                    insertbackground='#ffffff',
                                                    selectbackground='#3498db',
                                                    relief='flat', borderwidth=0)
        self.debug_text.pack(fill=tk.BOTH, expand=True)

        # ===== Statistics Section =====
        self.stats_frame = ttk.LabelFrame(main_frame, text="📈 Statistics & Info",
                                          padding=15, style='Card.TFrame')
        self.stats_frame.pack(fill=tk.X)

        self.stats_label = ttk.Label(self.stats_frame, text=self.t('no_data'),
                                     style='Status.TLabel',
                                     font=('Segoe UI', 11, 'bold'))
        self.stats_label.pack()

        self._configure_treeview_style()

        # อัพเดทสถานะทั้ง 2 ประเภทเมื่อเริ่มโปรแกรม
        self.update_all_base_status()
        self.update_base_status()

    def _configure_treeview_style(self):
        """กำหนดสไตล์ modern สำหรับ treeview"""
        style = ttk.Style()

        style.configure('Modern.Treeview',
                        background='#ffffff',
                        foreground='#2c3e50',
                        fieldbackground='#ffffff',
                        font=('Segoe UI', 10))

        style.configure('Modern.Treeview.Heading',
                        background='#3498db',
                        foreground='#ffffff',
                        font=('Segoe UI', 10, 'bold'),
                        relief='flat')

        style.map('Modern.Treeview',
                  background=[('selected', '#3498db')],
                  foreground=[('selected', '#ffffff')])

        self.tree.configure(style='Modern.Treeview')

    def on_base_type_change(self):
        """เมื่อเปลี่ยน base type"""
        selected = self.selected_base_type.get()
        self.debug_log(f"เปลี่ยน Base Type เป็น: {selected.upper()}")
        self.update_base_status()

    def update_base_status(self):
        """อัพเดทสถานะไฟล์ base ที่เลือก (แสดงใน selection area)"""
        selected = self.selected_base_type.get()
        base_df = config_manager.base_files.get(selected)

        if base_df is not None and not base_df.empty:
            menu_count = len(base_df)
            status_text = f"🗂️ {selected.upper()} Cost: พร้อมใช้งาน ({menu_count} เมนู)"
            color = self.colors['success']
        else:
            status_text = f"🗂️ {selected.upper()} Cost: ไม่พบข้อมูล"
            color = self.colors['danger']

        self.base_status_label.config(text=status_text, foreground=color)

    def update_all_base_status(self):
        """อัพเดทสถานะไฟล์ base ทั้งหมดในพื้นที่ status"""
        # Hashira Status
        hashira_df = config_manager.base_files.get('hashira')
        if hashira_df is not None and not hashira_df.empty:
            hashira_count = len(hashira_df)
            hashira_text = f"🏯 Hashira Cost: พร้อมใช้งาน ({hashira_count} เมนู) [Real Data]"
            hashira_color = self.colors['success']
        else:
            hashira_text = f"🏯 Hashira Cost: ไม่พบข้อมูล"
            hashira_color = self.colors['danger']

        # Hamada Status
        hamada_df = config_manager.base_files.get('hamada')
        if hamada_df is not None and not hamada_df.empty:
            hamada_count = len(hamada_df)
            hamada_text = f"🍜 Hamada Cost: พร้อมใช้งาน ({hamada_count} เมนู) [Real Data]"
            hamada_color = self.colors['success']
        else:
            hamada_text = f"🍜 Hamada Cost: ไม่พบข้อมูล"
            hamada_color = self.colors['danger']

        self.hashira_status_label.config(text=hashira_text, foreground=hashira_color)
        self.hamada_status_label.config(text=hamada_text, foreground=hamada_color)

    def debug_log(self, message, level="INFO"):
        """เพิ่มข้อความใน debug log"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        log_message = f"[{timestamp}] {level}: {message}\n"

        self.debug_text.insert(tk.END, log_message)
        self.debug_text.see(tk.END)

        if level == "ERROR":
            logger.error(message)
        elif level == "WARNING":
            logger.warning(message)
        else:
            logger.info(message)

    def clear_debug(self):
        """เคลียร์ debug log"""
        self.debug_text.delete(1.0, tk.END)
        self.debug_log("Debug log cleared")
        self.debug_data = {
            'not_found_menus': [],
            'zero_qty_items': [],
            'invalid_qty_items': [],
            'nan_cost_items': [],
            'matched_menus': [],
            'not_sold_menus': [],
            'processing_summary': {}
        }

    def load_import_file(self):
        """โหลดไฟล์ import"""
        file_path = filedialog.askopenfilename(
            title="เลือกไฟล์ Import (ข้อมูลที่ต้องการคำนวณ)",
            filetypes=[("Excel Files", "*.xlsx *.xls")]
        )
        if not file_path:
            return

        self.import_file = file_path
        self.import_status_label.config(
            text=f"📂 Import: {os.path.basename(file_path)}",
            foreground=self.colors['success']
        )

        self.debug_log(f"เลือกไฟล์ Import: {file_path}")
        self._preview_file(file_path, "Import")

    def load_template_file(self):
        """โหลดไฟล์ template เพิ่มเติม"""
        file_path = filedialog.askopenfilename(
            title="เลือกไฟล์ Template เพิ่มเติม",
            filetypes=[("Excel Files", "*.xlsx *.xls")]
        )
        if not file_path:
            return

        self.template_file = file_path
        self.template_status_label.config(
            text=f"📄 Template: {os.path.basename(file_path)}",
            foreground=self.colors['success']
        )

        self.debug_log(f"เลือกไฟล์ Template: {file_path}")
        self._preview_file(file_path, "Template")

    def _preview_file(self, file_path, file_type):
        """แสดงตัวอย่างข้อมูลในไฟล์"""
        try:
            df_preview = pd.read_excel(file_path)
            self.debug_log(f"ไฟล์ {file_type} มี {len(df_preview)} แถว, {len(df_preview.columns)} คอลัมน์")
            self.debug_log(f"คอลัมน์: {df_preview.columns.tolist()}")

            # ตรวจสอบคอลัมน์ที่จำเป็น
            required_cols = ["MENU NAME", "Qty"] if file_type != "Base Cost" else ["MENU NAME", "Material Cost"]
            missing_cols = [col for col in required_cols if col not in df_preview.columns]
            if missing_cols:
                self.debug_log(f"คำเตือน: ไฟล์ขาดคอลัมน์ {missing_cols}", "WARNING")

            # แสดงตัวอย่างข้อมูล
            self.debug_log(f"ตัวอย่างข้อมูล {file_type} 5 แถวแรก:")
            for i, row in df_preview.head(5).iterrows():
                menu = row.get("MENU NAME", "N/A")
                if file_type == "Base Cost":
                    cost = row.get("Material Cost", "N/A")
                    self.debug_log(f"  แถว {i + 1}: {menu} = {cost}")
                else:
                    qty = row.get("Qty", "N/A")
                    self.debug_log(f"  แถว {i + 1}: {menu} = {qty}")

        except Exception as e:
            self.debug_log(f"Error ตรวจสอบไฟล์ {file_type}: {str(e)}", "ERROR")

    # ===== ฟังก์ชันใหม่สำหรับอัพเดท Base Cost แยกกัน =====
    def update_hashira_cost(self):
        """อัพเดท Hashira Cost เท่านั้น"""
        file_path = filedialog.askopenfilename(
            title="เลือกไฟล์ Hashira Cost ใหม่",
            filetypes=[("Excel Files", "*.xlsx *.xls")]
        )
        if not file_path:
            return

        self.debug_log("=== อัพเดท Hashira Cost ===")
        self.debug_log(f"ไฟล์ที่เลือก: {os.path.basename(file_path)}")

        try:
            # ตรวจสอบว่ามี sheet ชื่ออะไรใช้ได้บ้าง
            excel_file = pd.ExcelFile(file_path)
            self.debug_log(f"Sheets ที่มี: {excel_file.sheet_names}")

            # ลองอ่านจาก sheet ที่เป็นไปได้
            df_new = None
            possible_sheets = ["Hashira Cost", "Sheet1", excel_file.sheet_names[0]]

            for sheet_name in possible_sheets:
                if sheet_name in excel_file.sheet_names:
                    try:
                        df_new = pd.read_excel(file_path, sheet_name=sheet_name)
                        self.debug_log(f"อ่านข้อมูลจาก sheet: {sheet_name}")
                        break
                    except Exception as e:
                        self.debug_log(f"ไม่สามารถอ่าน sheet {sheet_name}: {e}", "WARNING")
                        continue

            if df_new is None:
                raise ValueError("ไม่สามารถอ่านข้อมูลจากไฟล์ได้")

            # ตรวจสอบข้อมูลและอัพเดท
            self._validate_and_update_base(df_new, 'hashira', file_path)

        except Exception as e:
            error_msg = f"ไม่สามารถอัพเดท Hashira Cost ได้: {str(e)}"
            self.debug_log(error_msg, "ERROR")
            self.debug_log(traceback.format_exc(), "ERROR")
            messagebox.showerror("Error", error_msg)

    def update_hamada_cost(self):
        """อัพเดท Hamada Cost เท่านั้น"""
        file_path = filedialog.askopenfilename(
            title="เลือกไฟล์ Hamada Cost ใหม่",
            filetypes=[("Excel Files", "*.xlsx *.xls")]
        )
        if not file_path:
            return

        self.debug_log("=== อัพเดท Hamada Cost ===")
        self.debug_log(f"ไฟล์ที่เลือก: {os.path.basename(file_path)}")

        try:
            # ตรวจสอบว่ามี sheet ชื่ออะไรใช้ได้บ้าง
            excel_file = pd.ExcelFile(file_path)
            self.debug_log(f"Sheets ที่มี: {excel_file.sheet_names}")

            # ลองอ่านจาก sheet ที่เป็นไปได้
            df_new = None
            possible_sheets = ["Hamada Cost", "Sheet1", excel_file.sheet_names[0]]

            for sheet_name in possible_sheets:
                if sheet_name in excel_file.sheet_names:
                    try:
                        df_new = pd.read_excel(file_path, sheet_name=sheet_name)
                        self.debug_log(f"อ่านข้อมูลจาก sheet: {sheet_name}")
                        break
                    except Exception as e:
                        self.debug_log(f"ไม่สามารถอ่าน sheet {sheet_name}: {e}", "WARNING")
                        continue

            if df_new is None:
                raise ValueError("ไม่สามารถอ่านข้อมูลจากไฟล์ได้")

            # ตรวจสอบข้อมูลและอัพเดท
            self._validate_and_update_base(df_new, 'hamada', file_path)

        except Exception as e:
            error_msg = f"ไม่สามารถอัพเดท Hamada Cost ได้: {str(e)}"
            self.debug_log(error_msg, "ERROR")
            self.debug_log(traceback.format_exc(), "ERROR")
            messagebox.showerror("Error", error_msg)

    def _validate_and_update_base(self, df_new, base_type, file_path):
        """ตรวจสอบและอัพเดท Base Cost"""
        try:
            # ลบช่องว่างจากชื่อคอลัมน์
            df_new.columns = df_new.columns.str.strip()

            self.debug_log(f"คอลัมน์ในไฟล์: {df_new.columns.tolist()}")
            self.debug_log(f"จำนวนแถว: {len(df_new)}")

            # ตรวจสอบคอลัมน์ที่จำเป็น
            required_columns = ['MENU NAME', 'Material Cost']
            missing_columns = [col for col in required_columns if col not in df_new.columns]

            if missing_columns:
                raise ValueError(f"ไฟล์ขาดคอลัมน์ที่จำเป็น: {missing_columns}")

            # ลบแถวที่ MENU NAME เป็นว่าง
            original_count = len(df_new)
            df_new = df_new.dropna(subset=['MENU NAME'])
            df_new = df_new[df_new['MENU NAME'] != '']
            cleaned_count = len(df_new)

            if original_count != cleaned_count:
                self.debug_log(f"ลบแถวที่ชื่อเมนูว่าง: {original_count - cleaned_count} แถว")

            # แสดงตัวอย่างข้อมูล
            self.debug_log("ตัวอย่างข้อมูลใหม่ 10 แถวแรก:")
            for i, (idx, row) in enumerate(df_new.head(10).iterrows()):
                menu = row['MENU NAME']
                cost = row.get('Material Cost', 'N/A')
                self.debug_log(f"  {i + 1}. {menu}: {cost}")

            # ตรวจสอบข้อมูลซ้ำ
            duplicate_menus = df_new['MENU NAME'].duplicated().sum()
            if duplicate_menus > 0:
                self.debug_log(f"คำเตือน: พบเมนูซ้ำ {duplicate_menus} รายการ", "WARNING")
                # ลบข้อมูลซ้ำ (เก็บรายการแรก)
                df_new = df_new.drop_duplicates(subset=['MENU NAME'], keep='first')
                self.debug_log(f"ลบข้อมูลซ้ำแล้ว เหลือ: {len(df_new)} เมนู")

            # อัพเดทข้อมูล
            config_manager.update_single_base(base_type, df_new.copy())

            # อัพเดทสถานะใน UI
            self.update_base_status()
            self.update_all_base_status()

            success_msg = f"อัพเดท {base_type.upper()} Cost สำเร็จ!"
            detail_msg = f"""อัพเดทข้อมูล {base_type.upper()} Cost เรียบร้อยแล้ว:

📁 ไฟล์: {os.path.basename(file_path)}
📊 จำนวนเมนู: {len(df_new)} รายการ
💾 บันทึกลงระบบแล้ว
🏪 สถานะ: สามารถใช้งานได้ทันที"""

            self.debug_log(success_msg)
            messagebox.showinfo("สำเร็จ!", detail_msg)

        except Exception as e:
            raise e

    def save_current_base(self):
        """บันทึกข้อมูล Base Cost ปัจจุบันเป็นไฟล์"""
        try:
            result = config_manager.save_base_cost_file()
            if result:
                self.debug_log("บันทึกไฟล์ Base Cost.xlsx สำเร็จ")

                # อัพเดท status เป็นการแสดงว่าข้อมูลมาจากไฟล์แล้ว
                hashira_count = len(config_manager.base_files['hashira'])
                hamada_count = len(config_manager.base_files['hamada'])

                self.hashira_status_label.config(
                    text=f"🏯 Hashira Cost: พร้อมใช้งาน ({hashira_count} เมนู) [Saved]",
                    foreground=self.colors['success'])

                self.hamada_status_label.config(
                    text=f"🍜 Hamada Cost: พร้อมใช้งาน ({hamada_count} เมนู) [Saved]",
                    foreground=self.colors['success'])

                messagebox.showinfo("สำเร็จ!", f"""บันทึก Base Cost.xlsx เรียบร้อยแล้ว!

📁 ไฟล์: Base Cost.xlsx
🏯 Hashira Cost: {hashira_count} เมนู  
🍜 Hamada Cost: {hamada_count} เมนู
📍 ตำแหน่ง: {os.path.abspath('Base Cost.xlsx')}""")
            else:
                self.debug_log("ไม่สามารถบันทึกไฟล์ Base Cost.xlsx ได้", "ERROR")
                messagebox.showerror("Error", "ไม่สามารถบันทึกไฟล์ได้")

        except Exception as e:
            error_msg = f"Error บันทึกไฟล์ Base: {str(e)}"
            self.debug_log(error_msg, "ERROR")
            messagebox.showerror("Error", error_msg)

    def get_current_base_df(self):
        """ดึง DataFrame ของ base ที่เลือกอยู่"""
        selected = self.selected_base_type.get()
        return config_manager.base_files.get(selected)

    def check_base_file(self):
        """ตรวจสอบและแสดงข้อมูลไฟล์ base ที่เลือก"""
        selected = self.selected_base_type.get()
        df_base = self.get_current_base_df()

        if df_base is None:
            messagebox.showerror("Error", f"ไม่สามารถโหลด {selected.upper()} Cost ได้")
            return

        self.debug_log(f"=== ตรวจสอบ {selected.upper()} Cost ===")
        self.debug_log(f"จำนวนเมนู: {len(df_base)}")
        self.debug_log(f"คอลัมน์: {df_base.columns.tolist()}")

        # สถิติราคา
        try:
            if "Material Cost" in df_base.columns:
                # ลบช่องว่างออกจากชื่อคอลัมน์
                df_base.columns = df_base.columns.str.strip()
                cost_column = "Material Cost"

                cost_stats = pd.to_numeric(df_base[cost_column], errors='coerce').describe()
                self.debug_log("สถิติราคา Material Cost:")
                self.debug_log(f"  ต่ำสุด: {cost_stats['min']:.2f} บาท")
                self.debug_log(f"  สูงสุด: {cost_stats['max']:.2f} บาท")
                self.debug_log(f"  เฉลี่ย: {cost_stats['mean']:.2f} บาท")
                self.debug_log(f"  จำนวนรายการ: {int(cost_stats['count'])}")
        except Exception as e:
            self.debug_log(f"Error คำนวณสถิติ: {str(e)}", "ERROR")

        # ตัวอย่างเมนู
        self.debug_log(f"ตัวอย่างเมนู 10 รายการแรก ({selected.upper()}):")
        for i, (menu_name, row) in enumerate(df_base.head(10).iterrows()):
            try:
                material_cost = pd.to_numeric(row.get("Material Cost", 0), errors='coerce')
                if pd.isna(material_cost):
                    material_cost = 0
                self.debug_log(f"  {i + 1}. {menu_name}: {material_cost:.2f} บาท")
            except:
                self.debug_log(f"  {i + 1}. {menu_name}: ERROR")

    def calculate(self):
        """คำนวณต้นทุนแบบ Enhanced"""
        df_base = self.get_current_base_df()
        selected_base = self.selected_base_type.get()

        if df_base is None:
            messagebox.showerror("Error", f"ไฟล์ {selected_base.upper()} Cost ไม่สามารถโหลดได้")
            return

        if not self.import_file:
            messagebox.showwarning("Warning", "กรุณาเลือกไฟล์ Import ก่อน")
            return

        self.debug_log(f"=== เริ่มการคำนวณด้วย {selected_base.upper()} Cost ===")

        # เคลียร์ debug data เก่า
        self.debug_data = {
            'not_found_menus': [],
            'zero_qty_items': [],
            'invalid_qty_items': [],
            'nan_cost_items': [],
            'matched_menus': [],
            'not_sold_menus': [],
            'processing_summary': {}
        }

        try:
            # อ่านไฟล์ import
            df_import = pd.read_excel(self.import_file)
            self.debug_log(f"อ่านไฟล์ import สำเร็จ: {len(df_import)} แถว")

            # ลบช่องว่างจากชื่อคอลัมน์
            df_base.columns = df_base.columns.str.strip()
            df_import.columns = df_import.columns.str.strip()

        except Exception as e:
            error_msg = f"ไม่สามารถอ่านไฟล์ได้: {str(e)}"
            self.debug_log(error_msg, "ERROR")
            messagebox.showerror("Error", error_msg)
            return

        results = []
        matched_count = 0

        self.debug_log(f"เริ่มประมวลผล {len(df_import)} รายการ")

        for idx, row in df_import.iterrows():
            menu = row.get("MENU NAME")
            qty = row.get("Qty", 0)

            if self.debug_mode.get():
                self.debug_log(f"ประมวลผลแถว {idx + 1}: {menu} = {qty}")

            # ตรวจสอบ menu name
            if pd.isna(menu) or menu == "":
                self.debug_log(f"แถว {idx + 1}: ชื่อเมนูว่าง", "WARNING")
                continue

            # ตรวจสอบ quantity
            original_qty = qty
            try:
                qty = float(qty) if not pd.isna(qty) else 0
                if qty == 0:
                    self.debug_data['zero_qty_items'].append({
                        'row': idx + 1,
                        'menu': menu,
                        'original_qty': original_qty
                    })
                    if self.debug_mode.get():
                        self.debug_log(f"แถว {idx + 1}: Qty = 0 สำหรับ {menu}", "WARNING")
            except (ValueError, TypeError):
                self.debug_data['invalid_qty_items'].append({
                    'row': idx + 1,
                    'menu': menu,
                    'invalid_qty': original_qty
                })
                self.debug_log(f"แถว {idx + 1}: Qty ไม่ใช่ตัวเลข ({original_qty}) สำหรับ {menu}", "WARNING")
                qty = 0

            # ค้นหาใน base file
            if menu in df_base.index:
                try:
                    material_cost = df_base.at[menu, "Material Cost"]

                    # ตรวจสอบ material cost
                    if pd.isna(material_cost):
                        self.debug_data['nan_cost_items'].append({
                            'menu': menu,
                            'row': idx + 1
                        })
                        self.debug_log(f"แถว {idx + 1}: Material Cost เป็น NaN สำหรับ {menu}", "WARNING")
                        material_cost = 0
                    else:
                        material_cost = float(material_cost)

                    total_cost = qty * material_cost
                    results.append([menu, qty, material_cost, total_cost])

                    self.debug_data['matched_menus'].append({
                        'row': idx + 1,
                        'menu': menu,
                        'qty': qty,
                        'material_cost': material_cost,
                        'total_cost': total_cost
                    })

                    matched_count += 1

                    if self.debug_mode.get():
                        self.debug_log(f"  ✓ พบเมนู: {material_cost:.2f} x {qty} = {total_cost:.2f}")

                except Exception as e:
                    self.debug_log(f"Error คำนวณ {menu}: {str(e)}", "ERROR")
            else:
                self.debug_data['not_found_menus'].append({
                    'row': idx + 1,
                    'menu': menu,
                    'qty': qty
                })
                if self.debug_mode.get():
                    self.debug_log(f"  ✗ ไม่พบเมนู: {menu}", "WARNING")

        # หาเมนูที่ไม่ได้ขาย
        import_menus = set(df_import["MENU NAME"].dropna().tolist())
        base_menus = set(df_base.index.tolist())
        not_sold_menus = base_menus - import_menus

        self.debug_data['not_sold_menus'] = []
        for menu in not_sold_menus:
            try:
                cost = df_base.at[menu, "Material Cost"]
                if pd.isna(cost):
                    cost = 0
                else:
                    cost = float(cost)

                self.debug_data['not_sold_menus'].append({
                    'menu': menu,
                    'material_cost': cost
                })
            except:
                self.debug_data['not_sold_menus'].append({
                    'menu': menu,
                    'material_cost': 0
                })

        # บันทึกสรุปการประมวลผล
        self.debug_data['processing_summary'] = {
            'total_rows': len(df_import),
            'matched_count': matched_count,
            'not_found_count': len(self.debug_data['not_found_menus']),
            'zero_qty_count': len(self.debug_data['zero_qty_items']),
            'invalid_qty_count': len(self.debug_data['invalid_qty_items']),
            'nan_cost_count': len(self.debug_data['nan_cost_items']),
            'not_sold_count': len(self.debug_data['not_sold_menus']),
            'base_type': selected_base.upper(),
            'timestamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }

        # สรุปผลการประมวลผล
        self.debug_log(f"=== สรุปผลการประมวลผล ({selected_base.upper()}) ===")
        summary = self.debug_data['processing_summary']
        self.debug_log(f"รายการทั้งหมด: {summary['total_rows']}")
        self.debug_log(f"พบเมนูที่ตรงกัน: {summary['matched_count']}")
        self.debug_log(f"ไม่พบเมนู: {summary['not_found_count']}")
        self.debug_log(f"Qty = 0: {summary['zero_qty_count']}")
        self.debug_log(f"Qty ไม่ถูกต้อง: {summary['invalid_qty_count']}")
        self.debug_log(f"Material Cost NaN: {summary['nan_cost_count']}")
        self.debug_log(f"รายการที่ไม่ได้ขาย: {summary['not_sold_count']}")

        # สร้าง DataFrame ผลลัพธ์
        self.df_result = pd.DataFrame(results, columns=["MENU NAME", "Qty", "Material Cost", "Total Cost"])

        if self.df_result.empty:
            error_msg = f"ไม่มีชื่อเมนูที่ตรงกับ {selected_base.upper()} Cost"
            self.debug_log(error_msg, "WARNING")
            messagebox.showwarning("ไม่พบข้อมูล", error_msg)
            return

        # คำนวณ Grand Total
        grand_total = self.df_result["Total Cost"].sum()
        self.debug_log(f"Grand Total: {grand_total:.2f}")

        # เพิ่มแถว Grand Total
        grand_total_row = pd.DataFrame([["Grand Total", "", "", grand_total]],
                                       columns=["MENU NAME", "Qty", "Material Cost", "Total Cost"])
        self.df_result = pd.concat([self.df_result, grand_total_row], ignore_index=True)

        # แสดงในตาราง
        self._update_result_table()

        # อัพเดทสถิติ
        self._update_statistics(matched_count, summary['not_found_count'], grand_total, selected_base)

        success_msg = f"คำนวณเสร็จสิ้น - พบ {matched_count} รายการ (ใช้ {selected_base.upper()} Cost)"
        self.debug_log(success_msg)
        messagebox.showinfo("Success", success_msg)

    def _update_result_table(self):
        """อัพเดทตารางผลลัพธ์"""
        # เคลียร์ตารางเก่า
        for row in self.tree.get_children():
            self.tree.delete(row)

        # เพิ่มข้อมูลใหม่
        for _, r in self.df_result.iterrows():
            qty_display = int(r["Qty"]) if r["Qty"] != "" else ""
            cost_display = f"{r['Material Cost']:.2f}" if r["Material Cost"] != "" else ""
            total_display = f"{r['Total Cost']:.2f}" if r["Total Cost"] != "" else ""

            # ไฮไลท์แถว Grand Total
            tags = ("grand_total",) if r["MENU NAME"] == "Grand Total" else ()

            self.tree.insert("", tk.END,
                             values=(r["MENU NAME"], qty_display, cost_display, total_display),
                             tags=tags)

        # จัดรูปแบบแถว Grand Total
        self.tree.tag_configure("grand_total",
                                background="#3498db",
                                foreground="#ffffff",
                                font=("Segoe UI", 10, "bold"))

    def _update_statistics(self, matched, not_found, grand_total, base_type):
        """อัพเดทสถิติ"""
        stats_text = (f"✅ พบเมนู: {matched} | ❌ ไม่พบ: {not_found} | "
                      f"💰 รวม: {grand_total:,.2f} บาท | 🏪 Base: {base_type.upper()}")
        self.stats_label.config(text=stats_text)

    def export_excel(self):
        """บันทึก Excel แบบปกติ"""
        if self.df_result is None or self.df_result.empty:
            messagebox.showwarning("Warning", "กรุณาคำนวณก่อนบันทึก")
            return

        selected_base = self.selected_base_type.get()
        default_name = f"Cost_Calculation_{selected_base.upper()}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

        file_path = filedialog.asksaveasfilename(
            title="บันทึกผลลัพธ์",
            defaultextension=".xlsx",
            initialname=default_name,
            filetypes=[("Excel Files", "*.xlsx")]
        )
        if not file_path:
            return

        try:
            df_export = self.df_result.copy()
            df_export.to_excel(file_path, index=False)

            # จัดรูปแบบไฟล์ Excel
            wb = load_workbook(file_path)
            ws = wb.active

            # จัดรูปแบบ header
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

            # จัดรูปแบบตัวเลข
            for row in ws.iter_rows(min_row=2):
                if len(row) >= 3 and isinstance(row[2].value, (int, float)):
                    row[2].number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
                if len(row) >= 4 and isinstance(row[3].value, (int, float)):
                    row[3].number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1

            # ไฮไลท์แถว Grand Total
            for row in ws.iter_rows(min_row=2):
                if row[0].value == "Grand Total":
                    for cell in row:
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")

            wb.save(file_path)

            success_msg = f"บันทึกไฟล์สำเร็จ: {os.path.basename(file_path)}"
            self.debug_log(success_msg)
            messagebox.showinfo("Success", f"บันทึกผลลัพธ์แล้วที่:\n{file_path}")

        except Exception as e:
            error_msg = f"ไม่สามารถบันทึกไฟล์ได้: {str(e)}"
            self.debug_log(error_msg, "ERROR")
            messagebox.showerror("Error", error_msg)

    def export_excel_with_debug(self):
        """บันทึก Excel พร้อม Debug sheets และไฟล์เพิ่มเติม"""
        if self.df_result is None or self.df_result.empty:
            messagebox.showwarning("Warning", "กรุณาคำนวณก่อนบันทึก")
            return

        selected_base = self.selected_base_type.get()
        default_name = f"Cost_Calculation_Debug_{selected_base.upper()}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

        file_path = filedialog.asksaveasfilename(
            title="บันทึกผลลัพธ์พร้อม Debug",
            defaultextension=".xlsx",
            initialname=default_name,
            filetypes=[("Excel Files", "*.xlsx")]
        )
        if not file_path:
            return

        self.debug_log("=== เริ่มการบันทึกไฟล์พร้อม Debug และไฟล์เพิ่มเติม ===")

        try:
            wb = Workbook()

            # Sheet 1: ผลลัพธ์หลัก
            ws_main = wb.active
            ws_main.title = "ผลลัพธ์"

            df_export = self.df_result.copy()
            df_export["Material Cost"] = pd.to_numeric(df_export["Material Cost"], errors="coerce").round(2)
            df_export["Total Cost"] = pd.to_numeric(df_export["Total Cost"], errors="coerce").round(2)

            for r in dataframe_to_rows(df_export, index=False, header=True):
                ws_main.append(r)

            # จัดรูปแบบ Sheet หลัก
            self._format_main_sheet(ws_main)

            # Sheet 2: ข้อมูล Import
            if self.import_file:
                self._add_import_data_sheet(wb)

            # Sheet 3: ข้อมูล Template (ถ้ามี)
            if self.template_file:
                self._add_template_data_sheet(wb)

            # Sheet 4: ข้อมูล Base Cost ที่ใช้
            self._add_base_data_sheet(wb)

            # Sheet 5: เมนูที่ไม่พบ
            if self.debug_data['not_found_menus']:
                self._add_not_found_sheet(wb)

            # Sheet 6: สรุปการประมวลผล
            self._add_summary_sheet(wb)

            # ปรับความกว้างคอลัมน์
            self._adjust_column_width(wb)

            wb.save(file_path)

            success_msg = f"บันทึกไฟล์พร้อม Debug สำเร็จ: {os.path.basename(file_path)}"
            self.debug_log(success_msg)

            sheet_info = f"สร้าง {len(wb.worksheets)} sheets:\n"
            for ws in wb.worksheets:
                sheet_info += f"- {ws.title}\n"

            messagebox.showinfo("Success", f"บันทึกผลลัพธ์พร้อม Debug แล้ว:\n{file_path}\n\n{sheet_info}")

        except Exception as e:
            error_msg = f"ไม่สามารถบันทึกไฟล์ได้: {str(e)}"
            self.debug_log(error_msg, "ERROR")
            self.debug_log(traceback.format_exc(), "ERROR")
            messagebox.showerror("Error", error_msg)

    def _add_import_data_sheet(self, wb):
        """เพิ่ม sheet ข้อมูล Import"""
        try:
            df_import = pd.read_excel(self.import_file)
            ws_import = wb.create_sheet("ข้อมูล Import")

            # เพิ่ม header พิเศษ
            ws_import.append(["ข้อมูล Import ที่ใช้ในการคำนวณ"])
            ws_import.append([f"ไฟล์: {os.path.basename(self.import_file)}"])
            ws_import.append([f"วันที่: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
            ws_import.append([])  # บรรทัดว่าง

            # เพิ่มข้อมูล
            for r in dataframe_to_rows(df_import, index=False, header=True):
                ws_import.append(r)

            # จัดรูปแบบ
            ws_import[1][0].font = Font(bold=True, size=14)
            ws_import[1][0].fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")

            # จัดรูปแบบ header ของข้อมูล
            header_row = 5  # แถวที่เป็น header ของข้อมูล
            for cell in ws_import[header_row]:
                if cell.value:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

            self.debug_log("เพิ่ม Sheet ข้อมูล Import สำเร็จ")

        except Exception as e:
            self.debug_log(f"Error เพิ่ม Import sheet: {e}", "ERROR")

    def _add_template_data_sheet(self, wb):
        """เพิ่ม sheet ข้อมูล Template"""
        try:
            df_template = pd.read_excel(self.template_file)
            ws_template = wb.create_sheet("ข้อมูล Template")

            # เพิ่ม header พิเศษ
            ws_template.append(["ข้อมูล Template เพิ่มเติม"])
            ws_template.append([f"ไฟล์: {os.path.basename(self.template_file)}"])
            ws_template.append([f"วันที่: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
            ws_template.append([])

            # เพิ่มข้อมูล
            for r in dataframe_to_rows(df_template, index=False, header=True):
                ws_template.append(r)

            # จัดรูปแบบ
            ws_template[1][0].font = Font(bold=True, size=14)
            ws_template[1][0].fill = PatternFill(start_color="F0E6FF", end_color="F0E6FF", fill_type="solid")

            # จัดรูปแบบ header ของข้อมูล
            header_row = 5
            for cell in ws_template[header_row]:
                if cell.value:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

            self.debug_log("เพิ่ม Sheet ข้อมูล Template สำเร็จ")

        except Exception as e:
            self.debug_log(f"Error เพิ่ม Template sheet: {e}", "ERROR")

    def _add_base_data_sheet(self, wb):
        """เพิ่ม sheet ข้อมูล Base Cost ที่ใช้"""
        try:
            selected_base = self.selected_base_type.get()
            df_base = self.get_current_base_df()

            if df_base is None:
                return

            ws_base = wb.create_sheet(f"ข้อมูล {selected_base.upper()} Cost")

            # เพิ่ม header พิเศษ
            ws_base.append([f"ข้อมูล {selected_base.upper()} Cost ที่ใช้ในการคำนวณ"])
            ws_base.append([f"ประเภท: {selected_base.upper()} Cost"])
            ws_base.append([f"จำนวนเมนู: {len(df_base)}"])
            ws_base.append([f"วันที่: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
            ws_base.append([])

            # เพิ่มข้อมูล base (reset index เพื่อให้ MENU NAME กลับมาเป็นคอลัมน์)
            df_base_export = df_base.reset_index()
            for r in dataframe_to_rows(df_base_export, index=False, header=True):
                ws_base.append(r)

            # จัดรูปแบบ
            ws_base[1][0].font = Font(bold=True, size=14)
            ws_base[1][0].fill = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")

            # จัดรูปแบบ header ของข้อมูล
            header_row = 6
            for cell in ws_base[header_row]:
                if cell.value:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

            self.debug_log(f"เพิ่ม Sheet ข้อมูล {selected_base.upper()} Cost สำเร็จ")

        except Exception as e:
            self.debug_log(f"Error เพิ่ม Base sheet: {e}", "ERROR")

    def _add_not_found_sheet(self, wb):
        """เพิ่ม sheet เมนูที่ไม่พบ"""
        ws_not_found = wb.create_sheet("เมนูที่ไม่พบ")

        # Header
        ws_not_found.append(["เมนูที่ไม่พบในฐานข้อมูล"])
        ws_not_found.append([f"Base Cost: {self.debug_data['processing_summary']['base_type']}"])
        ws_not_found.append([])
        ws_not_found.append(["แถวที่", "ชื่อเมนู", "จำนวน"])

        # Header formatting
        for i in range(1, 5):
            if i <= 2:
                ws_not_found[i][0].font = Font(bold=True, size=12)
                ws_not_found[i][0].fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
            elif i == 4:
                for j, cell in enumerate(ws_not_found[i]):
                    if cell.value:
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

        # ข้อมูล
        for item in self.debug_data['not_found_menus']:
            ws_not_found.append([item['row'], item['menu'], item['qty']])

    def _add_summary_sheet(self, wb):
        """เพิ่ม sheet สรุปการประมวลผล"""
        ws_summary = wb.create_sheet("สรุปการประมวลผล")

        summary = self.debug_data['processing_summary']

        summary_data = [
            ["🔍 สรุปการประมวลผล - Enhanced Version", ""],
            ["", ""],
            ["📊 ข้อมูลทั่วไป", ""],
            ["Base Cost ที่ใช้", summary['base_type']],
            ["รายการทั้งหมด", summary['total_rows']],
            ["พบเมนูที่ตรงกัน", summary['matched_count']],
            ["ไม่พบเมนู", summary['not_found_count']],
            ["Qty = 0", summary['zero_qty_count']],
            ["Qty ไม่ถูกต้อง", summary['invalid_qty_count']],
            ["Material Cost NaN", summary['nan_cost_count']],
            ["รายการที่ไม่ได้ขาย", summary['not_sold_count']],
            ["", ""],
            ["📅 ข้อมูลการประมวลผล", ""],
            ["เวลาประมวลผล", summary['timestamp']],
            ["ไฟล์ Import", os.path.basename(self.import_file) if self.import_file else "N/A"],
            ["ไฟล์ Template", os.path.basename(self.template_file) if self.template_file else "ไม่มี"],
            ["", ""],
        ]

        # เพิ่มรายละเอียดเมนูที่พบ
        if self.debug_data['matched_menus']:
            summary_data.append(["✅ รายการเมนูที่พบ", ""])
            summary_data.append(["แถวที่", "ชื่อเมนู", "จำนวน", "ต้นทุน", "รวม"])
            for item in self.debug_data['matched_menus']:
                summary_data.append([
                    item['row'],
                    item['menu'],
                    item['qty'],
                    f"{item['material_cost']:.2f}",
                    f"{item['total_cost']:.2f}"
                ])
            summary_data.append(["", "", "", "", ""])

        # เพิ่มรายละเอียดเมนูที่ไม่พบ
        if self.debug_data['not_found_menus']:
            summary_data.append(["❌ รายการเมนูที่ไม่พบ", ""])
            summary_data.append(["แถวที่", "ชื่อเมนู", "จำนวน"])
            for item in self.debug_data['not_found_menus']:
                summary_data.append([item['row'], item['menu'], item['qty']])
            summary_data.append(["", "", ""])

        # เพิ่มรายการที่ไม่ได้ขาย
        if self.debug_data['not_sold_menus']:
            summary_data.append(["📋 รายการที่ไม่ได้ขาย (อยู่ใน Base แต่ไม่มีใน Import)", ""])
            summary_data.append(["ชื่อเมนู", "ต้นทุน"])
            for item in self.debug_data['not_sold_menus']:
                summary_data.append([item['menu'], f"{item['material_cost']:.2f}"])

        # เขียนข้อมูลลง sheet
        for row_data in summary_data:
            ws_summary.append(row_data)

        # จัดรูปแบบ summary sheet
        self._format_summary_sheet(ws_summary)

        self.debug_log("บันทึกไฟล์พร้อม Debug sheets สำเร็จ")

    def _format_main_sheet(self, ws):
        """จัดรูปแบบ sheet หลัก"""
        # Header formatting
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        # Number formatting
        for row in ws.iter_rows(min_row=2):
            if len(row) >= 3 and isinstance(row[2].value, (int, float)):
                row[2].number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
            if len(row) >= 4 and isinstance(row[3].value, (int, float)):
                row[3].number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1

        # Highlight Grand Total row
        for row in ws.iter_rows(min_row=2):
            if row[0].value == "Grand Total":
                for cell in row:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")

    def _format_summary_sheet(self, ws):
        """จัดรูปแบบ summary sheet"""
        # Main title
        ws[1][0].font = Font(bold=True, size=14)
        ws[1][0].fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")

        # Section headers
        for row_idx, row in enumerate(ws.iter_rows(), 1):
            if row[0].value in ["📊 ข้อมูลทั่วไป", "📅 ข้อมูลการประมวลผล",
                                "✅ รายการเมนูที่พบ", "❌ รายการเมนูที่ไม่พบ",
                                "📋 รายการที่ไม่ได้ขาย (อยู่ใน Base แต่ไม่มีใน Import)"]:
                row[0].font = Font(bold=True, size=12)
                row[0].fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")

    def _adjust_column_width(self, wb):
        """ปรับความกว้างคอลัมน์"""
        for ws in wb.worksheets:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 60)
                ws.column_dimensions[column_letter].width = adjusted_width

    def import_new_base(self):
        """Import ไฟล์ Base Cost ใหม่ (ทั้งหมด)"""
        file_path = filedialog.askopenfilename(
            title="เลือกไฟล์ Base Cost ใหม่ (ต้องมี Sheet ทั้ง Hashira และ Hamada)",
            filetypes=[("Excel Files", "*.xlsx *.xls")]
        )
        if not file_path:
            return

        try:
            # อ่านไฟล์ Excel ที่มีหลาย sheet
            excel_file = pd.ExcelFile(file_path)
            self.debug_log(f"อ่านไฟล์ Base ใหม่: {os.path.basename(file_path)}")
            self.debug_log(f"Sheets ที่มี: {excel_file.sheet_names}")

            # ตรวจสอบว่ามี sheet ที่ต้องการหรือไม่
            required_sheets = ["Hashira Cost", "Hamada Cost"]
            found_sheets = []

            for sheet in required_sheets:
                if sheet in excel_file.sheet_names:
                    found_sheets.append(sheet)
                    try:
                        df_sheet = pd.read_excel(file_path, sheet_name=sheet)

                        # ตรวจสอบคอลัมน์ที่จำเป็น
                        if "MENU NAME" not in df_sheet.columns:
                            self.debug_log(f"Sheet {sheet} ขาดคอลัมน์ MENU NAME", "ERROR")
                            continue

                        # ลบช่องว่างจากชื่อคอลัมน์
                        df_sheet.columns = df_sheet.columns.str.strip()
                        df_sheet = df_sheet.set_index("MENU NAME")

                        # อัพเดท config
                        if sheet == "Hashira Cost":
                            config_manager.base_files['hashira'] = df_sheet
                        elif sheet == "Hamada Cost":
                            config_manager.base_files['hamada'] = df_sheet

                        self.debug_log(f"โหลด {sheet} สำเร็จ: {len(df_sheet)} เมนู")

                    except Exception as e:
                        self.debug_log(f"Error โหลด sheet {sheet}: {e}", "ERROR")

            if found_sheets:
                # อัพเดท status
                self.update_base_status()
                self.update_all_base_status()

                success_msg = f"อัพเดทไฟล์ Base สำเร็จ\nSheets ที่โหลด: {', '.join(found_sheets)}"
                self.debug_log(success_msg)
                messagebox.showinfo("Success", success_msg)
            else:
                error_msg = "ไม่พบ Sheet ที่ต้องการ (Hashira Cost, Hamada Cost)"
                self.debug_log(error_msg, "ERROR")
                messagebox.showerror("Error", error_msg)

        except Exception as e:
            error_msg = f"ไม่สามารถ import ไฟล์ Base ใหม่ได้: {str(e)}"
            self.debug_log(error_msg, "ERROR")
            self.debug_log(traceback.format_exc(), "ERROR")
            messagebox.showerror("Error", error_msg)

    def export_base_file(self):
        """Export ไฟล์ Base Cost ปัจจุบัน"""
        if not any(config_manager.base_files.values()):
            messagebox.showerror("Error", "ไม่มีข้อมูล Base Cost ให้ export")
            return

        file_path = filedialog.asksaveasfilename(
            title="Export Base Cost Files",
            defaultextension=".xlsx",
            initialname=f"Base_Cost_Export_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
            filetypes=[("Excel Files", "*.xlsx")]
        )
        if not file_path:
            return

        try:
            wb = Workbook()
            # ลบ sheet default
            wb.remove(wb.active)

            sheets_created = 0

            # Export Hashira Cost
            if config_manager.base_files['hashira'] is not None:
                ws_hashira = wb.create_sheet("Hashira Cost")
                df_hashira = config_manager.base_files['hashira'].reset_index()

                for r in dataframe_to_rows(df_hashira, index=False, header=True):
                    ws_hashira.append(r)

                # จัดรูปแบบ header
                for cell in ws_hashira[1]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")

                sheets_created += 1
                self.debug_log(f"Export Hashira Cost: {len(df_hashira)} เมนู")

            # Export Hamada Cost
            if config_manager.base_files['hamada'] is not None:
                ws_hamada = wb.create_sheet("Hamada Cost")
                df_hamada = config_manager.base_files['hamada'].reset_index()

                for r in dataframe_to_rows(df_hamada, index=False, header=True):
                    ws_hamada.append(r)

                # จัดรูปแบบ header
                for cell in ws_hamada[1]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")

                sheets_created += 1
                self.debug_log(f"Export Hamada Cost: {len(df_hamada)} เมนู")

            # ปรับความกว้างคอลัมน์
            self._adjust_column_width(wb)

            wb.save(file_path)

            success_msg = f"Export ไฟล์ Base สำเร็จ: {os.path.basename(file_path)}"
            self.debug_log(success_msg)
            messagebox.showinfo("Success", f"Export ไฟล์ Base แล้ว ({sheets_created} sheets):\n{file_path}")

        except Exception as e:
            error_msg = f"ไม่สามารถ export ไฟล์ Base ได้: {str(e)}"
            self.debug_log(error_msg, "ERROR")
            messagebox.showerror("Error", error_msg)


def main():
    """ฟังก์ชันหลักพร้อม error handling"""
    try:
        root = tk.Tk()

        # Set modern window styling
        try:
            root.iconphoto(True, tk.PhotoImage(data=''))
        except:
            pass

        app = EnhancedCostCalculatorApp(root)
        logger.info("Starting Enhanced Cost Calculator with Real Data")

        # แสดง Welcome Message
        app.debug_log("=== Enhanced Cost Calculator เริ่มทำงาน ===")
        app.debug_log("✅ ข้อมูลจริง Hamada Cost (87 เมนู) และ Hashira Cost (68 เมนู) พร้อมใช้งาน")
        app.debug_log("💡 เลือกประเภท Base Cost แล้วเริ่มการคำนวณได้เลย")
        app.debug_log("🔧 สามารถอัพเดท Base Cost แยกกันได้ด้วยปุ่มสีแดง")
        app.debug_log("📋 ทั้งหมด 155 เมนูพร้อมใช้งาน")

        root.mainloop()

    except Exception as e:
        logger.error(f"Critical error in main: {str(e)}")
        logger.error(traceback.format_exc())
        messagebox.showerror("ผิดพลาดร้ายแรง", f"เกิดข้อผิดพลาดร้ายแรง:\n{str(e)}")


if __name__ == "__main__":
    main()