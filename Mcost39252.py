import pandas as pd
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import numbers, Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import logging
import traceback
import os

# ===== EMBEDDED BASE COST DATA =====
# ข้อมูล Hashira Cost (ฝังในโปรแกรม) - 68 เมนู (อัพเดทแล้ว)
HASHIRA_COST_DATA = {
    "Pork Cut Steak Set": 46.53,
    "Tonteki Set": 131.99,
    "Beef Cut Steak Set (Medium Rare / Well done)": 107.87,
    "Buta Teriyaki Set": 41.77,
    "Tonkatsu Set": 58.39,
    "Tonkatsu Tamagotoji Set": 60.91,
    "Beef Hamburg Set": 76.4,
    "Salmon Set (Teriyaki / Shioyaki)": 68.02,
    "Kaisen Teppan Yaki Set": 121.23,
    "Topping Fried Egg": 3.5,
    "Gokuatsu Tonkatsu Set": 131.77,
    "Gokuatsu Katsu Don Set": 133,
    "Gokuatsu Tonkatsu Tamagotoji Set": 131.46,
    "Tonkotsu Chashumen": 64.23,
    "Tonkotsu Niku-Niku Chashumen": 77.98,
    "Tonkotsu Spicy Ramen": 54.05,
    "Tonkotsu Hashira Ramen": 74.33,
    "Miso Chashumen": 56.52,
    "Miso Hokkaido Ramen": 70.08,
    "Shoyu Chashumen": 63.1,
    "Shoyu Tokyo Ramen": 64.44,
    "Tom Yum Tonkotsu Kaisen Ramen": 55.55,
    "Tom Yum Kung Ramen": 116.08,
    "Tori Paitan Karaage Ramen": 42.88,
    "Tori Paitan Kaisen Ramen": 50.08,
    "Tonkatsu": 37.4,
    "Hotate": 33.33,
    "Ebi Tempura": 26,
    "Chashu": 27.3,
    "Spicy Nikumiso": 7.97,
    "Karashinegi": 10.54,
    "Negi": 6.43,
    "Naruto": 12.92,
    "Ajitama": 3.84,
    "Menma": 9.6,
    "Nori": 5.4,
    "Corn": 6,
    "Butter": 1.31,
    "Boiled Cabbage": 2.73,
    "Boiled Bean Sprout": 1.98,
    "Karashi Takana": 6.45,
    "Kikurage": 3.41,
    "Kanikama": 7.92,
    "Kaedama": 7.5,
    "GG Sauce Katsudon": 67.94,
    "GG Buta Teriyaki Don": 45.93,
    "GG Tonkatsu Bento": 79.95,
    "GG Yakisoba": 45.61,
    "GG Yakisoba with Sunny Side Egg": 49.11,
    "GG Pork Okonomiyaki": 52.4,
    "Coca-Cola Original": 12.54,
    "Coca-Cola Zero Sugar": 12.54,
    "Sprite": 12.54,
    "Soda": 7.59,
    "Mineral Water": 7,
    "HBD Sparkling Water Lemon (No Sugar, No Calories)": 17,
    "HBD Sparkling Water Honey Yuzu (No Sugar, No Calories)": 17,
    "HBD Sparkling Water Peach (No Sugar, No Calories)": 17,
    "HBD Sparkling Water Kyoho (No Sugar, No Calories)": 17,
    "Asahi Beer": 41.92,
    "Heineken Beer": 38.43,
    "Singha Beer": 31.71,
    "Drinking Water": 3.74,
    "Melon Ball": 25,
    "Vanilla Monaka": 27,
    "Azuki Bar": 27,
    "Rice": 5.85,
    "Miso Soup": 3.04,
}

# ข้อมูล Hamada Cost (ฝังในโปรแกรม) - 87 เมนู (ไม่เปลี่ยน)
HAMADA_COST_DATA = {
    "Ebi Tempura Set": 67.53,
    "Salmon Tempura Set": 51.84,
    "Ebi & Salmon Tempura Set": 70.42,
    "Ebi Tempura Don Set": 69.33,
    "Salmon Tempura Don Set": 53.64,
    "Ebi & Salmon Tempura Don Set": 69.29,
    "Ebi Tempura Udon": 56.03,
    "Salmon Tempura Udon": 54.34,
    "Ebi & Salmon Tempura Udon": 73.57,
    "Yakiudon Buta": 37.21,
    "Yakiudon Kaisen": 48.67,
    "Ebi Fry Curry Rice": 62.31,
    "Tonkatsu Curry Rice": 69.17,
    "Salmonkatsu Curry Rice": 60.62,
    "Ebi Tempura（1pc）": 16.16,
    "Salmon Tempura（1pc）": 10.92,
    "Takoyaki": 33.71,
    "Sushi Yorokobi": 63.56,
    "Sushi Tokujo": 203.12,
    "Maguro Zanmai": 89.93,
    "Salmon Zanmai": 98.14,
    "Hamachi Zanmai": 168.1,
    "Duo Sushi": 108.73,
    "Salmon Mountain": 200.7,
    "Salmon Sushi & Ikura Gunkan": 193.91,
    "California Roll": 50.76,
    "Salmon Roll": 87.59,
    "Unagi Roll": 104.1,
    "Hokkaido Hotate Roll": 133.97,
    "Salmon Chirashi Sushi": 48.53,
    "Salmon Tobiko Chirashi Sushi": 66.95,
    "Barachirashi Sushi": 77.77,
    "Salmon Don": 79.45,
    "Salmon Ikura Don": 108.68,
    "Maguro Don": 73.91,
    "Negitoro Ikura Don": 111.38,
    "Kaisen Don": 243.48,
    "Unagi Don": 246.42,
    "Salmon Sashimi": 42.71,
    "Maguro Sashimi": 39.59,
    "Hamachi Sashimi": 60.89,
    "Hotate Sashimi": 52.93,
    "Salmon Sushi (2pcs)": 29.95,
    "Maguro Sushi (2pcs)": 27.87,
    "Hamachi Sushi (2pcs)": 42.07,
    "Tamago Sushi (2pcs)": 9.63,
    "Ebi Sushi (2pcs)": 19.94,
    "Ika Sushi (2pcs)": 19.87,
    "Hokki Sushi (2pcs)": 19.96,
    "Hotate Sushi (2pcs)": 51.71,
    "Unagi Sushi (2pcs)": 56.12,
    "Anago Sushi (2pcs)": 112.92,
    "Negitoro Gunkan (2pcs)": 21.27,
    "Tobiko Gunkan (2pcs)": 18.32,
    "Ikura Gunkan (2pcs)": 68.24,
    "GG California Roll": 51.43,
    "GG Salmon Sushi": 93.14,
    "GG Tamago & Ebi Sushi": 56.38,
    "GG Ebi & Tobiko Don": 68.9,
    "GG Kanikama & Tobiko Don": 48.9,
    "GG Triple Mix Don": 58.8,
    "GG Ebiten Don": 63.34,
    "GG Salmon & Ebiten Don": 74.35,
    "Coca-Cola Original": 12.54,
    "Coke-Cola Zero Sugar": 12.54,
    "Sprite": 12.54,
    "Soda": 7.59,
    "Mineral Water": 7,
    "HBD Sparkling Water Lemon (No Sugar, No Calories)": 17,
    "HBD Sparkling Water Honey Yuzu (No Sugar, No Calories)": 17,
    "HBD Sparkling Water Peach (No Sugar, No Calories)": 17,
    "HBD Sparkling Water Kyoho (No Sugar, No Calories)": 17,
    "Asahi Beer": 41.92,
    "Heineken Beer": 38.43,
    "Singha Beer": 31.71,
    "Drinking Water": 3.74,
    "Matcha Soft Serve": 15.09,
    "Yogurt Soft Serve": 17.72,
    "Two Tone Soft Serve": 16.41,
    "Matcha Japanese": 22.6,
    "Matcha Marshmallow": 30.78,
    "Mix Berry Yogurt": 34.94,
    "Brownie Yogurt": 41.79,
    "Two Tone Brownie": 37.45,
    "Saikyo Matcha": 31.54,
    "Rice": 5.85,
    "Miso Soup": 3.58,
}


# ===== Enhanced Configuration Manager =====
class EmbeddedConfigManager:
    def __init__(self):
        self.base_data = {
            'hashira': self._create_hashira_dataframe(),
            'hamada': self._create_hamada_dataframe()
        }
        self.config_file = "mapcost_config.ini"

    def _create_hashira_dataframe(self):
        """สร้าง DataFrame สำหรับ Hashira Cost จากข้อมูลที่ฝังไว้"""
        if not HASHIRA_COST_DATA:
            return None

        df = pd.DataFrame(list(HASHIRA_COST_DATA.items()),
                          columns=['MENU NAME', 'Material Cost'])
        return df.set_index('MENU NAME')

    def _create_hamada_dataframe(self):
        """สร้าง DataFrame สำหรับ Hamada Cost จากข้อมูลที่ฝังไว้"""
        if not HAMADA_COST_DATA:
            return None

        df = pd.DataFrame(list(HAMADA_COST_DATA.items()),
                          columns=['MENU NAME', 'Material Cost'])
        return df.set_index('MENU NAME')

    def get_base_data(self, base_type):
        """ดึงข้อมูล base ตามประเภท"""
        return self.base_data.get(base_type)

    def update_base_data_from_excel(self, file_path):
        """อัพเดทข้อมูล base จากไฟล์ Excel ภายนอก"""
        try:
            excel_file = pd.ExcelFile(file_path)

            # อ่าน Hashira Cost
            if "Hashira Cost" in excel_file.sheet_names:
                df_hashira = pd.read_excel(file_path, sheet_name="Hashira Cost")
                if "MENU NAME" in df_hashira.columns:
                    df_hashira.columns = df_hashira.columns.str.strip()
                    self.base_data['hashira'] = df_hashira.set_index("MENU NAME")

            # อ่าน Hamada Cost
            if "Hamada Cost" in excel_file.sheet_names:
                df_hamada = pd.read_excel(file_path, sheet_name="Hamada Cost")
                if "MENU NAME" in df_hamada.columns:
                    df_hamada.columns = df_hamada.columns.str.strip()
                    self.base_data['hamada'] = df_hamada.set_index("MENU NAME")

            return True
        except Exception as e:
            print(f"Error updating base data: {e}")
            return False

    def get_statistics(self):
        """ดึงสถิติของข้อมูล base"""
        stats = {}
        for base_type, df in self.base_data.items():
            if df is not None:
                stats[base_type] = {
                    'total_menus': len(df),
                    'avg_cost': df['Material Cost'].mean(),
                    'min_cost': df['Material Cost'].min(),
                    'max_cost': df['Material Cost'].max()
                }
        return stats


# ===== Modern GUI Styling =====
def configure_modern_style():
    """กำหนดธีมสีสวยงามและทันสมัย"""
    style = ttk.Style()

    try:
        style.theme_use('clam')
    except:
        style.theme_use('default')

    colors = {
        'bg_primary': '#2c3e50',
        'bg_secondary': '#34495e',
        'accent': '#3498db',
        'success': '#27ae60',
        'warning': '#f39c12',
        'danger': '#e74c3c',
        'light': '#ecf0f1',
        'white': '#ffffff',
        'text_dark': '#2c3e50',
        'text_light': '#ffffff'
    }

    # Configure styles
    style.configure('Title.TLabel', font=('Segoe UI', 16, 'bold'), foreground=colors['text_dark'])
    style.configure('Heading.TLabel', font=('Segoe UI', 12, 'bold'), foreground=colors['accent'])
    style.configure('Status.TLabel', font=('Segoe UI', 10), padding=(5, 2))

    style.configure('Accent.TButton', font=('Segoe UI', 9, 'bold'), padding=(10, 5))
    style.configure('Success.TButton', font=('Segoe UI', 9, 'bold'), padding=(10, 5))
    style.configure('Warning.TButton', font=('Segoe UI', 9, 'bold'), padding=(10, 5))

    style.configure('Card.TFrame', relief='solid', borderwidth=1, padding=10)
    style.configure('Modern.TNotebook', tabposition='n')
    style.configure('Modern.TNotebook.Tab', font=('Segoe UI', 10, 'bold'), padding=(20, 8))

    return colors


# ===== Logging Setup =====
logging.basicConfig(
    level=logging.DEBUG,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('hamada_calculator_embedded.log', encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# ===== Language System =====
LANGUAGES = {
    'th': {
        'app_title': 'Hamada & Hashira Cost Calculator - Embedded Edition',
        'select_import_file': 'เลือกไฟล์ Import',
        'calculate': 'คำนวณ',
        'save_excel': 'บันทึก Excel',
        'save_excel_debug': 'บันทึก Excel + Debug',
        'check_base_file': 'ตรวจสอบข้อมูล Base',
        'export_base_file': 'Export ข้อมูล Base',
        'import_new_base': 'Import Base จากไฟล์',
        'import_template_file': 'Import Template เพิ่มเติม',
        'clear_debug': 'เคลียร์ Debug',
        'debug_mode': 'เปิด Debug Mode',
        'select_base_type': 'เลือกประเภท Base Cost',
        'hashira_cost': 'Hashira Cost',
        'hamada_cost': 'Hamada Cost',
        'no_file_selected': 'ยังไม่ได้เลือกไฟล์',
        'file_selected': 'ไฟล์',
        'base_data_ready': 'ข้อมูล Base: พร้อมใช้งาน (ฝังในโปรแกรม)',
        'results_tab': 'ผลลัพธ์',
        'debug_tab': 'Debug Log',
        'statistics': 'สถิติ',
        'no_data': 'ยังไม่มีข้อมูล',
        'menu_name': 'MENU NAME',
        'qty': 'Qty',
        'material_cost': 'Material Cost',
        'total_cost': 'Total Cost',
        'grand_total': 'Grand Total',
        'matched_menus': 'พบเมนู',
        'not_found_menus': 'ไม่พบ',
        'total_amount': 'รวม',
        'baht': 'บาท',
        'error': 'ผิดพลาด',
        'warning': 'คำเตือน',
        'success': 'สำเร็จ',
        'please_select_file': 'กรุณาเลือกไฟล์ Import ก่อน',
        'calculation_complete': 'คำนวณเสร็จสิ้น - พบ',
        'items': 'รายการ',
        'template_file': 'ไฟล์ Template',
        'embedded_data': 'ข้อมูลฝังในโปรแกรม'
    },
    'en': {
        'app_title': 'Hamada & Hashira Cost Calculator - Embedded Edition',
        'select_import_file': 'Select Import File',
        'calculate': 'Calculate',
        'save_excel': 'Save Excel',
        'save_excel_debug': 'Save Excel + Debug',
        'check_base_file': 'Check Base Data',
        'export_base_file': 'Export Base Data',
        'import_new_base': 'Import Base from File',
        'import_template_file': 'Import Additional Template',
        'clear_debug': 'Clear Debug',
        'debug_mode': 'Enable Debug Mode',
        'select_base_type': 'Select Base Cost Type',
        'hashira_cost': 'Hashira Cost',
        'hamada_cost': 'Hamada Cost',
        'no_file_selected': 'No file selected',
        'file_selected': 'File',
        'base_data_ready': 'Base Data: Ready (Embedded)',
        'results_tab': 'Results',
        'debug_tab': 'Debug Log',
        'statistics': 'Statistics',
        'no_data': 'No data available',
        'menu_name': 'MENU NAME',
        'qty': 'Qty',
        'material_cost': 'Material Cost',
        'total_cost': 'Total Cost',
        'grand_total': 'Grand Total',
        'matched_menus': 'Found',
        'not_found_menus': 'Not Found',
        'total_amount': 'Total',
        'baht': 'Baht',
        'error': 'Error',
        'warning': 'Warning',
        'success': 'Success',
        'please_select_file': 'Please select import file first',
        'calculation_complete': 'Calculation complete - found',
        'items': 'items',
        'template_file': 'Template File',
        'embedded_data': 'Embedded Data'
    },
    'jp': {
        'app_title': 'Hamada & Hashira コスト計算機 - 組み込み版',
        'select_import_file': 'インポートファイル選択',
        'calculate': '計算',
        'save_excel': 'Excel保存',
        'save_excel_debug': 'Excel + デバッグ保存',
        'check_base_file': 'ベースデータ確認',
        'export_base_file': 'ベースデータ出力',
        'import_new_base': 'ファイルからベース取込',
        'import_template_file': '追加テンプレート取込',
        'clear_debug': 'デバッグクリア',
        'debug_mode': 'デバッグモード有効',
        'select_base_type': 'ベースコストタイプ選択',
        'hashira_cost': 'Hashira Cost',
        'hamada_cost': 'Hamada Cost',
        'no_file_selected': 'ファイル未選択',
        'file_selected': 'ファイル',
        'base_data_ready': 'ベースデータ: 準備完了（組込済）',
        'results_tab': '結果',
        'debug_tab': 'デバッグログ',
        'statistics': '統計',
        'no_data': 'データなし',
        'menu_name': 'メニュー名',
        'qty': '数量',
        'material_cost': '材料費',
        'total_cost': '合計金額',
        'grand_total': '総合計',
        'matched_menus': '一致',
        'not_found_menus': '未発見',
        'total_amount': '合計',
        'baht': 'バーツ',
        'error': 'エラー',
        'warning': '警告',
        'success': '成功',
        'please_select_file': 'インポートファイルを選択してください',
        'calculation_complete': '計算完了 - 発見',
        'items': '項目',
        'template_file': 'テンプレートファイル',
        'embedded_data': '組み込みデータ'
    }
}

# Initialize embedded configuration
config_manager = EmbeddedConfigManager()


class EmbeddedCostCalculatorApp:
    def __init__(self, root):
        self.root = root
        self.current_language = 'th'
        self.colors = configure_modern_style()

        self.root.title(self.t('app_title'))
        self.root.geometry("1400x900")
        self.root.configure(bg=self.colors['light'])
        self.root.minsize(1200, 800)

        # ตัวแปรสำหรับไฟล์
        self.import_file = None
        self.template_file = None
        self.selected_base_type = tk.StringVar(value='hamada')  # default เป็น hamada
        self.df_result = None
        self.debug_mode = tk.BooleanVar(value=False)

        # ข้อมูล debug
        self.debug_data = {
            'not_found_menus': [],
            'zero_qty_items': [],
            'invalid_qty_items': [],
            'nan_cost_items': [],
            'matched_menus': [],
            'not_sold_menus': [],
            'processing_summary': {}
        }

        self._build_ui()
        self._show_embedded_info()
        logger.info("Embedded Cost Calculator initialized")

    def t(self, key):
        """แปลข้อความตามภาษาที่เลือก"""
        return LANGUAGES.get(self.current_language, {}).get(key, key)

    def _show_embedded_info(self):
        """แสดงข้อมูลเกี่ยวกับข้อมูลที่ฝังไว้"""
        stats = config_manager.get_statistics()
        self.debug_log("🚀 Hamada & Hashira Cost Calculator - Embedded Edition (Updated)")
        self.debug_log("=" * 60)
        self.debug_log("📊 ข้อมูล Base Cost ที่ฝังในโปรแกรม (อัพเดทแล้ว):")

        for base_type, stat in stats.items():
            self.debug_log(f"🏪 {base_type.upper()} Cost:")
            self.debug_log(f"   📋 จำนวนเมนู: {stat['total_menus']}")
            self.debug_log(f"   💰 ราคาเฉลี่ย: {stat['avg_cost']:.2f} บาท")
            self.debug_log(f"   📉 ราคาต่ำสุด: {stat['min_cost']:.2f} บาท")
            self.debug_log(f"   📈 ราคาสูงสุด: {stat['max_cost']:.2f} บาท")
            self.debug_log("")

    def _build_ui(self):
        """สร้าง UI แบบ Enhanced"""
        # Main container
        main_frame = ttk.Frame(self.root, style='Card.TFrame')
        main_frame.pack(fill=tk.BOTH, expand=True, padx=15, pady=15)

        # Enhanced Header
        header_frame = ttk.Frame(main_frame)
        header_frame.pack(fill=tk.X, pady=(0, 20))

        # Title with updated indicator
        title_label = ttk.Label(header_frame,
                                text="🍱 Hamada & Hashira Cost Calculator - Updated Edition",
                                style='Title.TLabel')
        title_label.pack(pady=(0, 5))

        subtitle = ttk.Label(header_frame,
                             text="💾 ข้อมูล Hashira Cost อัพเดทแล้ว | ข้อมูลฝังในโปรแกรม",
                             font=('Segoe UI', 11), foreground=self.colors['success'])
        subtitle.pack()

        # Language Selection
        lang_frame = ttk.LabelFrame(main_frame, text="🌐 Language / ภาษา / 言語", padding=10)
        lang_frame.pack(fill=tk.X, pady=(0, 15))

        self.lang_var = tk.StringVar(value='th')
        lang_buttons_frame = ttk.Frame(lang_frame)
        lang_buttons_frame.pack()

        lang_buttons = [
            ('🇹🇭 ไทย', 'th'),
            ('🇺🇸 English', 'en'),
            ('🇯🇵 日本語', 'jp')
        ]

        for text, code in lang_buttons:
            ttk.Radiobutton(lang_buttons_frame, text=text,
                            variable=self.lang_var, value=code,
                            command=lambda: self.change_language(self.lang_var.get())).pack(side=tk.LEFT, padx=10)

        # ===== Base Type Selection Card =====
        base_selection_frame = ttk.LabelFrame(main_frame, text="🏪 เลือกประเภท Base Cost",
                                              padding=15, style='Card.TFrame')
        base_selection_frame.pack(fill=tk.X, pady=(0, 15))

        # Base type info และ selection
        base_info_frame = ttk.Frame(base_selection_frame)
        base_info_frame.pack(fill=tk.X)

        # แสดงข้อมูลสถิติ base
        stats = config_manager.get_statistics()

        # Hashira info
        hashira_frame = ttk.Frame(base_info_frame)
        hashira_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 10))

        hashira_radio = ttk.Radiobutton(hashira_frame, text="🏯 Hashira Cost (อัพเดทแล้ว)",
                                        variable=self.selected_base_type, value='hashira',
                                        command=self.on_base_type_change)
        hashira_radio.pack(anchor="w")

        if 'hashira' in stats:
            hashira_info = f"📋 {stats['hashira']['total_menus']} เมนู | 💰 เฉลี่ย {stats['hashira']['avg_cost']:.0f} บาท"
            ttk.Label(hashira_frame, text=hashira_info,
                      font=('Segoe UI', 9), foreground=self.colors['bg_secondary']).pack(anchor="w", padx=(20, 0))

        # Hamada info
        hamada_frame = ttk.Frame(base_info_frame)
        hamada_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(10, 0))

        hamada_radio = ttk.Radiobutton(hamada_frame, text="🍜 Hamada Cost",
                                       variable=self.selected_base_type, value='hamada',
                                       command=self.on_base_type_change)
        hamada_radio.pack(anchor="w")

        if 'hamada' in stats:
            hamada_info = f"📋 {stats['hamada']['total_menus']} เมนู | 💰 เฉลี่ย {stats['hamada']['avg_cost']:.0f} บาท"
            ttk.Label(hamada_frame, text=hamada_info,
                      font=('Segoe UI', 9), foreground=self.colors['bg_secondary']).pack(anchor="w", padx=(20, 0))

        # Base status
        status_frame = ttk.Frame(base_selection_frame)
        status_frame.pack(fill=tk.X, pady=(10, 0))

        self.base_status_label = ttk.Label(status_frame, text="", style='Status.TLabel')
        self.base_status_label.pack()
        self.update_base_status()

        # ===== File Management Section =====
        file_frame = ttk.LabelFrame(main_frame, text="📁 File Management",
                                    padding=15, style='Card.TFrame')
        file_frame.pack(fill=tk.X, pady=(0, 15))

        # File selection buttons - Row 1
        file_btn_frame1 = ttk.Frame(file_frame)
        file_btn_frame1.pack(fill=tk.X, pady=(0, 8))

        ttk.Button(file_btn_frame1, text="📁 เลือกไฟล์ Import",
                   command=self.load_import_file, style='Accent.TButton').pack(
            side=tk.LEFT, padx=5, fill=tk.X, expand=True)

        ttk.Button(file_btn_frame1, text="📄 Import Template เพิ่มเติม",
                   command=self.load_template_file, style='Warning.TButton').pack(
            side=tk.LEFT, padx=5, fill=tk.X, expand=True)

        ttk.Button(file_btn_frame1, text="🔍 ตรวจสอบข้อมูล Base",
                   command=self.check_base_data, style='Accent.TButton').pack(
            side=tk.LEFT, padx=5, fill=tk.X, expand=True)

        # File selection buttons - Row 2
        file_btn_frame2 = ttk.Frame(file_frame)
        file_btn_frame2.pack(fill=tk.X)

        ttk.Button(file_btn_frame2, text="📥 Import Base จากไฟล์",
                   command=self.import_external_base, style='Warning.TButton').pack(
            side=tk.LEFT, padx=5, fill=tk.X, expand=True)

        ttk.Button(file_btn_frame2, text="📤 Export ข้อมูล Base",
                   command=self.export_base_data, style='Accent.TButton').pack(
            side=tk.LEFT, padx=5, fill=tk.X, expand=True)

        # File status display
        status_frame = ttk.Frame(file_frame)
        status_frame.pack(fill=tk.X, pady=(10, 0))

        self.import_status_label = ttk.Label(status_frame, text="📂 Import: ยังไม่ได้เลือกไฟล์",
                                             style='Status.TLabel', foreground=self.colors['danger'])
        self.import_status_label.pack(anchor="w", pady=2)

        self.template_status_label = ttk.Label(status_frame, text="📄 Template: ยังไม่ได้เลือกไฟล์",
                                               style='Status.TLabel', foreground=self.colors['warning'])
        self.template_status_label.pack(anchor="w", pady=2)

        # ===== Action Buttons =====
        action_frame = ttk.LabelFrame(main_frame, text="🚀 Actions",
                                      padding=15, style='Card.TFrame')
        action_frame.pack(fill=tk.X, pady=(0, 15))

        # Debug mode checkbox
        debug_frame = ttk.Frame(action_frame)
        debug_frame.pack(fill=tk.X, pady=(0, 10))

        self.debug_checkbox = ttk.Checkbutton(debug_frame, text="🔍 เปิด Debug Mode (แสดงรายละเอียด)",
                                              variable=self.debug_mode)
        self.debug_checkbox.pack(side=tk.LEFT)

        ttk.Button(debug_frame, text="🗑️ เคลียร์ Debug",
                   command=self.clear_debug, style='Warning.TButton').pack(side=tk.RIGHT)

        # Action buttons
        btn_frame = ttk.Frame(action_frame)
        btn_frame.pack(fill=tk.X)

        ttk.Button(btn_frame, text="⚡ คำนวณ", command=self.calculate,
                   style='Success.TButton').pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)

        ttk.Button(btn_frame, text="💾 บันทึก Excel", command=self.export_excel,
                   style='Success.TButton').pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)

        ttk.Button(btn_frame, text="🔍💾 บันทึก Excel + Debug", command=self.export_excel_with_debug,
                   style='Warning.TButton').pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)

        # ===== Content Area =====
        content_frame = ttk.Frame(main_frame, style='Card.TFrame')
        content_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 15))

        # Notebook for tabs
        self.notebook = ttk.Notebook(content_frame, style='Modern.TNotebook')
        self.notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # ===== Results Tab =====
        result_frame = ttk.Frame(self.notebook)
        self.notebook.add(result_frame, text=f"📋 {self.t('results_tab')}")

        table_frame = ttk.Frame(result_frame, padding=10)
        table_frame.pack(fill=tk.BOTH, expand=True)

        table_header = ttk.Label(table_frame, text="💰 ผลการคำนวณต้นทุน",
                                 style='Heading.TLabel')
        table_header.pack(pady=(0, 10))

        # Results table
        self.tree = ttk.Treeview(table_frame, columns=("menu", "qty", "cost", "total"),
                                 show="headings", height=16)

        self.tree.heading("menu", text=f"🍽️ {self.t('menu_name')}")
        self.tree.heading("qty", text=f"📊 {self.t('qty')}")
        self.tree.heading("cost", text=f"💵 {self.t('material_cost')}")
        self.tree.heading("total", text=f"💰 {self.t('total_cost')}")

        self.tree.column("menu", width=400, anchor="w")
        self.tree.column("qty", width=100, anchor="center")
        self.tree.column("cost", width=150, anchor="e")
        self.tree.column("total", width=150, anchor="e")

        tree_scroll = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=tree_scroll.set)

        tree_scroll.pack(side=tk.RIGHT, fill=tk.Y, padx=(5, 0))
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # ===== Debug Tab =====
        debug_frame = ttk.Frame(self.notebook)
        self.notebook.add(debug_frame, text=f"🔍 {self.t('debug_tab')}")

        debug_header_frame = ttk.Frame(debug_frame, padding=10)
        debug_header_frame.pack(fill=tk.X)

        debug_title = ttk.Label(debug_header_frame, text="🔧 Enhanced Debug Console",
                                style='Heading.TLabel')
        debug_title.pack()

        debug_text_frame = ttk.Frame(debug_frame, padding=(10, 0, 10, 10))
        debug_text_frame.pack(fill=tk.BOTH, expand=True)

        self.debug_text = scrolledtext.ScrolledText(debug_text_frame, wrap=tk.WORD, height=18,
                                                    font=('Consolas', 10),
                                                    bg='#1e1e1e', fg='#ffffff',
                                                    insertbackground='#ffffff',
                                                    selectbackground='#3498db',
                                                    relief='flat', borderwidth=0)
        self.debug_text.pack(fill=tk.BOTH, expand=True)

        # ===== Statistics Section =====
        self.stats_frame = ttk.LabelFrame(main_frame, text="📈 Statistics & Info",
                                          padding=15, style='Card.TFrame')
        self.stats_frame.pack(fill=tk.X)

        self.stats_label = ttk.Label(self.stats_frame, text=self.t('no_data'),
                                     style='Status.TLabel',
                                     font=('Segoe UI', 11, 'bold'))
        self.stats_label.pack()

        self._configure_treeview_style()

    def change_language(self, lang_code):
        """เปลี่ยนภาษา"""
        self.current_language = lang_code
        self.debug_log(f"🌍 เปลี่ยนภาษาเป็น: {lang_code}")
        # รีเฟรช UI text ตามภาษาใหม่
        self.root.title(self.t('app_title'))

    def _configure_treeview_style(self):
        """กำหนดสไตล์ modern สำหรับ treeview"""
        style = ttk.Style()

        style.configure('Modern.Treeview',
                        background='#ffffff',
                        foreground='#2c3e50',
                        fieldbackground='#ffffff',
                        font=('Segoe UI', 10))

        style.configure('Modern.Treeview.Heading',
                        background='#3498db',
                        foreground='#ffffff',
                        font=('Segoe UI', 10, 'bold'),
                        relief='flat')

        style.map('Modern.Treeview',
                  background=[('selected', '#3498db')],
                  foreground=[('selected', '#ffffff')])

        self.tree.configure(style='Modern.Treeview')

    def on_base_type_change(self):
        """เมื่อเปลี่ยน base type"""
        selected = self.selected_base_type.get()
        self.debug_log(f"🔄 เปลี่ยน Base Type เป็น: {selected.upper()}")
        self.update_base_status()

    def update_base_status(self):
        """อัพเดทสถานะข้อมูล base"""
        selected = self.selected_base_type.get()
        df_base = config_manager.get_base_data(selected)

        if df_base is not None and not df_base.empty:
            menu_count = len(df_base)
            avg_cost = df_base['Material Cost'].mean()

            if selected == 'hashira':
                status_text = f"💾 {selected.upper()} Cost: ข้อมูลฝังในโปรแกรม (อัพเดทแล้ว) ({menu_count} เมนู, เฉลี่ย {avg_cost:.0f} บาท)"
            else:
                status_text = f"💾 {selected.upper()} Cost: ข้อมูลฝังในโปรแกรม ({menu_count} เมนู, เฉลี่ย {avg_cost:.0f} บาท)"

            color = self.colors['success']
        else:
            status_text = f"❌ {selected.upper()} Cost: ไม่มีข้อมูล"
            color = self.colors['danger']

        self.base_status_label.config(text=status_text, foreground=color)

    def debug_log(self, message, level="INFO"):
        """เพิ่มข้อความใน debug log"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        log_message = f"[{timestamp}] {level}: {message}\n"

        self.debug_text.insert(tk.END, log_message)
        self.debug_text.see(tk.END)

        if level == "ERROR":
            logger.error(message)
        elif level == "WARNING":
            logger.warning(message)
        else:
            logger.info(message)

    def clear_debug(self):
        """เคลียร์ debug log"""
        self.debug_text.delete(1.0, tk.END)
        self.debug_log("🗑️ Debug log cleared")
        self._show_embedded_info()  # แสดงข้อมูล embedded อีกครั้ง

        self.debug_data = {
            'not_found_menus': [],
            'zero_qty_items': [],
            'invalid_qty_items': [],
            'nan_cost_items': [],
            'matched_menus': [],
            'not_sold_menus': [],
            'processing_summary': {}
        }

    def load_import_file(self):
        """โหลดไฟล์ import"""
        file_path = filedialog.askopenfilename(
            title="เลือกไฟล์ Import (ข้อมูลที่ต้องการคำนวณ)",
            filetypes=[("Excel Files", "*.xlsx *.xls")]
        )
        if not file_path:
            return

        self.import_file = file_path
        self.import_status_label.config(
            text=f"📂 Import: {os.path.basename(file_path)}",
            foreground=self.colors['success']
        )

        self.debug_log(f"📁 เลือกไฟล์ Import: {file_path}")
        self._preview_file(file_path, "Import")

    def load_template_file(self):
        """โหลดไฟล์ template เพิ่มเติม"""
        file_path = filedialog.askopenfilename(
            title="เลือกไฟล์ Template เพิ่มเติม",
            filetypes=[("Excel Files", "*.xlsx *.xls")]
        )
        if not file_path:
            return

        self.template_file = file_path
        self.template_status_label.config(
            text=f"📄 Template: {os.path.basename(file_path)}",
            foreground=self.colors['success']
        )

        self.debug_log(f"📄 เลือกไฟล์ Template: {file_path}")
        self._preview_file(file_path, "Template")

    def _preview_file(self, file_path, file_type):
        """แสดงตัวอย่างข้อมูลในไฟล์"""
        try:
            df_preview = pd.read_excel(file_path)
            self.debug_log(f"📊 ไฟล์ {file_type} มี {len(df_preview)} แถว, {len(df_preview.columns)} คอลัมน์")
            self.debug_log(f"📋 คอลัมน์: {df_preview.columns.tolist()}")

            # ตรวจสอบคอลัมน์ที่จำเป็น
            required_cols = ["MENU NAME", "Qty"]
            missing_cols = [col for col in required_cols if col not in df_preview.columns]
            if missing_cols:
                self.debug_log(f"⚠️ คำเตือน: ไฟล์ขาดคอลัมน์ {missing_cols}", "WARNING")

            # แสดงตัวอย่างข้อมูล
            self.debug_log(f"📋 ตัวอย่างข้อมูล {file_type} 5 แถวแรก:")
            for i, row in df_preview.head(5).iterrows():
                menu = row.get("MENU NAME", "N/A")
                qty = row.get("Qty", "N/A")
                self.debug_log(f"   {i + 1}. {menu} = {qty}")

        except Exception as e:
            self.debug_log(f"❌ Error ตรวจสอบไฟล์ {file_type}: {str(e)}", "ERROR")

    def get_current_base_df(self):
        """ดึง DataFrame ของ base ที่เลือกอยู่"""
        selected = self.selected_base_type.get()
        return config_manager.get_base_data(selected)

    def check_base_data(self):
        """ตรวจสอบและแสดงข้อมูล base ที่เลือก"""
        selected = self.selected_base_type.get()
        df_base = self.get_current_base_df()

        if df_base is None:
            messagebox.showerror("Error", f"ไม่มีข้อมูล {selected.upper()} Cost")
            return

        self.debug_log(f"🔍 === ตรวจสอบข้อมูล {selected.upper()} Cost (Embedded) ===")
        self.debug_log(f"📊 จำนวนเมนู: {len(df_base)}")
        self.debug_log(f"📋 คอลัมน์: {df_base.columns.tolist()}")

        # สถิติราคา
        try:
            cost_stats = df_base["Material Cost"].describe()
            self.debug_log("💰 สถิติราคา Material Cost:")
            self.debug_log(f"   📉 ต่ำสุด: {cost_stats['min']:.2f} บาท")
            self.debug_log(f"   📈 สูงสุด: {cost_stats['max']:.2f} บาท")
            self.debug_log(f"   💰 เฉลี่ย: {cost_stats['mean']:.2f} บาท")
            self.debug_log(f"   📊 จำนวนรายการ: {cost_stats['count']}")

            # แสดงข้อมูลพิเศษสำหรับ Hashira ที่อัพเดท
            if selected == 'hashira':
                self.debug_log(f"✅ ข้อมูล Hashira Cost ได้รับการอัพเดทแล้ว!")

        except Exception as e:
            self.debug_log(f"❌ Error คำนวณสถิติ: {str(e)}", "ERROR")

        # ตัวอย่างเมนู
        self.debug_log(f"🍽️ ตัวอย่างเมนู 10 รายการแรก ({selected.upper()}):")
        for i, (menu_name, row) in enumerate(df_base.head(10).iterrows()):
            try:
                material_cost = row["Material Cost"]
                self.debug_log(f"   {i + 1}. {menu_name}: {material_cost:.2f} บาท")
            except:
                self.debug_log(f"   {i + 1}. {menu_name}: ERROR")

    def calculate(self):
        """คำนวณต้นทุนแบบ Enhanced ด้วยข้อมูลที่ฝังไว้"""
        df_base = self.get_current_base_df()
        selected_base = self.selected_base_type.get()

        if df_base is None:
            messagebox.showerror("Error", f"ไม่มีข้อมูล {selected_base.upper()} Cost")
            return

        if not self.import_file:
            messagebox.showwarning("Warning", "กรุณาเลือกไฟล์ Import ก่อน")
            return

        self.debug_log(f"⚡ === เริ่มการคำนวณด้วย {selected_base.upper()} Cost (Updated Embedded Data) ===")

        # เคลียร์ debug data เก่า
        self.debug_data = {
            'not_found_menus': [],
            'zero_qty_items': [],
            'invalid_qty_items': [],
            'nan_cost_items': [],
            'matched_menus': [],
            'not_sold_menus': [],
            'processing_summary': {}
        }

        try:
            # อ่านไฟล์ import
            df_import = pd.read_excel(self.import_file)
            self.debug_log(f"📁 อ่านไฟล์ import สำเร็จ: {len(df_import)} แถว")

            # ลบช่องว่างจากชื่อคอลัมน์
            df_import.columns = df_import.columns.str.strip()

        except Exception as e:
            error_msg = f"❌ ไม่สามารถอ่านไฟล์ได้: {str(e)}"
            self.debug_log(error_msg, "ERROR")
            messagebox.showerror("Error", error_msg)
            return

        results = []
        matched_count = 0

        self.debug_log(f"🔄 เริ่มประมวลผล {len(df_import)} รายการ")

        for idx, row in df_import.iterrows():
            menu = row.get("MENU NAME")
            qty = row.get("Qty", 0)

            if self.debug_mode.get():
                self.debug_log(f"   ประมวลผลแถว {idx + 1}: {menu} = {qty}")

            # ตรวจสอบ menu name
            if pd.isna(menu) or menu == "":
                self.debug_log(f"⚠️ แถว {idx + 1}: ชื่อเมนูว่าง", "WARNING")
                continue

            # ตรวจสอบ quantity
            original_qty = qty
            try:
                qty = float(qty) if not pd.isna(qty) else 0
                if qty == 0:
                    self.debug_data['zero_qty_items'].append({
                        'row': idx + 1,
                        'menu': menu,
                        'original_qty': original_qty
                    })
                    if self.debug_mode.get():
                        self.debug_log(f"⚠️ แถว {idx + 1}: Qty = 0 สำหรับ {menu}", "WARNING")
            except (ValueError, TypeError):
                self.debug_data['invalid_qty_items'].append({
                    'row': idx + 1,
                    'menu': menu,
                    'invalid_qty': original_qty
                })
                self.debug_log(f"⚠️ แถว {idx + 1}: Qty ไม่ใช่ตัวเลข ({original_qty}) สำหรับ {menu}", "WARNING")
                qty = 0

            # ค้นหาใน base data (embedded)
            if menu in df_base.index:
                try:
                    material_cost = df_base.at[menu, "Material Cost"]

                    # ตรวจสอบ material cost
                    if pd.isna(material_cost):
                        self.debug_data['nan_cost_items'].append({
                            'menu': menu,
                            'row': idx + 1
                        })
                        self.debug_log(f"⚠️ แถว {idx + 1}: Material Cost เป็น NaN สำหรับ {menu}", "WARNING")
                        material_cost = 0
                    else:
                        material_cost = float(material_cost)

                    total_cost = qty * material_cost
                    results.append([menu, qty, material_cost, total_cost])

                    self.debug_data['matched_menus'].append({
                        'row': idx + 1,
                        'menu': menu,
                        'qty': qty,
                        'material_cost': material_cost,
                        'total_cost': total_cost
                    })

                    matched_count += 1

                    if self.debug_mode.get():
                        self.debug_log(f"   ✅ พบเมนู: {material_cost:.2f} x {qty} = {total_cost:.2f}")

                except Exception as e:
                    self.debug_log(f"❌ Error คำนวณ {menu}: {str(e)}", "ERROR")
            else:
                self.debug_data['not_found_menus'].append({
                    'row': idx + 1,
                    'menu': menu,
                    'qty': qty
                })
                if self.debug_mode.get():
                    self.debug_log(f"   ❌ ไม่พบเมนู: {menu}", "WARNING")

        # หาเมนูที่ไม่ได้ขาย (อยู่ใน base แต่ไม่มีใน import)
        import_menus = set(df_import["MENU NAME"].dropna().tolist())
        base_menus = set(df_base.index.tolist())
        not_sold_menus = base_menus - import_menus

        self.debug_data['not_sold_menus'] = []
        for menu in not_sold_menus:
            try:
                cost = df_base.at[menu, "Material Cost"]
                if pd.isna(cost):
                    cost = 0
                else:
                    cost = float(cost)

                self.debug_data['not_sold_menus'].append({
                    'menu': menu,
                    'material_cost': cost
                })
            except:
                self.debug_data['not_sold_menus'].append({
                    'menu': menu,
                    'material_cost': 0
                })

        # บันทึกสรุปการประมวลผล
        self.debug_data['processing_summary'] = {
            'total_rows': len(df_import),
            'matched_count': matched_count,
            'not_found_count': len(self.debug_data['not_found_menus']),
            'zero_qty_count': len(self.debug_data['zero_qty_items']),
            'invalid_qty_count': len(self.debug_data['invalid_qty_items']),
            'nan_cost_count': len(self.debug_data['nan_cost_items']),
            'not_sold_count': len(self.debug_data['not_sold_menus']),
            'base_type': selected_base.upper(),
            'data_source': 'EMBEDDED_UPDATED',
            'timestamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }

        # สรุปผลการประมวลผล
        self.debug_log(f"📊 === สรุปผลการประมวลผล ({selected_base.upper()} - UPDATED EMBEDDED) ===")
        summary = self.debug_data['processing_summary']
        self.debug_log(f"📋 รายการทั้งหมด: {summary['total_rows']}")
        self.debug_log(f"✅ พบเมนูที่ตรงกัน: {summary['matched_count']}")
        self.debug_log(f"❌ ไม่พบเมนู: {summary['not_found_count']}")
        self.debug_log(f"⚠️ Qty = 0: {summary['zero_qty_count']}")
        self.debug_log(f"⚠️ Qty ไม่ถูกต้อง: {summary['invalid_qty_count']}")
        self.debug_log(f"⚠️ Material Cost NaN: {summary['nan_cost_count']}")
        self.debug_log(f"📋 รายการที่ไม่ได้ขาย: {summary['not_sold_count']}")

        # สร้าง DataFrame ผลลัพธ์
        self.df_result = pd.DataFrame(results, columns=["MENU NAME", "Qty", "Material Cost", "Total Cost"])

        if self.df_result.empty:
            error_msg = f"ไม่มีชื่อเมนูที่ตรงกับ {selected_base.upper()} Cost"
            self.debug_log(error_msg, "WARNING")
            messagebox.showwarning("ไม่พบข้อมูล", error_msg)
            return

        # คำนวณ Grand Total
        grand_total = self.df_result["Total Cost"].sum()
        self.debug_log(f"💰 Grand Total: {grand_total:.2f} บาท")

        # เพิ่มแถว Grand Total
        grand_total_row = pd.DataFrame([["Grand Total", "", "", grand_total]],
                                       columns=["MENU NAME", "Qty", "Material Cost", "Total Cost"])
        self.df_result = pd.concat([self.df_result, grand_total_row], ignore_index=True)

        # แสดงในตาราง
        self._update_result_table()

        # อัพเดทสถิติ
        self._update_statistics(matched_count, summary['not_found_count'], grand_total, selected_base)

        success_msg = f"✅ คำนวณเสร็จสิ้น - พบ {matched_count} รายการ (ใช้ {selected_base.upper()} Cost - Updated)"
        self.debug_log(success_msg)
        messagebox.showinfo("Success", success_msg)

    def _update_result_table(self):
        """อัพเดทตารางผลลัพธ์"""
        # เคลียร์ตารางเก่า
        for row in self.tree.get_children():
            self.tree.delete(row)

        # เพิ่มข้อมูลใหม่
        for _, r in self.df_result.iterrows():
            qty_display = int(r["Qty"]) if r["Qty"] != "" else ""
            cost_display = f"{r['Material Cost']:.2f}" if r["Material Cost"] != "" else ""
            total_display = f"{r['Total Cost']:.2f}" if r["Total Cost"] != "" else ""

            # ไฮไลท์แถว Grand Total
            tags = ("grand_total",) if r["MENU NAME"] == "Grand Total" else ()

            self.tree.insert("", tk.END,
                             values=(r["MENU NAME"], qty_display, cost_display, total_display),
                             tags=tags)

        # จัดรูปแบบแถว Grand Total
        self.tree.tag_configure("grand_total",
                                background="#3498db",
                                foreground="#ffffff",
                                font=("Segoe UI", 10, "bold"))

    def _update_statistics(self, matched, not_found, grand_total, base_type):
        """อัพเดทสถิติ"""
        update_indicator = " (Updated)" if base_type.lower() == 'hashira' else ""
        stats_text = (f"✅ พบเมนู: {matched} | ❌ ไม่พบ: {not_found} | "
                      f"💰 รวม: {grand_total:,.2f} บาท | 🏪 Base: {base_type.upper()}{update_indicator}")
        self.stats_label.config(text=stats_text)

    def export_excel(self):
        """บันทึก Excel แบบปกติ"""
        self.debug_log("🔄 เริ่มกระบวนการบันทึก Excel...")

        if self.df_result is None or self.df_result.empty:
            self.debug_log("⚠️ ไม่มีผลลัพธ์ให้บันทึก", "WARNING")
            messagebox.showwarning("Warning", "กรุณาคำนวณก่อนบันทึก")
            return

        selected_base = self.selected_base_type.get()
        default_name = f"Cost_Calculation_{selected_base.upper()}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

        try:
            file_path = filedialog.asksaveasfilename(
                title="บันทึกผลลัพธ์การคำนวณ",
                defaultextension=".xlsx",
                initialname=default_name,
                filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")]
            )

            if not file_path:
                self.debug_log("ℹ️ ผู้ใช้ยกเลิกการบันทึก")
                return

            self.debug_log(f"📂 ไฟล์ที่เลือก: {file_path}")

            # เตรียมข้อมูล
            df_export = self.df_result.copy()
            df_export['Material Cost'] = pd.to_numeric(df_export['Material Cost'], errors='coerce').fillna(0)
            df_export['Total Cost'] = pd.to_numeric(df_export['Total Cost'], errors='coerce').fillna(0)

            # บันทึกไฟล์ Excel
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                df_export.to_excel(writer, sheet_name='ผลลัพธ์', index=False)

                # จัดรูปแบบ
                workbook = writer.book
                worksheet = writer.sheets['ผลลัพธ์']

                # จัดรูปแบบ header
                header_font = Font(bold=True)
                header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill

                # จัดรูปแบบตัวเลข
                for row in worksheet.iter_rows(min_row=2):
                    if len(row) >= 3 and isinstance(row[2].value, (int, float)):
                        row[2].number_format = '#,##0.00'
                    if len(row) >= 4 and isinstance(row[3].value, (int, float)):
                        row[3].number_format = '#,##0.00'

                # ไฮไลท์แถว Grand Total
                grand_total_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
                for row in worksheet.iter_rows(min_row=2):
                    if row[0].value == "Grand Total":
                        for cell in row:
                            cell.font = Font(bold=True)
                            cell.fill = grand_total_fill

            success_msg = f"💾 บันทึกไฟล์สำเร็จ: {os.path.basename(file_path)}"
            self.debug_log(success_msg)
            messagebox.showinfo("Success",
                                f"บันทึกผลลัพธ์สำเร็จ!\n\n📁 ไฟล์: {os.path.basename(file_path)}\n📍 ที่อยู่: {file_path}")

        except Exception as e:
            error_msg = f"❌ ไม่สามารถบันทึกไฟล์ได้: {str(e)}"
            self.debug_log(error_msg, "ERROR")
            messagebox.showerror("Error", f"เกิดข้อผิดพลาดในการบันทึก:\n{error_msg}")

    def export_excel_with_debug(self):
        """บันทึก Excel พร้อม Debug sheets"""
        self.debug_log("🔄 เริ่มกระบวนการบันทึก Excel พร้อม Debug...")

        if self.df_result is None or self.df_result.empty:
            messagebox.showwarning("Warning", "กรุณาคำนวณก่อนบันทึก")
            return

        selected_base = self.selected_base_type.get()
        default_name = f"Cost_Calculation_Debug_{selected_base.upper()}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

        try:
            file_path = filedialog.asksaveasfilename(
                title="บันทึกผลลัพธ์พร้อม Debug Information",
                defaultextension=".xlsx",
                initialname=default_name,
                filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")]
            )

            if not file_path:
                return

            self.debug_log(f"📂 ไฟล์ที่เลือก: {file_path}")
            wb = Workbook()

            # Sheet 1: ผลลัพธ์หลัก
            self.debug_log("🔄 สร้าง Sheet 1: ผลลัพธ์การคำนวณ...")
            ws_main = wb.active
            ws_main.title = "ผลลัพธ์การคำนวณ"

            df_export = self.df_result.copy()
            df_export["Material Cost"] = pd.to_numeric(df_export["Material Cost"], errors="coerce").fillna(0).round(2)
            df_export["Total Cost"] = pd.to_numeric(df_export["Total Cost"], errors="coerce").fillna(0).round(2)

            for r in dataframe_to_rows(df_export, index=False, header=True):
                ws_main.append(r)

            self._format_main_sheet(ws_main)
            self.debug_log("✅ Sheet ผลลัพธ์การคำนวณ สร้างเสร็จ")

            # Sheet 2: ข้อมูล Base Cost ที่ใช้
            self._add_base_data_sheet(wb)

            # Sheet 3: ข้อมูล Import
            if self.import_file and os.path.exists(self.import_file):
                self._add_import_data_sheet(wb)

            # Sheet 4: เมนูที่ไม่พบ
            if self.debug_data['not_found_menus']:
                self._add_not_found_sheet(wb)

            # Sheet 5: สรุปการประมวลผล
            self._add_summary_sheet(wb)

            # ปรับความกว้างคอลัมน์
            self._adjust_column_width(wb)

            wb.save(file_path)

            success_msg = f"💾 บันทึกไฟล์พร้อม Debug สำเร็จ: {os.path.basename(file_path)}"
            self.debug_log(success_msg)

            sheet_info = f"สร้าง {len(wb.worksheets)} sheets:\n"
            for i, ws in enumerate(wb.worksheets, 1):
                sheet_info += f"{i}. {ws.title}\n"

            messagebox.showinfo("Success",
                                f"บันทึกไฟล์ Debug สำเร็จ!\n\n"
                                f"📁 ไฟล์: {os.path.basename(file_path)}\n"
                                f"📊 จำนวน Sheets: {len(wb.worksheets)}\n\n{sheet_info}")

        except Exception as e:
            error_msg = f"❌ ไม่สามารถบันทึกไฟล์ Debug ได้: {str(e)}"
            self.debug_log(error_msg, "ERROR")
            messagebox.showerror("Error", f"เกิดข้อผิดพลาด:\n{error_msg}")

    def _add_base_data_sheet(self, wb):
        """เพิ่ม sheet ข้อมูล Base Cost ที่ใช้"""
        try:
            selected_base = self.selected_base_type.get()
            df_base = self.get_current_base_df()

            if df_base is None:
                return

            update_indicator = " (อัพเดทแล้ว)" if selected_base == 'hashira' else ""
            ws_base = wb.create_sheet(f"ข้อมูล {selected_base.upper()} Cost{update_indicator}")

            # Header พิเศษ
            ws_base.append([f"🏪 ข้อมูล {selected_base.upper()} Cost ที่ใช้ในการคำนวณ{update_indicator}"])
            ws_base.append([f"ประเภท: {selected_base.upper()} Cost"])
            ws_base.append([f"แหล่งข้อมูล: Embedded in Program (Updated)"])
            ws_base.append([f"จำนวนเมนู: {len(df_base)}"])
            ws_base.append([f"วันที่: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
            ws_base.append([])

            # ข้อมูล base
            df_base_export = df_base.reset_index()
            for r in dataframe_to_rows(df_base_export, index=False, header=True):
                ws_base.append(r)

            # จัดรูปแบบ
            ws_base[1][0].font = Font(bold=True, size=14)
            color = "FFE6E6" if selected_base == 'hashira' else "E6FFE6"  # สีแดงอ่อนสำหรับ hashira ที่อัพเดท
            ws_base[1][0].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

            # Header ของข้อมูล
            header_row = 7
            for cell in ws_base[header_row]:
                if cell.value:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

            self.debug_log(f"✅ เพิ่ม Sheet ข้อมูล {selected_base.upper()} Cost สำเร็จ")

        except Exception as e:
            self.debug_log(f"❌ Error เพิ่ม Base sheet: {e}", "ERROR")

    def _add_import_data_sheet(self, wb):
        """เพิ่ม sheet ข้อมูล Import"""
        try:
            df_import = pd.read_excel(self.import_file)
            ws_import = wb.create_sheet("ข้อมูล Import")

            # Header พิเศษ
            ws_import.append(["📂 ข้อมูล Import ที่ใช้ในการคำนวณ"])
            ws_import.append([f"ไฟล์: {os.path.basename(self.import_file)}"])
            ws_import.append([f"วันที่: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
            ws_import.append([f"จำนวนแถว: {len(df_import)}"])
            ws_import.append([f"Base Cost: {self.debug_data['processing_summary']['base_type']} (Updated Embedded)"])
            ws_import.append([])

            # ข้อมูล
            for r in dataframe_to_rows(df_import, index=False, header=True):
                ws_import.append(r)

            # จัดรูปแบบ
            ws_import[1][0].font = Font(bold=True, size=14)
            ws_import[1][0].fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")

            self.debug_log("✅ เพิ่ม Sheet ข้อมูล Import สำเร็จ")

        except Exception as e:
            self.debug_log(f"❌ Error เพิ่ม Import sheet: {e}", "ERROR")

    def _add_not_found_sheet(self, wb):
        """เพิ่ม sheet เมนูที่ไม่พบ"""
        ws_not_found = wb.create_sheet("เมนูที่ไม่พบ")

        # Header
        ws_not_found.append(["❌ เมนูที่ไม่พบในฐานข้อมูล"])
        ws_not_found.append([f"Base Cost: {self.debug_data['processing_summary']['base_type']} (Updated Embedded)"])
        ws_not_found.append([f"จำนวน: {len(self.debug_data['not_found_menus'])} รายการ"])
        ws_not_found.append([])
        ws_not_found.append(["แถวที่", "ชื่อเมนู", "จำนวน"])

        # ข้อมูล
        for item in self.debug_data['not_found_menus']:
            ws_not_found.append([item['row'], item['menu'], item['qty']])

        # จัดรูปแบบ
        for i in range(1, 6):
            if i <= 3:
                ws_not_found[i][0].font = Font(bold=True, size=12)
                ws_not_found[i][0].fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

    def _add_summary_sheet(self, wb):
        """เพิ่ม sheet สรุปการประมวลผล"""
        ws_summary = wb.create_sheet("สรุปการประมวลผล")
        summary = self.debug_data['processing_summary']

        # ข้อมูลสรุป
        summary_data = [
            ["🔍 สรุปการประมวลผล - Updated Embedded Edition", ""],
            ["", ""],
            ["📊 ข้อมูลทั่วไป", ""],
            ["Base Cost ที่ใช้", f"{summary['base_type']} (Updated Embedded)"],
            ["แหล่งข้อมูล", "ข้อมูลฝังในโปรแกรม (อัพเดทแล้ว)"],
            ["รายการทั้งหมด", summary['total_rows']],
            ["พบเมนูที่ตรงกัน", summary['matched_count']],
            ["ไม่พบเมนู", summary['not_found_count']],
            ["Qty = 0", summary['zero_qty_count']],
            ["Qty ไม่ถูกต้อง", summary['invalid_qty_count']],
            ["Material Cost NaN", summary['nan_cost_count']],
            ["รายการที่ไม่ได้ขาย", summary['not_sold_count']],
            ["", ""],
            ["📅 ข้อมูลการประมวลผล", ""],
            ["เวลาประมวลผล", summary['timestamp']],
            ["ไฟล์ Import", os.path.basename(self.import_file) if self.import_file else "N/A"],
            ["ไฟล์ Template", os.path.basename(self.template_file) if self.template_file else "ไม่มี"]
        ]

        # เขียนข้อมูลลง sheet
        for row_data in summary_data:
            ws_summary.append(row_data)

        # จัดรูปแบบ
        ws_summary[1][0].font = Font(bold=True, size=14)
        ws_summary[1][0].fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")

    def _format_main_sheet(self, ws):
        """จัดรูปแบบ sheet หลัก"""
        # Header formatting
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        # Number formatting
        for row in ws.iter_rows(min_row=2):
            if len(row) >= 3 and isinstance(row[2].value, (int, float)):
                row[2].number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
            if len(row) >= 4 and isinstance(row[3].value, (int, float)):
                row[3].number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1

        # Highlight Grand Total row
        for row in ws.iter_rows(min_row=2):
            if row[0].value == "Grand Total":
                for cell in row:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")

    def _adjust_column_width(self, wb):
        """ปรับความกว้างคอลัมน์"""
        for ws in wb.worksheets:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 60)
                ws.column_dimensions[column_letter].width = adjusted_width

    def import_external_base(self):
        """Import ไฟล์ Base Cost ภายนอกเพื่ออัพเดทข้อมูลฝัง"""
        file_path = filedialog.askopenfilename(
            title="เลือกไฟล์ Base Cost ภายนอกเพื่ออัพเดท",
            filetypes=[("Excel Files", "*.xlsx *.xls")]
        )
        if not file_path:
            return

        try:
            success = config_manager.update_base_data_from_excel(file_path)

            if success:
                self.debug_log(f"🔄 อัพเดทข้อมูล Base จากไฟล์: {os.path.basename(file_path)}")

                # แสดงสถิติใหม่
                stats = config_manager.get_statistics()
                for base_type, stat in stats.items():
                    self.debug_log(f"📊 {base_type.upper()}: {stat['total_menus']} เมนู")

                self.update_base_status()
                messagebox.showinfo("Success", f"อัพเดทข้อมูล Base สำเร็จ:\n{os.path.basename(file_path)}")
            else:
                raise Exception("ไม่สามารถอ่านไฟล์ได้")

        except Exception as e:
            error_msg = f"❌ ไม่สามารถ import ไฟล์ Base ได้: {str(e)}"
            self.debug_log(error_msg, "ERROR")
            messagebox.showerror("Error", error_msg)

    def export_base_data(self):
        """Export ข้อมูล Base Cost ปัจจุบัน"""
        file_path = filedialog.asksaveasfilename(
            title="Export ข้อมูล Base Cost",
            defaultextension=".xlsx",
            initialname=f"Base_Cost_Export_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
            filetypes=[("Excel Files", "*.xlsx")]
        )
        if not file_path:
            return

        try:
            wb = Workbook()
            wb.remove(wb.active)  # ลบ sheet default

            sheets_created = 0

            # Export Hashira Cost (Updated)
            if config_manager.base_data['hashira'] is not None:
                ws_hashira = wb.create_sheet("Hashira Cost (Updated)")
                df_hashira = config_manager.base_data['hashira'].reset_index()

                # เพิ่ม info header
                ws_hashira.append(["🏯 Hashira Cost - อัพเดทแล้ว"])
                ws_hashira.append([f"จำนวนเมนู: {len(df_hashira)}"])
                ws_hashira.append([f"วันที่ Export: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
                ws_hashira.append([])

                for r in dataframe_to_rows(df_hashira, index=False, header=True):
                    ws_hashira.append(r)

                # จัดรูปแบบ header
                ws_hashira[1][0].font = Font(bold=True, size=14)
                ws_hashira[1][0].fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")

                for cell in ws_hashira[5]:  # Data header
                    if cell.value:
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

                sheets_created += 1
                self.debug_log(f"📤 Export Hashira Cost (Updated): {len(df_hashira)} เมนู")

            # Export Hamada Cost
            if config_manager.base_data['hamada'] is not None:
                ws_hamada = wb.create_sheet("Hamada Cost")
                df_hamada = config_manager.base_data['hamada'].reset_index()

                # เพิ่ม info header
                ws_hamada.append(["🍜 Hamada Cost"])
                ws_hamada.append([f"จำนวนเมนู: {len(df_hamada)}"])
                ws_hamada.append([f"วันที่ Export: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
                ws_hamada.append([])

                for r in dataframe_to_rows(df_hamada, index=False, header=True):
                    ws_hamada.append(r)

                # จัดรูปแบบ header
                ws_hamada[1][0].font = Font(bold=True, size=14)
                ws_hamada[1][0].fill = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")

                for cell in ws_hamada[5]:  # Data header
                    if cell.value:
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

                sheets_created += 1
                self.debug_log(f"📤 Export Hamada Cost: {len(df_hamada)} เมนู")

            # ปรับความกว้างคอลัมน์
            self._adjust_column_width(wb)
            wb.save(file_path)

            success_msg = f"📤 Export ข้อมูล Base สำเร็จ: {os.path.basename(file_path)}"
            self.debug_log(success_msg)
            messagebox.showinfo("Success", f"Export ข้อมูล Base แล้ว ({sheets_created} sheets):\n{file_path}")

        except Exception as e:
            error_msg = f"❌ ไม่สามารถ export ข้อมูล Base ได้: {str(e)}"
            self.debug_log(error_msg, "ERROR")
            messagebox.showerror("Error", error_msg)


def main():
    """ฟังก์ชันหลักพร้อม error handling"""
    try:
        root = tk.Tk()

        # Set window icon
        try:
            root.iconphoto(True, tk.PhotoImage(data=''))
        except:
            pass

        app = EmbeddedCostCalculatorApp(root)
        logger.info("Starting Updated Embedded Cost Calculator")
        root.mainloop()

    except Exception as e:
        logger.error(f"Critical error in main: {str(e)}")
        logger.error(traceback.format_exc())
        messagebox.showerror("ผิดพลาดร้ายแรง", f"เกิดข้อผิดพลาดร้ายแรง:\n{str(e)}")


if __name__ == "__main__":
    main()