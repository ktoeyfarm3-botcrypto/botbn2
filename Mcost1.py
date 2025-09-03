import pandas as pd
import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext, filedialog
import tkinter.simpledialog
from datetime import datetime
import logging
import traceback
import os

# Enhanced imports with error handling
try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import numbers, Font, PatternFill, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows

    OPENPYXL_AVAILABLE = True
except ImportError as e:
    OPENPYXL_AVAILABLE = False
    print(f"Warning: openpyxl not available: {e}")

try:
    import xlsxwriter

    XLSXWRITER_AVAILABLE = True
except ImportError:
    XLSXWRITER_AVAILABLE = False


# ===== Enhanced Configuration with Built-in Data =====
class ConfigManager:
    def __init__(self):
        self.config_file = "mapcost_config.ini"
        self.base_files = {
            'hashira': None,
            'hamada': None
        }
        self.initialize_default_data()
        self.load_config()

    def initialize_default_data(self):
        """สร้างข้อมูล Hamada Cost และ Hashira Cost จากข้อมูลจริง"""

        # ข้อมูลจริงจาก Hamada Cost.xlsx
        hamada_menu_names = [
            "Ebi Tempura Set", "Salmon Tempura Set", "Ebi & Salmon Tempura Set",
            "Ebi Tempura Don Set", "Salmon Tempura Don Set", "Ebi & Salmon Tempura Don Set",
            "Ebi Tempura Udon", "Salmon Tempura Udon", "Ebi & Salmon Tempura Udon",
            "Yakiudon Buta", "Yakiudon Kaisen", "Ebi Fry Curry Rice",
            "Tonkatsu Curry Rice", "Salmonkatsu Curry Rice", "Ebi Tempura（1pc）",
            "Salmon Tempura（1pc）", "Takoyaki", "Sushi Yorokobi",
            "Sushi Tokujo", "Maguro Zanmai", "Salmon Zanmai",
            "Hamachi Zanmai", "Duo Sushi", "Salmon Mountain",
            "Salmon Sushi & Ikura Gunkan", "California Roll", "Salmon Roll",
            "Unagi Roll", "Hokkaido Hotate Roll", "Salmon Chirashi Sushi",
            "Salmon Tobiko Chirashi Sushi", "Barachirashi Sushi", "Salmon Don",
            "Salmon Ikura Don", "Maguro Don", "Negitoro Ikura Don",
            "Kaisen Don", "Unagi Don", "Salmon Sashimi",
            "Maguro Sashimi", "Hamachi Sashimi", "Hotate Sashimi",
            "Salmon Sushi (2pcs)", "Maguro Sushi (2pcs)", "Hamachi Sushi (2pcs)",
            "Tamago Sushi (2pcs)", "Ebi Sushi (2pcs)", "Ika Sushi (2pcs)",
            "Hokki Sushi (2pcs)", "Hotate Sushi (2pcs)", "Unagi Sushi (2pcs)",
            "Anago Sushi (2pcs)", "Negitoro Gunkan (2pcs)", "Tobiko Gunkan (2pcs)",
            "Ikura Gunkan (2pcs)", "GG California Roll", "GG Salmon Sushi",
            "GG Tamago & Ebi Sushi", "GG Ebi & Tobiko Don", "GG Kanikama & Tobiko Don",
            "GG Triple Mix Don", "GG Ebiten Don", "GG Salmon & Ebiten Don",
            "Coca-Cola Original", "Coke-Cola Zero Sugar", "Sprite",
            "Soda", "Mineral Water", "HBD Sparkling Water Lemon (No Sugar, No Calories)",
            "HBD Sparkling Water Honey Yuzu (No Sugar, No Calories)",
            "HBD Sparkling Water Peach (No Sugar, No Calories)",
            "HBD Sparkling Water Kyoho (No Sugar, No Calories)", "Asahi Beer",
            "Heineken Beer ", "Singha Beer", "Drinking Water",
            "Matcha Soft Serve", "Yogurt Soft Serve", "Two Tone Soft Serve",
            "Matcha Japanese", "Matcha Marshmallow", "Mix Berry Yogurt",
            "Brownie Yogurt", "Two Tone Brownie", "Saikyo Matcha",
            "Rice", "Miso Soup"
        ]

        hamada_costs = [
            67.53, 51.84, 70.42, 69.33, 53.64, 69.29,
            56.03, 54.34, 73.57, 37.21, 48.67, 62.31,
            69.17, 60.62, 16.16, 10.92, 33.71, 63.56,
            203.12, 89.93, 98.14, 168.1, 108.73, 200.7,
            193.91, 50.76, 87.59, 104.1, 133.97, 48.53,
            66.95, 77.77, 79.45, 108.68, 73.91, 111.38,
            243.48, 246.42, 42.71, 39.59, 60.89, 52.93,
            29.95, 27.87, 42.07, 9.63, 19.94, 19.87,
            19.96, 51.71, 56.12, 112.92, 21.27, 18.32,
            68.24, 51.43, 93.14, 56.38, 68.9, 48.9,
            58.8, 63.34, 74.35, 12.54, 12.54, 12.54,
            7.59, 7, 17, 17, 17, 17, 41.92,
            38.43, 31.71, 3.74, 15.09, 17.72, 16.41,
            22.6, 30.78, 34.94, 41.79, 37.45, 31.54,
            5.85, 3.58
        ]

        # ข้อมูลจริงจาก Hashira Cost.xlsx
        hashira_menu_names = [
            "Pork Cut Steak Set", "Tonteki Set", "Beef Cut Steak Set (Medium Rare / Well done)",
            "Buta Teriyaki Set", "Tonkatsu Set", "Tonkatsu Tamagotoji Set",
            "Beef Hamburg Set", "Salmon Set (Teriyaki / Shioyaki)", "Kaisen Teppan Yaki Set",
            "Topping Fried Egg", "Gokuatsu Tonkatsu Set", "Gokuatsu Katsu Don Set",
            "Gokuatsu Tonkatsu Tamagotoji Set", "Tonkotsu Chashumen", "Tonkotsu Niku-Niku Chashumen",
            "Tonkotsu Spicy Ramen", "Tonkotsu Hashira Ramen", "Miso Chashumen",
            "Miso Hokkaido Ramen", "Shoyu Chashumen", "Shoyu Tokyo Ramen",
            "Tom Yum Tonkotsu Kaisen Ramen", "Tom Yum Kung Ramen", "Tori Paitan Karaage Ramen",
            "Tori Paitan Kaisen Ramen", "Tonkatsu", "Hotate",
            "Ebi Tempura", "Chashu", "Spicy Nikumiso",
            "Karashinegi", "Negi", "Naruto",
            "Ajitama", "Menma", "Nori",
            "Corn", "Butter", "Boiled Cabbage",
            "Boiled Bean Sprout", "Karashi Takana", "Kikurage",
            "Kanikama", "Kaedama", "GG Sauce Katsudon",
            "GG Buta Teriyaki Don", "GG Tonkatsu Bento", "GG Yakisoba",
            "GG Yakisoba with Sunny Side Egg", "GG Pork Okonomiyaki", "Coca-Cola Original",
            "Coca-Cola Zero Sugar", "Sprite", "Soda",
            "Mineral Water", "HBD Sparkling Water Lemon (No Sugar, No Calories)",
            "HBD Sparkling Water Honey Yuzu (No Sugar, No Calories)",
            "HBD Sparkling Water Peach (No Sugar, No Calories)",
            "HBD Sparkling Water Kyoho (No Sugar, No Calories)", "Asahi Beer",
            "Heineken Beer ", "Singha Beer", "Drinking Water",
            "Melon Ball", "Vanilla Monaka", "Azuki Bar",
            "Rice", "Miso Soup"
        ]

        hashira_costs = [
            46.53, 131.99, 107.87, 41.77, 58.39, 60.91,
            76.4, 68.02, 121.23, 3.5, 131.77, 133,
            131.46, 64.23, 77.98, 54.05, 74.33, 56.52,
            70.08, 63.1, 64.44, 55.55, 116.08, 42.88,
            50.08, 37.4, 33.33, 26, 27.3, 7.97,
            10.54, 6.43, 12.92, 3.84, 9.6, 5.4,
            6, 1.31, 2.73, 1.98, 6.45, 3.41,
            7.92, 7.5, 67.94, 45.93, 79.95, 45.61,
            49.11, 52.4, 12.54, 12.54, 12.54, 7.59,
            7, 17, 17, 17, 17, 41.92,
            38.43, 31.71, 3.74, 25, 27, 27,
            5.85, 3.04
        ]

        # สร้าง DataFrame
        hamada_real_data = {
            "MENU NAME": hamada_menu_names,
            "Material Cost": hamada_costs
        }

        hashira_real_data = {
            "MENU NAME": hashira_menu_names,
            "Material Cost": hashira_costs
        }

        # สร้าง DataFrame และ set index
        self.base_files['hamada'] = pd.DataFrame(hamada_real_data).set_index('MENU NAME')
        self.base_files['hashira'] = pd.DataFrame(hashira_real_data).set_index('MENU NAME')

        logger.info("โหลดข้อมูลจริง Base Cost เรียบร้อย")
        logger.info(f"Hamada Cost: {len(self.base_files['hamada'])} เมนู")
        logger.info(f"Hashira Cost: {len(self.base_files['hashira'])} เมนู")

    def load_config(self):
        """โหลดการตั้งค่าจากไฟล์"""
        # โหลดไฟล์ Base Cost.xlsx หากมี (จะ override ข้อมูลตัวอย่าง)
        self.load_base_cost_file()

    def load_base_cost_file(self):
        """โหลดไฟล์ Base Cost.xlsx"""
        base_file = "Base Cost.xlsx"
        if os.path.exists(base_file):
            try:
                # อ่าน Hashira Cost sheet
                hashira_df = pd.read_excel(base_file, sheet_name="Hashira Cost")
                if 'MENU NAME' in hashira_df.columns:
                    self.base_files['hashira'] = hashira_df.set_index("MENU NAME")

                # อ่าน Hamada Cost sheet
                hamada_df = pd.read_excel(base_file, sheet_name="Hamada Cost")
                if 'MENU NAME' in hamada_df.columns:
                    self.base_files['hamada'] = hamada_df.set_index("MENU NAME")

                logger.info(f"โหลดไฟล์ {base_file} สำเร็จ (แทนที่ข้อมูลตัวอย่าง)")
                logger.info(f"Hashira Cost: {len(self.base_files['hashira'])} เมนู")
                logger.info(f"Hamada Cost: {len(self.base_files['hamada'])} เมนู")

            except Exception as e:
                logger.error(f"Error loading base file: {e}")
                logger.info("ใช้ข้อมูลตัวอย่างต่อไป")

    def save_base_cost_file(self):
        """บันทึกไฟล์ Base Cost.xlsx - แก้ไขแล้ว"""
        if not OPENPYXL_AVAILABLE:
            logger.error("openpyxl not available for saving")
            return False

        try:
            wb = Workbook()
            # ลบ sheet default
            if wb.worksheets:
                wb.remove(wb.active)

            # สร้าง Hashira Cost sheet
            if self.base_files.get('hashira') is not None:
                ws_hashira = wb.create_sheet("Hashira Cost")
                df_hashira = self.base_files['hashira'].reset_index()

                # เขียนข้อมูลลง worksheet
                for r_idx, r in enumerate(dataframe_to_rows(df_hashira, index=False, header=True)):
                    for c_idx, value in enumerate(r):
                        cell = ws_hashira.cell(row=r_idx + 1, column=c_idx + 1, value=value)
                        # Format header
                        if r_idx == 0:
                            cell.font = Font(bold=True)
                            cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")

            # สร้าง Hamada Cost sheet
            if self.base_files.get('hamada') is not None:
                ws_hamada = wb.create_sheet("Hamada Cost")
                df_hamada = self.base_files['hamada'].reset_index()

                # เขียนข้อมูลลง worksheet
                for r_idx, r in enumerate(dataframe_to_rows(df_hamada, index=False, header=True)):
                    for c_idx, value in enumerate(r):
                        cell = ws_hamada.cell(row=r_idx + 1, column=c_idx + 1, value=value)
                        # Format header
                        if r_idx == 0:
                            cell.font = Font(bold=True)
                            cell.fill = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")

            # ปรับความกว้างคอลัมน์
            for ws in wb.worksheets:
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if cell.value and len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 60)
                    ws.column_dimensions[column_letter].width = adjusted_width

            wb.save("Base Cost.xlsx")
            logger.info("บันทึก Base Cost.xlsx สำเร็จ")
            return True

        except Exception as e:
            logger.error(f"Error saving base file: {e}")
            return False

    def update_single_base(self, base_type, new_df):
        """อัพเดท Base Cost แยกกันตามประเภท"""
        try:
            # ลบช่องว่างจากชื่อคอลัมน์
            new_df.columns = new_df.columns.str.strip()

            # ตรวจสอบคอลัมน์ที่จำเป็น
            if 'MENU NAME' not in new_df.columns:
                raise ValueError(f"ไฟล์ไม่มีคอลัมน์ 'MENU NAME'")

            if 'Material Cost' not in new_df.columns:
                raise ValueError(f"ไฟล์ไม่มีคอลัมน์ 'Material Cost'")

            # ตั้งค่า index
            new_df = new_df.set_index("MENU NAME")

            # อัพเดทข้อมูล
            self.base_files[base_type] = new_df

            # บันทึกไฟล์ Base Cost.xlsx
            self.save_base_cost_file()

            logger.info(f"อัพเดท {base_type.upper()} Cost สำเร็จ: {len(new_df)} เมนู")
            return True

        except Exception as e:
            logger.error(f"Error updating {base_type} base: {e}")
            raise e


# ===== Modern GUI Styling =====
def configure_modern_style():
    """กำหนดธีมสีสวยงามและทันสมัย"""
    style = ttk.Style()

    try:
        style.theme_use('clam')
    except:
        style.theme_use('default')

    colors = {
        'bg_primary': '#2c3e50',
        'bg_secondary': '#34495e',
        'accent': '#3498db',
        'success': '#27ae60',
        'warning': '#f39c12',
        'danger': '#e74c3c',
        'light': '#ecf0f1',
        'white': '#ffffff',
        'text_dark': '#2c3e50',
        'text_light': '#ffffff'
    }

    # Configure styles
    style.configure('Title.TLabel', font=('Segoe UI', 16, 'bold'), foreground=colors['text_dark'])
    style.configure('Heading.TLabel', font=('Segoe UI', 12, 'bold'), foreground=colors['accent'])
    style.configure('Status.TLabel', font=('Segoe UI', 10), padding=(5, 2))

    style.configure('Accent.TButton', font=('Segoe UI', 9, 'bold'), padding=(10, 5))
    style.configure('Success.TButton', font=('Segoe UI', 9, 'bold'), padding=(10, 5))
    style.configure('Warning.TButton', font=('Segoe UI', 9, 'bold'), padding=(10, 5))
    style.configure('Danger.TButton', font=('Segoe UI', 9, 'bold'), padding=(10, 5))

    style.configure('Card.TFrame', relief='solid', borderwidth=1, padding=10)
    style.configure('Modern.TNotebook', tabposition='n')
    style.configure('Modern.TNotebook.Tab', font=('Segoe UI', 10, 'bold'), padding=(20, 8))

    return colors


# ===== Logging Setup =====
logging.basicConfig(
    level=logging.DEBUG,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('hamada_calculator_debug.log', encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# Initialize configuration
config_manager = ConfigManager()


class EnhancedCostCalculatorApp:
    def __init__(self, root):
        self.root = root
        self.colors = configure_modern_style()

        self.root.title("🍱 Hamada & Hashira Cost Calculator - FIXED")
        self.root.geometry("1500x950")
        self.root.configure(bg=self.colors['light'])
        self.root.minsize(1300, 850)

        # ตัวแปรสำหรับไฟล์
        self.import_file = None
        self.template_file = None
        self.selected_base_type = tk.StringVar(value='hamada')
        self.df_result = None
        self.debug_mode = tk.BooleanVar(value=False)

        # ข้อมูล debug แยกตามประเภท
        self.debug_data = {
            'not_found_menus': [],
            'zero_qty_items': [],
            'invalid_qty_items': [],
            'nan_cost_items': [],
            'matched_menus': [],
            'not_sold_menus': [],
            'processing_summary': {}
        }

        self._build_ui()
        logger.info("Enhanced Cost Calculator initialized with built-in data")

    def _build_ui(self):
        """สร้าง UI แบบ Enhanced"""
        # Main container
        main_frame = ttk.Frame(self.root, style='Card.TFrame')
        main_frame.pack(fill=tk.BOTH, expand=True, padx=15, pady=15)

        # Enhanced Header
        header_frame = ttk.Frame(main_frame)
        header_frame.pack(fill=tk.X, pady=(0, 20))

        title_label = ttk.Label(header_frame, text="🍱 Hamada & Hashira Cost Calculator",
                                style='Title.TLabel')
        title_label.pack(pady=(0, 5))

        # สถานะ libraries
        libs_status = "✅ Excel Export Ready" if OPENPYXL_AVAILABLE or XLSXWRITER_AVAILABLE else "❌ Excel Libraries Missing"
        subtitle = ttk.Label(header_frame, text=f"Enhanced Multi-Base Cost Management System - {libs_status}",
                             font=('Segoe UI', 11), foreground=self.colors['bg_secondary'])
        subtitle.pack()

        # ===== Base Type Selection Card =====
        base_selection_frame = ttk.LabelFrame(main_frame, text="🏪 Base Cost Selection",
                                              padding=15, style='Card.TFrame')
        base_selection_frame.pack(fill=tk.X, pady=(0, 15))

        base_frame = ttk.Frame(base_selection_frame)
        base_frame.pack(fill=tk.X)

        ttk.Label(base_frame, text="เลือกประเภท Base Cost:",
                  font=('Segoe UI', 11, 'bold')).pack(side=tk.LEFT, padx=(0, 20))

        # Radio buttons
        hashira_radio = ttk.Radiobutton(base_frame, text="🏯 Hashira Cost",
                                        variable=self.selected_base_type, value='hashira',
                                        command=self.on_base_type_change)
        hashira_radio.pack(side=tk.LEFT, padx=10)

        hamada_radio = ttk.Radiobutton(base_frame, text="🍜 Hamada Cost",
                                       variable=self.selected_base_type, value='hamada',
                                       command=self.on_base_type_change)
        hamada_radio.pack(side=tk.LEFT, padx=10)

        self.base_status_label = ttk.Label(base_frame, text="", style='Status.TLabel')
        self.base_status_label.pack(side=tk.RIGHT, padx=20)

        # ===== File Management Section =====
        file_frame = ttk.LabelFrame(main_frame, text="📁 File Management",
                                    padding=15, style='Card.TFrame')
        file_frame.pack(fill=tk.X, pady=(0, 15))

        # Import buttons
        import_frame = ttk.Frame(file_frame)
        import_frame.pack(fill=tk.X, pady=(0, 10))

        ttk.Button(import_frame, text="📁 เลือกไฟล์ Import",
                   command=self.load_import_file, style='Accent.TButton').pack(
            side=tk.LEFT, padx=5, fill=tk.X, expand=True)

        ttk.Button(import_frame, text="📄 Import Template เพิ่มเติม",
                   command=self.load_template_file, style='Warning.TButton').pack(
            side=tk.LEFT, padx=5, fill=tk.X, expand=True)

        ttk.Button(import_frame, text="🔍 ตรวจสอบไฟล์ Base",
                   command=self.check_base_file, style='Accent.TButton').pack(
            side=tk.LEFT, padx=5, fill=tk.X, expand=True)

        # File status display
        status_frame = ttk.Frame(file_frame)
        status_frame.pack(fill=tk.X)

        self.import_status_label = ttk.Label(status_frame, text="📂 Import: ยังไม่ได้เลือกไฟล์",
                                             style='Status.TLabel', foreground=self.colors['danger'])
        self.import_status_label.pack(anchor="w", pady=2)

        self.template_status_label = ttk.Label(status_frame, text="📄 Template: ยังไม่ได้เลือกไฟล์",
                                               style='Status.TLabel', foreground=self.colors['warning'])
        self.template_status_label.pack(anchor="w", pady=2)

        self.hashira_status_label = ttk.Label(status_frame, text="🏯 Hashira Cost: กำลังตรวจสอบ...",
                                              style='Status.TLabel', foreground=self.colors['warning'])
        self.hashira_status_label.pack(anchor="w", pady=2)

        self.hamada_status_label = ttk.Label(status_frame, text="🍜 Hamada Cost: กำลังตรวจสอบ...",
                                             style='Status.TLabel', foreground=self.colors['warning'])
        self.hamada_status_label.pack(anchor="w", pady=2)

        # ===== Action Buttons =====
        action_frame = ttk.LabelFrame(main_frame, text="🚀 Actions",
                                      padding=15, style='Card.TFrame')
        action_frame.pack(fill=tk.X, pady=(0, 15))

        # Debug mode
        debug_frame = ttk.Frame(action_frame)
        debug_frame.pack(fill=tk.X, pady=(0, 10))

        self.debug_checkbox = ttk.Checkbutton(debug_frame, text="🔍 เปิด Debug Mode",
                                              variable=self.debug_mode)
        self.debug_checkbox.pack(side=tk.LEFT)

        ttk.Button(debug_frame, text="🗑️ เคลียร์ Debug",
                   command=self.clear_debug, style='Warning.TButton').pack(side=tk.RIGHT)

        # Calculate button
        calc_frame = ttk.Frame(action_frame)
        calc_frame.pack(fill=tk.X, pady=(0, 10))

        ttk.Button(calc_frame, text="⚡ คำนวณ", command=self.calculate,
                   style='Success.TButton').pack(fill=tk.X)

        # Export buttons - แก้ไขแล้ว
        export_frame = ttk.Frame(action_frame)
        export_frame.pack(fill=tk.X)

        ttk.Button(export_frame, text="💾 บันทึก Excel (Auto)",
                   command=self.export_excel_enhanced,
                   style='Success.TButton').pack(side=tk.LEFT, padx=2, fill=tk.X, expand=True)

        ttk.Button(export_frame, text="🔍💾 บันทึก Debug (Auto)",
                   command=self.export_debug_enhanced,
                   style='Warning.TButton').pack(side=tk.LEFT, padx=2, fill=tk.X, expand=True)

        ttk.Button(export_frame, text="📤 Export Base (Auto)",
                   command=self.export_base_enhanced,
                   style='Accent.TButton').pack(side=tk.LEFT, padx=2, fill=tk.X, expand=True)

        # ===== Content Area =====
        content_frame = ttk.Frame(main_frame, style='Card.TFrame')
        content_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 15))

        self.notebook = ttk.Notebook(content_frame, style='Modern.TNotebook')
        self.notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # Results Tab
        result_frame = ttk.Frame(self.notebook)
        self.notebook.add(result_frame, text="📋 ผลลัพธ์")

        table_frame = ttk.Frame(result_frame, padding=10)
        table_frame.pack(fill=tk.BOTH, expand=True)

        table_header = ttk.Label(table_frame, text="💰 Calculation Results",
                                 style='Heading.TLabel')
        table_header.pack(pady=(0, 10))

        # Results table
        self.tree = ttk.Treeview(table_frame, columns=("menu", "qty", "cost", "total"),
                                 show="headings", height=16)

        self.tree.heading("menu", text="🍽️ MENU NAME")
        self.tree.heading("qty", text="📊 Qty")
        self.tree.heading("cost", text="💵 Material Cost")
        self.tree.heading("total", text="💰 Total Cost")

        self.tree.column("menu", width=400, anchor="w")
        self.tree.column("qty", width=100, anchor="center")
        self.tree.column("cost", width=150, anchor="e")
        self.tree.column("total", width=150, anchor="e")

        tree_scroll = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=tree_scroll.set)

        tree_scroll.pack(side=tk.RIGHT, fill=tk.Y, padx=(5, 0))
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # Debug Tab
        debug_frame = ttk.Frame(self.notebook)
        self.notebook.add(debug_frame, text="🔍 Debug Log")

        debug_header_frame = ttk.Frame(debug_frame, padding=10)
        debug_header_frame.pack(fill=tk.X)

        debug_title = ttk.Label(debug_header_frame, text="🔧 Enhanced Debug Console",
                                style='Heading.TLabel')
        debug_title.pack()

        debug_text_frame = ttk.Frame(debug_frame, padding=(10, 0, 10, 10))
        debug_text_frame.pack(fill=tk.BOTH, expand=True)

        self.debug_text = scrolledtext.ScrolledText(debug_text_frame, wrap=tk.WORD, height=18,
                                                    font=('Consolas', 10),
                                                    bg='#1e1e1e', fg='#ffffff',
                                                    insertbackground='#ffffff',
                                                    selectbackground='#3498db',
                                                    relief='flat', borderwidth=0)
        self.debug_text.pack(fill=tk.BOTH, expand=True)

        # Statistics Section
        self.stats_frame = ttk.LabelFrame(main_frame, text="📈 Statistics & Info",
                                          padding=15, style='Card.TFrame')
        self.stats_frame.pack(fill=tk.X)

        self.stats_label = ttk.Label(self.stats_frame, text="ยังไม่มีข้อมูล",
                                     style='Status.TLabel',
                                     font=('Segoe UI', 11, 'bold'))
        self.stats_label.pack()

        self._configure_treeview_style()
        self.update_all_base_status()
        self.update_base_status()

    def _configure_treeview_style(self):
        """กำหนดสไตล์ modern สำหรับ treeview"""
        style = ttk.Style()
        style.configure('Modern.Treeview',
                        background='#ffffff',
                        foreground='#2c3e50',
                        fieldbackground='#ffffff',
                        font=('Segoe UI', 10))
        style.configure('Modern.Treeview.Heading',
                        background='#3498db',
                        foreground='#ffffff',
                        font=('Segoe UI', 10, 'bold'),
                        relief='flat')
        style.map('Modern.Treeview',
                  background=[('selected', '#3498db')],
                  foreground=[('selected', '#ffffff')])
        self.tree.configure(style='Modern.Treeview')

    def on_base_type_change(self):
        """เมื่อเปลี่ยน base type"""
        selected = self.selected_base_type.get()
        self.debug_log(f"เปลี่ยน Base Type เป็น: {selected.upper()}")
        self.update_base_status()

    def update_base_status(self):
        """อัพเดทสถานะไฟล์ base ที่เลือก"""
        selected = self.selected_base_type.get()
        base_df = config_manager.base_files.get(selected)

        if base_df is not None and not base_df.empty:
            menu_count = len(base_df)
            status_text = f"🗂️ {selected.upper()} Cost: พร้อมใช้งาน ({menu_count} เมนู)"
            color = self.colors['success']
        else:
            status_text = f"🗂️ {selected.upper()} Cost: ไม่พบข้อมูล"
            color = self.colors['danger']

        self.base_status_label.config(text=status_text, foreground=color)

    def update_all_base_status(self):
        """อัพเดทสถานะไฟล์ base ทั้งหมด"""
        # Hashira Status
        hashira_df = config_manager.base_files.get('hashira')
        if hashira_df is not None and not hashira_df.empty:
            hashira_count = len(hashira_df)
            hashira_text = f"🏯 Hashira Cost: พร้อมใช้งาน ({hashira_count} เมนู) [Real Data]"
            hashira_color = self.colors['success']
        else:
            hashira_text = f"🏯 Hashira Cost: ไม่พบข้อมูล"
            hashira_color = self.colors['danger']

        # Hamada Status
        hamada_df = config_manager.base_files.get('hamada')
        if hamada_df is not None and not hamada_df.empty:
            hamada_count = len(hamada_df)
            hamada_text = f"🍜 Hamada Cost: พร้อมใช้งาน ({hamada_count} เมนู) [Real Data]"
            hamada_color = self.colors['success']
        else:
            hamada_text = f"🍜 Hamada Cost: ไม่พบข้อมูล"
            hamada_color = self.colors['danger']

        self.hashira_status_label.config(text=hashira_text, foreground=hashira_color)
        self.hamada_status_label.config(text=hamada_text, foreground=hamada_color)

    def debug_log(self, message, level="INFO"):
        """เพิ่มข้อความใน debug log"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        log_message = f"[{timestamp}] {level}: {message}\n"

        self.debug_text.insert(tk.END, log_message)
        self.debug_text.see(tk.END)

        if level == "ERROR":
            logger.error(message)
        elif level == "WARNING":
            logger.warning(message)
        else:
            logger.info(message)

    def clear_debug(self):
        """เคลียร์ debug log"""
        self.debug_text.delete(1.0, tk.END)
        self.debug_log("Debug log cleared")
        self.debug_data = {
            'not_found_menus': [],
            'zero_qty_items': [],
            'invalid_qty_items': [],
            'nan_cost_items': [],
            'matched_menus': [],
            'not_sold_menus': [],
            'processing_summary': {}
        }

    def load_import_file(self):
        """โหลดไฟล์ import - แก้ไขแล้ว"""
        try:
            file_path = filedialog.askopenfilename(
                title="เลือกไฟล์ Import",
                filetypes=[
                    ("Excel Files", "*.xlsx"),
                    ("Excel Files", "*.xls"),
                    ("All Files", "*.*")
                ]
            )

            if not file_path:
                return

            # ตรวจสอบไฟล์
            if not os.path.exists(file_path):
                messagebox.showerror("Error", "ไฟล์ที่เลือกไม่พบ")
                return

            self.import_file = file_path
            self.import_status_label.config(
                text=f"📂 Import: {os.path.basename(file_path)}",
                foreground=self.colors['success']
            )

            self.debug_log(f"เลือกไฟล์ Import: {file_path}")
            self._preview_file(file_path, "Import")

        except Exception as e:
            error_msg = f"Error เลือกไฟล์: {str(e)}"
            self.debug_log(error_msg, "ERROR")
            messagebox.showerror("Error", error_msg)

    def load_template_file(self):
        """โหลดไฟล์ template - แก้ไขแล้ว"""
        try:
            file_path = filedialog.askopenfilename(
                title="เลือกไฟล์ Template",
                filetypes=[
                    ("Excel Files", "*.xlsx"),
                    ("Excel Files", "*.xls"),
                    ("All Files", "*.*")
                ]
            )

            if not file_path:
                return

            if not os.path.exists(file_path):
                messagebox.showerror("Error", "ไฟล์ที่เลือกไม่พบ")
                return

            self.template_file = file_path
            self.template_status_label.config(
                text=f"📄 Template: {os.path.basename(file_path)}",
                foreground=self.colors['success']
            )

            self.debug_log(f"เลือกไฟล์ Template: {file_path}")
            self._preview_file(file_path, "Template")

        except Exception as e:
            error_msg = f"Error เลือกไฟล์: {str(e)}"
            self.debug_log(error_msg, "ERROR")
            messagebox.showerror("Error", error_msg)

    def _preview_file(self, file_path, file_type):
        """แสดงตัวอย่างข้อมูลในไฟล์ - แก้ไขแล้ว"""
        try:
            # ตรวจสอบไฟล์ก่อนอ่าน
            if not os.path.exists(file_path):
                self.debug_log(f"ไฟล์ไม่พบ: {file_path}", "ERROR")
                return

            df_preview = pd.read_excel(file_path, nrows=10)  # อ่านแค่ 10 แถวแรก
            self.debug_log(f"ไฟล์ {file_type} มี {len(df_preview)} แถว (preview), {len(df_preview.columns)} คอลัมน์")
            self.debug_log(f"คอลัมน์: {df_preview.columns.tolist()}")

            # ตรวจสอบคอลัมน์ที่จำเป็น
            required_columns = ["MENU NAME", "Qty"]
            missing_columns = [col for col in required_columns if col not in df_preview.columns]

            if missing_columns:
                self.debug_log(f"⚠️ คอลัมน์ที่หายไป: {missing_columns}", "WARNING")
            else:
                self.debug_log("✅ คอลัมน์ครบถ้วน")

            # แสดงตัวอย่างข้อมูล
            self.debug_log(f"ตัวอย่างข้อมูล {file_type} 5 แถวแรก:")
            for i, row in df_preview.head(5).iterrows():
                menu = row.get("MENU NAME", "N/A")
                qty = row.get("Qty", "N/A")
                self.debug_log(f"  แถว {i + 1}: {menu} = {qty}")

        except Exception as e:
            self.debug_log(f"Error ตรวจสอบไฟล์ {file_type}: {str(e)}", "ERROR")

    def check_base_file(self):
        """ตรวจสอบและแสดงข้อมูลไฟล์ base ที่เลือก"""
        selected = self.selected_base_type.get()
        df_base = config_manager.base_files.get(selected)

        if df_base is None:
            messagebox.showerror("Error", f"ไม่สามารถโหลด {selected.upper()} Cost ได้")
            return

        self.debug_log(f"=== ตรวจสอบ {selected.upper()} Cost ===")
        self.debug_log(f"จำนวนเมนู: {len(df_base)}")
        self.debug_log(f"คอลัมน์: {df_base.columns.tolist()}")

        # สถิติราคา
        try:
            if "Material Cost" in df_base.columns:
                cost_stats = pd.to_numeric(df_base["Material Cost"], errors='coerce').describe()
                self.debug_log("สถิติราคา Material Cost:")
                self.debug_log(f"  ต่ำสุด: {cost_stats['min']:.2f} บาท")
                self.debug_log(f"  สูงสุด: {cost_stats['max']:.2f} บาท")
                self.debug_log(f"  เฉลี่ย: {cost_stats['mean']:.2f} บาท")
                self.debug_log(f"  จำนวนรายการ: {int(cost_stats['count'])}")

            # แสดงตัวอย่างเมนู 10 รายการแรก
            self.debug_log(f"ตัวอย่างเมนู 10 รายการแรก:")
            for i, (menu_name, row) in enumerate(df_base.head(10).iterrows()):
                cost = row.get('Material Cost', 'N/A')
                self.debug_log(f"  {i + 1}. {menu_name}: {cost} บาท")

        except Exception as e:
            self.debug_log(f"Error คำนวณสถิติ: {str(e)}", "ERROR")

    def calculate(self):
        """คำนวณต้นทุน - แก้ไขแล้ว"""
        df_base = config_manager.base_files.get(self.selected_base_type.get())
        selected_base = self.selected_base_type.get()

        if df_base is None or df_base.empty:
            messagebox.showerror("Error", f"ไฟล์ {selected_base.upper()} Cost ไม่สามารถโหลดได้")
            return

        if not self.import_file:
            messagebox.showwarning("Warning", "กรุณาเลือกไฟล์ Import ก่อน")
            return

        if not os.path.exists(self.import_file):
            messagebox.showerror("Error", "ไฟล์ Import ที่เลือกไม่พบ")
            return

        self.debug_log(f"=== เริ่มการคำนวณด้วย {selected_base.upper()} Cost ===")

        try:
            # อ่านไฟล์ import
            self.debug_log(f"กำลังอ่านไฟล์: {self.import_file}")
            df_import = pd.read_excel(self.import_file)
            self.debug_log(f"อ่านไฟล์ import สำเร็จ: {len(df_import)} แถว")

            # ทำความสะอาดชื่อคอลัมน์
            df_base.columns = df_base.columns.str.strip()
            df_import.columns = df_import.columns.str.strip()

            self.debug_log(f"คอลัมน์ import: {df_import.columns.tolist()}")
            self.debug_log(f"คอลัมน์ base: {df_base.columns.tolist()}")

            # ตรวจสอบคอลัมน์ที่จำเป็น
            required_columns = ["MENU NAME", "Qty"]
            missing_columns = [col for col in required_columns if col not in df_import.columns]

            if missing_columns:
                error_msg = f"ไฟล์ Import ขาดคอลัมน์: {missing_columns}"
                self.debug_log(error_msg, "ERROR")
                messagebox.showerror("Error", error_msg)
                return

        except Exception as e:
            error_msg = f"ไม่สามารถอ่านไฟล์ได้: {str(e)}"
            self.debug_log(error_msg, "ERROR")
            messagebox.showerror("Error", error_msg)
            return

        # เริ่มคำนวณ
        results = []
        matched_count = 0
        not_found_count = 0
        invalid_qty_count = 0

        self.debug_log("เริ่มประมวลผลข้อมูล...")

        for idx, row in df_import.iterrows():
            menu = row.get("MENU NAME")
            qty = row.get("Qty", 0)

            # ข้าม row ที่ไม่มีชื่อเมนู
            if pd.isna(menu) or menu == "":
                continue

            # แปลง quantity เป็นตัวเลข
            try:
                if pd.isna(qty) or qty == "" or qty == 0:
                    qty = 0
                    invalid_qty_count += 1
                    if self.debug_mode.get():
                        self.debug_log(f"⚠️ Qty = 0 สำหรับ: {menu}", "WARNING")
                else:
                    qty = float(qty)
            except (ValueError, TypeError):
                qty = 0
                invalid_qty_count += 1
                if self.debug_mode.get():
                    self.debug_log(f"⚠️ Qty ไม่ถูกต้องสำหรับ: {menu}", "WARNING")

            # หาเมนูใน base cost
            if menu in df_base.index:
                try:
                    material_cost = df_base.at[menu, "Material Cost"]

                    # ตรวจสอบและแปลง material cost
                    if pd.isna(material_cost):
                        material_cost = 0
                        self.debug_log(f"⚠️ Material Cost = NaN สำหรับ: {menu}", "WARNING")
                    else:
                        try:
                            material_cost = float(material_cost)
                        except (ValueError, TypeError):
                            material_cost = 0
                            self.debug_log(f"⚠️ Material Cost ไม่ถูกต้องสำหรับ: {menu}", "WARNING")

                    total_cost = qty * material_cost
                    results.append([menu, qty, material_cost, total_cost])
                    matched_count += 1

                    if self.debug_mode.get():
                        self.debug_log(f"✓ พบเมนู: {menu} = {material_cost:.2f} x {qty} = {total_cost:.2f}")

                except Exception as e:
                    self.debug_log(f"Error คำนวณ {menu}: {str(e)}", "ERROR")
            else:
                not_found_count += 1
                if self.debug_mode.get():
                    self.debug_log(f"❌ ไม่พบเมนู: {menu}", "WARNING")

        # รายงานผลการประมวลผล
        self.debug_log(f"=== สรุปผลการประมวลผล ===")
        self.debug_log(f"✅ พบเมนู: {matched_count}")
        self.debug_log(f"❌ ไม่พบเมนู: {not_found_count}")
        self.debug_log(f"⚠️ Qty ไม่ถูกต้อง: {invalid_qty_count}")

        # สร้าง DataFrame ผลลัพธ์
        if not results:
            error_msg = f"ไม่มีชื่อเมนูที่ตรงกับ {selected_base.upper()} Cost"
            self.debug_log(error_msg, "WARNING")
            messagebox.showwarning("ไม่พบข้อมูล", error_msg)
            return

        self.df_result = pd.DataFrame(results, columns=["MENU NAME", "Qty", "Material Cost", "Total Cost"])

        # คำนวณ Grand Total
        grand_total = self.df_result["Total Cost"].sum()
        self.debug_log(f"💰 Grand Total: {grand_total:,.2f} บาท")

        # เพิ่ม Grand Total row
        grand_total_row = pd.DataFrame([["== GRAND TOTAL ==", "", "", grand_total]],
                                       columns=["MENU NAME", "Qty", "Material Cost", "Total Cost"])
        self.df_result = pd.concat([self.df_result, grand_total_row], ignore_index=True)

        # แสดงผล
        self._update_result_table()
        self._update_statistics(matched_count, grand_total, selected_base)

        success_msg = f"คำนวณเสร็จสิ้น - พบ {matched_count} รายการ (ใช้ {selected_base.upper()} Cost)"
        self.debug_log(success_msg)
        messagebox.showinfo("Success", success_msg)

    def _update_result_table(self):
        """อัพเดทตารางผลลัพธ์"""
        # เคลียร์ข้อมูลเก่า
        for row in self.tree.get_children():
            self.tree.delete(row)

        if self.df_result is None or self.df_result.empty:
            return

        # เพิ่มข้อมูลใหม่
        for _, r in self.df_result.iterrows():
            try:
                # จัดรูปแบบการแสดงผล
                menu_name = r["MENU NAME"]

                if r["Qty"] == "" or pd.isna(r["Qty"]):
                    qty_display = ""
                else:
                    qty_display = f"{float(r['Qty']):g}"  # ใช้ :g เพื่อลบ .0 ที่ไม่จำเป็น

                if r["Material Cost"] == "" or pd.isna(r["Material Cost"]):
                    cost_display = ""
                else:
                    cost_display = f"{float(r['Material Cost']):.2f}"

                if pd.isna(r["Total Cost"]):
                    total_display = "0.00"
                else:
                    total_display = f"{float(r['Total Cost']):.2f}"

                # กำหนด tags
                tags = ()
                if "GRAND TOTAL" in str(menu_name).upper():
                    tags = ("grand_total",)

                self.tree.insert("", tk.END,
                                 values=(menu_name, qty_display, cost_display, total_display),
                                 tags=tags)
            except Exception as e:
                self.debug_log(f"Error แสดงผล row: {e}", "ERROR")

        # จัดรูปแบบ Grand Total
        self.tree.tag_configure("grand_total",
                                background="#3498db",
                                foreground="#ffffff",
                                font=("Segoe UI", 10, "bold"))

    def _update_statistics(self, matched, grand_total, base_type):
        """อัพเดทสถิติ"""
        stats_text = (f"✅ พบเมนู: {matched} | 💰 รวม: {grand_total:,.2f} บาท | 🏪 Base: {base_type.upper()}")
        self.stats_label.config(text=stats_text)

    # ===== ฟังก์ชันบันทึก Excel แก้ไขใหม่ =====
    def export_excel_enhanced(self):
        """บันทึก Excel แบบ Enhanced - ใช้ได้แน่นอน"""
        self.debug_log("🔄 เริ่มบันทึก Excel แบบ Enhanced...")

        if self.df_result is None or self.df_result.empty:
            messagebox.showwarning("Warning", "กรุณาคำนวณก่อนบันทึก")
            return

        selected_base = self.selected_base_type.get()
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"Cost_Calculation_{selected_base.upper()}_{timestamp}.xlsx"
        file_path = os.path.join(os.getcwd(), filename)

        self.debug_log(f"📂 กำลังบันทึกไฟล์: {filename}")
        self.debug_log(f"📍 ตำแหน่งไฟล์: {file_path}")

        try:
            # วิธีที่ 1: ใช้ openpyxl (ถ้ามี)
            if OPENPYXL_AVAILABLE:
                self.debug_log("ใช้ openpyxl engine")
                success = self._export_with_openpyxl(file_path, selected_base)
                if success:
                    self._show_export_success_dialog(filename, file_path, selected_base)
                    return

            # วิธีที่ 2: ใช้ xlsxwriter (ถ้ามี)
            if XLSXWRITER_AVAILABLE:
                self.debug_log("ใช้ xlsxwriter engine")
                success = self._export_with_xlsxwriter(file_path, selected_base)
                if success:
                    self._show_export_success_dialog(filename, file_path, selected_base)
                    return

            # วิธีที่ 3: ใช้ pandas default
            self.debug_log("ใช้ pandas default engine")
            success = self._export_with_pandas(file_path, selected_base)
            if success:
                self._show_export_success_dialog(filename, file_path, selected_base)
                return

            # ถ้าทุกวิธีล้มเหลว
            raise Exception("ทุกวิธีการ export ล้มเหลว")

        except Exception as e:
            error_msg = f"Error บันทึกไฟล์: {str(e)}"
            self.debug_log(error_msg, "ERROR")
            self.debug_log(f"Traceback: {traceback.format_exc()}", "ERROR")

            # แสดง error dialog พร้อมคำแนะนำ
            error_dialog_msg = f"""❌ ไม่สามารถบันทึกไฟล์ Excel ได้

🔍 สาเหตุที่เป็นไปได้:
• ไฟล์ถูกเปิดอยู่ในโปรแกรมอื่น
• ไม่มีสิทธิ์เขียนไฟล์ในโฟลเดอร์นี้
• Library สำหรับ Excel ขาดหายไป

💡 วิธีแก้ไข:
1. ปิดไฟล์ Excel ที่อาจเปิดอยู่
2. เรียกใช้โปรแกรมในฐานะ Administrator
3. ลองบันทึกในโฟลเดอร์อื่น

🛠️ Error Details:
{str(e)}"""

            messagebox.showerror("Export Error", error_dialog_msg)

    def _export_with_openpyxl(self, file_path, selected_base):
        """บันทึกด้วย openpyxl"""
        try:
            # ตรวจสอบและทำความสะอาดข้อมูล
            df_clean = self.df_result.copy()

            # แปลงข้อมูลให้เป็นรูปแบบที่ถูกต้อง
            for col in ['Qty', 'Material Cost', 'Total Cost']:
                if col in df_clean.columns:
                    df_clean[col] = pd.to_numeric(df_clean[col], errors='coerce').fillna(0)

            wb = Workbook()
            ws = wb.active
            ws.title = "Cost Calculation"

            # เขียน header
            headers = df_clean.columns.tolist()
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

            # เขียนข้อมูล
            for row_num, (_, row) in enumerate(df_clean.iterrows(), 2):
                for col_num, value in enumerate(row, 1):
                    cell = ws.cell(row=row_num, column=col_num, value=value)

                    # ไฮไลท์ Grand Total
                    if "GRAND TOTAL" in str(row.iloc[0]).upper():
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")

            # ปรับความกว้างคอลัมน์
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter

                for cell in column:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass

                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            wb.save(file_path)
            self.debug_log(f"✅ บันทึกด้วย openpyxl สำเร็จ")
            return True

        except Exception as e:
            self.debug_log(f"Error openpyxl export: {str(e)}", "ERROR")
            return False

    def _export_with_xlsxwriter(self, file_path, selected_base):
        """บันทึกด้วย xlsxwriter"""
        try:
            # ตรวจสอบและทำความสะอาดข้อมูล
            df_clean = self.df_result.copy()

            # แปลงข้อมูลให้เป็นรูปแบบที่ถูกต้อง
            for col in ['Qty', 'Material Cost', 'Total Cost']:
                if col in df_clean.columns:
                    df_clean[col] = pd.to_numeric(df_clean[col], errors='coerce').fillna(0)

            with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
                df_clean.to_excel(writer, sheet_name='Cost Calculation', index=False)

                # จัดรูปแบบ
                workbook = writer.book
                worksheet = writer.sheets['Cost Calculation']

                # Header format
                header_format = workbook.add_format({
                    'bold': True,
                    'bg_color': '#CCCCCC',
                    'border': 1
                })

                # Grand total format
                grand_total_format = workbook.add_format({
                    'bold': True,
                    'bg_color': '#FFFF99',
                    'border': 1
                })

                # Apply header format
                for col_num, header in enumerate(df_clean.columns):
                    worksheet.write(0, col_num, header, header_format)

                # Apply grand total format
                for row_num, (_, row) in enumerate(df_clean.iterrows(), 1):
                    if "GRAND TOTAL" in str(row.iloc[0]).upper():
                        for col_num, value in enumerate(row):
                            worksheet.write(row_num, col_num, value, grand_total_format)

                # Auto-fit columns
                for col_num, header in enumerate(df_clean.columns):
                    worksheet.set_column(col_num, col_num, 15)

            self.debug_log(f"✅ บันทึกด้วย xlsxwriter สำเร็จ")
            return True

        except Exception as e:
            self.debug_log(f"Error xlsxwriter export: {str(e)}", "ERROR")
            return False

    def _export_with_pandas(self, file_path, selected_base):
        """บันทึกด้วย pandas default"""
        try:
            # ตรวจสอบและทำความสะอาดข้อมูล
            df_clean = self.df_result.copy()

            # แปลงข้อมูลให้เป็นรูปแบบที่ถูกต้อง
            for col in ['Qty', 'Material Cost', 'Total Cost']:
                if col in df_clean.columns:
                    df_clean[col] = pd.to_numeric(df_clean[col], errors='coerce').fillna(0)

            # บันทึกแบบ basic
            df_clean.to_excel(file_path, index=False, sheet_name='Cost Calculation')

            self.debug_log(f"✅ บันทึกด้วย pandas default สำเร็จ")
            return True

        except Exception as e:
            self.debug_log(f"Error pandas export: {str(e)}", "ERROR")
            return False

    def _show_export_success_dialog(self, filename, file_path, selected_base):
        """แสดง dialog สำเร็จ"""
        record_count = len(self.df_result) if self.df_result is not None else 0
        grand_total = self.df_result[
            "Total Cost"].sum() if self.df_result is not None and not self.df_result.empty else 0

        success_msg = f"""✅ บันทึกไฟล์เรียบร้อย!

📁 ชื่อไฟล์: {filename}
📍 ตำแหน่ง: {file_path}
📊 จำนวนข้อมูล: {record_count} รายการ
💰 ยอดรวม: {grand_total:,.2f} บาท
🏪 ใช้ Base: {selected_base.upper()} Cost

✅ ไฟล์พร้อมใช้งาน"""

        messagebox.showinfo("สำเร็จ!", success_msg)

    def export_debug_enhanced(self):
        """บันทึก Debug แบบ Enhanced"""
        self.debug_log("🔄 เริ่มบันทึก Debug แบบ Enhanced...")

        if self.df_result is None or self.df_result.empty:
            messagebox.showwarning("Warning", "กรุณาคำนวณก่อนบันทึก Debug")
            return

        selected_base = self.selected_base_type.get()
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"Cost_Debug_{selected_base.upper()}_{timestamp}.xlsx"
        file_path = os.path.join(os.getcwd(), filename)

        self.debug_log(f"📂 บันทึกไฟล์ Debug ที่: {file_path}")

        try:
            if OPENPYXL_AVAILABLE:
                # ใช้ openpyxl สำหรับ debug
                wb = Workbook()

                # ลบ default sheet
                if wb.worksheets:
                    wb.remove(wb.active)

                # Sheet 1: ผลลัพธ์
                ws_main = wb.create_sheet("ผลลัพธ์")
                df_clean = self.df_result.copy()

                # เขียนข้อมูลผลลัพธ์
                for r_idx, r in enumerate(dataframe_to_rows(df_clean, index=False, header=True)):
                    for c_idx, value in enumerate(r):
                        cell = ws_main.cell(row=r_idx + 1, column=c_idx + 1, value=value)
                        if r_idx == 0:  # Header
                            cell.font = Font(bold=True)
                            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

                # Sheet 2: Base Data ที่ใช้
                if config_manager.base_files.get(selected_base) is not None:
                    ws_base = wb.create_sheet(f"{selected_base.upper()} Cost")
                    df_base = config_manager.base_files[selected_base].reset_index()

                    for r_idx, r in enumerate(dataframe_to_rows(df_base, index=False, header=True)):
                        for c_idx, value in enumerate(r):
                            cell = ws_base.cell(row=r_idx + 1, column=c_idx + 1, value=value)
                            if r_idx == 0:  # Header
                                cell.font = Font(bold=True)
                                cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")

                # Sheet 3: Debug Info
                ws_debug = wb.create_sheet("Debug Info")
                debug_info = [
                    ["Debug Information", ""],
                    ["Timestamp", datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
                    ["Base Type Used", selected_base.upper()],
                    ["Total Records", len(df_clean)],
                    ["Grand Total", f"{df_clean['Total Cost'].sum():.2f} บาท"],
                    ["Import File", os.path.basename(self.import_file) if self.import_file else "N/A"],
                ]

                for r_idx, (key, value) in enumerate(debug_info, 1):
                    ws_debug.cell(row=r_idx, column=1, value=key).font = Font(bold=True)
                    ws_debug.cell(row=r_idx, column=2, value=value)

                # ปรับความกว้างคอลัมน์ทุก sheet
                for ws in wb.worksheets:
                    for column in ws.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if cell.value and len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 60)
                        ws.column_dimensions[column_letter].width = adjusted_width

                wb.save(file_path)

            else:
                # ใช้ pandas สำหรับ debug
                with pd.ExcelWriter(file_path, engine='openpyxl' if OPENPYXL_AVAILABLE else 'xlsxwriter') as writer:
                    self.df_result.to_excel(writer, sheet_name='ผลลัพธ์', index=False)

                    if config_manager.base_files.get(selected_base) is not None:
                        df_base = config_manager.base_files[selected_base].reset_index()
                        df_base.to_excel(writer, sheet_name=f'{selected_base.upper()} Cost', index=False)

            self.debug_log(f"✅ บันทึก Debug สำเร็จ: {filename}")

            success_msg = f"""✅ บันทึกไฟล์ Debug เรียบร้อย!

📁 ชื่อไฟล์: {filename}  
📍 ตำแหน่ง: {file_path}
📊 Sheets: ผลลัพธ์, {selected_base.upper()} Cost, Debug Info
🏪 Base Type: {selected_base.upper()}

✅ ไฟล์พร้อมใช้งาน"""

            messagebox.showinfo("สำเร็จ!", success_msg)
            return True

        except Exception as e:
            self.debug_log(f"Error debug export: {str(e)}", "ERROR")
            return False

    def export_base_enhanced(self):
        """Export Base Cost แบบ Enhanced"""
        self.debug_log("🔄 เริ่ม Export Base Cost แบบ Enhanced...")

        if not any(config_manager.base_files.values()):
            messagebox.showerror("Error", "ไม่มีข้อมูล Base Cost ให้ export")
            return

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"Base_Cost_Export_{timestamp}.xlsx"
        file_path = os.path.join(os.getcwd(), filename)

        self.debug_log(f"📂 Export Base ที่: {file_path}")

        try:
            if OPENPYXL_AVAILABLE:
                wb = Workbook()
                if wb.worksheets:
                    wb.remove(wb.active)

                sheets_created = 0

                # Export Hashira Cost
                if config_manager.base_files.get('hashira') is not None:
                    ws_hashira = wb.create_sheet("Hashira Cost")
                    df_hashira = config_manager.base_files['hashira'].reset_index()

                    for r_idx, r in enumerate(dataframe_to_rows(df_hashira, index=False, header=True)):
                        for c_idx, value in enumerate(r):
                            cell = ws_hashira.cell(row=r_idx + 1, column=c_idx + 1, value=value)
                            if r_idx == 0:
                                cell.font = Font(bold=True)
                                cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")

                    sheets_created += 1
                    self.debug_log(f"✅ Export Hashira: {len(df_hashira)} เมนู")

                # Export Hamada Cost
                if config_manager.base_files.get('hamada') is not None:
                    ws_hamada = wb.create_sheet("Hamada Cost")
                    df_hamada = config_manager.base_files['hamada'].reset_index()

                    for r_idx, r in enumerate(dataframe_to_rows(df_hamada, index=False, header=True)):
                        for c_idx, value in enumerate(r):
                            cell = ws_hamada.cell(row=r_idx + 1, column=c_idx + 1, value=value)
                            if r_idx == 0:
                                cell.font = Font(bold=True)
                                cell.fill = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")

                    sheets_created += 1
                    self.debug_log(f"✅ Export Hamada: {len(df_hamada)} เมนู")

                # ปรับความกว้างคอลัมน์
                for ws in wb.worksheets:
                    for column in ws.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if cell.value and len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 60)
                        ws.column_dimensions[column_letter].width = adjusted_width

                wb.save(file_path)

            else:
                # ใช้ pandas สำหรับ export
                with pd.ExcelWriter(file_path) as writer:
                    if config_manager.base_files.get('hashira') is not None:
                        df_hashira = config_manager.base_files['hashira'].reset_index()
                        df_hashira.to_excel(writer, sheet_name='Hashira Cost', index=False)
                        sheets_created += 1

                    if config_manager.base_files.get('hamada') is not None:
                        df_hamada = config_manager.base_files['hamada'].reset_index()
                        df_hamada.to_excel(writer, sheet_name='Hamada Cost', index=False)
                        sheets_created += 1

            hashira_count = len(config_manager.base_files['hashira']) if config_manager.base_files.get(
                'hashira') is not None else 0
            hamada_count = len(config_manager.base_files['hamada']) if config_manager.base_files.get(
                'hamada') is not None else 0

            self.debug_log(f"✅ Export Base สำเร็จ: {filename}")

            success_msg = f"""✅ Export ไฟล์ Base เรียบร้อย!

📁 ชื่อไฟล์: {filename}
📍 ตำแหน่ง: {file_path}  
📊 Sheets: {sheets_created}
🏯 Hashira: {hashira_count} เมนู
🍜 Hamada: {hamada_count} เมนู

✅ ไฟล์พร้อมใช้งาน"""

            messagebox.showinfo("สำเร็จ!", success_msg)
            return True

        except Exception as e:
            error_msg = f"Error export Base: {str(e)}"
            self.debug_log(error_msg, "ERROR")
            messagebox.showerror("Error", error_msg)
            return False

    def check_excel_libraries(self):
        """ตรวจสอบ Excel libraries ที่มีอยู่"""
        self.debug_log("=== ตรวจสอบ Excel Libraries ===")

        if OPENPYXL_AVAILABLE:
            self.debug_log("✅ openpyxl: พร้อมใช้งาน")
        else:
            self.debug_log("❌ openpyxl: ไม่พบ")

        if XLSXWRITER_AVAILABLE:
            self.debug_log("✅ xlsxwriter: พร้อมใช้งาน")
        else:
            self.debug_log("❌ xlsxwriter: ไม่พบ")

        try:
            import xlwt
            self.debug_log("✅ xlwt: พร้อมใช้งาน")
        except ImportError:
            self.debug_log("❌ xlwt: ไม่พบ")

        if not (OPENPYXL_AVAILABLE or XLSXWRITER_AVAILABLE):
            self.debug_log("⚠️ แนะนำให้ติดตั้ง: pip install openpyxl xlsxwriter", "WARNING")

    # Alternative export methods เผื่อ main methods ไม่ได้
    def export_excel_simple_fallback(self):
        """วิธีบันทึก Excel แบบ fallback"""
        if self.df_result is None or self.df_result.empty:
            messagebox.showwarning("Warning", "กรุณาคำนวณก่อนบันทึก")
            return

        selected_base = self.selected_base_type.get()
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

        try:
            # ลองบันทึกเป็น CSV ก่อน
            csv_filename = f"Cost_Calculation_{selected_base.upper()}_{timestamp}.csv"
            csv_path = os.path.join(os.getcwd(), csv_filename)

            self.df_result.to_csv(csv_path, index=False, encoding='utf-8-sig')
            self.debug_log(f"✅ บันทึก CSV สำเร็จ: {csv_filename}")

            # ลองแปลงเป็น Excel
            try:
                excel_filename = f"Cost_Calculation_{selected_base.upper()}_{timestamp}.xlsx"
                excel_path = os.path.join(os.getcwd(), excel_filename)

                # อ่าน CSV แล้วบันทึกเป็น Excel
                df_from_csv = pd.read_csv(csv_path)
                df_from_csv.to_excel(excel_path, index=False, sheet_name='Results')

                self.debug_log(f"✅ แปลง CSV เป็น Excel สำเร็จ: {excel_filename}")

                # ลบไฟล์ CSV
                os.remove(csv_path)

                messagebox.showinfo("สำเร็จ!", f"""✅ บันทึกไฟล์เรียบร้อย!

📁 ชื่อไฟล์: {excel_filename}
📍 ตำแหน่ง: {excel_path}
📊 จำนวนข้อมูล: {len(df_from_csv)} รายการ
🏪 ใช้ Base: {selected_base.upper()} Cost

✅ ไฟล์พร้อมใช้งาน""")

            except Exception as e2:
                # ถ้า Excel ไม่ได้ ให้ CSV
                self.debug_log(f"Excel ไม่ได้ ใช้ CSV: {str(e2)}", "WARNING")
                messagebox.showinfo("บันทึก CSV สำเร็จ", f"""📄 บันทึกเป็น CSV แทน:

📁 ชื่อไฟล์: {csv_filename}
📍 ตำแหน่ง: {csv_path}
📊 จำนวนข้อมูล: {len(self.df_result)} รายการ

💡 เปิดด้วย Excel ได้ปกติ""")

        except Exception as e:
            error_msg = f"Error บันทึกไฟล์: {str(e)}"
            self.debug_log(error_msg, "ERROR")
            messagebox.showerror("Error", error_msg)


def main():
    """ฟังก์ชันหลัก - แก้ไขแล้ว"""
    try:
        root = tk.Tk()
        app = EnhancedCostCalculatorApp(root)

        app.debug_log("=== Enhanced Cost Calculator เริ่มทำงาน ===")
        app.debug_log("✅ ข้อมูลจริง Hamada (87 เมนู) และ Hashira (68 เมนู) พร้อมใช้งาน")
        app.debug_log("💾 การบันทึกไฟล์ใช้ระบบ Enhanced (หลายวิธี)")
        app.debug_log("📂 ไฟล์จะถูกบันทึกในโฟลเดอร์เดียวกับโปรแกรม")

        # ตรวจสอบ libraries
        app.check_excel_libraries()

        root.mainloop()

    except Exception as e:
        logger.error(f"Critical error: {str(e)}")
        print(f"Error: {str(e)}")
        print(traceback.format_exc())
        messagebox.showerror("Error", f"เกิดข้อผิดพลาด: {str(e)}")


if __name__ == "__main__":
    main()