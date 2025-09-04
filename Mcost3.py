import pandas as pd
import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext
import tkinter.simpledialog
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import numbers, Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import logging
import traceback
import os


# ===== Enhanced Configuration with Built-in Data =====
class ConfigManager:
    def __init__(self):
        self.config_file = "mapcost_config.ini"
        self.base_files = {
            'hashira': None,
            'hamada': None
        }
        self.initialize_default_data()
        self.load_config()

    def initialize_default_data(self):
        """สร้างข้อมูล Hamada Cost และ Hashira Cost จากข้อมูลจริง"""

        # ข้อมูลจริงจาก Hamada Cost.xlsx
        hamada_menu_names = [
            "Ebi Tempura Set", "Salmon Tempura Set", "Ebi & Salmon Tempura Set",
            "Ebi Tempura Don Set", "Salmon Tempura Don Set", "Ebi & Salmon Tempura Don Set",
            "Ebi Tempura Udon", "Salmon Tempura Udon", "Ebi & Salmon Tempura Udon",
            "Yakiudon Buta", "Yakiudon Kaisen", "Ebi Fry Curry Rice",
            "Tonkatsu Curry Rice", "Salmonkatsu Curry Rice", "Ebi Tempura（1pc）",
            "Salmon Tempura（1pc）", "Takoyaki", "Sushi Yorokobi",
            "Sushi Tokujo", "Maguro Zanmai", "Salmon Zanmai",
            "Hamachi Zanmai", "Duo Sushi", "Salmon Mountain",
            "Salmon Sushi & Ikura Gunkan", "California Roll", "Salmon Roll",
            "Unagi Roll", "Hokkaido Hotate Roll", "Salmon Chirashi Sushi",
            "Salmon Tobiko Chirashi Sushi", "Barachirashi Sushi", "Salmon Don",
            "Salmon Ikura Don", "Maguro Don", "Negitoro Ikura Don",
            "Kaisen Don", "Unagi Don", "Salmon Sashimi",
            "Maguro Sashimi", "Hamachi Sashimi", "Hotate Sashimi",
            "Salmon Sushi (2pcs)", "Maguro Sushi (2pcs)", "Hamachi Sushi (2pcs)",
            "Tamago Sushi (2pcs)", "Ebi Sushi (2pcs)", "Ika Sushi (2pcs)",
            "Hokki Sushi (2pcs)", "Hotate Sushi (2pcs)", "Unagi Sushi (2pcs)",
            "Anago Sushi (2pcs)", "Negitoro Gunkan (2pcs)", "Tobiko Gunkan (2pcs)",
            "Ikura Gunkan (2pcs)", "GG California Roll", "GG Salmon Sushi",
            "GG Tamago & Ebi Sushi", "GG Ebi & Tobiko Don", "GG Kanikama & Tobiko Don",
            "GG Triple Mix Don", "GG Ebiten Don", "GG Salmon & Ebiten Don",
            "Coca-Cola Original", "Coke-Cola Zero Sugar", "Sprite",
            "Soda", "Mineral Water", "HBD Sparkling Water Lemon (No Sugar, No Calories)",
            "HBD Sparkling Water Honey Yuzu (No Sugar, No Calories)",
            "HBD Sparkling Water Peach (No Sugar, No Calories)",
            "HBD Sparkling Water Kyoho (No Sugar, No Calories)", "Asahi Beer",
            "Heineken Beer ", "Singha Beer", "Drinking Water",
            "Matcha Soft Serve", "Yogurt Soft Serve", "Two Tone Soft Serve",
            "Matcha Japanese", "Matcha Marshmallow", "Mix Berry Yogurt",
            "Brownie Yogurt", "Two Tone Brownie", "Saikyo Matcha",
            "Rice", "Miso Soup"
        ]
        hamada_costs = [
            67.53, 51.84, 70.42, 69.33, 53.64, 69.29,
            56.03, 54.34, 73.57, 37.21, 48.67, 62.31,
            69.17, 60.62, 16.16, 10.92, 33.71, 63.56,
            203.12, 89.93, 98.14, 168.1, 108.73, 200.7,
            193.91, 50.76, 87.59, 104.1, 133.97, 48.53,
            66.95, 77.77, 79.45, 108.68, 73.91, 111.38,
            243.48, 246.42, 42.71, 39.59, 60.89, 52.93,
            29.95, 27.87, 42.07, 9.63, 19.94, 19.87,
            19.96, 51.71, 56.12, 112.92, 21.27, 18.32,
            68.24, 51.43, 93.14, 56.38, 68.9, 48.9,
            58.8, 63.34, 74.35, 12.54, 12.54, 12.54,
            7.59, 7, 17, 17, 17, 17, 41.92,
            38.43, 31.71, 3.74, 15.09, 17.72, 16.41,
            22.6, 30.78, 34.94, 41.79, 37.45, 31.54,
            5.85, 3.58
        ]

        # ข้อมูลจริงจาก Hashira Cost.xlsx (ครบทั้ง 68 รายการ)
        hashira_menu_names = [
            "Pork Cut Steak Set", "Tonteki Set", "Beef Cut Steak Set (Medium Rare / Well done)",
            "Buta Teriyaki Set", "Tonkatsu Set", "Tonkatsu Tamagotoji Set",
            "Beef Hamburg Set", "Salmon Set (Teriyaki / Shioyaki)", "Kaisen Teppan Yaki Set",
            "Topping Fried Egg", "Gokuatsu Tonkatsu Set", "Gokuatsu Katsu Don Set",
            "Gokuatsu Tonkatsu Tamagotoji Set", "Tonkotsu Chashumen", "Tonkotsu Niku-Niku Chashumen",
            "Tonkotsu Spicy Ramen", "Tonkotsu Hashira Ramen", "Miso Chashumen",
            "Miso Hokkaido Ramen", "Shoyu Chashumen", "Shoyu Tokyo Ramen",
            "Tom Yum Tonkotsu Kaisen Ramen", "Tom Yum Kung Ramen", "Tori Paitan Karaage Ramen",
            "Tori Paitan Kaisen Ramen", "Tonkatsu", "Hotate",
            "Ebi Tempura", "Chashu", "Spicy Nikumiso",
            "Karashinegi", "Negi", "Naruto",
            "Ajitama", "Menma", "Nori",
            "Corn", "Butter", "Boiled Cabbage",
            "Boiled Bean Sprout", "Karashi Takana", "Kikurage",
            "Kanikama", "Kaedama", "GG Sauce Katsudon",
            "GG Buta Teriyaki Don", "GG Tonkatsu Bento", "GG Yakisoba",
            "GG Yakisoba with Sunny Side Egg", "GG Pork Okonomiyaki", "Coca-Cola Original",
            "Coca-Cola Zero Sugar", "Sprite", "Soda",
            "Mineral Water", "HBD Sparkling Water Lemon (No Sugar, No Calories)",
            "HBD Sparkling Water Honey Yuzu (No Sugar, No Calories)",
            "HBD Sparkling WatCalorieser Peach (No Sugar, No )",
            "HBD Sparkling Water Kyoho (No Sugar, No Calories)", "Asahi Beer",
            "Heineken Beer ", "Singha Beer", "Drinking Water",
            "Melon Ball", "Vanilla Monaka", "Azuki Bar",
            "Rice", "Miso Soup"
        ]

        hashira_costs = [
            46.53, 131.99, 107.87, 41.77, 58.39, 60.91,
            76.4, 68.02, 121.23, 3.5, 131.77, 133,
            131.46, 64.23, 77.98, 54.05, 74.33, 56.52,
            70.08, 63.1, 64.44, 55.55, 116.08, 42.88,
            50.08, 37.4, 33.33, 26, 27.3, 7.97,
            10.54, 6.43, 12.92, 3.84, 9.6, 5.4,
            6, 1.31, 2.73, 1.98, 6.45, 3.41,
            7.92, 7.5, 67.94, 45.93, 79.95, 45.61,
            49.11, 52.4, 12.54, 12.54, 12.54, 7.59,
            7, 17, 17, 17, 17, 41.92,
            38.43, 31.71, 3.74, 25, 27, 27,
            5.85, 3.04
        ]

        # สร้าง DataFrame
        hamada_real_data = {
            "MENU NAME": hamada_menu_names,
            "Material Cost": hamada_costs
        }

        hashira_real_data = {
            "MENU NAME": hashira_menu_names,
            "Material Cost": hashira_costs
        }

        # สร้าง DataFrame และ set index
        self.base_files['hamada'] = pd.DataFrame(hamada_real_data).set_index('MENU NAME')
        self.base_files['hashira'] = pd.DataFrame(hashira_real_data).set_index('MENU NAME')

        logger.info("โหลดข้อมูลจริง Base Cost เรียบร้อย")
        logger.info(f"Hamada Cost: {len(self.base_files['hamada'])} เมนู")
        logger.info(f"Hashira Cost: {len(self.base_files['hashira'])} เมนู")

    def load_config(self):
        """โหลดการตั้งค่าจากไฟล์"""
        # โหลดไฟล์ Base Cost.xlsx หากมี (จะ override ข้อมูลตัวอย่าง)
        self.load_base_cost_file()

    def load_base_cost_file(self):
        """โหลดไฟล์ Base Cost.xlsx"""
        base_file = "Base Cost.xlsx"
        if os.path.exists(base_file):
            try:
                # อ่าน Hashira Cost sheet
                hashira_df = pd.read_excel(base_file, sheet_name="Hashira Cost")
                if 'MENU NAME' in hashira_df.columns:
                    self.base_files['hashira'] = hashira_df.set_index("MENU NAME")

                # อ่าน Hamada Cost sheet
                hamada_df = pd.read_excel(base_file, sheet_name="Hamada Cost")
                if 'MENU NAME' in hamada_df.columns:
                    self.base_files['hamada'] = hamada_df.set_index("MENU NAME")

                logger.info(f"โหลดไฟล์ {base_file} สำเร็จ (แทนที่ข้อมูลตัวอย่าง)")
                logger.info(f"Hashira Cost: {len(self.base_files['hashira'])} เมนู")
                logger.info(f"Hamada Cost: {len(self.base_files['hamada'])} เมนู")

            except Exception as e:
                logger.error(f"Error loading base file: {e}")
                logger.info("ใช้ข้อมูลตัวอย่างต่อไป")

    def save_base_cost_file(self):
        """บันทึกไฟล์ Base Cost.xlsx"""
        try:
            wb = Workbook()
            # ลบ sheet default
            wb.remove(wb.active)

            # สร้าง Hashira Cost sheet
            if self.base_files['hashira'] is not None:
                ws_hashira = wb.create_sheet("Hashira Cost")
                df_hashira = self.base_files['hashira'].reset_index()

                for r in dataframe_to_rows(df_hashira, index=False, header=True):
                    ws_hashira.append(r)

                # จัดรูปแบบ header
                for cell in ws_hashira[1]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")

            # สร้าง Hamada Cost sheet
            if self.base_files['hamada'] is not None:
                ws_hamada = wb.create_sheet("Hamada Cost")
                df_hamada = self.base_files['hamada'].reset_index()

                for r in dataframe_to_rows(df_hamada, index=False, header=True):
                    ws_hamada.append(r)

                # จัดรูปแบบ header
                for cell in ws_hamada[1]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")

            # ปรับความกว้างคอลัมน์
            for ws in wb.worksheets:
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 60)
                    ws.column_dimensions[column_letter].width = adjusted_width

            wb.save("Base Cost.xlsx")
            logger.info("บันทึก Base Cost.xlsx สำเร็จ")
            return True

        except Exception as e:
            logger.error(f"Error saving base file: {e}")
            return False

    def update_single_base(self, base_type, new_df):
        """อัพเดท Base Cost แยกกันตามประเภท"""
        try:
            # ลบช่องว่างจากชื่อคอลัมน์
            new_df.columns = new_df.columns.str.strip()

            # ตรวจสอบคอลัมน์ที่จำเป็น
            if 'MENU NAME' not in new_df.columns:
                raise ValueError(f"ไฟล์ไม่มีคอลัมน์ 'MENU NAME'")

            if 'Material Cost' not in new_df.columns:
                raise ValueError(f"ไฟล์ไม่มีคอลัมน์ 'Material Cost'")

            # ตั้งค่า index
            new_df = new_df.set_index("MENU NAME")

            # อัพเดทข้อมูล
            self.base_files[base_type] = new_df

            # บันทึกไฟล์ Base Cost.xlsx
            self.save_base_cost_file()

            logger.info(f"อัพเดท {base_type.upper()} Cost สำเร็จ: {len(new_df)} เมนู")
            return True

        except Exception as e:
            logger.error(f"Error updating {base_type} base: {e}")
            raise e


# ===== Modern GUI Styling =====
def configure_modern_style():
    """กำหนดธีมสีสวยงามและทันสมัย"""
    style = ttk.Style()

    try:
        style.theme_use('clam')
    except:
        style.theme_use('default')

    colors = {
        'bg_primary': '#2c3e50',
        'bg_secondary': '#34495e',
        'accent': '#3498db',
        'success': '#27ae60',
        'warning': '#f39c12',
        'danger': '#e74c3c',
        'light': '#ecf0f1',
        'white': '#ffffff',
        'text_dark': '#2c3e50',
        'text_light': '#ffffff'
    }

    # Configure styles
    style.configure('Title.TLabel', font=('Segoe UI', 16, 'bold'), foreground=colors['text_dark'])
    style.configure('Heading.TLabel', font=('Segoe UI', 12, 'bold'), foreground=colors['accent'])
    style.configure('Status.TLabel', font=('Segoe UI', 10), padding=(5, 2))

    style.configure('Accent.TButton', font=('Segoe UI', 9, 'bold'), padding=(10, 5))
    style.configure('Success.TButton', font=('Segoe UI', 9, 'bold'), padding=(10, 5))
    style.configure('Warning.TButton', font=('Segoe UI', 9, 'bold'), padding=(10, 5))
    style.configure('Danger.TButton', font=('Segoe UI', 9, 'bold'), padding=(10, 5))

    style.configure('Card.TFrame', relief='solid', borderwidth=1, padding=10)
    style.configure('Modern.TNotebook', tabposition='n')
    style.configure('Modern.TNotebook.Tab', font=('Segoe UI', 10, 'bold'), padding=(20, 8))

    return colors


# ===== Logging Setup =====
logging.basicConfig(
    level=logging.DEBUG,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('hamada_calculator_debug.log', encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# Initialize configuration
config_manager = ConfigManager()


class EnhancedCostCalculatorApp:
    def __init__(self, root):
        self.root = root
        self.colors = configure_modern_style()

        self.root.title("🍱 Hamada & Hashira Cost Calculator - Fixed")
        self.root.geometry("1500x950")
        self.root.configure(bg=self.colors['light'])
        self.root.minsize(1300, 850)

        # ตัวแปรสำหรับไฟล์
        self.import_file = None
        self.template_file = None
        self.selected_base_type = tk.StringVar(value='hamada')
        self.df_result = None
        self.df_import = None  # เก็บข้อมูลไฟล์ import
        self.df_template = None  # เก็บข้อมูลไฟล์ template
        self.debug_mode = tk.BooleanVar(value=False)

        # ข้อมูล debug แยกตามประเภท
        self.debug_data = {
            'not_found_menus': [],
            'zero_qty_items': [],
            'invalid_qty_items': [],
            'nan_cost_items': [],
            'matched_menus': [],
            'not_sold_menus': [],
            'processing_summary': {}
        }

        self._build_ui()
        logger.info("Enhanced Cost Calculator initialized with built-in data")

    def _build_ui(self):
        """สร้าง UI แบบ Enhanced"""
        # Main container
        main_frame = ttk.Frame(self.root, style='Card.TFrame')
        main_frame.pack(fill=tk.BOTH, expand=True, padx=15, pady=15)

        # Enhanced Header
        header_frame = ttk.Frame(main_frame)
        header_frame.pack(fill=tk.X, pady=(0, 20))

        title_label = ttk.Label(header_frame, text="🍱 Hamada & Hashira Cost Calculator",
                                style='Title.TLabel')
        title_label.pack(pady=(0, 5))

        subtitle = ttk.Label(header_frame, text="Enhanced Multi-Base Cost Management System (Real Data Loaded)",
                             font=('Segoe UI', 11), foreground=self.colors['bg_secondary'])
        subtitle.pack()

        # ===== Base Type Selection Card =====
        base_selection_frame = ttk.LabelFrame(main_frame, text="🏪 Base Cost Selection",
                                              padding=15, style='Card.TFrame')
        base_selection_frame.pack(fill=tk.X, pady=(0, 15))

        base_frame = ttk.Frame(base_selection_frame)
        base_frame.pack(fill=tk.X)

        ttk.Label(base_frame, text="เลือกประเภท Base Cost:",
                  font=('Segoe UI', 11, 'bold')).pack(side=tk.LEFT, padx=(0, 20))

        # Radio buttons
        hashira_radio = ttk.Radiobutton(base_frame, text="🏯 Hashira Cost",
                                        variable=self.selected_base_type, value='hashira',
                                        command=self.on_base_type_change)
        hashira_radio.pack(side=tk.LEFT, padx=10)

        hamada_radio = ttk.Radiobutton(base_frame, text="🍜 Hamada Cost",
                                       variable=self.selected_base_type, value='hamada',
                                       command=self.on_base_type_change)
        hamada_radio.pack(side=tk.LEFT, padx=10)

        self.base_status_label = ttk.Label(base_frame, text="", style='Status.TLabel')
        self.base_status_label.pack(side=tk.RIGHT, padx=20)

        # ===== File Management Section =====
        file_frame = ttk.LabelFrame(main_frame, text="📁 File Management",
                                    padding=15, style='Card.TFrame')
        file_frame.pack(fill=tk.X, pady=(0, 15))

        # Import buttons
        import_frame = ttk.Frame(file_frame)
        import_frame.pack(fill=tk.X, pady=(0, 10))

        ttk.Button(import_frame, text="📁 เลือกไฟล์ Import",
                   command=self.load_import_file, style='Accent.TButton').pack(
            side=tk.LEFT, padx=5, fill=tk.X, expand=True)

        ttk.Button(import_frame, text="📄 Import Template เพิ่มเติม",
                   command=self.load_template_file, style='Warning.TButton').pack(
            side=tk.LEFT, padx=5, fill=tk.X, expand=True)

        ttk.Button(import_frame, text="🔍 ตรวจสอบไฟล์ Base",
                   command=self.check_base_file, style='Accent.TButton').pack(
            side=tk.LEFT, padx=5, fill=tk.X, expand=True)

        # File status display
        status_frame = ttk.Frame(file_frame)
        status_frame.pack(fill=tk.X)

        self.import_status_label = ttk.Label(status_frame, text="📂 Import: ยังไม่ได้เลือกไฟล์",
                                             style='Status.TLabel', foreground=self.colors['danger'])
        self.import_status_label.pack(anchor="w", pady=2)

        self.template_status_label = ttk.Label(status_frame, text="📄 Template: ยังไม่ได้เลือกไฟล์",
                                               style='Status.TLabel', foreground=self.colors['warning'])
        self.template_status_label.pack(anchor="w", pady=2)

        self.hashira_status_label = ttk.Label(status_frame, text="🏯 Hashira Cost: กำลังตรวจสอบ...",
                                              style='Status.TLabel', foreground=self.colors['warning'])
        self.hashira_status_label.pack(anchor="w", pady=2)

        self.hamada_status_label = ttk.Label(status_frame, text="🍜 Hamada Cost: กำลังตรวจสอบ...",
                                             style='Status.TLabel', foreground=self.colors['warning'])
        self.hamada_status_label.pack(anchor="w", pady=2)

        # ===== Action Buttons =====
        action_frame = ttk.LabelFrame(main_frame, text="🚀 Actions",
                                      padding=15, style='Card.TFrame')
        action_frame.pack(fill=tk.X, pady=(0, 15))

        # Debug mode
        debug_frame = ttk.Frame(action_frame)
        debug_frame.pack(fill=tk.X, pady=(0, 10))

        self.debug_checkbox = ttk.Checkbutton(debug_frame, text="🔍 เปิด Debug Mode",
                                              variable=self.debug_mode)
        self.debug_checkbox.pack(side=tk.LEFT)

        ttk.Button(debug_frame, text="🗑️ เคลียร์ Debug",
                   command=self.clear_debug, style='Warning.TButton').pack(side=tk.RIGHT)

        # Calculate button
        calc_frame = ttk.Frame(action_frame)
        calc_frame.pack(fill=tk.X, pady=(0, 10))

        ttk.Button(calc_frame, text="⚡ คำนวณ", command=self.calculate,
                   style='Success.TButton').pack(fill=tk.X)

        # ===== SINGLE EXPORT BUTTON =====
        export_frame = ttk.Frame(action_frame)
        export_frame.pack(fill=tk.X)

        # ปุ่ม Export เดียวที่รวมทุกอย่าง
        ttk.Button(export_frame, text="💾 Export ไฟล์ Excel ครบชุด (4 Sheets)",
                   command=self.export_complete_excel,
                   style='Success.TButton').pack(fill=tk.X, ipady=8)

        # ===== Content Area =====
        content_frame = ttk.Frame(main_frame, style='Card.TFrame')
        content_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 15))

        self.notebook = ttk.Notebook(content_frame, style='Modern.TNotebook')
        self.notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # Results Tab
        result_frame = ttk.Frame(self.notebook)
        self.notebook.add(result_frame, text="📋 ผลลัพธ์")

        table_frame = ttk.Frame(result_frame, padding=10)
        table_frame.pack(fill=tk.BOTH, expand=True)

        table_header = ttk.Label(table_frame, text="💰 Calculation Results",
                                 style='Heading.TLabel')
        table_header.pack(pady=(0, 10))

        # Results table
        self.tree = ttk.Treeview(table_frame, columns=("menu", "qty", "cost", "total"),
                                 show="headings", height=16)

        self.tree.heading("menu", text="🍽️ MENU NAME")
        self.tree.heading("qty", text="📊 Qty")
        self.tree.heading("cost", text="💵 Material Cost")
        self.tree.heading("total", text="💰 Total Cost")

        self.tree.column("menu", width=400, anchor="w")
        self.tree.column("qty", width=100, anchor="center")
        self.tree.column("cost", width=150, anchor="e")
        self.tree.column("total", width=150, anchor="e")

        tree_scroll = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=tree_scroll.set)

        tree_scroll.pack(side=tk.RIGHT, fill=tk.Y, padx=(5, 0))
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # Debug Tab
        debug_frame = ttk.Frame(self.notebook)
        self.notebook.add(debug_frame, text="🔍 Debug Log")

        debug_header_frame = ttk.Frame(debug_frame, padding=10)
        debug_header_frame.pack(fill=tk.X)

        debug_title = ttk.Label(debug_header_frame, text="🔧 Enhanced Debug Console",
                                style='Heading.TLabel')
        debug_title.pack()

        debug_text_frame = ttk.Frame(debug_frame, padding=(10, 0, 10, 10))
        debug_text_frame.pack(fill=tk.BOTH, expand=True)

        self.debug_text = scrolledtext.ScrolledText(debug_text_frame, wrap=tk.WORD, height=18,
                                                    font=('Consolas', 10),
                                                    bg='#1e1e1e', fg='#ffffff',
                                                    insertbackground='#ffffff',
                                                    selectbackground='#3498db',
                                                    relief='flat', borderwidth=0)
        self.debug_text.pack(fill=tk.BOTH, expand=True)

        # Statistics Section
        self.stats_frame = ttk.LabelFrame(main_frame, text="📈 Statistics & Info",
                                          padding=15, style='Card.TFrame')
        self.stats_frame.pack(fill=tk.X)

        self.stats_label = ttk.Label(self.stats_frame, text="ยังไม่มีข้อมูล",
                                     style='Status.TLabel',
                                     font=('Segoe UI', 11, 'bold'))
        self.stats_label.pack()

        self._configure_treeview_style()
        self.update_all_base_status()
        self.update_base_status()

    def _configure_treeview_style(self):
        """กำหนดสไตล์ modern สำหรับ treeview"""
        style = ttk.Style()
        style.configure('Modern.Treeview',
                        background='#ffffff',
                        foreground='#2c3e50',
                        fieldbackground='#ffffff',
                        font=('Segoe UI', 10))
        style.configure('Modern.Treeview.Heading',
                        background='#3498db',
                        foreground='#ffffff',
                        font=('Segoe UI', 10, 'bold'),
                        relief='flat')
        style.map('Modern.Treeview',
                  background=[('selected', '#3498db')],
                  foreground=[('selected', '#ffffff')])
        self.tree.configure(style='Modern.Treeview')

    def on_base_type_change(self):
        """เมื่อเปลี่ยน base type"""
        selected = self.selected_base_type.get()
        self.debug_log(f"เปลี่ยน Base Type เป็น: {selected.upper()}")
        self.update_base_status()

    def update_base_status(self):
        """อัพเดทสถานะไฟล์ base ที่เลือก"""
        selected = self.selected_base_type.get()
        base_df = config_manager.base_files.get(selected)

        if base_df is not None and not base_df.empty:
            menu_count = len(base_df)
            status_text = f"🗂️ {selected.upper()} Cost: พร้อมใช้งาน ({menu_count} เมนู)"
            color = self.colors['success']
        else:
            status_text = f"🗂️ {selected.upper()} Cost: ไม่พบข้อมูล"
            color = self.colors['danger']

        self.base_status_label.config(text=status_text, foreground=color)

    def update_all_base_status(self):
        """อัพเดทสถานะไฟล์ base ทั้งหมด"""
        # Hashira Status
        hashira_df = config_manager.base_files.get('hashira')
        if hashira_df is not None and not hashira_df.empty:
            hashira_count = len(hashira_df)
            hashira_text = f"🏯 Hashira Cost: พร้อมใช้งาน ({hashira_count} เมนู) [Real Data]"
            hashira_color = self.colors['success']
        else:
            hashira_text = f"🏯 Hashira Cost: ไม่พบข้อมูล"
            hashira_color = self.colors['danger']

        # Hamada Status
        hamada_df = config_manager.base_files.get('hamada')
        if hamada_df is not None and not hamada_df.empty:
            hamada_count = len(hamada_df)
            hamada_text = f"🍜 Hamada Cost: พร้อมใช้งาน ({hamada_count} เมนู) [Real Data]"
            hamada_color = self.colors['success']
        else:
            hamada_text = f"🍜 Hamada Cost: ไม่พบข้อมูล"
            hamada_color = self.colors['danger']

        self.hashira_status_label.config(text=hashira_text, foreground=hashira_color)
        self.hamada_status_label.config(text=hamada_text, foreground=hamada_color)

    def debug_log(self, message, level="INFO"):
        """เพิ่มข้อความใน debug log"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        log_message = f"[{timestamp}] {level}: {message}\n"

        self.debug_text.insert(tk.END, log_message)
        self.debug_text.see(tk.END)

        if level == "ERROR":
            logger.error(message)
        elif level == "WARNING":
            logger.warning(message)
        else:
            logger.info(message)

    def clear_debug(self):
        """เคลียร์ debug log"""
        self.debug_text.delete(1.0, tk.END)
        self.debug_log("Debug log cleared")
        self.debug_data = {
            'not_found_menus': [],
            'zero_qty_items': [],
            'invalid_qty_items': [],
            'nan_cost_items': [],
            'matched_menus': [],
            'not_sold_menus': [],
            'processing_summary': {}
        }

    def load_import_file(self):
        """โหลดไฟล์ import"""
        from tkinter import filedialog

        try:
            file_path = filedialog.askopenfilename(
                title="เลือกไฟล์ Import",
                filetypes=[("Excel Files", "*.xlsx"), ("Excel Files", "*.xls")]
            )

            if not file_path:
                return

            # โหลดและเก็บข้อมูลไฟล์
            self.df_import = pd.read_excel(file_path)
            self.import_file = file_path

            self.import_status_label.config(
                text=f"📂 Import: {os.path.basename(file_path)} ({len(self.df_import)} แถว)",
                foreground=self.colors['success']
            )

            self.debug_log(f"เลือกไฟล์ Import: {file_path}")
            self._preview_file(self.df_import, "Import")

        except Exception as e:
            self.debug_log(f"Error เลือกไฟล์: {e}", "ERROR")
            messagebox.showerror("Error", f"ไม่สามารถเลือกไฟล์ได้: {str(e)}")

    def load_template_file(self):
        """โหลดไฟล์ template"""
        from tkinter import filedialog

        try:
            file_path = filedialog.askopenfilename(
                title="เลือกไฟล์ Template",
                filetypes=[("Excel Files", "*.xlsx"), ("Excel Files", "*.xls")]
            )

            if not file_path:
                return

            # โหลดและเก็บข้อมูลไฟล์ template
            self.df_template = pd.read_excel(file_path)
            self.template_file = file_path

            self.template_status_label.config(
                text=f"📄 Template: {os.path.basename(file_path)} ({len(self.df_template)} แถว)",
                foreground=self.colors['success']
            )

            self.debug_log(f"เลือกไฟล์ Template: {file_path}")
            self._preview_file(self.df_template, "Template")

        except Exception as e:
            self.debug_log(f"Error เลือกไฟล์ Template: {e}", "ERROR")
            messagebox.showerror("Error", f"ไม่สามารถเลือกไฟล์ได้: {str(e)}")

    def _preview_file(self, df_data, file_type):
        """แสดงตัวอย่างข้อมูลในไฟล์"""
        try:
            self.debug_log(f"ไฟล์ {file_type} มี {len(df_data)} แถว, {len(df_data.columns)} คอลัมน์")
            self.debug_log(f"คอลัมน์: {df_data.columns.tolist()}")

            # แสดงตัวอย่างข้อมูล
            self.debug_log(f"ตัวอย่างข้อมูล {file_type} 5 แถวแรก:")
            for i, (idx, row) in enumerate(df_data.head(5).iterrows()):
                menu = row.get("MENU NAME", "N/A")
                qty = row.get("Qty", "N/A")
                self.debug_log(f"  แถว {i + 1}: {menu} = {qty}")

        except Exception as e:
            self.debug_log(f"Error ตรวจสอบไฟล์ {file_type}: {str(e)}", "ERROR")

    def show_all_base_menus(self):
        """แสดงรายการเมนูทั้งหมดใน Base Cost ที่เลือก"""
        selected = self.selected_base_type.get()
        df_base = config_manager.base_files.get(selected)

        if df_base is None:
            messagebox.showerror("Error", f"ไม่สามารถโหลด {selected.upper()} Cost ได้")
            return

        # สร้างหน้าต่างใหม่
        menu_window = tk.Toplevel(self.root)
        menu_window.title(f"📋 รายการเมนูทั้งหมดใน {selected.upper()} Cost")
        menu_window.geometry("800x600")
        menu_window.configure(bg='white')

        # Header
        header_frame = tk.Frame(menu_window, bg='#3498db', height=60)
        header_frame.pack(fill='x')
        header_frame.pack_propagate(False)

        title_label = tk.Label(header_frame,
                               text=f"📋 {selected.upper()} Cost - ทั้งหมด {len(df_base)} เมนู",
                               font=('Segoe UI', 14, 'bold'),
                               fg='white', bg='#3498db')
        title_label.pack(pady=15)

        # ค้นหา
        search_frame = tk.Frame(menu_window, bg='white')
        search_frame.pack(fill='x', padx=20, pady=10)

        tk.Label(search_frame, text="🔍 ค้นหาเมนู:",
                 font=('Segoe UI', 11, 'bold')).pack(side='left', padx=(0, 10))

        search_var = tk.StringVar()
        search_entry = tk.Entry(search_frame, textvariable=search_var,
                                font=('Segoe UI', 10), width=40)
        search_entry.pack(side='left', padx=(0, 10))

        # รายการเมนู
        list_frame = tk.Frame(menu_window, bg='white')
        list_frame.pack(fill='both', expand=True, padx=20, pady=(0, 20))

        # Listbox กับ Scrollbar
        listbox_frame = tk.Frame(list_frame)
        listbox_frame.pack(fill='both', expand=True)

        scrollbar = tk.Scrollbar(listbox_frame)
        scrollbar.pack(side='right', fill='y')

        menu_listbox = tk.Listbox(listbox_frame,
                                  yscrollcommand=scrollbar.set,
                                  font=('Consolas', 10),
                                  selectmode='extended')
        menu_listbox.pack(side='left', fill='both', expand=True)
        scrollbar.config(command=menu_listbox.yview)

        # ใส่ข้อมูลเมนู
        all_menus = []
        for idx, menu_name in enumerate(df_base.index):
            cost = df_base.at[menu_name, "Material Cost"]
            menu_display = f"{idx + 1:3d}. {menu_name} (Cost: {cost:.2f} บาท)"
            all_menus.append((menu_name, menu_display))
            menu_listbox.insert('end', menu_display)

        # ฟังก์ชันค้นหา
        def filter_menus():
            search_text = search_var.get().lower()
            menu_listbox.delete(0, 'end')

            filtered_count = 0
            for menu_name, menu_display in all_menus:
                if search_text in menu_name.lower():
                    menu_listbox.insert('end', menu_display)
                    filtered_count += 1

            title_label.config(text=f"📋 {selected.upper()} Cost - แสดง {filtered_count}/{len(df_base)} เมนู")

        search_entry.bind('<KeyRelease>', lambda e: filter_menus())

        # ปุ่มยัดเฉพาะเมนูที่ค้นหา
        button_frame = tk.Frame(menu_window, bg='white')
        button_frame.pack(fill='x', padx=20, pady=10)

        def copy_selected_menus():
            """Copy เมนูที่เลือกไปยัง clipboard"""
            selected_indices = menu_listbox.curselection()
            if not selected_indices:
                messagebox.showinfo("Info", "กรุณาเลือกเมนูที่จะ copy")
                return

            selected_menus = []
            for i in selected_indices:
                menu_text = menu_listbox.get(i)
                # ดึงเฉพาะชื่อเมนู (ตัด index และราคาออก)
                menu_name = menu_text.split('. ', 1)[1].split(' (Cost:')[0]
                selected_menus.append(menu_name)

            # Copy ไป clipboard
            menu_window.clipboard_clear()
            menu_window.clipboard_append('\n'.join(selected_menus))

            messagebox.showinfo("Success", f"Copy {len(selected_menus)} เมนูเรียบร้อย")

        tk.Button(button_frame, text="📋 Copy เมนูที่เลือก",
                  command=copy_selected_menus,
                  bg='#27ae60', fg='white', font=('Segoe UI', 10, 'bold'),
                  padx=15, pady=5).pack(side='left')

        tk.Button(button_frame, text="❌ ปิด",
                  command=menu_window.destroy,
                  bg='#e74c3c', fg='white', font=('Segoe UI', 10, 'bold'),
                  padx=15, pady=5).pack(side='right')

        # Focus ที่ search box
        search_entry.focus_set()

    def check_base_file(self):
        """ตรวจสอบและแสดงข้อมูลไฟล์ base ที่เลือก"""
        selected = self.selected_base_type.get()
        df_base = config_manager.base_files.get(selected)

        if df_base is None:
            messagebox.showerror("Error", f"ไม่สามารถโหลด {selected.upper()} Cost ได้")
            return

        self.debug_log(f"=== ตรวจสอบ {selected.upper()} Cost ===")
        self.debug_log(f"จำนวนเมนู: {len(df_base)}")
        self.debug_log(f"คอลัมน์: {df_base.columns.tolist()}")

        # สถิติราคา
        try:
            if "Material Cost" in df_base.columns:
                cost_stats = pd.to_numeric(df_base["Material Cost"], errors='coerce').describe()
                self.debug_log("สถิติราคา Material Cost:")
                self.debug_log(f"  ต่ำสุด: {cost_stats['min']:.2f} บาท")
                self.debug_log(f"  สูงสุด: {cost_stats['max']:.2f} บาท")
                self.debug_log(f"  เฉลี่ย: {cost_stats['mean']:.2f} บาท")
                self.debug_log(f"  จำนวนรายการ: {int(cost_stats['count'])}")
        except Exception as e:
            self.debug_log(f"Error คำนวณสถิติ: {str(e)}", "ERROR")

    def calculate(self):
        """คำนวณต้นทุน - Enhanced Version"""
        df_base = config_manager.base_files.get(self.selected_base_type.get())
        selected_base = self.selected_base_type.get()

        # Validation checks
        if df_base is None or df_base.empty:
            messagebox.showerror("Error", f"ไฟล์ {selected_base.upper()} Cost ไม่สามารถโหลดได้")
            return

        if self.df_import is None:
            messagebox.showwarning("Warning", "กรุณาเลือกไฟล์ Import ก่อน")
            return

        self.debug_log(f"=== เริ่มการคำนวณด้วย {selected_base.upper()} Cost ===")

        try:
            # ทำความสะอาดข้อมูล
            df_base.columns = df_base.columns.str.strip()
            self.df_import.columns = self.df_import.columns.str.strip()

            # ตรวจสอบคอลัมน์ที่จำเป็น
            required_columns = ["MENU NAME", "Qty"]
            missing_columns = [col for col in required_columns if col not in self.df_import.columns]

            if missing_columns:
                error_msg = f"ไฟล์ Import ไม่มีคอลัมน์: {', '.join(missing_columns)}"
                self.debug_log(error_msg, "ERROR")
                messagebox.showerror("Error", error_msg)
                return

            self.debug_log(f"อ่านไฟล์ import สำเร็จ: {len(self.df_import)} แถว")

        except Exception as e:
            error_msg = f"ไม่สามารถอ่านไฟล์ได้: {str(e)}"
            self.debug_log(error_msg, "ERROR")
            messagebox.showerror("Error", error_msg)
            return

        # เริ่มการคำนวณ
        results = []
        matched_count = 0
        self.debug_data['not_found_menus'] = []
        self.debug_data['matched_menus'] = []

        for idx, row in self.df_import.iterrows():
            menu = row.get("MENU NAME")
            qty = row.get("Qty", 0)

            if pd.isna(menu) or menu == "":
                continue

            try:
                qty = float(qty) if not pd.isna(qty) else 0
            except:
                qty = 0
                self.debug_data['invalid_qty_items'].append(menu)

            if menu in df_base.index:
                try:
                    material_cost = df_base.at[menu, "Material Cost"]
                    if pd.isna(material_cost):
                        material_cost = 0
                    else:
                        material_cost = float(material_cost)

                    total_cost = qty * material_cost
                    results.append([menu, qty, material_cost, total_cost])
                    matched_count += 1
                    self.debug_data['matched_menus'].append(menu)

                    if self.debug_mode.get():
                        self.debug_log(f"✓ พบเมนู: {menu} = {material_cost:.2f} x {qty} = {total_cost:.2f}")

                except Exception as e:
                    self.debug_log(f"Error คำนวณ {menu}: {str(e)}", "ERROR")
                    self.debug_data['nan_cost_items'].append(menu)
            else:
                self.debug_data['not_found_menus'].append(menu)
                if self.debug_mode.get():
                    self.debug_log(f"❌ ไม่พบเมนู: {menu}", "WARNING")

        # สร้าง DataFrame ผลลัพธ์
        self.df_result = pd.DataFrame(results, columns=["MENU NAME", "Qty", "Material Cost", "Total Cost"])

        if self.df_result.empty:
            error_msg = f"ไม่มีชื่อเมนูที่ตรงกับ {selected_base.upper()} Cost"
            self.debug_log(error_msg, "WARNING")
            messagebox.showwarning("ไม่พบข้อมูล", error_msg)
            return

        # คำนวณผลรวมทุกคอลัมน์
        total_qty = self.df_result["Qty"].sum()
        total_material_cost = self.df_result["Material Cost"].sum()
        grand_total = self.df_result["Total Cost"].sum()

        # เพิ่มแถว TOTAL
        total_row = pd.DataFrame([["TOTAL", total_qty, total_material_cost, grand_total]],
                                 columns=["MENU NAME", "Qty", "Material Cost", "Total Cost"])
        self.df_result = pd.concat([self.df_result, total_row], ignore_index=True)

        # บันทึก summary ใน debug_data
        self.debug_data['processing_summary'] = {
            'total_import_items': len(self.df_import),
            'matched_items': matched_count,
            'not_found_items': len(self.debug_data['not_found_menus']),
            'total_qty': total_qty,
            'total_material_cost': total_material_cost,
            'grand_total': grand_total,
            'base_type_used': selected_base.upper()
        }

        # แสดงผล
        self._update_result_table()
        self._update_statistics(matched_count, grand_total, selected_base)

        success_msg = f"คำนวณเสร็จสิ้น - พบ {matched_count} รายการ (ใช้ {selected_base.upper()} Cost)"
        self.debug_log(success_msg)
        messagebox.showinfo("Success", success_msg)

    def _update_result_table(self):
        """อัพเดทตารางผลลัพธ์"""
        for row in self.tree.get_children():
            self.tree.delete(row)

        for _, r in self.df_result.iterrows():
            qty_display = int(r["Qty"]) if r["Qty"] != "" else ""
            cost_display = f"{r['Material Cost']:.2f}" if r["Material Cost"] != "" else ""
            total_display = f"{r['Total Cost']:.2f}" if r["Total Cost"] != "" else ""

            tags = ("total_row",) if r["MENU NAME"] == "TOTAL" else ()

            self.tree.insert("", tk.END,
                             values=(r["MENU NAME"], qty_display, cost_display, total_display),
                             tags=tags)

        self.tree.tag_configure("total_row",
                                background="#3498db",
                                foreground="#ffffff",
                                font=("Segoe UI", 10, "bold"))

    def _update_statistics(self, matched, grand_total, base_type):
        """อัพเดทสถิติ"""
        stats_text = (f"✅ พบเมนู: {matched} | 💰 รวม: {grand_total:,.2f} บาท | 🏪 Base: {base_type.upper()}")
        self.stats_label.config(text=stats_text)

    # ===== FIXED EXPORT FUNCTION - ไฟล์เดียว 4 Sheets =====
    def export_complete_excel(self):
        """Export ไฟล์ Excel เดียวที่มี 4 Sheets ครบชุด - Fixed Version"""
        self.debug_log("🔄 เริ่ม Export ไฟล์ Excel ครบชุด (Fixed Version)...")

        # ตรวจสอบข้อมูลพื้นฐาน
        if self.df_result is None or self.df_result.empty:
            messagebox.showwarning("Warning", "กรุณาคำนวณก่อน Export")
            return

        selected_base = self.selected_base_type.get()
        filename = f"Cost_Analysis_Complete_{selected_base.upper()}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        file_path = os.path.join(os.getcwd(), filename)

        try:
            self.debug_log(f"📂 สร้างไฟล์ Excel ที่: {file_path}")

            # สร้าง Excel Workbook
            wb = Workbook()
            wb.remove(wb.active)  # ลบ sheet default

            sheets_created = 0

            # ===== Sheet 1: "Calculation results" (แก้ไขชื่อและเพิ่มผลรวม) =====
            if self.df_result is not None:
                ws_result = wb.create_sheet("Calculation results")

                # เพิ่มหัวข้อและข้อมูลสรุป
                ws_result.append(["📊 Cost Calculation Results"])
                ws_result.append([])
                ws_result.append([f"🏪 Base Cost Used: {selected_base.upper()}"])
                ws_result.append([f"📅 Calculation Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
                ws_result.append([f"📋 Total Items: {len(self.df_result) - 1}"])  # ลบ TOTAL row
                ws_result.append([])

                # เพิ่มตาราง results (ผลรวมทุกคอลัมน์รวมอยู่แล้วใน df_result)
                for r in dataframe_to_rows(self.df_result, index=False, header=True):
                    ws_result.append(r)

                # จัดรูปแบบ
                # Header ใหญ่
                ws_result['A1'].font = Font(bold=True, size=14, color="FFFFFF")
                ws_result['A1'].fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")

                # Table headers
                header_row = 7
                for cell in ws_result[header_row]:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="2ECC71", end_color="2ECC71", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")

                # ไฮไลท์ TOTAL row
                for row in range(header_row + 1, ws_result.max_row + 1):
                    if ws_result[f'A{row}'].value == "TOTAL":
                        for col in range(1, 5):
                            cell = ws_result.cell(row=row, column=col)
                            cell.font = Font(bold=True, color="FFFFFF")
                            cell.fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")

                sheets_created += 1
                self.debug_log(f"✅ Sheet 1: Calculation results ({len(self.df_result)} รายการ)")

            # ===== Sheet 2: "Base Cost ({selected_base.upper()})" (ตัดเลข "2." ออก) =====
            df_base = config_manager.base_files.get(selected_base)
            if df_base is not None:
                ws_base = wb.create_sheet(f"Base Cost ({selected_base.upper()})")
                df_base_export = df_base.reset_index()

                # เพิ่มหัวข้อ
                ws_base.append([f"🏪 {selected_base.upper()} Base Cost Data"])
                ws_base.append([])
                ws_base.append([f"📅 Export Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
                ws_base.append([f"📋 Total Menus: {len(df_base_export)}"])
                ws_base.append([])

                # เพิ่มข้อมูล
                for r in dataframe_to_rows(df_base_export, index=False, header=True):
                    ws_base.append(r)

                # จัดรูปแบบ
                ws_base['A1'].font = Font(bold=True, size=14, color="FFFFFF")
                if selected_base == 'hashira':
                    ws_base['A1'].fill = PatternFill(start_color="9B59B6", end_color="9B59B6", fill_type="solid")
                else:
                    ws_base['A1'].fill = PatternFill(start_color="E67E22", end_color="E67E22", fill_type="solid")

                # Table headers
                header_row = 6
                for cell in ws_base[header_row]:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")

                sheets_created += 1
                self.debug_log(f"✅ Sheet 2: Base Cost {selected_base.upper()} ({len(df_base_export)} เมนู)")

            # ===== Sheet 3: "Import (By User)" (แก้ไขชื่อ) =====
            if self.df_import is not None:
                ws_import = wb.create_sheet("Import (By User)")

                # เพิ่มหัวข้อ (แก้ไขชื่อ)
                ws_import.append(["📁 Import (By User) Data"])
                ws_import.append([])
                ws_import.append([f"📂 File: {os.path.basename(self.import_file) if self.import_file else 'N/A'}"])
                ws_import.append([f"📅 Import Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
                ws_import.append([f"📋 Total Rows: {len(self.df_import)}"])
                ws_import.append([])

                # เพิ่มข้อมูล
                for r in dataframe_to_rows(self.df_import, index=False, header=True):
                    ws_import.append(r)

                # จัดรูปแบบ
                ws_import['A1'].font = Font(bold=True, size=14, color="FFFFFF")
                ws_import['A1'].fill = PatternFill(start_color="16A085", end_color="16A085", fill_type="solid")

                # Table headers
                header_row = 7
                for cell in ws_import[header_row]:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="1ABC9C", end_color="1ABC9C", fill_type="solid")

                sheets_created += 1
                self.debug_log(f"✅ Sheet 3: Import (By User) ({len(self.df_import)} แถว)")

            # ===== Sheet 4: "Raw file(POS)" (แก้ไขชื่อ) =====
            if self.df_template is not None:
                ws_template = wb.create_sheet("Raw file(POS)")

                # เพิ่มหัวข้อ (แก้ไขชื่อ)
                ws_template.append(["📄 Raw file(POS) Data"])
                ws_template.append([])
                ws_template.append([f"📂 File: {os.path.basename(self.template_file) if self.template_file else 'N/A'}"])
                ws_template.append([f"📅 Import Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
                ws_template.append([f"📋 Total Rows: {len(self.df_template)}"])
                ws_template.append([])

                # เพิ่มข้อมูล
                for r in dataframe_to_rows(self.df_template, index=False, header=True):
                    ws_template.append(r)

                # จัดรูปแบบ
                ws_template['A1'].font = Font(bold=True, size=14, color="FFFFFF")
                ws_template['A1'].fill = PatternFill(start_color="F39C12", end_color="F39C12", fill_type="solid")

                # Table headers
                header_row = 7
                for cell in ws_template[header_row]:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="E67E22", end_color="E67E22", fill_type="solid")

                sheets_created += 1
                self.debug_log(f"✅ Sheet 4: Raw file(POS) ({len(self.df_template)} แถว)")
            else:
                # สร้าง sheet ว่างถ้าไม่มี template
                ws_template = wb.create_sheet("Raw file(POS)")
                ws_template.append(["📄 Raw file(POS) Data"])
                ws_template.append([])
                ws_template.append(["ℹ️ No template file imported"])

                ws_template['A1'].font = Font(bold=True, size=14, color="FFFFFF")
                ws_template['A1'].fill = PatternFill(start_color="95A5A6", end_color="95A5A6", fill_type="solid")

                sheets_created += 1
                self.debug_log("✅ Sheet 4: Raw file(POS) (ไม่มีข้อมูล)")

            # ปรับความกว้างคอลัมน์ทุก sheet
            for ws in wb.worksheets:
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width

            # บันทึกไฟล์
            wb.save(file_path)

            # สรุปผลการ Export (อัพเดทข้อความ)
            summary_msg = f"""✅ Export ไฟล์ Excel สำเร็จ!

📁 ชื่อไฟล์: {filename}
📍 ตำแหน่ง: {file_path}
📊 จำนวน Sheets: {sheets_created}

📋 รายการ Sheets:
1. Calculation results ({len(self.df_result)} รายการ) ✨ มีผลรวมทุกคอลัมน์
2. Base Cost ({selected_base.upper()}) ({len(df_base)} เมนู)  
3. Import (By User) ({len(self.df_import) if self.df_import is not None else 0} แถว)
4. Raw file(POS) ({len(self.df_template) if self.df_template is not None else 0} แถว)

💰 สรุปผลรวม:
📊 Total Qty: {self.debug_data['processing_summary']['total_qty']:,.0f}
💵 Total Material Cost: {self.debug_data['processing_summary']['total_material_cost']:,.2f} บาท
💰 Grand Total: {self.debug_data['processing_summary']['grand_total']:,.2f} บาท
🏪 Base: {selected_base.upper()}

✅ ไฟล์พร้อมใช้งาน"""

            self.debug_log(f"✅ Export สำเร็จ: {filename}")
            self.debug_log(f"📊 Sheets: {sheets_created}")
            self.debug_log(f"📊 Total Qty: {self.debug_data['processing_summary']['total_qty']:,.0f}")
            self.debug_log(
                f"💵 Total Material Cost: {self.debug_data['processing_summary']['total_material_cost']:,.2f} บาท")
            self.debug_log(f"💰 Grand Total: {self.debug_data['processing_summary']['grand_total']:,.2f} บาท")

            messagebox.showinfo("Export สำเร็จ!", summary_msg)

        except Exception as e:
            error_msg = f"Error Export ไฟล์: {str(e)}"
            self.debug_log(error_msg, "ERROR")
            self.debug_log(f"Traceback: {traceback.format_exc()}", "ERROR")
            messagebox.showerror("Error", error_msg)


def main():
    """ฟังก์ชันหลัก"""
    try:
        root = tk.Tk()
        app = EnhancedCostCalculatorApp(root)

        app.debug_log("=== Enhanced Cost Calculator เริ่มทำงาน ===")
        app.debug_log("✅ ข้อมูลจริง Hamada (87 เมนู) และ Hashira (68 เมนู) พร้อมใช้งาน")
        app.debug_log("💾 Export ไฟล์เดียว 4 Sheets: Calculation results | Base Cost | Import (By User) | Raw file(POS)")
        app.debug_log("📊 Sheet 1 มีผลรวม Qty, Material Cost, Total Cost แล้ว")
        app.debug_log("📂 ไฟล์จะถูกบันทึกในโฟลเดอร์เดียวกับโปรแกรม")

        root.mainloop()

    except Exception as e:
        logger.error(f"Critical error: {str(e)}")
        messagebox.showerror("Error", f"เกิดข้อผิดพลาด: {str(e)}")


if __name__ == "__main__":
    main()